from fastapi import FastAPI, APIRouter, HTTPException, File, UploadFile, Form, Cookie, Response, Request
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import json
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
# import google.generativeai as genai
import aiofiles
import base64

import random
import requests

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

GOOGLE_GEMINI_API_KEY = os.environ.get('GOOGLE_GEMINI_API_KEY', '')

GEMINI_MODEL = "gemini-2.5-flash"
GEMINI_FALLBACK_MODEL = "gemini-2.0-flash"
gemini_client = None

FREETTS_URL = os.environ.get('FREETTS_URL', 'https://api.freetts.org')
FREE_TTS_VOICE = os.environ.get('FREE_TTS_VOICE', 'pt-BR-FranciscaNeural')

EIDOS_URL = os.environ.get('EIDOS_URL', 'https://eidosspeech.xyz/api/v1/tts')
EIDOS_API_KEY = os.environ.get('EIDOS_API_KEY', '')

async def get_user_api_key(user_id: str) -> Optional[str]:
    try:
        user_doc = await db.users.find_one({"user_id": user_id}, {"gemini_api_key": 1})
        if user_doc:
            key = user_doc.get("gemini_api_key")
            logging.info(f"Found API key for user {user_id}: {'Yes' if key else 'No'}")
            return key
        logging.warning(f"User {user_id} not found")
    except Exception as e:
        logging.error(f"Error fetching user API key: {e}")
    return None

async def get_freellm_api_key(user_id: str) -> Optional[str]:
    return None

# ========== LLM CALLS ==========

async def call_llm(prompt: str, session_id: str = "default", system_message: str = "Você é um assistente útil.", user_id: Optional[str] = None) -> str:
    """Call Gemini API - user must configure their own API key"""
    
    if not user_id:
        return "⚠️ Serviço de IA indisponível. Faça login e configure sua chave Gemini no perfil."
    
    user_api_key = await get_user_api_key(user_id)
    if not user_api_key:
        return "⚠️ Configure sua chave de API Gemini nas configurações do perfil para usar IA."
    
    result, error_type = await call_gemini(prompt, system_message, user_api_key)
    if result:
        return result
    
    if error_type == "quota":
        return "⚠️ Sua cota da API Gemini esgotou. Acesse https://makersuite.google.com/app/apikey para verificar seu plano ou crie uma nova chave."
    if error_type == "invalid":
        return "⚠️ Sua chave de API Gemini é inválida. Verifique em https://makersuite.google.com/app/apikey"
    return "⚠️ Erro ao contactar API Gemini. Verifique se sua chave é válida em https://makersuite.google.com/app/apikey"

async def call_gemini(prompt: str, system_message: str, api_key: str) -> tuple[Optional[str], Optional[str]]:
    """Call Gemini API, returns (response_text, error_type).
    error_type: None on success, 'quota' on 429, 'invalid' on 400/401/403, 'other' otherwise."""
    from urllib.parse import quote
    models_to_try = ["gemini-2.5-flash", "gemini-2.0-flash", "gemini-2.0-flash-lite"]
    if GEMINI_MODEL not in models_to_try:
        models_to_try.insert(0, GEMINI_MODEL)
    
    last_error = None
    import time
    for model in models_to_try:
        timeout = 90 if "2.5" in model else 30
        url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={quote(api_key)}"
        payload = {"contents": [{"parts": [{"text": prompt}]}]}
        if system_message:
            payload["systemInstruction"] = {"parts": [{"text": system_message}]}
        
        try:
            resp = requests.post(url, json=payload, timeout=timeout)
            logging.info(f"Gemini call (model={model} timeout={timeout}s): status={resp.status_code}")
            if resp.status_code == 200:
                data = resp.json()
                text = data.get("candidates", [{}])[0].get("content", {}).get("parts", [{}])[0].get("text", "")
                if text:
                    return text, None
                logging.warning(f"Gemini 200 OK but no text in response (model={model})")
            else:
                logging.error(f"Gemini API error (model={model}): {resp.status_code} - {resp.text[:500]}")
            if resp.status_code in (400, 401, 403):
                return None, "invalid"
            if resp.status_code == 429:
                last_error = "quota"
            elif resp.status_code != 200:
                last_error = "other"
        except Exception as e:
            logging.error(f"Gemini error (model={model}): {e}")
            last_error = "other"
    return None, last_error



app = FastAPI()
api_router = APIRouter(prefix="/api")

# ========== FREE TTS ==========

async def call_free_tts(text: str) -> Optional[str]:
    """Try FreeTTS first (might need specific voice format)"""
    try:
        payload = {
            "text": text,
            "voice": FREE_TTS_VOICE,
            "rate": "+0%",
            "pitch": "+0Hz"
        }
        
        resp = requests.post(f"{FREETTS_URL}/tts", json=payload, timeout=30)
        
        if resp.status_code == 200:
            result = resp.json()
            file_id = result.get("file_id")
            if file_id:
                return f"{FREETTS_URL}/download/{file_id}"
    except Exception as e:
        logging.error(f"FreeTTS error: {e}")
    return None

async def call_eidos_tts(text: str) -> Optional[str]:
    """Try eidosSpeech (Edge TTS) - needs API key"""
    if not EIDOS_API_KEY:
        return None
    
    try:
        payload = {
            "text": text,
            "voice": "pt-BR-FranciscaNeural"
        }
        headers = {
            "X-API-Key": EIDOS_API_KEY,
            "Content-Type": "application/json"
        }
        
        resp = requests.post(EIDOS_URL, json=payload, headers=headers, timeout=30)
        
        if resp.status_code == 200:
            # Returns audio directly
            return resp.content
    except Exception as e:
        logging.error(f"eidosSpeech error: {e}")
    return None

@api_router.post("/tts")
async def text_to_speech(request: Request, data: dict, session_token: Optional[str] = Cookie(None)):
    """Text to speech - tries multiple free providers"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    text = data.get("text", "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="Texto vazio")
    
    text = text[:1000]
    
    # Try FreeTTS first
    audio_url = await call_free_tts(text)
    if audio_url:
        return {"audio_url": audio_url}
    
    # Try eidosSpeech if configured
    audio_data = await call_eidos_tts(text)
    if audio_data:
        # Return base64 audio
        import base64
        b64 = base64.b64encode(audio_data).decode()
        return {"audio_data": b64, "format": "mp3"}
    
    return None

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    user_id: str
    email: str
    name: str
    picture: Optional[str] = None
    xp: int = 0
    rank: str = "Recruta"
    birth_date: Optional[str] = None
    bio: Optional[str] = None
    gemini_api_key: Optional[str] = None
    created_at: datetime

class UserCreate(BaseModel):
    email: str
    password: str
    name: str
    gemini_api_key: Optional[str] = None

class UserLogin(BaseModel):
    email: str
    password: str

class SessionData(BaseModel):
    user_id: str
    session_token: str
    expires_at: datetime
    created_at: datetime

class Task(BaseModel):
    model_config = ConfigDict(extra="ignore")
    task_id: str
    user_id: str
    title: str
    description: Optional[str] = None
    completed: bool = False
    date: str
    priority: str = "medium"
    xp_reward: int = 10
    recurrence: str = "once"  # once, daily, weekly, monthly
    is_template: bool = True
    created_at: datetime

class TaskCreate(BaseModel):
    title: str
    description: Optional[str] = None
    date: str
    priority: str = "medium"
    recurrence: str = "once"  # once, daily, weekly, monthly

class Habit(BaseModel):
    model_config = ConfigDict(extra="ignore")
    habit_id: str
    user_id: str
    name: str
    description: Optional[str] = None
    color: str
    streak: int = 0
    best_streak: int = 0
    completions: List[str] = []
    created_at: datetime

class HabitCreate(BaseModel):
    name: str
    description: Optional[str] = None
    color: str = "#007AFF"

class Transaction(BaseModel):
    model_config = ConfigDict(extra="ignore")
    transaction_id: str
    user_id: str
    type: str
    amount: float
    category: str
    description: Optional[str] = None
    date: str
    created_at: datetime

class TransactionCreate(BaseModel):
    type: str
    amount: float
    category: str
    description: Optional[str] = None
    date: str

class Budget(BaseModel):
    model_config = ConfigDict(extra="ignore")
    budget_id: str
    user_id: str
    category: str
    limit: float
    spent: float = 0
    month: str
    created_at: datetime

class BudgetCreate(BaseModel):
    category: str
    limit: float
    month: str

class Goal(BaseModel):
    model_config = ConfigDict(extra="ignore")
    goal_id: str
    user_id: str
    title: str
    description: Optional[str] = None
    target_date: str
    progress: float = 0
    sprint_duration: int = 60
    daily_checks: List[str] = []
    sprints: List[Dict[str, Any]] = []
    created_at: datetime

class GoalCreate(BaseModel):
    title: str
    description: Optional[str] = None
    target_date: str
    sprint_duration: int = 60

class Challenge(BaseModel):
    model_config = ConfigDict(extra="ignore")
    challenge_id: str
    title: str
    description: str
    xp_reward: int
    week_start: str
    week_end: str
    completed_by: List[str] = []
    created_at: datetime

class Achievement(BaseModel):
    model_config = ConfigDict(extra="ignore")
    achievement_id: str
    user_id: str
    title: str
    description: str
    icon: str
    unlocked_at: datetime

class ChatMessage(BaseModel):
    model_config = ConfigDict(extra="ignore")
    message_id: str
    user_id: str
    role: str
    content: str
    transaction_data: Optional[Dict[str, Any]] = None
    created_at: datetime

class Report(BaseModel):
    model_config = ConfigDict(extra="ignore")
    report_id: str
    user_id: str
    type: str
    period: str
    data: Dict[str, Any]
    insights: str
    created_at: datetime

# ========== WORKOUT MODELS ==========
class WorkoutPlan(BaseModel):
    model_config = ConfigDict(extra="ignore")
    plan_id: str
    user_id: str
    name: str
    description: Optional[str] = None
    exercises: List[Dict[str, Any]] = []  # [{name, sets, reps, weight, notes, tutorial, video_url, muscle_group, rest_seconds}]
    plan_duration: str = "dia"  # dia, semana, mes, ciclo
    generated_by_ai: bool = False
    days: Optional[List[Dict[str, Any]]] = None  # For multi-day plans: [{day_name, day_label, exercises}]
    objective: Optional[str] = None
    level: Optional[str] = None
    created_at: datetime

class WorkoutPlanCreate(BaseModel):
    name: str
    description: Optional[str] = None
    exercises: List[Dict[str, Any]] = []
    plan_duration: str = "dia"
    days: Optional[List[Dict[str, Any]]] = None

class WorkoutPlanGenerate(BaseModel):
    objective: str  # hipertrofia, emagrecimento, condicionamento, forca, flexibilidade
    level: str  # iniciante, intermediario, avancado
    muscle_groups: Optional[List[str]] = None  # peito, costas, pernas, ombros, biceps, triceps, abdomen, gluteos, trapezio, antebraco, panturrilha
    duration: str = "dia"  # dia, semana, mes, ciclo
    # New fields for split-based generation
    generation_mode: str = "periodo"  # "periodo" or "tipo_treino"
    split_type: Optional[str] = None  # "AB", "ABC", "ABCD", "ABCDE"
    split_config: Optional[List[Dict[str, Any]]] = None  # [{label: "A", name: "Peito e Tríceps", muscle_groups: ["peito", "triceps"]}]
    training_days_per_week: Optional[int] = None  # 2-7
    cycle_weeks: Optional[int] = None  # 1-12
    include_cardio: bool = False
    cardio_type: Optional[str] = None  # "corrida", "bike", "HIIT", "caminhada", "natacao", "pular_corda"
    cardio_mode: Optional[str] = None  # "hibrido", "hibrido_alternado"
    health_condition: Optional[str] = None  # user health conditions/injuries to consider

class WorkoutSession(BaseModel):
    model_config = ConfigDict(extra="ignore")
    session_id: str
    user_id: str
    plan_id: str
    plan_name: str
    status: str = "active"  # active, completed, abandoned
    started_at: str
    completed_at: Optional[str] = None
    total_duration_seconds: int = 0
    exercises: List[Dict[str, Any]] = []  # [{name, sets, reps, weight, completed, time_spent_seconds, sets_completed}]
    current_exercise_idx: int = 0
    rest_timer_seconds: int = 60
    feedback: Optional[Dict[str, Any]] = None  # {difficulty: 1-5, feeling: str, notes: str}
    day_index: Optional[int] = None  # For multi-day plans

class WorkoutLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    log_id: str
    user_id: str
    plan_id: Optional[str] = None
    activity_type: str  # running, weightlifting, cycling, swimming, etc.
    name: str
    duration_minutes: int = 0
    distance_km: Optional[float] = None
    calories: Optional[int] = None
    exercises_completed: List[Dict[str, Any]] = []
    notes: Optional[str] = None
    xp_earned: int = 20
    completed: bool = True
    date: str
    created_at: datetime

class WorkoutLogCreate(BaseModel):
    plan_id: Optional[str] = None
    activity_type: str
    name: str
    duration_minutes: int = 0
    distance_km: Optional[float] = None
    calories: Optional[int] = None
    exercises_completed: List[Dict[str, Any]] = []
    notes: Optional[str] = None
    date: str

# ========== NOTIFICATION MODELS ==========
class Notification(BaseModel):
    model_config = ConfigDict(extra="ignore")
    notification_id: str
    user_id: str
    title: str
    message: str
    type: str  # reminder, achievement, alert, system
    category: str  # workout, habit, task, hydration, custom
    scheduled_time: Optional[str] = None
    repeat: str = "none"  # none, daily, weekly, custom
    repeat_days: List[str] = []  # ["monday", "tuesday", etc.]
    enabled: bool = True
    channels: List[str] = ["in_app"]  # in_app, browser, email, whatsapp, telegram
    last_sent: Optional[str] = None
    created_at: datetime

class NotificationCreate(BaseModel):
    title: str
    message: str
    type: str = "reminder"
    category: str = "custom"
    scheduled_time: Optional[str] = None
    repeat: str = "none"
    repeat_days: List[str] = []
    channels: List[str] = ["in_app"]

class NotificationLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    log_id: str
    notification_id: str
    user_id: str
    sent_at: datetime
    channel: str
    status: str  # sent, read, dismissed

# ========== BODY MEASUREMENT MODELS ==========
class BodyMeasurement(BaseModel):
    model_config = ConfigDict(extra="ignore")
    measurement_id: str
    user_id: str
    date: str
    # Peso e composição corporal
    weight_kg: Optional[float] = None
    body_fat_percentage: Optional[float] = None
    muscle_mass_kg: Optional[float] = None
    bone_mass_kg: Optional[float] = None
    water_percentage: Optional[float] = None
    visceral_fat: Optional[int] = None
    metabolic_age: Optional[int] = None
    bmr_kcal: Optional[int] = None  # Taxa metabólica basal
    # Medidas corporais (cm)
    height_cm: Optional[float] = None
    neck_cm: Optional[float] = None
    shoulders_cm: Optional[float] = None
    chest_cm: Optional[float] = None
    waist_cm: Optional[float] = None
    abdomen_cm: Optional[float] = None
    hips_cm: Optional[float] = None
    left_arm_cm: Optional[float] = None
    right_arm_cm: Optional[float] = None
    left_forearm_cm: Optional[float] = None
    right_forearm_cm: Optional[float] = None
    left_thigh_cm: Optional[float] = None
    right_thigh_cm: Optional[float] = None
    left_calf_cm: Optional[float] = None
    right_calf_cm: Optional[float] = None
    # Calculados
    bmi: Optional[float] = None  # IMC
    # Notas e observações
    notes: Optional[str] = None
    source: str = "manual"  # manual, pdf_import, bioimpedance
    created_at: datetime

class BodyMeasurementCreate(BaseModel):
    date: str
    weight_kg: Optional[float] = None
    body_fat_percentage: Optional[float] = None
    muscle_mass_kg: Optional[float] = None
    bone_mass_kg: Optional[float] = None
    water_percentage: Optional[float] = None
    visceral_fat: Optional[int] = None
    metabolic_age: Optional[int] = None
    bmr_kcal: Optional[int] = None
    height_cm: Optional[float] = None
    neck_cm: Optional[float] = None
    shoulders_cm: Optional[float] = None
    chest_cm: Optional[float] = None
    waist_cm: Optional[float] = None
    abdomen_cm: Optional[float] = None
    hips_cm: Optional[float] = None
    left_arm_cm: Optional[float] = None
    right_arm_cm: Optional[float] = None
    left_forearm_cm: Optional[float] = None
    right_forearm_cm: Optional[float] = None
    left_thigh_cm: Optional[float] = None
    right_thigh_cm: Optional[float] = None
    left_calf_cm: Optional[float] = None
    right_calf_cm: Optional[float] = None
    notes: Optional[str] = None
    source: str = "manual"

# ========== DAILY WORKOUT TRACKING ==========
class DailyWorkoutStatus(BaseModel):
    model_config = ConfigDict(extra="ignore")
    status_id: str
    user_id: str
    plan_id: str
    date: str
    exercises_status: Dict[int, bool] = {}  # {exercise_index: completed}
    completed: bool = False
    created_at: datetime
    updated_at: datetime

async def get_current_user(authorization: Optional[str] = None, session_token: Optional[str] = Cookie(None)) -> User:
    token = session_token or (authorization.replace("Bearer ", "") if authorization else None)
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    session = await db.user_sessions.find_one({"session_token": token}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    expires_at = session["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="Session expired")
    
    user_doc = await db.users.find_one({"user_id": session["user_id"]}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Ensure all required fields have defaults
    user_doc.setdefault("name", "Usuário")
    user_doc.setdefault("xp", 0)
    user_doc.setdefault("rank", "Recruta")
    user_doc.setdefault("picture", None)
    user_doc.setdefault("birth_date", None)
    user_doc.setdefault("bio", None)
    user_doc.setdefault("gemini_api_key", None)
    
    if isinstance(user_doc['created_at'], str):
        user_doc['created_at'] = datetime.fromisoformat(user_doc['created_at'])
    
    return User(**user_doc)

@api_router.get("/")
async def root():
    return {"message": "Sirius API - Discipline is Destiny"}

class SyncPayload(BaseModel):
    record_id: str
    operation: str
    data: Dict[str, Any]
    timestamp: str

@api_router.post("/sync/{table_name}")
async def sync_table(
    table_name: str,
    payload: SyncPayload,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Sync data from mobile app"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    collection = db[table_name]
    record_id = payload.record_id
    data = payload.data
    
    if payload.operation == "DELETE":
        await collection.delete_one({f"{table_name[:-1]}_id": record_id})
    elif payload.operation == "INSERT":
        data["synced_at"] = datetime.now(timezone.utc).isoformat()
        await collection.update_one(
            {f"{table_name[:-1]}_id": record_id},
            {"$set": data},
            upsert=True
        )
    elif payload.operation == "UPDATE":
        data["updated_at"] = datetime.now(timezone.utc).isoformat()
        data["synced_at"] = datetime.now(timezone.utc).isoformat()
        await collection.update_one(
            {f"{table_name[:-1]}_id": record_id},
            {"$set": data},
            upsert=True
        )
    
    return {"success": True}

@api_router.get("/sync/{table_name}/{user_id}")
async def get_sync_data(
    table_name: str,
    user_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Get data for sync to mobile app"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if user.user_id != user_id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    collection = db[table_name]
    user_field = "user_id"
    
    records = await collection.find(
        {user_field: user_id},
        {"_id": 0}
    ).to_list(1000)
    
    return records

@api_router.post("/auth/register")
async def register(user_data: UserCreate, response: Response):
    existing = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    hashed_password = bcrypt.hashpw(user_data.password.encode('utf-8'), bcrypt.gensalt())
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    
    user_doc = {
        "user_id": user_id,
        "email": user_data.email,
        "name": user_data.name,
        "password": hashed_password.decode('utf-8'),
        "picture": None,
        "xp": 0,
        "rank": "Recruta",
        "gemini_api_key": user_data.gemini_api_key,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    session_token = f"session_{uuid.uuid4().hex}"
    session_doc = {
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.user_sessions.insert_one(session_doc)
    
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=7*24*60*60
    )
    
    return {"session_token": session_token, "user": {"user_id": user_id, "email": user_data.email, "name": user_data.name}}

@api_router.post("/auth/login")
async def login(credentials: UserLogin, response: Response):
    user_doc = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not bcrypt.checkpw(credentials.password.encode('utf-8'), user_doc['password'].encode('utf-8')):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    session_token = f"session_{uuid.uuid4().hex}"
    session_doc = {
        "user_id": user_doc["user_id"],
        "session_token": session_token,
        "expires_at": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.user_sessions.insert_one(session_doc)
    
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=7*24*60*60
    )
    
    return {"session_token": session_token, "user": {"user_id": user_doc["user_id"], "email": user_doc["email"], "name": user_doc["name"]}}

@api_router.get("/auth/google-session")
async def process_google_session(session_id: str, response: Response):
    try:
        headers = {"X-Session-ID": session_id}
        res = requests.get("https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data", headers=headers)
        
        if res.status_code != 200:
            raise HTTPException(status_code=400, detail="Invalid session ID")
        
        data = res.json()
        user_doc = await db.users.find_one({"email": data["email"]}, {"_id": 0})
        
        if user_doc:
            user_id = user_doc["user_id"]
        else:
            user_id = f"user_{uuid.uuid4().hex[:12]}"
            user_doc = {
                "user_id": user_id,
                "email": data["email"],
                "name": data["name"],
                "picture": data.get("picture"),
                "xp": 0,
                "rank": "Recruta",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.users.insert_one(user_doc)
        
        session_token = data["session_token"]
        session_doc = {
            "user_id": user_id,
            "session_token": session_token,
            "expires_at": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.user_sessions.insert_one(session_doc)
        
        response.set_cookie(
            key="session_token",
            value=session_token,
            httponly=True,
            secure=True,
            samesite="none",
            path="/",
            max_age=7*24*60*60
        )
        
        return {"session_token": session_token, "user": {"user_id": user_id, "email": data["email"], "name": data["name"], "picture": data.get("picture")}}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/auth/me")
async def get_me(request: Request, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    return user

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    token = session_token or (auth_header.replace("Bearer ", "") if auth_header else None)
    
    if token:
        await db.user_sessions.delete_one({"session_token": token})
    
    response.delete_cookie("session_token", path="/")
    return {"message": "Logged out"}

@api_router.post("/auth/upload-picture")
async def upload_profile_picture(
    request: Request,
    file: UploadFile = File(...),
    session_token: Optional[str] = Cookie(None)
):
    """Upload a profile picture for the user"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Validate file type
    allowed_types = ["image/jpeg", "image/png", "image/gif", "image/webp"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Only JPEG, PNG, GIF or WebP images are allowed")
    
    # Read and encode file
    content = await file.read()
    
    # Check file size (max 5MB)
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File size must be less than 5MB")
    
    # Store as base64 data URL
    file_base64 = base64.b64encode(content).decode('utf-8')
    data_url = f"data:{file.content_type};base64,{file_base64}"
    
    # Update user picture
    await db.users.update_one(
        {"user_id": user.user_id},
        {"$set": {"picture": data_url}}
    )
    
    return {"message": "Profile picture updated", "picture": data_url}

@api_router.delete("/auth/remove-picture")
async def remove_profile_picture(request: Request, session_token: Optional[str] = Cookie(None)):
    """Remove the user's profile picture"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    await db.users.update_one(
        {"user_id": user.user_id},
        {"$set": {"picture": None}}
    )
    
    return {"message": "Profile picture removed"}

@api_router.patch("/auth/profile")
async def update_profile(request: Request, data: dict, session_token: Optional[str] = Cookie(None)):
    """Update user profile info (name, birth_date, bio)"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    update_fields = {}
    for field in ["name", "birth_date", "bio", "health_condition", "gemini_api_key"]:
        if field in data:
            update_fields[field] = data[field]
    
    if not update_fields:
        raise HTTPException(status_code=400, detail="Nenhum campo para atualizar")
    
    await db.users.update_one(
        {"user_id": user.user_id},
        {"$set": update_fields}
    )
    
    updated_user = await db.users.find_one({"user_id": user.user_id}, {"_id": 0, "password": 0})
    updated_user = updated_user or {}
    updated_user.setdefault("gemini_api_key", None)
    return updated_user

@api_router.get("/auth/birthday-check")
async def check_birthday(request: Request, session_token: Optional[str] = Cookie(None)):
    """Check if today is user's birthday"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    user_doc = await db.users.find_one({"user_id": user.user_id}, {"_id": 0})
    birth_date = user_doc.get("birth_date")
    
    if not birth_date:
        return {"is_birthday": False, "age": None}
    
    try:
        bd = datetime.strptime(birth_date, "%Y-%m-%d")
        today = datetime.now(timezone.utc)
        is_birthday = bd.month == today.month and bd.day == today.day
        age = today.year - bd.year - ((today.month, today.day) < (bd.month, bd.day))
        return {"is_birthday": is_birthday, "age": age, "birth_date": birth_date}
    except Exception:
        return {"is_birthday": False, "age": None}

@api_router.post("/study/notebooks/{notebook_id}/topic-progress")
async def update_topic_progress(request: Request, notebook_id: str, data: dict, session_token: Optional[str] = Cookie(None)):
    """Mark a topic/subtopic as studied, reviewed, etc."""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    topic_key = data.get("topic_key", "")  # e.g. "0" for assunto index or "0_1" for subtopic
    status = data.get("status", "studied")  # studied, reviewed, mastered
    checked = data.get("checked", True)
    
    if not topic_key:
        raise HTTPException(status_code=400, detail="topic_key é obrigatório")
    
    progress_id = f"tp_{notebook_id}_{user.user_id}"
    progress_doc = await db.topic_progress.find_one({"progress_id": progress_id}, {"_id": 0})
    
    if not progress_doc:
        progress_doc = {
            "progress_id": progress_id,
            "notebook_id": notebook_id,
            "user_id": user.user_id,
            "topics": {},
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.topic_progress.insert_one(progress_doc)
    
    field_key = f"topics.{topic_key}.{status}"
    if checked:
        await db.topic_progress.update_one(
            {"progress_id": progress_id},
            {"$set": {field_key: True, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    else:
        await db.topic_progress.update_one(
            {"progress_id": progress_id},
            {"$unset": {field_key: ""}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    
    updated = await db.topic_progress.find_one({"progress_id": progress_id}, {"_id": 0})
    return updated

@api_router.get("/study/notebooks/{notebook_id}/topic-progress")
async def get_topic_progress(request: Request, notebook_id: str, session_token: Optional[str] = Cookie(None)):
    """Get topic progress for a notebook"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    progress_doc = await db.topic_progress.find_one(
        {"notebook_id": notebook_id, "user_id": user.user_id}, {"_id": 0}
    )
    
    return progress_doc or {"topics": {}}

@api_router.get("/tasks")
async def get_tasks(request: Request, date: Optional[str] = None, recurrence: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if not date:
        date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    query = {"user_id": user.user_id}
    if recurrence:
        query["recurrence"] = recurrence
    
    all_tasks = await db.tasks.find({"user_id": user.user_id, "is_template": True}, {"_id": 0}).to_list(1000)
    
    result_tasks = []
    for task in all_tasks:
        if recurrence and task.get('recurrence') != recurrence:
            continue
            
        instance = await db.task_instances.find_one({
            "task_id": task["task_id"],
            "date": date
        }, {"_id": 0})
        
        task_copy = task.copy()
        task_copy["date"] = date
        task_copy["completed"] = instance["completed"] if instance else False
        task_copy["instance_id"] = instance["instance_id"] if instance else None
        
        if isinstance(task_copy['created_at'], str):
            task_copy['created_at'] = datetime.fromisoformat(task_copy['created_at'])
        
        result_tasks.append(task_copy)
    
    return result_tasks

@api_router.post("/tasks")
async def create_task(request: Request, task_data: TaskCreate, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    task_id = f"task_{uuid.uuid4().hex[:12]}"
    task_doc = {
        "task_id": task_id,
        "user_id": user.user_id,
        "title": task_data.title,
        "description": task_data.description,
        "priority": task_data.priority,
        "xp_reward": 5 if task_data.priority == "low" else 10 if task_data.priority == "medium" else 15,
        "recurrence": task_data.recurrence,
        "is_template": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.tasks.insert_one(task_doc)
    task_doc.pop('_id', None)  # Remove MongoDB ObjectId
    task_doc['created_at'] = datetime.fromisoformat(task_doc['created_at'])
    task_doc['date'] = task_data.date
    task_doc['completed'] = False
    task_doc['instance_id'] = None
    return Task(**task_doc)

@api_router.patch("/tasks/{task_id}")
async def update_task(request: Request, task_id: str, completed: bool, date: str, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    task = await db.tasks.find_one({"task_id": task_id, "user_id": user.user_id}, {"_id": 0})
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    instance = await db.task_instances.find_one({"task_id": task_id, "date": date}, {"_id": 0})
    
    if not instance:
        instance_id = f"inst_{uuid.uuid4().hex[:12]}"
        instance_doc = {
            "instance_id": instance_id,
            "task_id": task_id,
            "user_id": user.user_id,
            "date": date,
            "completed": completed,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.task_instances.insert_one(instance_doc)
        was_completed = False
    else:
        await db.task_instances.update_one(
            {"instance_id": instance["instance_id"]},
            {"$set": {"completed": completed}}
        )
        was_completed = instance["completed"]
    
    # Completing task - award XP
    if completed and not was_completed:
        new_xp = user.xp + task['xp_reward']
        new_rank = calculate_rank(new_xp)
        await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
        return {"message": "Task completed", "xp_earned": task['xp_reward'], "new_xp": new_xp, "new_rank": new_rank}
    
    # Uncompleting task - deduct XP
    if not completed and was_completed:
        new_xp = max(0, user.xp - task['xp_reward'])
        new_rank = calculate_rank(new_xp)
        await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
        return {"message": "Task uncompleted", "xp_earned": -task['xp_reward'], "new_xp": new_xp, "new_rank": new_rank}
    
    return {"message": "Task updated"}

@api_router.delete("/tasks/{task_id}")
async def delete_task(request: Request, task_id: str, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.tasks.delete_one({"task_id": task_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Task not found")
    return {"message": "Task deleted"}


@api_router.patch("/tasks/{task_id}/status")
async def update_task_status(request: Request, task_id: str, session_token: Optional[str] = Cookie(None)):
    """Update task kanban status (todo, in_progress, done)"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    body = await request.json()
    new_status = body.get("status", "todo")
    date = body.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d"))
    
    if new_status not in ("todo", "in_progress", "done"):
        raise HTTPException(status_code=400, detail="Status must be todo, in_progress, or done")
    
    task = await db.tasks.find_one({"task_id": task_id, "user_id": user.user_id}, {"_id": 0})
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    completed = new_status == "done"
    instance = await db.task_instances.find_one({"task_id": task_id, "date": date}, {"_id": 0})
    
    if not instance:
        instance_id = f"inst_{uuid.uuid4().hex[:12]}"
        instance_doc = {
            "instance_id": instance_id,
            "task_id": task_id,
            "user_id": user.user_id,
            "date": date,
            "completed": completed,
            "status": new_status,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.task_instances.insert_one(instance_doc)
        was_completed = False
    else:
        was_completed = instance.get("completed", False)
        await db.task_instances.update_one(
            {"instance_id": instance["instance_id"]},
            {"$set": {"completed": completed, "status": new_status}}
        )
    
    xp_earned = 0
    new_xp = user.xp
    new_rank = user.rank
    
    if completed and not was_completed:
        new_xp = user.xp + task['xp_reward']
        new_rank = calculate_rank(new_xp)
        await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
        xp_earned = task['xp_reward']
    elif not completed and was_completed:
        new_xp = max(0, user.xp - task['xp_reward'])
        new_rank = calculate_rank(new_xp)
        await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
        xp_earned = -task['xp_reward']
    
    return {"message": "Status updated", "status": new_status, "xp_earned": xp_earned, "new_xp": new_xp, "new_rank": new_rank}



@api_router.get("/habits")
async def get_habits(request: Request, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    habits = await db.habits.find({"user_id": user.user_id}, {"_id": 0}).to_list(1000)
    for habit in habits:
        if isinstance(habit['created_at'], str):
            habit['created_at'] = datetime.fromisoformat(habit['created_at'])
    return habits

@api_router.post("/habits")
async def create_habit(request: Request, habit_data: HabitCreate, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    habit_id = f"habit_{uuid.uuid4().hex[:12]}"
    habit_doc = {
        "habit_id": habit_id,
        "user_id": user.user_id,
        "name": habit_data.name,
        "description": habit_data.description,
        "color": habit_data.color,
        "streak": 0,
        "best_streak": 0,
        "completions": [],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.habits.insert_one(habit_doc)
    habit_doc.pop('_id', None)  # Remove MongoDB ObjectId
    habit_doc['created_at'] = datetime.fromisoformat(habit_doc['created_at'])
    return Habit(**habit_doc)

@api_router.post("/habits/{habit_id}/complete")
async def complete_habit(request: Request, habit_id: str, date: str, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    habit = await db.habits.find_one({"habit_id": habit_id, "user_id": user.user_id}, {"_id": 0})
    if not habit:
        raise HTTPException(status_code=404, detail="Habit not found")
    
    # Toggle: if already completed, uncomplete it
    if date in habit['completions']:
        # Uncomplete - remove date and deduct XP
        completions = [d for d in habit['completions'] if d != date]
        completions.sort()
        
        streak = calculate_streak(completions)
        # Recalculate best_streak from all completions
        best_streak = calculate_best_streak(completions)
        
        await db.habits.update_one(
            {"habit_id": habit_id},
            {"$set": {"completions": completions, "streak": streak, "best_streak": best_streak}}
        )
        
        # Deduct XP
        new_xp = max(0, user.xp - 8)
        new_rank = calculate_rank(new_xp)
        await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
        
        return {"message": "Habit uncompleted", "streak": streak, "best_streak": best_streak, "xp_earned": -8, "new_xp": new_xp, "uncompleted": True}
    
    # Complete - add date and award XP
    completions = habit['completions'] + [date]
    completions.sort()
    
    streak = calculate_streak(completions)
    best_streak = max(habit.get('best_streak', 0), streak)
    
    await db.habits.update_one(
        {"habit_id": habit_id},
        {"$set": {"completions": completions, "streak": streak, "best_streak": best_streak}}
    )
    
    new_xp = user.xp + 8
    new_rank = calculate_rank(new_xp)
    await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
    
    return {"message": "Habit completed", "streak": streak, "best_streak": best_streak, "xp_earned": 8, "new_xp": new_xp, "uncompleted": False}

@api_router.delete("/habits/{habit_id}")
async def delete_habit(request: Request, habit_id: str, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.habits.delete_one({"habit_id": habit_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Habit not found")
    return {"message": "Habit deleted"}

@api_router.get("/transactions")
async def get_transactions(request: Request, month: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    query = {"user_id": user.user_id}
    if month:
        query["date"] = {"$regex": f"^{month}"}
    
    transactions = await db.transactions.find(query, {"_id": 0}).to_list(1000)
    for transaction in transactions:
        if isinstance(transaction['created_at'], str):
            transaction['created_at'] = datetime.fromisoformat(transaction['created_at'])
    return transactions

@api_router.post("/transactions")
async def create_transaction(request: Request, transaction_data: TransactionCreate, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    transaction_id = f"trans_{uuid.uuid4().hex[:12]}"
    transaction_doc = {
        "transaction_id": transaction_id,
        "user_id": user.user_id,
        "type": transaction_data.type,
        "amount": transaction_data.amount,
        "category": transaction_data.category,
        "description": transaction_data.description,
        "date": transaction_data.date,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.transactions.insert_one(transaction_doc)
    transaction_doc.pop('_id', None)  # Remove MongoDB ObjectId
    
    if transaction_data.type == "expense":
        month = transaction_data.date[:7]
        budget = await db.budgets.find_one(
            {"user_id": user.user_id, "category": transaction_data.category, "month": month},
            {"_id": 0}
        )
        if budget:
            new_spent = budget['spent'] + transaction_data.amount
            await db.budgets.update_one(
                {"budget_id": budget['budget_id']},
                {"$set": {"spent": new_spent}}
            )
    
    transaction_doc['created_at'] = datetime.fromisoformat(transaction_doc['created_at'])
    return Transaction(**transaction_doc)

@api_router.delete("/transactions/{transaction_id}")
async def delete_transaction(request: Request, transaction_id: str, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Primeiro verificar se a transação existe e pegar seus dados
    transaction = await db.transactions.find_one({"transaction_id": transaction_id, "user_id": user.user_id}, {"_id": 0})
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    # Deletar projeções relacionadas (parcelas futuras)
    await db.projections.delete_many({"source_transaction_id": transaction_id, "user_id": user.user_id})
    
    # Deletar a transação
    await db.transactions.delete_one({"transaction_id": transaction_id, "user_id": user.user_id})
    
    return {"message": "Transaction and related projections deleted"}

@api_router.get("/budgets")
async def get_budgets(request: Request, month: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    query = {"user_id": user.user_id}
    if month:
        query["month"] = month
    
    budgets = await db.budgets.find(query, {"_id": 0}).to_list(1000)
    for budget in budgets:
        if isinstance(budget['created_at'], str):
            budget['created_at'] = datetime.fromisoformat(budget['created_at'])
    return budgets

@api_router.post("/budgets")
async def create_budget(request: Request, budget_data: BudgetCreate, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    existing = await db.budgets.find_one(
        {"user_id": user.user_id, "category": budget_data.category, "month": budget_data.month},
        {"_id": 0}
    )
    if existing:
        raise HTTPException(status_code=400, detail="Budget already exists for this category and month")
    
    budget_id = f"budget_{uuid.uuid4().hex[:12]}"
    budget_doc = {
        "budget_id": budget_id,
        "user_id": user.user_id,
        "category": budget_data.category,
        "limit": budget_data.limit,
        "spent": 0,
        "month": budget_data.month,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.budgets.insert_one(budget_doc)
    budget_doc.pop('_id', None)  # Remove MongoDB ObjectId
    budget_doc['created_at'] = datetime.fromisoformat(budget_doc['created_at'])
    return Budget(**budget_doc)

@api_router.get("/goals")
async def get_goals(request: Request, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    goals = await db.goals.find({"user_id": user.user_id}, {"_id": 0}).to_list(1000)
    for goal in goals:
        if isinstance(goal['created_at'], str):
            goal['created_at'] = datetime.fromisoformat(goal['created_at'])
    return goals

@api_router.post("/goals")
async def create_goal(request: Request, goal_data: GoalCreate, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    goal_id = f"goal_{uuid.uuid4().hex[:12]}"
    goal_doc = {
        "goal_id": goal_id,
        "user_id": user.user_id,
        "title": goal_data.title,
        "description": goal_data.description,
        "target_date": goal_data.target_date,
        "progress": 0,
        "sprint_duration": goal_data.sprint_duration,
        "sprints": [],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.goals.insert_one(goal_doc)
    goal_doc.pop('_id', None)  # Remove MongoDB ObjectId
    goal_doc['created_at'] = datetime.fromisoformat(goal_doc['created_at'])
    return Goal(**goal_doc)

@api_router.patch("/goals/{goal_id}")
async def update_goal(request: Request, goal_id: str, progress: float, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.goals.update_one(
        {"goal_id": goal_id, "user_id": user.user_id},
        {"$set": {"progress": progress}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Goal not found")
    return {"message": "Goal updated"}

@api_router.delete("/goals/{goal_id}")
async def delete_goal(request: Request, goal_id: str, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.goals.delete_one({"goal_id": goal_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Goal not found")
    return {"message": "Goal deleted"}

@api_router.get("/achievements")
async def get_achievements(request: Request, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    achievements = await db.achievements.find({"user_id": user.user_id}, {"_id": 0}).to_list(1000)
    for achievement in achievements:
        if isinstance(achievement.get('unlocked_at'), str):
            achievement['unlocked_at'] = datetime.fromisoformat(achievement['unlocked_at'])
    return achievements


@api_router.get("/achievements/full")
async def get_full_achievements(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get all possible achievements with progress tracking"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Fetch data for progress calculation
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    tasks_completed = await db.task_instances.count_documents({"user_id": user.user_id, "completed": True})
    habits = await db.habits.find({"user_id": user.user_id}, {"_id": 0}).to_list(1000)
    transactions = await db.transactions.find({"user_id": user.user_id}, {"_id": 0}).to_list(5000)
    study_sessions = await db.study_sessions.find({"user_id": user.user_id}, {"_id": 0}).to_list(5000)
    workout_logs = await db.workout_logs.find({"user_id": user.user_id, "completed": True}, {"_id": 0}).to_list(1000)
    flashcards = await db.flashcards.find({"user_id": user.user_id}, {"_id": 0}).to_list(5000)
    meals = await db.meals.find({"user_id": user.user_id}, {"_id": 0}).to_list(5000)
    goals = await db.goals.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    study_streak = await db.study_streaks.find_one({"user_id": user.user_id}, {"_id": 0})
    
    total_study_minutes = sum(s.get("duration_minutes", 0) for s in study_sessions)
    total_workout_minutes = sum(w.get("duration_minutes", 0) for w in workout_logs)
    max_habit_streak = max((h.get("current_streak", 0) for h in habits), default=0)
    longest_study_streak = study_streak.get("longest_streak", 0) if study_streak else 0
    
    # Define all achievements
    all_achievements = [
        # Tasks
        {"id": "task_1", "title": "Primeira Missão", "description": "Complete sua primeira tarefa", "icon": "check", "category": "tasks", "color": "#007AFF", "target": 1, "current": min(tasks_completed, 1)},
        {"id": "task_10", "title": "Executor", "description": "Complete 10 tarefas", "icon": "check-double", "category": "tasks", "color": "#007AFF", "target": 10, "current": min(tasks_completed, 10)},
        {"id": "task_50", "title": "Produtivo", "description": "Complete 50 tarefas", "icon": "list-checks", "category": "tasks", "color": "#007AFF", "target": 50, "current": min(tasks_completed, 50)},
        {"id": "task_200", "title": "Imparável", "description": "Complete 200 tarefas", "icon": "rocket", "category": "tasks", "color": "#007AFF", "target": 200, "current": min(tasks_completed, 200)},
        
        # Habits
        {"id": "habit_create", "title": "Novo Hábito", "description": "Crie seu primeiro hábito", "icon": "trending-up", "category": "habits", "color": "#39FF14", "target": 1, "current": min(len(habits), 1)},
        {"id": "habit_streak_7", "title": "Semana Perfeita", "description": "Mantenha um streak de 7 dias em um hábito", "icon": "flame", "category": "habits", "color": "#39FF14", "target": 7, "current": min(max_habit_streak, 7)},
        {"id": "habit_streak_30", "title": "Mês de Ferro", "description": "Mantenha um streak de 30 dias", "icon": "flame", "category": "habits", "color": "#39FF14", "target": 30, "current": min(max_habit_streak, 30)},
        {"id": "habit_streak_100", "title": "Disciplina Absoluta", "description": "100 dias de streak em um hábito", "icon": "crown", "category": "habits", "color": "#39FF14", "target": 100, "current": min(max_habit_streak, 100)},
        
        # Finance
        {"id": "fin_first", "title": "Primeiro Registro", "description": "Registre sua primeira transação", "icon": "dollar", "category": "finance", "color": "#FF9500", "target": 1, "current": min(len(transactions), 1)},
        {"id": "fin_50", "title": "Controlador", "description": "Registre 50 transações", "icon": "wallet", "category": "finance", "color": "#FF9500", "target": 50, "current": min(len(transactions), 50)},
        {"id": "fin_200", "title": "Mestre das Finanças", "description": "Registre 200 transações", "icon": "bar-chart", "category": "finance", "color": "#FF9500", "target": 200, "current": min(len(transactions), 200)},
        
        # Study
        {"id": "study_first", "title": "Primeira Sessão", "description": "Realize sua primeira sessão de estudo", "icon": "book", "category": "study", "color": "#A855F7", "target": 1, "current": min(len(study_sessions), 1)},
        {"id": "study_hours_10", "title": "Estudioso", "description": "Acumule 10 horas de estudo", "icon": "clock", "category": "study", "color": "#A855F7", "target": 600, "current": min(total_study_minutes, 600)},
        {"id": "study_hours_50", "title": "Acadêmico", "description": "Acumule 50 horas de estudo", "icon": "graduation-cap", "category": "study", "color": "#A855F7", "target": 3000, "current": min(total_study_minutes, 3000)},
        {"id": "study_streak_14", "title": "Foco Total", "description": "14 dias consecutivos de estudo", "icon": "target", "category": "study", "color": "#A855F7", "target": 14, "current": min(longest_study_streak, 14)},
        {"id": "flash_100", "title": "Memorização", "description": "Crie 100 flashcards", "icon": "brain", "category": "study", "color": "#A855F7", "target": 100, "current": min(len(flashcards), 100)},
        
        # Workouts
        {"id": "gym_first", "title": "Primeiro Treino", "description": "Complete seu primeiro treino", "icon": "dumbbell", "category": "workouts", "color": "#EF4444", "target": 1, "current": min(len(workout_logs), 1)},
        {"id": "gym_20", "title": "Atleta", "description": "Complete 20 treinos", "icon": "medal", "category": "workouts", "color": "#EF4444", "target": 20, "current": min(len(workout_logs), 20)},
        {"id": "gym_hours_10", "title": "Forte", "description": "Acumule 10 horas de treino", "icon": "timer", "category": "workouts", "color": "#EF4444", "target": 600, "current": min(total_workout_minutes, 600)},
        
        # Nutrition
        {"id": "meal_first", "title": "Primeira Refeição", "description": "Registre sua primeira refeição", "icon": "utensils", "category": "nutrition", "color": "#22C55E", "target": 1, "current": min(len(meals), 1)},
        {"id": "meal_50", "title": "Alimentação Consciente", "description": "Registre 50 refeições", "icon": "apple", "category": "nutrition", "color": "#22C55E", "target": 50, "current": min(len(meals), 50)},
        
        # Goals
        {"id": "goal_create", "title": "Visionário", "description": "Crie sua primeira meta", "icon": "target", "category": "goals", "color": "#F59E0B", "target": 1, "current": min(len(goals), 1)},
        {"id": "goal_5", "title": "Ambicioso", "description": "Tenha 5 metas ativas", "icon": "trophy", "category": "goals", "color": "#F59E0B", "target": 5, "current": min(len(goals), 5)},
        
        # XP / Rank
        {"id": "xp_100", "title": "Soldado", "description": "Alcance 100 XP", "icon": "zap", "category": "xp", "color": "#FFD700", "target": 100, "current": min(user.xp, 100)},
        {"id": "xp_500", "title": "Veterano", "description": "Alcance 500 XP", "icon": "star", "category": "xp", "color": "#FFD700", "target": 500, "current": min(user.xp, 500)},
        {"id": "xp_1000", "title": "Lenda", "description": "Alcance 1000 XP", "icon": "crown", "category": "xp", "color": "#FFD700", "target": 1000, "current": min(user.xp, 1000)},
        {"id": "xp_3000", "title": "Supremo", "description": "Alcance 3000 XP", "icon": "shield", "category": "xp", "color": "#FFD700", "target": 3000, "current": min(user.xp, 3000)},
    ]
    
    # Unlocked achievements from DB
    unlocked = await db.achievements.find({"user_id": user.user_id}, {"_id": 0}).to_list(1000)
    unlocked_titles = set(a.get("title", "") for a in unlocked)
    
    # Mark unlocked and auto-unlock new ones
    result = []
    newly_unlocked = []
    for ach in all_achievements:
        ach["progress"] = round((ach["current"] / ach["target"]) * 100, 1) if ach["target"] > 0 else 0
        ach["unlocked"] = ach["progress"] >= 100 or ach["title"] in unlocked_titles
        
        # Auto-unlock if progress is 100% but not yet in DB
        if ach["progress"] >= 100 and ach["title"] not in unlocked_titles:
            ach["unlocked"] = True
            newly_unlocked.append(ach)
            await db.achievements.insert_one({
                "achievement_id": f"ach_{uuid.uuid4().hex[:12]}",
                "user_id": user.user_id,
                "title": ach["title"],
                "description": ach["description"],
                "icon": ach["icon"],
                "category": ach["category"],
                "unlocked_at": datetime.now(timezone.utc).isoformat()
            })
        
        result.append(ach)
    
    total_unlocked = len([a for a in result if a["unlocked"]])
    
    return {
        "achievements": result,
        "total": len(result),
        "unlocked": total_unlocked,
        "locked": len(result) - total_unlocked,
        "completion_pct": round((total_unlocked / len(result)) * 100, 1) if result else 0,
        "newly_unlocked": [{"title": a["title"], "description": a["description"]} for a in newly_unlocked]
    }

@api_router.get("/chat/messages")
async def get_chat_messages(request: Request, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    messages = await db.chat_messages.find({"user_id": user.user_id}, {"_id": 0}).sort("created_at", 1).to_list(1000)
    for message in messages:
        if isinstance(message['created_at'], str):
            message['created_at'] = datetime.fromisoformat(message['created_at'])
    return messages

class ChatMessageCreate(BaseModel):
    content: str

@api_router.post("/chat/send")
async def send_chat_message(request: Request, message_data: ChatMessageCreate, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    content = message_data.content.strip()
    content_lower = content.lower()
    
    message_id = f"msg_{uuid.uuid4().hex[:12]}"
    user_message = {
        "message_id": message_id,
        "user_id": user.user_id,
        "role": "user",
        "content": content,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.chat_messages.insert_one(user_message.copy())
    
    try:
        # Get financial context
        current_month = datetime.now(timezone.utc).strftime("%Y-%m")
        transactions = await db.transactions.find(
            {"user_id": user.user_id, "date": {"$regex": f"^{current_month}"}},
            {"_id": 0}
        ).to_list(500)
        
        total_income = sum([t['amount'] for t in transactions if t['type'] == 'income'])
        total_expense = sum([t['amount'] for t in transactions if t['type'] == 'expense'])
        balance = total_income - total_expense
        
        # Get budgets
        budgets = await db.budgets.find({"user_id": user.user_id, "month": current_month}, {"_id": 0}).to_list(100)
        
        # Get categories usage
        expense_by_category = {}
        income_by_category = {}
        for t in transactions:
            if t['type'] == 'expense':
                expense_by_category[t['category']] = expense_by_category.get(t['category'], 0) + t['amount']
            else:
                income_by_category[t['category']] = income_by_category.get(t['category'], 0) + t['amount']
        
        context = f"""Contexto Financeiro do Usuário ({current_month}):
- Receitas Totais: R$ {total_income:.2f}
- Despesas Totais: R$ {total_expense:.2f}
- Saldo Atual: R$ {balance:.2f}
- Total de Transações: {len(transactions)}

Despesas por Categoria: {json.dumps(expense_by_category, ensure_ascii=False)}
Receitas por Categoria: {json.dumps(income_by_category, ensure_ascii=False)}
Orçamentos Definidos: {len(budgets)}"""

        # Detect user intent
        income_keywords = ['ganhei', 'recebi', 'entrou', 'salário', 'salario', 'renda', 'recebimento', 
                          'depósito', 'deposito', 'transferência recebida', 'pix recebido', 'crédito', 
                          'credito', 'freelance', 'bônus', 'bonus', 'comissão', 'comissao', 'vendi',
                          'receita', 'entrada', 'reembolso']
        
        expense_keywords = ['gastei', 'paguei', 'comprei', 'compra', 'gasto', 'despesa', 'conta', 
                           'boleto', 'parcela', 'débito', 'debito', 'saída', 'saida', 'pix enviado',
                           'transferi', 'aluguel', 'luz', 'água', 'agua', 'internet', 'supermercado',
                           'mercado', 'restaurante', 'uber', 'combustível', 'combustivel', 'gasolina']
        
        report_keywords = ['relatório', 'relatorio', 'resumo', 'analise', 'análise', 'como está', 
                          'como estão', 'situação', 'balanço', 'balanco', 'extrato', 'histórico',
                          'quanto gastei', 'quanto ganhei', 'quanto tenho', 'saldo']
        
        budget_keywords = ['orçamento', 'orcamento', 'limite', 'meta de gasto', 'definir limite',
                          'criar orçamento', 'criar orcamento', 'estabelecer limite']
        
        list_keywords = ['listar', 'mostrar', 'ver transações', 'ver gastos', 'ver receitas', 
                        'últimas transações', 'ultimas transacoes']
        
        help_keywords = ['ajuda', 'help', 'o que você pode fazer', 'comandos', 'funcionalidades']
        
        # Check for amount in message
        import re
        amount_pattern = r'(?:R\$\s*)?(\d+(?:[.,]\d{1,2})?)'
        amount_matches = re.findall(amount_pattern, content)
        has_amount = len(amount_matches) > 0
        
        is_income = any(word in content_lower for word in income_keywords)
        is_expense = any(word in content_lower for word in expense_keywords)
        is_report = any(word in content_lower for word in report_keywords)
        is_budget = any(word in content_lower for word in budget_keywords)
        is_list = any(word in content_lower for word in list_keywords)
        is_help = any(word in content_lower for word in help_keywords)
        
        ai_response = ""
        action_taken = None
        
        # HELP - Show available commands
        if is_help:
            ai_response = """🤖 **Olá! Sou o assistente financeiro do Sirius. Posso ajudar com:**

💰 **Registrar Transações (aceito várias de uma vez!):**
- "Recebi 5000 de salário"
- "Gastei 150 no supermercado"
- "Gastei 50 no mercado, 30 de uber, 200 de luz e 100 de internet"
- "Recebi 3000 de salário e 500 de freelance"
- "Paguei 200 de luz, 150 de água e 80 de internet"

📊 **Relatórios e Análises:**
- "Como estão minhas finanças?"
- "Resumo do mês"
- "Quanto gastei este mês?"
- "Qual meu saldo?"

📋 **Listar Transações:**
- "Mostrar últimas transações"
- "Ver gastos do mês"
- "Listar receitas"

💡 **Dicas e Insights:**
- "Me dê dicas de economia"
- "Analise meus gastos"

📝 **Orçamentos:**
- "Criar orçamento de 500 para alimentação"
- "Definir limite de 1000 para lazer"

**Categorias disponíveis:** alimentação, transporte, moradia, saúde, educação, lazer, salário, outros"""
        
        # LIST TRANSACTIONS
        elif is_list:
            recent = transactions[-10:] if len(transactions) > 10 else transactions
            if not recent:
                ai_response = "📋 Não há transações registradas este mês."
            else:
                ai_response = "📋 **Últimas Transações:**\n\n"
                for t in reversed(recent):
                    emoji = "💚" if t['type'] == 'income' else "🔴"
                    tipo = "+" if t['type'] == 'income' else "-"
                    ai_response += f"{emoji} {t['date']} | {tipo}R$ {t['amount']:.2f} | {t['category']}"
                    if t.get('description'):
                        ai_response += f" | {t['description']}"
                    ai_response += "\n"
        
        # MIXED TRANSACTIONS (both income and expense in same message)
        elif is_income and is_expense and has_amount:
            prompt = f'''Extraia TODAS as transações financeiras desta mensagem. A mensagem contém receitas E despesas misturadas.

Mensagem: "{content}"

Responda APENAS com um JSON array válido. Cada item deve ter o campo "type" como "income" ou "expense":
[{{"type": "income", "amount": 5000.0, "category": "salário", "description": "salário"}}, {{"type": "expense", "amount": 50.0, "category": "alimentação", "description": "mercado"}}]

Categorias para receita: salário, freelance, investimentos, vendas, reembolso, outros
Categorias para despesa: alimentação, transporte, moradia, saúde, educação, lazer, outros
Extraia o valor numérico exato de CADA transação. Responda SOMENTE com o JSON array.'''
            
            response = await call_llm(prompt, f"mixed_{user.user_id}", user_id=user.user_id)
            
            try:
                clean_response = response.strip()
                if "```" in clean_response:
                    clean_response = clean_response.split("```")[1].replace("json", "").strip()
                start_idx = clean_response.find('[')
                end_idx = clean_response.rfind(']') + 1
                transactions_list = []
                if start_idx != -1 and end_idx > start_idx:
                    transactions_list = json.loads(clean_response[start_idx:end_idx])
                else:
                    start_idx = clean_response.find('{')
                    end_idx = clean_response.rfind('}') + 1
                    if start_idx != -1 and end_idx > start_idx:
                        single = json.loads(clean_response[start_idx:end_idx])
                        transactions_list = [single]
                
                if not isinstance(transactions_list, list):
                    transactions_list = [transactions_list]
                
                registered = []
                running_balance = balance
                for td in transactions_list:
                    amt = float(td.get("amount", 0))
                    t_type = td.get("type", "expense")
                    if t_type not in ["income", "expense"]:
                        t_type = "expense"
                    if amt > 0:
                        transaction_id = f"trans_{uuid.uuid4().hex[:12]}"
                        transaction_doc = {
                            "transaction_id": transaction_id,
                            "user_id": user.user_id,
                            "type": t_type,
                            "amount": amt,
                            "category": td.get("category", "outros"),
                            "description": td.get("description", ""),
                            "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                            "created_at": datetime.now(timezone.utc).isoformat()
                        }
                        await db.transactions.insert_one(transaction_doc)
                        if t_type == "income":
                            running_balance += amt
                        else:
                            running_balance -= amt
                        registered.append({**td, "type": t_type})
                
                if registered:
                    action_taken = registered
                    total_income_added = sum(float(t['amount']) for t in registered if t['type'] == 'income')
                    total_expense_added = sum(float(t['amount']) for t in registered if t['type'] == 'expense')
                    
                    ai_response = f"""✅ **{len(registered)} Transações Registradas!**

"""
                    for i, td in enumerate(registered, 1):
                        emoji = "💰" if td['type'] == 'income' else "🔴"
                        tipo = "Receita" if td['type'] == 'income' else "Despesa"
                        ai_response += f"**{i}.** {emoji} R$ {float(td['amount']):.2f} | {tipo} | {td.get('category', 'outros')} | {td.get('description', '-')}\n"
                    
                    ai_response += f"\n💰 **Total Receitas:** R$ {total_income_added:.2f}"
                    ai_response += f"\n💸 **Total Despesas:** R$ {total_expense_added:.2f}"
                    ai_response += f"\n📊 **Novo Saldo:** R$ {running_balance:.2f}"
                else:
                    ai_response = "❌ Não consegui processar as transações. Tente separar receitas e despesas."
            except Exception as e:
                logging.error(f"Mixed transaction parse error: {e}")
                ai_response = "❌ Não consegui processar as transações. Tente novamente."
        
        # INCOME TRANSACTION (supports multiple in one message)
        elif is_income and has_amount:
            prompt = f'''Extraia TODAS as receitas/entradas financeiras desta mensagem. A mensagem pode conter UMA ou VÁRIAS receitas.

Mensagem: "{content}"

Responda APENAS com um JSON array válido. Mesmo que seja apenas uma receita, retorne como array:
[{{"type": "income", "amount": 1000.0, "category": "salário", "description": "descrição curta"}}]

Exemplo com múltiplas receitas:
[{{"type": "income", "amount": 5000.0, "category": "salário", "description": "salário mensal"}}, {{"type": "income", "amount": 500.0, "category": "freelance", "description": "freelance design"}}]

Categorias para receita: salário, freelance, investimentos, vendas, reembolso, outros
Extraia o valor numérico exato de CADA receita mencionada. Responda SOMENTE com o JSON array.'''
            
            response = await call_llm(prompt, f"income_{user.user_id}", user_id=user.user_id)
            
            try:
                clean_response = response.strip()
                if "```" in clean_response:
                    clean_response = clean_response.split("```")[1].replace("json", "").strip()
                # Try to parse as array first
                start_idx = clean_response.find('[')
                end_idx = clean_response.rfind(']') + 1
                transactions_list = []
                if start_idx != -1 and end_idx > start_idx:
                    transactions_list = json.loads(clean_response[start_idx:end_idx])
                else:
                    # Fallback: try single object
                    start_idx = clean_response.find('{')
                    end_idx = clean_response.rfind('}') + 1
                    if start_idx != -1 and end_idx > start_idx:
                        single = json.loads(clean_response[start_idx:end_idx])
                        transactions_list = [single]
                
                if not isinstance(transactions_list, list):
                    transactions_list = [transactions_list]
                
                registered = []
                running_balance = balance
                for td in transactions_list:
                    amt = float(td.get("amount", 0))
                    if amt > 0:
                        transaction_id = f"trans_{uuid.uuid4().hex[:12]}"
                        transaction_doc = {
                            "transaction_id": transaction_id,
                            "user_id": user.user_id,
                            "type": "income",
                            "amount": amt,
                            "category": td.get("category", "outros"),
                            "description": td.get("description", ""),
                            "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                            "created_at": datetime.now(timezone.utc).isoformat()
                        }
                        await db.transactions.insert_one(transaction_doc)
                        running_balance += amt
                        registered.append(td)
                
                if registered:
                    action_taken = registered if len(registered) > 1 else registered[0]
                    if len(registered) == 1:
                        td = registered[0]
                        ai_response = f"""✅ **Receita Registrada!**

💰 **Valor:** R$ {float(td['amount']):.2f}
📁 **Categoria:** {td.get('category', 'outros')}
📝 **Descrição:** {td.get('description', '-')}

📊 **Novo Saldo:** R$ {running_balance:.2f}"""
                    else:
                        total_added = sum(float(t['amount']) for t in registered)
                        ai_response = f"""✅ **{len(registered)} Receitas Registradas!**

"""
                        for i, td in enumerate(registered, 1):
                            ai_response += f"**{i}.** 💰 R$ {float(td['amount']):.2f} | {td.get('category', 'outros')} | {td.get('description', '-')}\n"
                        
                        ai_response += f"""
💰 **Total Adicionado:** R$ {total_added:.2f}
📊 **Novo Saldo:** R$ {running_balance:.2f}"""
                else:
                    ai_response = "❌ Não consegui processar a receita. Tente: 'Recebi 1000 de salário'"
            except Exception as e:
                logging.error(f"Income parse error: {e}")
                ai_response = "❌ Não consegui processar a receita. Tente: 'Recebi 1000 de salário'"
        
        # EXPENSE TRANSACTION (supports multiple in one message)
        elif is_expense and has_amount:
            prompt = f'''Extraia TODAS as despesas/gastos financeiros desta mensagem. A mensagem pode conter UMA ou VÁRIAS despesas.

Mensagem: "{content}"

Responda APENAS com um JSON array válido. Mesmo que seja apenas uma despesa, retorne como array:
[{{"type": "expense", "amount": 100.0, "category": "alimentação", "description": "descrição curta"}}]

Exemplo com múltiplas despesas:
[{{"type": "expense", "amount": 50.0, "category": "alimentação", "description": "supermercado"}}, {{"type": "expense", "amount": 30.0, "category": "transporte", "description": "uber"}}, {{"type": "expense", "amount": 200.0, "category": "moradia", "description": "conta de luz"}}]

Categorias para despesa: alimentação, transporte, moradia, saúde, educação, lazer, outros
Extraia o valor numérico exato de CADA despesa mencionada. Responda SOMENTE com o JSON array.'''
            
            response = await call_llm(prompt, f"expense_{user.user_id}", user_id=user.user_id)
            
            try:
                clean_response = response.strip()
                if "```" in clean_response:
                    clean_response = clean_response.split("```")[1].replace("json", "").strip()
                # Try to parse as array first
                start_idx = clean_response.find('[')
                end_idx = clean_response.rfind(']') + 1
                transactions_list = []
                if start_idx != -1 and end_idx > start_idx:
                    transactions_list = json.loads(clean_response[start_idx:end_idx])
                else:
                    # Fallback: try single object
                    start_idx = clean_response.find('{')
                    end_idx = clean_response.rfind('}') + 1
                    if start_idx != -1 and end_idx > start_idx:
                        single = json.loads(clean_response[start_idx:end_idx])
                        transactions_list = [single]
                
                if not isinstance(transactions_list, list):
                    transactions_list = [transactions_list]
                
                registered = []
                running_balance = balance
                budget_alerts = []
                for td in transactions_list:
                    amt = float(td.get("amount", 0))
                    if amt > 0:
                        transaction_id = f"trans_{uuid.uuid4().hex[:12]}"
                        transaction_doc = {
                            "transaction_id": transaction_id,
                            "user_id": user.user_id,
                            "type": "expense",
                            "amount": amt,
                            "category": td.get("category", "outros"),
                            "description": td.get("description", ""),
                            "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                            "created_at": datetime.now(timezone.utc).isoformat()
                        }
                        await db.transactions.insert_one(transaction_doc)
                        running_balance -= amt
                        registered.append(td)
                        
                        # Check budget for this category
                        cat = td.get("category", "outros")
                        for b in budgets:
                            if b['category'] == cat:
                                new_spent = expense_by_category.get(cat, 0) + sum(float(r['amount']) for r in registered if r.get('category') == cat)
                                pct = (new_spent / b['limit']) * 100
                                if pct >= 100:
                                    alert = f"⚠️ **ALERTA:** Orçamento de {cat} estourado! ({pct:.0f}%)"
                                    if alert not in budget_alerts:
                                        budget_alerts.append(alert)
                                elif pct >= 80:
                                    alert = f"⚠️ **Atenção:** {pct:.0f}% do orçamento de {cat} usado"
                                    if alert not in budget_alerts:
                                        budget_alerts.append(alert)
                
                if registered:
                    action_taken = registered if len(registered) > 1 else registered[0]
                    budget_alert_text = "\n".join(budget_alerts) if budget_alerts else ""
                    if budget_alert_text:
                        budget_alert_text = "\n\n" + budget_alert_text
                    
                    if len(registered) == 1:
                        td = registered[0]
                        ai_response = f"""✅ **Despesa Registrada!**

🔴 **Valor:** R$ {float(td['amount']):.2f}
📁 **Categoria:** {td.get('category', 'outros')}
📝 **Descrição:** {td.get('description', '-')}

📊 **Novo Saldo:** R$ {running_balance:.2f}{budget_alert_text}"""
                    else:
                        total_spent = sum(float(t['amount']) for t in registered)
                        ai_response = f"""✅ **{len(registered)} Despesas Registradas!**

"""
                        for i, td in enumerate(registered, 1):
                            ai_response += f"**{i}.** 🔴 R$ {float(td['amount']):.2f} | {td.get('category', 'outros')} | {td.get('description', '-')}\n"
                        
                        ai_response += f"""
💸 **Total Gasto:** R$ {total_spent:.2f}
📊 **Novo Saldo:** R$ {running_balance:.2f}{budget_alert_text}"""
                else:
                    ai_response = "❌ Não consegui processar a despesa. Tente: 'Gastei 50 no mercado'"
            except Exception as e:
                logging.error(f"Expense parse error: {e}")
                ai_response = "❌ Não consegui processar a despesa. Tente: 'Gastei 50 no mercado'"
        
        # CREATE BUDGET
        elif is_budget and has_amount:
            prompt = f'''Extraia os dados do orçamento desta mensagem.

Mensagem: "{content}"

Responda APENAS com JSON válido:
{{"category": "alimentação", "limit": 500.0}}

Categorias: alimentação, transporte, moradia, saúde, educação, lazer, outros
Responda SOMENTE com o JSON.'''
            
            response = await call_llm(prompt, f"budget_{user.user_id}", user_id=user.user_id)
            
            try:
                clean_response = response.strip()
                if "```" in clean_response:
                    clean_response = clean_response.split("```")[1].replace("json", "").strip()
                start_idx = clean_response.find('{')
                end_idx = clean_response.rfind('}') + 1
                if start_idx != -1 and end_idx > start_idx:
                    budget_data = json.loads(clean_response[start_idx:end_idx])
                    
                    # Check if budget exists
                    existing = await db.budgets.find_one({
                        "user_id": user.user_id, 
                        "category": budget_data['category'], 
                        "month": current_month
                    })
                    
                    if existing:
                        await db.budgets.update_one(
                            {"budget_id": existing['budget_id']},
                            {"$set": {"limit": float(budget_data['limit'])}}
                        )
                        ai_response = f"""✅ **Orçamento Atualizado!**

📁 **Categoria:** {budget_data['category']}
💰 **Novo Limite:** R$ {float(budget_data['limit']):.2f}
📅 **Mês:** {current_month}"""
                    else:
                        budget_id = f"budget_{uuid.uuid4().hex[:12]}"
                        budget_doc = {
                            "budget_id": budget_id,
                            "user_id": user.user_id,
                            "category": budget_data['category'],
                            "limit": float(budget_data['limit']),
                            "spent": expense_by_category.get(budget_data['category'], 0),
                            "month": current_month,
                            "created_at": datetime.now(timezone.utc).isoformat()
                        }
                        await db.budgets.insert_one(budget_doc)
                        ai_response = f"""✅ **Orçamento Criado!**

📁 **Categoria:** {budget_data['category']}
💰 **Limite:** R$ {float(budget_data['limit']):.2f}
📅 **Mês:** {current_month}"""
            except Exception:
                ai_response = "❌ Não consegui criar o orçamento. Tente: 'Criar orçamento de 500 para alimentação'"
        
        # REPORT / ANALYSIS
        elif is_report:
            # Generate detailed report
            top_expenses = sorted(expense_by_category.items(), key=lambda x: x[1], reverse=True)[:5]
            top_income = sorted(income_by_category.items(), key=lambda x: x[1], reverse=True)[:3]
            
            budget_status = ""
            for b in budgets:
                pct = (b['spent'] / b['limit']) * 100 if b['limit'] > 0 else 0
                status = "🔴 Estourado" if pct >= 100 else "🟡 Atenção" if pct >= 80 else "🟢 OK"
                budget_status += f"- {b['category']}: R$ {b['spent']:.2f} / R$ {b['limit']:.2f} ({pct:.0f}%) {status}\n"
            
            ai_response = f"""📊 **RELATÓRIO FINANCEIRO - {current_month}**

💰 **Resumo:**
- Receitas: R$ {total_income:.2f}
- Despesas: R$ {total_expense:.2f}
- **Saldo: R$ {balance:.2f}** {"✅" if balance >= 0 else "⚠️"}

📈 **Maiores Receitas:**
"""
            for cat, val in top_income:
                ai_response += f"- {cat}: R$ {val:.2f}\n"
            
            ai_response += "\n📉 **Maiores Despesas:**\n"
            for cat, val in top_expenses:
                ai_response += f"- {cat}: R$ {val:.2f}\n"
            
            if budget_status:
                ai_response += f"\n📋 **Status dos Orçamentos:**\n{budget_status}"
            
            # Add AI insights
            prompt = f"""Baseado nestes dados financeiros, dê 2-3 insights curtos e práticos:
{context}

Seja direto e objetivo. Foque em dicas acionáveis."""
            
            try:
                insights = await call_llm(prompt, f"insights_{user.user_id}", user_id=user.user_id)
                ai_response += f"\n💡 **Insights:**\n{insights}"
            except Exception:
                pass
        
        # GENERAL CONVERSATION - Use AI
        else:
            prompt = f"""Você é o assistente financeiro inteligente do Sirius. Ajude o usuário com finanças pessoais.

{context}

Mensagem do usuário: "{content}"

Responda de forma útil, amigável e em português. Se o usuário parecer querer registrar uma transação mas não ficou claro, pergunte os detalhes. Se for uma pergunta sobre finanças, responda com base no contexto. Se for uma saudação, seja amigável e ofereça ajuda.

Mantenha a resposta concisa (máximo 3-4 parágrafos)."""
            
            ai_response = await call_llm(prompt, f"chat_{user.user_id}", user_id=user.user_id)
        
        ai_message_id = f"msg_{uuid.uuid4().hex[:12]}"
        ai_message = {
            "message_id": ai_message_id,
            "user_id": user.user_id,
            "role": "assistant",
            "content": ai_response,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        if action_taken:
            ai_message["transaction_data"] = action_taken
        
        await db.chat_messages.insert_one(ai_message.copy())
        
        user_message['created_at'] = datetime.fromisoformat(user_message['created_at'])
        ai_message['created_at'] = datetime.fromisoformat(ai_message['created_at'])
        
        return {"user_message": user_message, "ai_message": ai_message}
    except Exception as e:
        logging.error(f"Chat error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/chat/analyze-image")
async def analyze_image_for_expenses(
    request: Request, 
    image: UploadFile = File(...),
    description: str = Form(""),
    session_token: Optional[str] = Cookie(None)
):
    """Analyze an image (receipt, invoice, etc.) and extract expense information"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if not gemini_client:
        raise HTTPException(status_code=503, detail="Serviço de IA não disponível")
    
    # Validate file type
    allowed_types = ["image/jpeg", "image/png", "image/webp", "image/heic", "image/heif", "image/jpg"]
    content_type = image.content_type or "image/jpeg"
    
    # Normalize content type
    if content_type == "image/jpg":
        content_type = "image/jpeg"
    
    if content_type not in allowed_types:
        raise HTTPException(status_code=400, detail=f"Tipo de imagem não suportado: {content_type}. Use JPEG, PNG, WebP ou HEIC.")
    
    try:
        # Read image content
        image_content = await image.read()
        
        # Check file size (max 20MB for inline)
        if len(image_content) > 20 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="Imagem muito grande. Máximo 20MB.")
        
        # Validate and normalize image using PIL
        from io import BytesIO
        from PIL import Image as PILImage
        try:
            pil_image = PILImage.open(BytesIO(image_content))
            # Convert RGBA/P/other modes to RGB for JPEG compatibility
            if pil_image.mode not in ('RGB', 'L'):
                pil_image = pil_image.convert('RGB')
            # Re-encode as JPEG for maximum compatibility with Gemini
            img_buffer = BytesIO()
            pil_image.save(img_buffer, format='JPEG', quality=90)
            image_content = img_buffer.getvalue()
            content_type = "image/jpeg"
            logging.info(f"Image validated and converted: {pil_image.size}, mode={pil_image.mode}")
        except Exception as pil_error:
            logging.warning(f"PIL image validation failed: {pil_error}, using original bytes")
            # If PIL can't open it, try sending original bytes
        
        # Create user message with image reference
        message_id = f"msg_{uuid.uuid4().hex[:12]}"
        user_message = {
            "message_id": message_id,
            "user_id": user.user_id,
            "role": "user",
            "content": f"[Imagem enviada] {description}" if description else "[Imagem enviada para análise de gastos]",
            "has_image": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.chat_messages.insert_one(user_message.copy())
        
        # Analyze image with Gemini Vision
        prompt = """Analise esta imagem de comprovante/nota fiscal/recibo e extraia as informações de gastos.

IMPORTANTE: Responda SEMPRE em formato JSON válido com a seguinte estrutura:
{
    "found_expenses": true/false,
    "expenses": [
        {
            "description": "descrição do item/gasto",
            "amount": 0.00,
            "category": "alimentação/transporte/moradia/saúde/educação/lazer/outros",
            "date": "YYYY-MM-DD ou null se não encontrar"
        }
    ],
    "total": 0.00,
    "establishment": "nome do estabelecimento ou null",
    "summary": "resumo breve da análise"
}

Se não conseguir identificar gastos na imagem, retorne:
{
    "found_expenses": false,
    "expenses": [],
    "total": 0,
    "establishment": null,
    "summary": "Não foi possível identificar gastos nesta imagem"
}"""

        # Call Gemini with image - use proper multimodal format
        try:
            image_part = types.Part.from_bytes(data=image_content, mime_type=content_type)
            response = gemini_client.models.generate_content(
                model=GEMINI_MODEL,
                contents=[prompt, image_part],
                config=types.GenerateContentConfig(
                    response_mime_type="application/json"
                )
            )
            ai_response_text = response.text
        except Exception as gemini_error:
            logging.error(f"Gemini Vision error (first attempt): {gemini_error}")
            # Second attempt: try with base64 encoding instead
            try:
                import base64 as b64
                b64_data = b64.standard_b64encode(image_content).decode("utf-8")
                image_part = types.Part.from_bytes(
                    data=base64.b64decode(b64_data) if isinstance(b64_data, str) else image_content,
                    mime_type="image/jpeg"
                )
                response = gemini_client.models.generate_content(
                    model=GEMINI_MODEL,
                    contents=[prompt, image_part],
                    config=types.GenerateContentConfig(
                        response_mime_type="application/json"
                    )
                )
                ai_response_text = response.text
            except Exception as retry_error:
                logging.error(f"Gemini Vision retry also failed: {retry_error}")
                # Fallback response if Gemini fails
                ai_response_text = json.dumps({
                    "found_expenses": False,
                    "expenses": [],
                    "total": 0,
                    "establishment": None,
                    "summary": "Não foi possível analisar a imagem. Por favor, tente com outra imagem ou formato diferente (JPEG/PNG)."
                })
        
        # Try to parse JSON from response
        transactions_created = []
        try:
            # Clean response - remove markdown code blocks if present
            cleaned_response = ai_response_text.strip()
            if cleaned_response.startswith("```json"):
                cleaned_response = cleaned_response[7:]
            if cleaned_response.startswith("```"):
                cleaned_response = cleaned_response[3:]
            if cleaned_response.endswith("```"):
                cleaned_response = cleaned_response[:-3]
            cleaned_response = cleaned_response.strip()
            
            parsed = json.loads(cleaned_response)
            
            if parsed.get("found_expenses") and parsed.get("expenses"):
                today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
                
                for expense in parsed["expenses"]:
                    if expense.get("amount") and float(expense["amount"]) > 0:
                        transaction_id = f"trans_{uuid.uuid4().hex[:12]}"
                        transaction_doc = {
                            "transaction_id": transaction_id,
                            "user_id": user.user_id,
                            "type": "expense",
                            "amount": float(expense["amount"]),
                            "category": expense.get("category", "outros"),
                            "description": expense.get("description", "Gasto via imagem"),
                            "date": expense.get("date") or today,
                            "source": "image_analysis",
                            "created_at": datetime.now(timezone.utc).isoformat()
                        }
                        await db.transactions.insert_one(transaction_doc)
                        transaction_doc.pop('_id', None)
                        transactions_created.append(transaction_doc)
                
                # Build response message
                if transactions_created:
                    response_parts = [f"✅ **{len(transactions_created)} gasto(s) registrado(s)!**\n"]
                    if parsed.get("establishment"):
                        response_parts.append(f"📍 **Local:** {parsed['establishment']}\n")
                    response_parts.append("\n**Itens:**")
                    for t in transactions_created:
                        response_parts.append(f"\n• {t['description']}: R$ {t['amount']:.2f} ({t['category']})")
                    response_parts.append(f"\n\n💰 **Total:** R$ {parsed.get('total', sum(t['amount'] for t in transactions_created)):.2f}")
                    
                    ai_response = "\n".join(response_parts)
                else:
                    ai_response = f"📝 **Análise da imagem:**\n\n{parsed.get('summary', 'Imagem analisada mas nenhum gasto foi registrado.')}"
            else:
                ai_response = f"📝 **Análise da imagem:**\n\n{parsed.get('summary', 'Não foi possível identificar gastos nesta imagem.')}"
                
        except json.JSONDecodeError:
            # If JSON parsing fails, return raw response
            ai_response = f"📝 **Análise da imagem:**\n\n{ai_response_text}"
        
        # Save AI response
        ai_message_id = f"msg_{uuid.uuid4().hex[:12]}"
        ai_message = {
            "message_id": ai_message_id,
            "user_id": user.user_id,
            "role": "assistant",
            "content": ai_response,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        if transactions_created:
            ai_message["transactions_created"] = [{"amount": t["amount"], "category": t["category"], "description": t["description"]} for t in transactions_created]
        
        await db.chat_messages.insert_one(ai_message.copy())
        
        user_message['created_at'] = datetime.fromisoformat(user_message['created_at'])
        ai_message['created_at'] = datetime.fromisoformat(ai_message['created_at'])
        
        return {
            "user_message": user_message, 
            "ai_message": ai_message,
            "transactions_created": transactions_created
        }
        
    except Exception as e:
        logging.error(f"Image analysis error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao analisar imagem: {str(e)}")

@api_router.get("/reports")
async def get_reports(request: Request, type: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    query = {"user_id": user.user_id}
    if type:
        query["type"] = type
    
    reports = await db.reports.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    for report in reports:
        if isinstance(report['created_at'], str):
            report['created_at'] = datetime.fromisoformat(report['created_at'])
    return reports

@api_router.post("/reports/generate")
async def generate_report(request: Request, report_type: str, period: str, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    tasks = await db.tasks.find({"user_id": user.user_id}, {"_id": 0}).to_list(1000)
    task_instances = await db.task_instances.find({"user_id": user.user_id}, {"_id": 0}).to_list(1000)
    habits = await db.habits.find({"user_id": user.user_id}, {"_id": 0}).to_list(1000)
    transactions = await db.transactions.find({"user_id": user.user_id}, {"_id": 0}).to_list(1000)
    goals = await db.goals.find({"user_id": user.user_id}, {"_id": 0}).to_list(1000)
    
    # Count completed task instances
    completed_task_instances = len([t for t in task_instances if t.get('completed', False)])
    
    data = {
        "tasks": len(tasks),
        "tasks_completed": completed_task_instances,
        "habits": len(habits),
        "total_habits_completions": sum([len(h['completions']) for h in habits]),
        "income": sum([t['amount'] for t in transactions if t['type'] == 'income']),
        "expenses": sum([t['amount'] for t in transactions if t['type'] == 'expense']),
        "goals": len(goals),
        "goals_progress": sum([g['progress'] for g in goals]) / len(goals) if goals else 0
    }
    
    try:
        prompt = f"""Você é um analista de produtividade e finanças. Gere um relatório {report_type} para o período {period} baseado nos seguintes dados:

Tarefas: {data['tasks']} total, {data['tasks_completed']} concluídas
Hábitos: {data['habits']} total, {data['total_habits_completions']} completações
Receitas: R$ {data['income']:.2f}
Despesas: R$ {data['expenses']:.2f}
Metas: {data['goals']} total, {data['goals_progress']:.1f}% progresso médio

Forneça insights, padrões identificados e sugestões de otimização em português."""
        
        # Use Emergent LLM API
        insights = await call_llm(prompt, f"report_{user.user_id}", user_id=user.user_id)
        
        report_id = f"report_{uuid.uuid4().hex[:12]}"
        report_doc = {
            "report_id": report_id,
            "user_id": user.user_id,
            "type": report_type,
            "period": period,
            "data": data,
            "insights": insights,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.reports.insert_one(report_doc)
        report_doc.pop('_id', None)  # Remove MongoDB ObjectId
        report_doc['created_at'] = datetime.fromisoformat(report_doc['created_at'])
        
        return Report(**report_doc)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/stats/dashboard")
async def get_dashboard_stats(request: Request, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    tasks_today = await db.tasks.count_documents({"user_id": user.user_id, "is_template": True})
    tasks_completed_today = await db.task_instances.count_documents({"user_id": user.user_id, "date": today, "completed": True})
    
    habits = await db.habits.find({"user_id": user.user_id}, {"_id": 0}).to_list(1000)
    habits_completed_today = len([h for h in habits if today in h['completions']])
    
    current_month = datetime.now(timezone.utc).strftime("%Y-%m")
    transactions = await db.transactions.find({"user_id": user.user_id, "date": {"$regex": f"^{current_month}"}}, {"_id": 0}).to_list(1000)
    income = sum([t['amount'] for t in transactions if t['type'] == 'income'])
    expenses = sum([t['amount'] for t in transactions if t['type'] == 'expense'])
    
    goals = await db.goals.find({"user_id": user.user_id}, {"_id": 0}).to_list(1000)
    avg_progress = sum([len(g.get('daily_checks', [])) for g in goals]) / len(goals) if goals else 0
    
    # ===== WORKOUT STATS =====
    week_start = (datetime.now(timezone.utc) - timedelta(days=7)).strftime("%Y-%m-%d")
    workouts_week = await db.workout_logs.find({
        "user_id": user.user_id,
        "date": {"$gte": week_start},
        "completed": True
    }, {"_id": 0}).to_list(100)
    
    workout_stats = {
        "workouts_this_week": len(workouts_week),
        "total_duration_minutes": sum(w.get("duration_minutes", 0) for w in workouts_week),
        "total_calories_burned": sum(w.get("calories", 0) or 0 for w in workouts_week),
        "total_xp_earned": sum(w.get("xp_earned", 0) for w in workouts_week)
    }
    
    # ===== NUTRITION STATS =====
    meals_today = await db.meals.find({"user_id": user.user_id, "date": today}, {"_id": 0}).to_list(100)
    water_today = await db.water_logs.find({"user_id": user.user_id, "date": today}, {"_id": 0}).to_list(100)
    nutrition_goals = await db.nutrition_goals.find_one({"user_id": user.user_id}, {"_id": 0})
    
    if not nutrition_goals:
        nutrition_goals = {"daily_calories": 2000, "daily_protein": 150, "daily_carbs": 250, "daily_fat": 65, "water_goal_ml": 2000}
    
    nutrition_stats = {
        "calories_consumed": sum(m.get("total_calories", 0) for m in meals_today),
        "calories_goal": nutrition_goals.get("daily_calories", 2000),
        "protein_consumed": round(sum(m.get("total_protein", 0) for m in meals_today), 1),
        "protein_goal": nutrition_goals.get("daily_protein", 150),
        "water_consumed_ml": sum(w.get("amount_ml", 0) for w in water_today),
        "water_goal_ml": nutrition_goals.get("water_goal_ml", 2000),
        "meals_count": len(meals_today)
    }
    
    # ===== STUDY STATS =====
    study_sessions_today = await db.study_sessions.find({"user_id": user.user_id, "date": today}, {"_id": 0}).to_list(100)
    study_streak = await db.study_streaks.find_one({"user_id": user.user_id}, {"_id": 0})
    flashcards = await db.flashcards.find({"user_id": user.user_id}, {"_id": 0}).to_list(1000)
    due_flashcards = [f for f in flashcards if f.get("next_review", "") <= today]
    notebooks = await db.notebooks.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    
    study_stats = {
        "study_time_today_minutes": sum(s.get("duration_minutes", 0) for s in study_sessions_today),
        "current_streak": study_streak.get("current_streak", 0) if study_streak else 0,
        "longest_streak": study_streak.get("longest_streak", 0) if study_streak else 0,
        "flashcards_due": len(due_flashcards),
        "total_flashcards": len(flashcards),
        "notebooks_count": len(notebooks)
    }
    
    # ===== SIMULADOS STATS =====
    simulados_list = await db.simulados.find({"user_id": user.user_id}, {"_id": 0, "simulado_id": 1}).to_list(200)
    simulado_attempts = await db.simulado_attempts.find({"user_id": user.user_id}, {"_id": 0}).to_list(500)
    sim_scores = [a.get("score", 0) for a in simulado_attempts]
    sim_correct = sum(a.get("correct_count", 0) for a in simulado_attempts)
    sim_total_q = sum(a.get("total_questions", 0) for a in simulado_attempts)
    
    simulado_stats = {
        "total_simulados": len(simulados_list),
        "total_attempts": len(simulado_attempts),
        "average_score": round(sum(sim_scores) / len(sim_scores), 1) if sim_scores else 0,
        "best_score": round(max(sim_scores), 1) if sim_scores else 0,
        "total_questions_answered": sim_total_q,
        "total_correct": sim_correct,
        "accuracy_rate": round(sim_correct / sim_total_q * 100, 1) if sim_total_q > 0 else 0
    }
    
    # ===== QUESTION LOGS FOR OVERVIEW =====
    all_question_logs = await db.question_logs.find({"user_id": user.user_id}, {"_id": 0}).to_list(5000)
    total_questions_all = sum(q.get("total", 0) for q in all_question_logs)
    total_correct_all = sum(q.get("correct", 0) for q in all_question_logs)
    
    question_overview = {
        "total_answered": total_questions_all,
        "total_correct": total_correct_all,
        "accuracy_rate": round(total_correct_all / total_questions_all * 100, 1) if total_questions_all > 0 else 0
    }
    
    return {
        "user": {"name": user.name, "xp": user.xp, "rank": user.rank, "picture": user.picture},
        "tasks_today": tasks_today,
        "tasks_completed_today": tasks_completed_today,
        "habits_total": len(habits),
        "habits_completed_today": habits_completed_today,
        "income": income,
        "expenses": expenses,
        "balance": income - expenses,
        "goals_total": len(goals),
        "goals_avg_progress": avg_progress,
        "workout_stats": workout_stats,
        "nutrition_stats": nutrition_stats,
        "study_stats": study_stats,
        "simulado_stats": simulado_stats,
        "question_overview": question_overview
    }


@api_router.get("/stats/analytics")
async def get_analytics_data(request: Request, days: int = 7, session_token: Optional[str] = Cookie(None)):
    """Get historical analytics data for dashboard charts"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if days > 90:
        days = 90
    
    # Generate date range
    today = datetime.now(timezone.utc)
    date_range = [(today - timedelta(days=i)).strftime("%Y-%m-%d") for i in range(days - 1, -1, -1)]
    
    # Fetch all required data in parallel
    habits = await db.habits.find({"user_id": user.user_id}, {"_id": 0}).to_list(1000)
    
    # Task instances for date range
    task_instances = await db.task_instances.find({
        "user_id": user.user_id,
        "date": {"$gte": date_range[0], "$lte": date_range[-1]}
    }, {"_id": 0}).to_list(5000)
    
    # Transactions for date range
    transactions = await db.transactions.find({
        "user_id": user.user_id,
        "date": {"$gte": date_range[0], "$lte": date_range[-1]}
    }, {"_id": 0}).to_list(5000)
    
    # Study sessions
    study_sessions = await db.study_sessions.find({
        "user_id": user.user_id,
        "date": {"$gte": date_range[0], "$lte": date_range[-1]}
    }, {"_id": 0}).to_list(5000)
    
    # Workout logs
    workout_logs = await db.workout_logs.find({
        "user_id": user.user_id,
        "date": {"$gte": date_range[0], "$lte": date_range[-1]},
        "completed": True
    }, {"_id": 0}).to_list(1000)
    
    # Question logs
    question_logs = await db.question_logs.find({
        "user_id": user.user_id,
        "date": {"$gte": date_range[0], "$lte": date_range[-1]}
    }, {"_id": 0}).to_list(5000)
    
    # XP history from various collections
    xp_logs = await db.xp_logs.find({
        "user_id": user.user_id,
        "date": {"$gte": date_range[0], "$lte": date_range[-1]}
    }, {"_id": 0}).to_list(5000)
    
    # Build daily data
    daily_data = []
    cumulative_xp = 0
    
    for date in date_range:
        day_label = date[5:]  # MM-DD format
        
        # Tasks
        tasks_done = len([t for t in task_instances if t.get("date") == date and t.get("completed")])
        
        # Habits
        habits_done = len([h for h in habits if date in h.get("completions", [])])
        habits_total = len(habits)
        
        # Finance
        day_income = sum(t["amount"] for t in transactions if t.get("date") == date and t.get("type") == "income")
        day_expenses = sum(t["amount"] for t in transactions if t.get("date") == date and t.get("type") == "expense")
        
        # Study
        study_minutes = sum(s.get("duration_minutes", 0) for s in study_sessions if s.get("date") == date)
        
        # Workouts
        workouts_done = len([w for w in workout_logs if w.get("date") == date])
        workout_minutes = sum(w.get("duration_minutes", 0) for w in workout_logs if w.get("date") == date)
        
        # Questions
        questions_answered = sum(q.get("total", 0) for q in question_logs if q.get("date") == date)
        questions_correct = sum(q.get("correct", 0) for q in question_logs if q.get("date") == date)
        
        # XP
        day_xp = sum(x.get("amount", 0) for x in xp_logs if x.get("date") == date)
        cumulative_xp += day_xp
        
        daily_data.append({
            "date": date,
            "label": day_label,
            "tasks": tasks_done,
            "habits": habits_done,
            "habits_total": habits_total,
            "income": round(day_income, 2),
            "expenses": round(day_expenses, 2),
            "balance": round(day_income - day_expenses, 2),
            "study_min": study_minutes,
            "workouts": workouts_done,
            "workout_min": workout_minutes,
            "questions": questions_answered,
            "correct": questions_correct,
            "xp": day_xp,
            "xp_cumulative": cumulative_xp,
        })
    
    return {
        "days": days,
        "data": daily_data,
        "totals": {
            "tasks": sum(d["tasks"] for d in daily_data),
            "habits_avg": round(sum(d["habits"] for d in daily_data) / max(len(daily_data), 1), 1),
            "income": round(sum(d["income"] for d in daily_data), 2),
            "expenses": round(sum(d["expenses"] for d in daily_data), 2),
            "study_hours": round(sum(d["study_min"] for d in daily_data) / 60, 1),
            "workouts": sum(d["workouts"] for d in daily_data),
            "questions": sum(d["questions"] for d in daily_data),
            "xp_earned": sum(d["xp"] for d in daily_data),
        }
    }


@api_router.post("/goals/{goal_id}/check")
async def check_goal_day(request: Request, goal_id: str, date: str, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    goal = await db.goals.find_one({"goal_id": goal_id, "user_id": user.user_id}, {"_id": 0})
    if not goal:
        raise HTTPException(status_code=404, detail="Goal not found")
    
    daily_checks = goal.get('daily_checks', [])
    was_checked = date in daily_checks
    
    if was_checked:
        daily_checks.remove(date)
        xp_change = -5
        new_xp = max(0, user.xp + xp_change)
        new_rank = calculate_rank(new_xp)
        await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
    else:
        daily_checks.append(date)
        xp_change = 5
        new_xp = user.xp + xp_change
        new_rank = calculate_rank(new_xp)
        await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
    
    await db.goals.update_one(
        {"goal_id": goal_id},
        {"$set": {"daily_checks": daily_checks}}
    )
    
    if xp_change > 0:
        return {"message": "Day checked", "xp_earned": xp_change, "new_xp": user.xp + xp_change}
    else:
        return {"message": "Day unchecked", "xp_earned": xp_change, "new_xp": user.xp + xp_change}

@api_router.get("/challenges/current")
async def get_current_challenges(request: Request, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    today = datetime.now(timezone.utc).date()
    week_start = (today - timedelta(days=today.weekday())).isoformat()
    
    challenges = await db.challenges.find({"week_start": week_start}, {"_id": 0}).to_list(100)
    
    if len(challenges) == 0:
        default_challenges = [
            {
                "challenge_id": f"chal_{uuid.uuid4().hex[:12]}",
                "title": "Mestre das Tarefas",
                "description": "Complete 10 tarefas esta semana",
                "xp_reward": 50,
                "week_start": week_start,
                "week_end": (today + timedelta(days=7-today.weekday())).isoformat(),
                "completed_by": [],
                "created_at": datetime.now(timezone.utc).isoformat()
            },
            {
                "challenge_id": f"chal_{uuid.uuid4().hex[:12]}",
                "title": "Guardião dos Hábitos",
                "description": "Mantenha 5 dias de streak em qualquer hábito",
                "xp_reward": 75,
                "week_start": week_start,
                "week_end": (today + timedelta(days=7-today.weekday())).isoformat(),
                "completed_by": [],
                "created_at": datetime.now(timezone.utc).isoformat()
            },
            {
                "challenge_id": f"chal_{uuid.uuid4().hex[:12]}",
                "title": "Controlador Financeiro",
                "description": "Registre todas as transações diárias por 5 dias",
                "xp_reward": 100,
                "week_start": week_start,
                "week_end": (today + timedelta(days=7-today.weekday())).isoformat(),
                "completed_by": [],
                "created_at": datetime.now(timezone.utc).isoformat()
            }
        ]
        await db.challenges.insert_many(default_challenges)
        challenges = default_challenges
    
    for challenge in challenges:
        if isinstance(challenge['created_at'], str):
            challenge['created_at'] = datetime.fromisoformat(challenge['created_at'])
        challenge['completed'] = user.user_id in challenge.get('completed_by', [])
    
    return challenges

@api_router.post("/challenges/{challenge_id}/complete")
async def complete_challenge(request: Request, challenge_id: str, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    challenge = await db.challenges.find_one({"challenge_id": challenge_id}, {"_id": 0})
    if not challenge:
        raise HTTPException(status_code=404, detail="Challenge not found")
    
    if user.user_id in challenge.get('completed_by', []):
        raise HTTPException(status_code=400, detail="Challenge already completed")
    
    await db.challenges.update_one(
        {"challenge_id": challenge_id},
        {"$push": {"completed_by": user.user_id}}
    )
    
    new_xp = user.xp + challenge['xp_reward']
    new_rank = calculate_rank(new_xp)
    await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
    
    achievement_id = f"ach_{uuid.uuid4().hex[:12]}"
    achievement_doc = {
        "achievement_id": achievement_id,
        "user_id": user.user_id,
        "title": challenge['title'],
        "description": challenge['description'],
        "icon": "trophy",
        "unlocked_at": datetime.now(timezone.utc).isoformat()
    }
    await db.achievements.insert_one(achievement_doc)
    
    return {"message": "Challenge completed", "xp_earned": challenge['xp_reward'], "new_xp": new_xp, "new_rank": new_rank}

@api_router.get("/alerts")
async def get_alerts(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get automatic system alerts (budget, habits, etc.)"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    notifications = []
    
    current_month = datetime.now(timezone.utc).strftime("%Y-%m")
    budgets = await db.budgets.find({"user_id": user.user_id, "month": current_month}, {"_id": 0}).to_list(100)
    
    for budget in budgets:
        percentage = (budget['spent'] / budget['limit']) * 100
        if percentage >= 90:
            notifications.append({
                "type": "budget_alert",
                "severity": "high" if percentage >= 100 else "warning",
                "title": "Orçamento Estourado" if percentage >= 100 else "Orçamento Quase Estourado",
                "message": f"Categoria {budget['category']}: {percentage:.0f}% do orçamento usado",
                "data": budget
            })
    
    habits = await db.habits.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    for habit in habits:
        if today not in habit['completions'] and habit['streak'] > 0:
            notifications.append({
                "type": "habit_reminder",
                "severity": "info",
                "title": "Hábito Pendente",
                "message": f"{habit['name']}: Não esqueça de marcar hoje! Streak: {habit['streak']} dias",
                "data": habit
            })
    
    return notifications

def calculate_rank(xp: int) -> str:
    ranks = [
        (0, "Recruta"),
        (200, "Soldado"),
        (500, "Cabo"),
        (1000, "Sargento"),
        (1800, "Subtenente"),
        (3000, "Tenente"),
        (4500, "Capitão"),
        (6500, "Major"),
        (9000, "Tenente-Coronel"),
        (12000, "Coronel"),
        (16000, "General de Brigada"),
        (21000, "General de Divisão"),
        (27000, "General de Exército"),
        (35000, "Marechal")
    ]
    for threshold, rank in reversed(ranks):
        if xp >= threshold:
            return rank
    return "Recruta"

async def award_xp(user_id: str, amount: int):
    """Award XP and update rank atomically"""
    user_doc = await db.users.find_one({"user_id": user_id}, {"xp": 1})
    current_xp = user_doc.get("xp", 0) if user_doc else 0
    new_xp = max(0, current_xp + amount)
    new_rank = calculate_rank(new_xp)
    await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"xp": new_xp, "rank": new_rank}}
    )
    return new_xp, new_rank

def calculate_streak(completions: List[str]) -> int:
    if not completions:
        return 0
    
    today = datetime.now(timezone.utc).date()
    completions_dates = [datetime.fromisoformat(d).date() for d in completions]
    completions_dates.sort(reverse=True)
    
    if completions_dates[0] != today and completions_dates[0] != today - timedelta(days=1):
        return 0
    
    streak = 1
    for i in range(len(completions_dates) - 1):
        if completions_dates[i] - completions_dates[i+1] == timedelta(days=1):
            streak += 1
        else:
            break
    return streak

def calculate_best_streak(completions: List[str]) -> int:
    """Calculate the longest streak ever from all completions"""
    if not completions:
        return 0
    
    completions_dates = sorted([datetime.fromisoformat(d).date() for d in completions])
    
    if len(completions_dates) == 1:
        return 1
    
    best_streak = 1
    current_streak = 1
    
    for i in range(1, len(completions_dates)):
        if completions_dates[i] - completions_dates[i-1] == timedelta(days=1):
            current_streak += 1
            best_streak = max(best_streak, current_streak)
        else:
            current_streak = 1
    
    return best_streak

# ========== FINANCE CATEGORIES ==========
DEFAULT_FINANCE_CATEGORIES = ["alimentação", "transporte", "moradia", "saúde", "educação", "lazer", "investimentos", "salário", "freelance", "outros"]

@api_router.get("/finance/categories")
async def get_finance_categories(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get user's finance categories (default + custom)"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    user_cats = await db.finance_categories.find(
        {"user_id": user.user_id},
        {"_id": 0}
    ).to_list(100)
    
    custom_names = [c["name"] for c in user_cats]
    
    # Build full category list: defaults + user custom ones
    all_categories = []
    for cat in DEFAULT_FINANCE_CATEGORIES:
        all_categories.append({"name": cat, "is_default": True})
    for c in user_cats:
        if c["name"] not in DEFAULT_FINANCE_CATEGORIES:
            all_categories.append({"name": c["name"], "is_default": False, "icon": c.get("icon", ""), "color": c.get("color", "")})
    
    return {"categories": all_categories}


@api_router.post("/finance/categories")
async def create_finance_category(request: Request, session_token: Optional[str] = Cookie(None)):
    """Create a custom finance category"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    body = await request.json()
    name = body.get("name", "").strip().lower()
    icon = body.get("icon", "")
    color = body.get("color", "")
    
    if not name:
        raise HTTPException(status_code=400, detail="Nome da categoria é obrigatório")
    
    if len(name) > 30:
        raise HTTPException(status_code=400, detail="Nome muito longo (máx 30 caracteres)")
    
    # Check if already exists (default or custom)
    all_defaults = [c.lower() for c in DEFAULT_FINANCE_CATEGORIES]
    if name in all_defaults:
        raise HTTPException(status_code=400, detail="Essa categoria já existe como padrão")
    
    existing = await db.finance_categories.find_one({"user_id": user.user_id, "name": name})
    if existing:
        raise HTTPException(status_code=400, detail="Você já tem uma categoria com esse nome")
    
    cat_doc = {
        "cat_id": f"cat_{uuid.uuid4().hex[:12]}",
        "user_id": user.user_id,
        "name": name,
        "icon": icon,
        "color": color,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.finance_categories.insert_one(cat_doc)
    
    return {"success": True, "category": {"name": name, "is_default": False, "icon": icon, "color": color}}


@api_router.delete("/finance/categories/{category_name}")
async def delete_finance_category(request: Request, category_name: str, session_token: Optional[str] = Cookie(None)):
    """Delete a custom finance category"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    decoded_name = category_name.lower().strip()
    
    # Can't delete default categories
    if decoded_name in [c.lower() for c in DEFAULT_FINANCE_CATEGORIES]:
        raise HTTPException(status_code=400, detail="Não é possível excluir categorias padrão")
    
    result = await db.finance_categories.delete_one({"user_id": user.user_id, "name": decoded_name})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Categoria não encontrada")
    
    return {"success": True, "message": "Categoria removida"}


@api_router.get("/finance/stats")
async def get_finance_stats(request: Request, month: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if not month:
        month = datetime.now(timezone.utc).strftime("%Y-%m")
    
    transactions = await db.transactions.find(
        {"user_id": user.user_id, "date": {"$regex": f"^{month}"}},
        {"_id": 0}
    ).to_list(1000)
    
    categories_expense = {}
    categories_income = {}
    
    for t in transactions:
        if t['type'] == 'expense':
            categories_expense[t['category']] = categories_expense.get(t['category'], 0) + t['amount']
        else:
            categories_income[t['category']] = categories_income.get(t['category'], 0) + t['amount']
    
    return {
        "expense_by_category": categories_expense,
        "income_by_category": categories_income,
        "total_expense": sum(categories_expense.values()),
        "total_income": sum(categories_income.values())
    }

@api_router.get("/finance/trend")
async def get_finance_trend(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get last 6 months income vs expense trend for charts"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    today = datetime.now(timezone.utc)
    months_data = []
    
    for i in range(5, -1, -1):
        month_date = today - timedelta(days=i*30)
        month_str = month_date.strftime("%Y-%m")
        month_label = month_date.strftime("%b/%y")
        
        trans = await db.transactions.find(
            {"user_id": user.user_id, "date": {"$regex": f"^{month_str}"}}, {"_id": 0}
        ).to_list(500)
        
        income = sum(t['amount'] for t in trans if t['type'] == 'income')
        expense = sum(t['amount'] for t in trans if t['type'] == 'expense')
        
        months_data.append({
            "month": month_label,
            "month_key": month_str,
            "receitas": round(income, 2),
            "despesas": round(expense, 2),
            "saldo": round(income - expense, 2),
            "economia": round((income - expense) / income * 100, 1) if income > 0 else 0
        })
    
    # Calculate overall stats
    total_income_6m = sum(m['receitas'] for m in months_data)
    total_expense_6m = sum(m['despesas'] for m in months_data)
    avg_monthly_expense = total_expense_6m / 6
    savings_rate = round((total_income_6m - total_expense_6m) / total_income_6m * 100, 1) if total_income_6m > 0 else 0
    
    return {
        "trend": months_data,
        "summary": {
            "total_income_6m": round(total_income_6m, 2),
            "total_expense_6m": round(total_expense_6m, 2),
            "avg_monthly_expense": round(avg_monthly_expense, 2),
            "savings_rate": savings_rate,
            "best_month": max(months_data, key=lambda m: m['saldo'])['month'] if months_data else None,
            "worst_month": min(months_data, key=lambda m: m['saldo'])['month'] if months_data else None
        }
    }

@api_router.get("/nutrition/weekly-trend")
async def get_nutrition_weekly_trend(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get last 7 days nutrition data for charts"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    today = datetime.now()
    days_data = []
    
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        date_str = day.strftime("%Y-%m-%d")
        day_label = day.strftime("%a")
        
        meals = await db.meals.find({"user_id": user.user_id, "date": date_str}, {"_id": 0}).to_list(50)
        water = await db.water_logs.find({"user_id": user.user_id, "date": date_str}, {"_id": 0}).to_list(50)
        
        cal = sum(m.get("total_calories", 0) for m in meals)
        prot = sum(m.get("total_protein", 0) for m in meals)
        carb = sum(m.get("total_carbs", 0) for m in meals)
        fat = sum(m.get("total_fat", 0) for m in meals)
        water_ml = sum(w.get("amount_ml", 0) for w in water)
        
        days_data.append({
            "day": day_label,
            "date": date_str,
            "calorias": round(cal),
            "proteina": round(prot, 1),
            "carboidratos": round(carb, 1),
            "gordura": round(fat, 1),
            "agua_ml": water_ml,
            "refeicoes": len(meals)
        })
    
    goals = await db.nutrition_goals.find_one({"user_id": user.user_id}, {"_id": 0})
    if not goals:
        goals = {"daily_calories": 2000, "daily_protein": 150, "daily_carbs": 250, "daily_fat": 65, "water_goal_ml": 2000}
    
    # Calculate averages
    days_with_data = [d for d in days_data if d['calorias'] > 0]
    avg_cal = round(sum(d['calorias'] for d in days_with_data) / max(len(days_with_data), 1))
    avg_prot = round(sum(d['proteina'] for d in days_with_data) / max(len(days_with_data), 1), 1)
    
    return {
        "daily": days_data,
        "goals": goals,
        "averages": {
            "calorias": avg_cal,
            "proteina": avg_prot,
            "dias_registrados": len(days_with_data)
        }
    }

@api_router.get("/study/overall-stats")
async def get_study_overall_stats(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get overall study statistics for charts"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Get all notebooks
    notebooks = await db.notebooks.find({"user_id": user.user_id}, {"_id": 0}).to_list(200)
    
    # Study time by discipline
    disc_data = []
    total_time = 0
    total_questions = 0
    total_correct = 0
    
    for nb in notebooks:
        time_min = nb.get("total_study_time_minutes", 0)
        questions = nb.get("total_questions", 0)
        correct = nb.get("correct_questions", 0)
        total_time += time_min
        total_questions += questions
        total_correct += correct
        
        if time_min > 0 or questions > 0:
            disc_data.append({
                "nome": nb.get("name", ""),
                "tempo_horas": round(time_min / 60, 1),
                "questoes": questions,
                "acertos": correct,
                "acuracia": round(correct / questions * 100, 1) if questions > 0 else 0,
                "color": nb.get("color", "#007AFF")
            })
    
    # Focus sessions (last 7 days)
    today = datetime.now()
    focus_daily = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        date_str = day.strftime("%Y-%m-%d")
        sessions = await db.focus_sessions.find({
            "user_id": user.user_id,
            "date": date_str
        }, {"_id": 0}).to_list(50)
        focus_min = sum(s.get("duration_minutes", 0) for s in sessions)
        focus_daily.append({
            "day": day.strftime("%a"),
            "date": date_str,
            "minutos": focus_min,
            "sessoes": len(sessions)
        })
    
    # Question logs (last 7 days)
    question_daily = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        date_str = day.strftime("%Y-%m-%d")
        q_logs = await db.question_logs.find({
            "user_id": user.user_id,
            "date": date_str
        }, {"_id": 0}).to_list(100)
        total_q = sum(q.get("total", 0) for q in q_logs)
        correct_q = sum(q.get("correct", 0) for q in q_logs)
        question_daily.append({
            "day": day.strftime("%a"),
            "date": date_str,
            "questoes": total_q,
            "acertos": correct_q,
            "acuracia": round(correct_q / total_q * 100, 1) if total_q > 0 else 0
        })
    
    return {
        "disciplinas": sorted(disc_data, key=lambda x: x["tempo_horas"], reverse=True),
        "focus_daily": focus_daily,
        "question_daily": question_daily,
        "totals": {
            "tempo_total_horas": round(total_time / 60, 1),
            "questoes_total": total_questions,
            "acertos_total": total_correct,
            "acuracia_geral": round(total_correct / total_questions * 100, 1) if total_questions > 0 else 0,
            "disciplinas_ativas": len(disc_data)
        }
    }



@api_router.get("/reports/{report_id}/download")
async def download_report(report_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    from fastapi.responses import StreamingResponse
    import io
    
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    report = await db.reports.find_one({"report_id": report_id, "user_id": user.user_id}, {"_id": 0})
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    
    content = f"""SIRIUS - RELATÓRIO {report['type'].upper()}
Período: {report['period']}
Gerado em: {report['created_at']}

{'='*60}
DADOS DO PERÍODO
{'='*60}

Tarefas: {report['data']['tasks_completed']}/{report['data']['tasks']}
Hábitos: {report['data']['total_habits_completions']} completações
Receitas: R$ {report['data']['income']:.2f}
Despesas: R$ {report['data']['expenses']:.2f}
Metas: {report['data']['goals']} total
Progresso Médio: {report['data']['goals_progress']:.1f}%

{'='*60}
INSIGHTS E SUGESTÕES
{'='*60}

{report['insights']}
"""
    
    buffer = io.BytesIO(content.encode('utf-8'))
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="text/plain",
        headers={"Content-Disposition": f"attachment; filename=sirius_relatorio_{report_id}.txt"}
    )

class CreditCard(BaseModel):
    model_config = ConfigDict(extra="ignore")
    card_id: str
    user_id: str
    name: str
    limit: float
    closing_day: int
    due_day: int
    created_at: datetime

class CreditCardCreate(BaseModel):
    name: str
    limit: float
    closing_day: int
    due_day: int

class Invoice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    invoice_id: str
    card_id: str
    user_id: str
    month: str
    amount: float
    paid: bool = False
    created_at: datetime

class Projection(BaseModel):
    model_config = ConfigDict(extra="ignore")
    projection_id: str
    user_id: str
    month: str
    description: str
    amount: float
    category: str
    projection_type: str  # "fixed", "installment", "manual"
    is_fixed: bool = False  # Despesa fixa que se repete todo mês
    repeat_count: Optional[int] = None  # Número de vezes que se repete (se não for fixa)
    remaining_repeats: Optional[int] = None  # Repetições restantes
    source_transaction_id: Optional[str] = None  # ID da transação original (para parcelas)
    installment_number: Optional[int] = None  # Número da parcela atual
    total_installments: Optional[int] = None  # Total de parcelas
    card_id: Optional[str] = None  # Cartão associado (se aplicável)
    created_at: datetime

class ProjectionCreate(BaseModel):
    description: str
    amount: float
    category: str
    month: str
    is_fixed: bool = False
    repeat_count: Optional[int] = None

class CardChargeRequest(BaseModel):
    amount: float
    description: str
    category: str
    payment_type: str = "vista"  # "vista" ou "parcelado"
    installments: Optional[int] = 1  # Número de parcelas
    start_month: str = "current"  # "current" ou "next" - quando começa a primeira parcela

@api_router.get("/credit-cards")
async def get_credit_cards(request: Request, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    cards = await db.credit_cards.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    for card in cards:
        if isinstance(card['created_at'], str):
            card['created_at'] = datetime.fromisoformat(card['created_at'])
    return cards

@api_router.post("/credit-cards")
async def create_credit_card(request: Request, card_data: CreditCardCreate, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    card_id = f"card_{uuid.uuid4().hex[:12]}"
    card_doc = {
        "card_id": card_id,
        "user_id": user.user_id,
        "name": card_data.name,
        "limit": card_data.limit,
        "closing_day": card_data.closing_day,
        "due_day": card_data.due_day,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.credit_cards.insert_one(card_doc)
    card_doc.pop('_id', None)  # Remove MongoDB ObjectId
    card_doc['created_at'] = datetime.fromisoformat(card_doc['created_at'])
    return CreditCard(**card_doc)

@api_router.get("/credit-cards/{card_id}/invoices")
async def get_card_invoices(request: Request, card_id: str, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    invoices = await db.invoices.find({"card_id": card_id, "user_id": user.user_id}, {"_id": 0}).to_list(100)
    for invoice in invoices:
        if isinstance(invoice['created_at'], str):
            invoice['created_at'] = datetime.fromisoformat(invoice['created_at'])
    return invoices

@api_router.post("/credit-cards/{card_id}/charge")
async def charge_to_card(request: Request, card_id: str, charge_data: CardChargeRequest, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    card = await db.credit_cards.find_one({"card_id": card_id, "user_id": user.user_id}, {"_id": 0})
    if not card:
        raise HTTPException(status_code=404, detail="Card not found")
    
    amount = charge_data.amount
    description = charge_data.description
    category = charge_data.category
    payment_type = charge_data.payment_type
    installments = charge_data.installments or 1
    start_month = charge_data.start_month  # "current" ou "next"
    
    if payment_type == "parcelado" and installments < 2:
        installments = 2  # Mínimo de 2 parcelas para parcelamento
    
    # Calcular valor da parcela
    installment_amount = amount / installments if payment_type == "parcelado" else amount
    
    # Determinar o mês de início baseado na escolha do usuário
    current_date = datetime.now(timezone.utc)
    if start_month == "next":
        # Primeira parcela no próximo mês (aritmética correta de mês)
        next_m = current_date.month + 1
        next_y = current_date.year + (next_m - 1) // 12
        next_m = ((next_m - 1) % 12) + 1
        first_month_date = current_date.replace(year=next_y, month=next_m, day=1)
        first_month = f"{next_y}-{next_m:02d}"
        transaction_date = first_month_date.strftime("%Y-%m-%d")
    else:
        # Primeira parcela no mês atual
        first_month = current_date.strftime("%Y-%m")
        transaction_date = current_date.strftime("%Y-%m-%d")
    
    transaction_id = f"trans_{uuid.uuid4().hex[:12]}"
    transaction_doc = {
        "transaction_id": transaction_id,
        "user_id": user.user_id,
        "type": "expense",
        "amount": installment_amount,  # Primeira parcela ou valor à vista
        "total_amount": amount,  # Valor total da compra
        "category": category,
        "description": f"{description} (Cartão: {card['name']})" + (f" - Parcela 1/{installments}" if payment_type == "parcelado" else ""),
        "date": transaction_date,
        "card_id": card_id,
        "payment_type": payment_type,
        "installments": installments,
        "installment_number": 1,
        "start_month": start_month,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Se for para o próximo mês, criar como projeção ao invés de transação
    if start_month == "next":
        # Criar projeção para primeira parcela
        projection_id = f"proj_{uuid.uuid4().hex[:12]}"
        projection_doc = {
            "projection_id": projection_id,
            "user_id": user.user_id,
            "month": first_month,
            "description": f"{description} (Cartão: {card['name']})" + (f" - Parcela 1/{installments}" if payment_type == "parcelado" else ""),
            "amount": installment_amount,
            "category": category,
            "projection_type": "installment",
            "is_fixed": False,
            "repeat_count": None,
            "remaining_repeats": None,
            "source_transaction_id": transaction_id,
            "installment_number": 1,
            "total_installments": installments,
            "card_id": card_id,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.projections.insert_one(projection_doc)
    else:
        # Mês atual - criar transação normalmente
        await db.transactions.insert_one(transaction_doc)
    
    # Atualizar fatura do mês correspondente
    invoice = await db.invoices.find_one({"card_id": card_id, "month": first_month}, {"_id": 0})
    
    if not invoice:
        invoice_id = f"inv_{uuid.uuid4().hex[:12]}"
        invoice_doc = {
            "invoice_id": invoice_id,
            "card_id": card_id,
            "user_id": user.user_id,
            "month": first_month,
            "amount": installment_amount,
            "paid": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.invoices.insert_one(invoice_doc)
    else:
        new_amount = invoice['amount'] + installment_amount
        await db.invoices.update_one(
            {"invoice_id": invoice['invoice_id']},
            {"$set": {"amount": new_amount}}
        )
    
    # Se for parcelado, criar projeções para os meses seguintes
    if payment_type == "parcelado" and installments > 1:
        # Determinar mês base para cálculo dos meses seguintes
        if start_month == "next":
            base_year = first_month_date.year
            base_month_num = first_month_date.month
        else:
            base_year = current_date.year
            base_month_num = current_date.month
        
        for i in range(2, installments + 1):  # Começar da parcela 2
            # Aritmética correta de meses (evita duplicação por uso de timedelta)
            total_months = base_month_num + (i - 1)
            future_year = base_year + (total_months - 1) // 12
            future_month_num = ((total_months - 1) % 12) + 1
            future_month = f"{future_year}-{future_month_num:02d}"
            
            projection_id = f"proj_{uuid.uuid4().hex[:12]}"
            projection_doc = {
                "projection_id": projection_id,
                "user_id": user.user_id,
                "month": future_month,
                "description": f"{description} (Cartão: {card['name']}) - Parcela {i}/{installments}",
                "amount": installment_amount,
                "category": category,
                "projection_type": "installment",
                "is_fixed": False,
                "repeat_count": None,
                "remaining_repeats": None,
                "source_transaction_id": transaction_id,
                "installment_number": i,
                "total_installments": installments,
                "card_id": card_id,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.projections.insert_one(projection_doc)
    
    return {
        "message": "Charged to card", 
        "transaction_id": transaction_id,
        "installment_amount": installment_amount,
        "total_amount": amount,
        "installments": installments
    }

@api_router.patch("/invoices/{invoice_id}/pay")
async def pay_invoice(request: Request, invoice_id: str, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.invoices.update_one(
        {"invoice_id": invoice_id, "user_id": user.user_id},
        {"$set": {"paid": True}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return {"message": "Invoice paid"}

# ========== PROJECTION ENDPOINTS ==========

@api_router.get("/projections")
async def get_projections(request: Request, month: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if not month:
        # Próximo mês por padrão
        next_month = datetime.now(timezone.utc) + timedelta(days=30)
        month = next_month.strftime("%Y-%m")
    
    projections = await db.projections.find(
        {"user_id": user.user_id, "month": month},
        {"_id": 0}
    ).to_list(1000)
    
    for proj in projections:
        if isinstance(proj['created_at'], str):
            proj['created_at'] = datetime.fromisoformat(proj['created_at'])
    
    return projections

@api_router.post("/projections")
async def create_projection(request: Request, projection_data: ProjectionCreate, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    projection_id = f"proj_{uuid.uuid4().hex[:12]}"
    projection_doc = {
        "projection_id": projection_id,
        "user_id": user.user_id,
        "month": projection_data.month,
        "description": projection_data.description,
        "amount": projection_data.amount,
        "category": projection_data.category,
        "projection_type": "manual",
        "is_fixed": projection_data.is_fixed,
        "repeat_count": projection_data.repeat_count if not projection_data.is_fixed else None,
        "remaining_repeats": projection_data.repeat_count if not projection_data.is_fixed else None,
        "source_transaction_id": None,
        "installment_number": None,
        "total_installments": None,
        "card_id": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.projections.insert_one(projection_doc)
    projection_doc.pop('_id', None)  # Remove MongoDB ObjectId
    
    # Se for despesa fixa ou com repetições, criar projeções para meses futuros
    if projection_data.is_fixed or (projection_data.repeat_count and projection_data.repeat_count > 1):
        base_year = int(projection_data.month.split("-")[0])
        base_month_num = int(projection_data.month.split("-")[1])
        repeat_times = 12 if projection_data.is_fixed else (projection_data.repeat_count - 1)
        
        for i in range(1, repeat_times + 1):
            # Aritmética correta de meses (evita duplicação por uso de timedelta)
            total_months = base_month_num + i
            future_year = base_year + (total_months - 1) // 12
            future_month_num = ((total_months - 1) % 12) + 1
            future_month = f"{future_year}-{future_month_num:02d}"
            
            future_proj_id = f"proj_{uuid.uuid4().hex[:12]}"
            future_proj_doc = {
                "projection_id": future_proj_id,
                "user_id": user.user_id,
                "month": future_month,
                "description": projection_data.description,
                "amount": projection_data.amount,
                "category": projection_data.category,
                "projection_type": "manual",
                "is_fixed": projection_data.is_fixed,
                "repeat_count": projection_data.repeat_count,
                "remaining_repeats": (projection_data.repeat_count - i - 1) if projection_data.repeat_count else None,
                "source_transaction_id": None,
                "installment_number": i + 1 if projection_data.repeat_count else None,
                "total_installments": projection_data.repeat_count,
                "card_id": None,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.projections.insert_one(future_proj_doc)
    
    projection_doc['created_at'] = datetime.fromisoformat(projection_doc['created_at'])
    return Projection(**projection_doc)

@api_router.patch("/projections/{projection_id}")
async def update_projection(request: Request, projection_id: str, amount: Optional[float] = None, description: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    update_data = {}
    if amount is not None:
        update_data["amount"] = amount
    if description is not None:
        update_data["description"] = description
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    result = await db.projections.update_one(
        {"projection_id": projection_id, "user_id": user.user_id},
        {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Projection not found")
    
    return {"message": "Projection updated"}

@api_router.delete("/projections/{projection_id}")
async def delete_projection(request: Request, projection_id: str, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.projections.delete_one({"projection_id": projection_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Projection not found")
    return {"message": "Projection deleted"}

@api_router.get("/projections/summary")
async def get_projection_summary(request: Request, month: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if not month:
        next_month = datetime.now(timezone.utc) + timedelta(days=30)
        month = next_month.strftime("%Y-%m")
    
    # Buscar projeções do mês
    projections = await db.projections.find(
        {"user_id": user.user_id, "month": month},
        {"_id": 0}
    ).to_list(1000)
    
    # Calcular totais por categoria
    categories_totals = {}
    total_projected = 0
    fixed_expenses = 0
    installment_expenses = 0
    manual_expenses = 0
    
    for proj in projections:
        cat = proj.get('category', 'outros')
        amount = proj.get('amount', 0)
        proj_type = proj.get('projection_type', 'manual')
        
        categories_totals[cat] = categories_totals.get(cat, 0) + amount
        total_projected += amount
        
        if proj.get('is_fixed'):
            fixed_expenses += amount
        elif proj_type == 'installment':
            installment_expenses += amount
        else:
            manual_expenses += amount
    
    # Buscar receitas recorrentes (estimativa baseada no mês atual)
    current_month = datetime.now(timezone.utc).strftime("%Y-%m")
    current_income = await db.transactions.find(
        {"user_id": user.user_id, "type": "income", "date": {"$regex": f"^{current_month}"}},
        {"_id": 0}
    ).to_list(1000)
    estimated_income = sum([t['amount'] for t in current_income])
    
    return {
        "month": month,
        "total_projected_expenses": total_projected,
        "fixed_expenses": fixed_expenses,
        "installment_expenses": installment_expenses,
        "manual_expenses": manual_expenses,
        "categories_totals": categories_totals,
        "estimated_income": estimated_income,
        "estimated_balance": estimated_income - total_projected,
        "projections_count": len(projections)
    }

@api_router.post("/projections/insights")
async def get_projection_insights(request: Request, month: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if not month:
        next_month = datetime.now(timezone.utc) + timedelta(days=30)
        month = next_month.strftime("%Y-%m")
    
    # Buscar resumo das projeções
    projections = await db.projections.find(
        {"user_id": user.user_id, "month": month},
        {"_id": 0}
    ).to_list(1000)
    
    total_projected = sum([p.get('amount', 0) for p in projections])
    
    # Buscar receitas do mês atual como base
    current_month = datetime.now(timezone.utc).strftime("%Y-%m")
    current_transactions = await db.transactions.find(
        {"user_id": user.user_id, "date": {"$regex": f"^{current_month}"}},
        {"_id": 0}
    ).to_list(1000)
    
    estimated_income = sum([t['amount'] for t in current_transactions if t['type'] == 'income'])
    
    # Categorias com maiores gastos
    categories = {}
    for p in projections:
        cat = p.get('category', 'outros')
        categories[cat] = categories.get(cat, 0) + p.get('amount', 0)
    
    top_categories = sorted(categories.items(), key=lambda x: x[1], reverse=True)[:3]
    
    prompt = f"""Você é um consultor financeiro pessoal. Analise a projeção de gastos para {month} e forneça insights e sugestões práticas.

Dados da Projeção:
- Receita Estimada: R$ {estimated_income:.2f}
- Total de Despesas Projetadas: R$ {total_projected:.2f}
- Saldo Estimado: R$ {estimated_income - total_projected:.2f}
- Número de Despesas Projetadas: {len(projections)}

Maiores Categorias de Gastos:
{chr(10).join([f"- {cat}: R$ {val:.2f}" for cat, val in top_categories])}

Despesas Parceladas: {len([p for p in projections if p.get('projection_type') == 'installment'])}
Despesas Fixas: {len([p for p in projections if p.get('is_fixed')])}

Forneça:
1. Uma análise do cenário financeiro projetado
2. Alertas se houver risco de saldo negativo
3. Sugestões de economia específicas para as maiores categorias
4. Dicas para melhorar a saúde financeira

Responda em português, de forma objetiva e prática."""
    
    try:
        insights = await call_llm(prompt, f"projection_insights_{user.user_id}", user_id=user.user_id)
        
        return {
            "month": month,
            "insights": insights,
            "summary": {
                "estimated_income": estimated_income,
                "total_projected": total_projected,
                "estimated_balance": estimated_income - total_projected,
                "top_categories": top_categories
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ========== WORKOUT ENDPOINTS ==========
@api_router.get("/workout-plans")
async def get_workout_plans(request: Request, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    plans = await db.workout_plans.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    for plan in plans:
        if isinstance(plan['created_at'], str):
            plan['created_at'] = datetime.fromisoformat(plan['created_at'])
    return plans

@api_router.post("/workout-plans")
async def create_workout_plan(request: Request, plan_data: WorkoutPlanCreate, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    plan_id = f"plan_{uuid.uuid4().hex[:12]}"
    plan_doc = {
        "plan_id": plan_id,
        "user_id": user.user_id,
        "name": plan_data.name,
        "description": plan_data.description,
        "exercises": plan_data.exercises,
        "plan_duration": plan_data.plan_duration or "dia",
        "generated_by_ai": False,
        "days": plan_data.days,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.workout_plans.insert_one(plan_doc)
    plan_doc.pop('_id', None)  # Remove MongoDB ObjectId
    plan_doc['created_at'] = datetime.fromisoformat(plan_doc['created_at'])
    return WorkoutPlan(**plan_doc)

@api_router.patch("/workout-plans/{plan_id}")
async def update_workout_plan(request: Request, plan_id: str, plan_data: WorkoutPlanCreate, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    update_data = {
        "name": plan_data.name,
        "description": plan_data.description,
        "exercises": plan_data.exercises,
        "plan_duration": plan_data.plan_duration or "dia",
        "days": plan_data.days
    }
    
    result = await db.workout_plans.update_one(
        {"plan_id": plan_id, "user_id": user.user_id},
        {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Workout plan not found")
    return {"message": "Workout plan updated"}

@api_router.delete("/workout-plans/{plan_id}")
async def delete_workout_plan(request: Request, plan_id: str, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.workout_plans.delete_one({"plan_id": plan_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Workout plan not found")
    return {"message": "Workout plan deleted"}

@api_router.get("/workouts")
async def get_workouts(request: Request, date: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    query = {"user_id": user.user_id}
    if date:
        query["date"] = date
    
    workouts = await db.workout_logs.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    for workout in workouts:
        if isinstance(workout['created_at'], str):
            workout['created_at'] = datetime.fromisoformat(workout['created_at'])
    return workouts


@api_router.get("/workouts/today-schedule")
async def get_today_workout_schedule(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get today's scheduled workout from the user's plan."""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Find the user's plans with days, ordered by creation date
    plans = await db.workout_plans.find(
        {"user_id": user.user_id, "days": {"$exists": True, "$ne": []}},
        {"_id": 0}
    ).sort("created_at", -1).to_list(20)
    
    if not plans:
        return {"scheduled": False, "message": "Nenhum plano de treino encontrado"}
    
    plan = plans[0]
    days = plan.get("days", [])
    if not days:
        return {"scheduled": False, "message": "Plano não possui dias definidos"}
    
    created_at = plan.get("created_at")
    if isinstance(created_at, str):
        created_at = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
    
    now = datetime.now(timezone.utc)
    days_since_start = (now - created_at).days
    day_index = days_since_start % len(days)
    
    today_workout = days[day_index]
    exercises = today_workout.get("exercises", [])
    
    # Check if already logged today
    today_str = now.strftime("%Y-%m-%d")
    already_logged = await db.workout_logs.find_one({
        "user_id": user.user_id,
        "plan_id": plan.get("plan_id"),
        "date": today_str
    })
    
    return {
        "scheduled": True,
        "plan_id": plan.get("plan_id"),
        "plan_name": plan.get("name", ""),
        "day_name": today_workout.get("day_name", ""),
        "day_label": today_workout.get("day_label", ""),
        "split_label": today_workout.get("split_label", ""),
        "week": today_workout.get("week", 0),
        "day_index": day_index,
        "total_days": len(days),
        "exercises": exercises,
        "exercise_count": len(exercises),
        "already_completed": already_logged is not None
    }

@api_router.post("/workouts")
async def log_workout(request: Request, workout_data: WorkoutLogCreate, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Calcular XP baseado no tipo de atividade e duração
    base_xp = 10
    duration_bonus = (workout_data.duration_minutes // 15) * 5  # +5 XP a cada 15 min
    xp_earned = base_xp + duration_bonus
    
    log_id = f"workout_{uuid.uuid4().hex[:12]}"
    workout_doc = {
        "log_id": log_id,
        "user_id": user.user_id,
        "plan_id": workout_data.plan_id,
        "activity_type": workout_data.activity_type,
        "name": workout_data.name,
        "duration_minutes": workout_data.duration_minutes,
        "distance_km": workout_data.distance_km,
        "calories": workout_data.calories,
        "exercises_completed": workout_data.exercises_completed,
        "notes": workout_data.notes,
        "xp_earned": xp_earned,
        "completed": True,
        "date": workout_data.date,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.workout_logs.insert_one(workout_doc)
    workout_doc.pop('_id', None)  # Remove MongoDB ObjectId
    
    # Award XP to user
    new_xp = user.xp + xp_earned
    new_rank = calculate_rank(new_xp)
    await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
    
    workout_doc['created_at'] = datetime.fromisoformat(workout_doc['created_at'])
    return {**workout_doc, "new_xp": new_xp, "new_rank": new_rank}

@api_router.patch("/workouts/{log_id}/toggle")
async def toggle_workout(request: Request, log_id: str, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    workout = await db.workout_logs.find_one({"log_id": log_id, "user_id": user.user_id}, {"_id": 0})
    if not workout:
        raise HTTPException(status_code=404, detail="Workout not found")
    
    new_completed = not workout['completed']
    xp_change = workout['xp_earned'] if new_completed else -workout['xp_earned']
    
    await db.workout_logs.update_one(
        {"log_id": log_id},
        {"$set": {"completed": new_completed}}
    )
    
    # Update user XP
    new_xp = max(0, user.xp + xp_change)
    new_rank = calculate_rank(new_xp)
    await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
    
    return {
        "message": "Workout toggled",
        "completed": new_completed,
        "xp_change": xp_change,
        "new_xp": new_xp,
        "new_rank": new_rank
    }

@api_router.delete("/workouts/{log_id}")
async def delete_workout(request: Request, log_id: str, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    workout = await db.workout_logs.find_one({"log_id": log_id, "user_id": user.user_id}, {"_id": 0})
    if not workout:
        raise HTTPException(status_code=404, detail="Workout not found")
    
    # Deduct XP if was completed
    if workout['completed']:
        new_xp = max(0, user.xp - workout['xp_earned'])
        new_rank = calculate_rank(new_xp)
        await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
    
    await db.workout_logs.delete_one({"log_id": log_id})
    return {"message": "Workout deleted"}

@api_router.get("/workout-stats")
async def get_workout_stats(request: Request, period: str = "week", session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Calculate date range
    today = datetime.now(timezone.utc)
    if period == "week":
        start_date = (today - timedelta(days=7)).strftime("%Y-%m-%d")
    elif period == "month":
        start_date = (today - timedelta(days=30)).strftime("%Y-%m-%d")
    else:
        start_date = (today - timedelta(days=365)).strftime("%Y-%m-%d")
    
    workouts = await db.workout_logs.find({
        "user_id": user.user_id,
        "date": {"$gte": start_date},
        "completed": True
    }, {"_id": 0}).to_list(1000)
    
    total_workouts = len(workouts)
    total_duration = sum([w.get('duration_minutes', 0) for w in workouts])
    total_distance = sum([w.get('distance_km', 0) or 0 for w in workouts])
    total_calories = sum([w.get('calories', 0) or 0 for w in workouts])
    total_xp = sum([w.get('xp_earned', 0) for w in workouts])
    
    # Count by activity type
    by_type = {}
    for w in workouts:
        t = w['activity_type']
        by_type[t] = by_type.get(t, 0) + 1
    
    return {
        "period": period,
        "total_workouts": total_workouts,
        "total_duration_minutes": total_duration,
        "total_distance_km": round(total_distance, 2),
        "total_calories": total_calories,
        "total_xp_earned": total_xp,
        "by_activity_type": by_type
    }

@api_router.get("/workout-stats/detailed")
async def get_detailed_workout_stats(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get detailed workout statistics for charts"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Get last 30 days of workouts
    today = datetime.now(timezone.utc)
    start_date = (today - timedelta(days=30)).strftime("%Y-%m-%d")
    
    workouts = await db.workout_logs.find({
        "user_id": user.user_id,
        "date": {"$gte": start_date},
        "completed": True
    }, {"_id": 0}).sort("date", 1).to_list(1000)
    
    # Daily workout data for chart
    daily_data = {}
    for i in range(30):
        date = (today - timedelta(days=29-i)).strftime("%Y-%m-%d")
        daily_data[date] = {"duration": 0, "calories": 0, "count": 0}
    
    for w in workouts:
        date = w['date']
        if date in daily_data:
            daily_data[date]['duration'] += w.get('duration_minutes', 0)
            daily_data[date]['calories'] += w.get('calories', 0) or 0
            daily_data[date]['count'] += 1
    
    # Calculate streak
    current_streak = 0
    best_streak = 0
    temp_streak = 0
    
    for i in range(30):
        date = (today - timedelta(days=i)).strftime("%Y-%m-%d")
        if daily_data.get(date, {}).get('count', 0) > 0:
            temp_streak += 1
            if i == 0 or (i > 0 and temp_streak > 0):
                current_streak = temp_streak
            best_streak = max(best_streak, temp_streak)
        else:
            if i == 0:
                current_streak = 0
            temp_streak = 0
    
    # Weekly consistency (how many days trained per week)
    weekly_consistency = {}
    for date, data in daily_data.items():
        week = datetime.strptime(date, "%Y-%m-%d").isocalendar()[1]
        if week not in weekly_consistency:
            weekly_consistency[week] = 0
        if data['count'] > 0:
            weekly_consistency[week] += 1
    
    # Calculate averages
    trained_days = sum(1 for d in daily_data.values() if d['count'] > 0)
    avg_duration = sum(d['duration'] for d in daily_data.values()) / max(trained_days, 1)
    avg_calories = sum(d['calories'] for d in daily_data.values()) / max(trained_days, 1)
    
    return {
        "daily_data": [{"date": k, **v} for k, v in sorted(daily_data.items())],
        "current_streak": current_streak,
        "best_streak": best_streak,
        "trained_days": trained_days,
        "total_days": 30,
        "consistency_percentage": round((trained_days / 30) * 100, 1),
        "avg_duration_minutes": round(avg_duration, 1),
        "avg_calories": round(avg_calories, 1),
        "weekly_consistency": weekly_consistency
    }

@api_router.get("/workout-stats/exercise-evolution")
async def get_exercise_evolution(request: Request, exercise_name: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    """Get weight/reps evolution history per exercise"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Aggregate from completed workout sessions
    sessions = await db.workout_sessions.find({
        "user_id": user.user_id,
        "status": "completed"
    }, {"_id": 0, "exercises": 1, "completed_at": 1, "plan_name": 1}).sort("completed_at", 1).to_list(500)
    
    # Aggregate from workout logs as well
    logs = await db.workout_logs.find({
        "user_id": user.user_id,
        "completed": True
    }, {"_id": 0, "exercises_completed": 1, "date": 1, "name": 1}).sort("date", 1).to_list(500)
    
    # Collect per-exercise data
    evolution = {}
    
    for s in sessions:
        date = s.get("completed_at", "")[:10]
        for ex in s.get("exercises", []):
            name = ex.get("name", "").strip()
            if not name:
                continue
            if exercise_name and exercise_name.lower() not in name.lower():
                continue
            if name not in evolution:
                evolution[name] = []
            
            # Get per-set data or use top-level values
            sets_data = ex.get("sets_data")
            if sets_data and isinstance(sets_data, list):
                for sd in sets_data:
                    if sd.get("completed"):
                        evolution[name].append({
                            "date": date,
                            "weight": sd.get("weight", ""),
                            "reps": sd.get("reps", 0),
                            "source": "session",
                            "plan_name": s.get("plan_name", "")
                        })
            else:
                evolution[name].append({
                    "date": date,
                    "weight": ex.get("weight", ""),
                    "reps": ex.get("reps", 0),
                    "sets": ex.get("sets_completed", 0),
                    "source": "session",
                    "plan_name": s.get("plan_name", "")
                })
    
    for w in logs:
        date = w.get("date", "")
        for ex in w.get("exercises_completed", []):
            name = ex.get("name", "").strip()
            if not name:
                continue
            if exercise_name and exercise_name.lower() not in name.lower():
                continue
            if name not in evolution:
                evolution[name] = []
            evolution[name].append({
                "date": date,
                "weight": ex.get("weight", ""),
                "reps": ex.get("reps", 0),
                "sets": ex.get("sets_completed", 0),
                "source": "log",
                "plan_name": w.get("name", "")
            })
    
    # Sort each exercise's data by date
    for name in evolution:
        evolution[name].sort(key=lambda x: x["date"])
    
    return {
        "exercises": evolution,
        "exercise_names": sorted(evolution.keys()) if not exercise_name else [exercise_name]
    }


@api_router.post("/workout-suggestions")
async def get_ai_workout_suggestions(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get AI-powered workout suggestions based on user's history and goals"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Get recent workouts
    today = datetime.now(timezone.utc)
    start_date = (today - timedelta(days=30)).strftime("%Y-%m-%d")
    
    workouts = await db.workout_logs.find({
        "user_id": user.user_id,
        "date": {"$gte": start_date}
    }, {"_id": 0}).to_list(100)
    
    # Get body measurements
    measurements = await db.body_measurements.find(
        {"user_id": user.user_id}, {"_id": 0}
    ).sort("date", -1).limit(1).to_list(1)
    
    latest_measurement = measurements[0] if measurements else None
    
    # Build prompt
    workout_summary = {}
    for w in workouts:
        t = w['activity_type']
        workout_summary[t] = workout_summary.get(t, 0) + 1
    
    prompt = f"""Com base no histórico de treinos e dados do usuário, sugira um plano de treino personalizado.

HISTÓRICO DE TREINOS (últimos 30 dias):
- Total de treinos: {len(workouts)}
- Por tipo: {json.dumps(workout_summary, indent=2)}

"""
    
    if latest_measurement:
        prompt += f"""MEDIDAS CORPORAIS:
- Peso: {latest_measurement.get('weight_kg', 'N/A')} kg
- Altura: {latest_measurement.get('height_cm', 'N/A')} cm
- Gordura corporal: {latest_measurement.get('body_fat_percentage', 'N/A')}%
- Massa muscular: {latest_measurement.get('muscle_mass_kg', 'N/A')} kg

"""
    
    prompt += """Por favor, forneça:
1. Análise do perfil de treino atual
2. Sugestão de treino para a próxima semana (com exercícios específicos)
3. Dicas de intensidade e progressão
4. Recomendações de descanso e recuperação
5. Sugestões de nutrição pré e pós-treino

Responda em português de forma prática e motivadora."""

    try:
        response = await call_llm(
            prompt=prompt,
            session_id=f"workout_suggestions_{user.user_id}",
            system_message="Você é um personal trainer experiente e nutricionista esportivo. Forneça sugestões personalizadas e práticas.",
            user_id=user.user_id
        )
        
        return {
            "suggestions": response,
            "based_on": {
                "total_workouts": len(workouts),
                "workout_types": workout_summary,
                "has_measurements": latest_measurement is not None
            }
        }
        
    except Exception as e:
        logging.error(f"Workout suggestions failed: {e}")
        raise HTTPException(status_code=500, detail="Não foi possível gerar sugestões. Tente novamente.")

# ========== AI WORKOUT GENERATION ==========
@api_router.post("/workout-plans/generate")
async def generate_workout_plan(request: Request, gen_data: WorkoutPlanGenerate, session_token: Optional[str] = Cookie(None)):
    """Generate a workout plan with AI including tutorials and YouTube video links"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Build health condition text if provided
    health_text = ""
    if gen_data.health_condition and gen_data.health_condition.strip():
        health_text = f"""
CONDIÇÃO DE SAÚDE / LESÕES DO USUÁRIO:
{gen_data.health_condition.strip()}

ATENÇÃO: Adapte TODOS os exercícios considerando esta condição. Evite exercícios que possam agravar a lesão/condição.
Inclua exercícios de fortalecimento e reabilitação quando apropriado.
Para cada exercício, adicione um campo "health_notes" com observações específicas sobre como adaptar o exercício à condição do usuário.
Se algum exercício for contraindicado, substitua por uma alternativa segura e explique por quê.
"""
        # Save health condition to user profile for future use
        await db.users.update_one(
            {"user_id": user.user_id},
            {"$set": {"health_condition": gen_data.health_condition.strip()}}
        )
    
    # ===== BUILD PROMPT BASED ON GENERATION MODE =====
    if gen_data.generation_mode == "tipo_treino" and gen_data.split_config:
        # --- SPLIT-BASED GENERATION (Tipo de Treino) ---
        # Strategy: Generate only BASE SPLITS (A, B, C...) + weekly progression notes
        # Then expand to full days on the server side to avoid huge AI responses
        split_description = []
        for split in gen_data.split_config:
            label = split.get("label", "?")
            name = split.get("name", "")
            groups = split.get("muscle_groups", [])
            split_description.append(f"  Treino {label}: {name} (Grupos: {', '.join(groups)})")
        split_text = "\n".join(split_description)
        
        days_per_week = gen_data.training_days_per_week or 5
        cycle_weeks = gen_data.cycle_weeks or 4
        split_type = gen_data.split_type or "ABC"
        split_labels = [s.get("label", "") for s in gen_data.split_config]
        
        cardio_text = ""
        if gen_data.include_cardio:
            cardio_name = gen_data.cardio_type or "corrida"
            cardio_labels = {
                "corrida": "Corrida", "bike": "Bike/Ciclismo", "HIIT": "HIIT",
                "caminhada": "Caminhada", "natacao": "Natação", "pular_corda": "Pular Corda",
                "eliptico": "Elíptico", "remo": "Remo"
            }
            cardio_display = cardio_labels.get(cardio_name, cardio_name)
            cardio_mode = gen_data.cardio_mode or "hibrido"
            
            if cardio_mode == "hibrido":
                cardio_text = f"""
TREINO HÍBRIDO (Musculação + Cardio no mesmo dia):
- NÃO crie split separada de cardio.
- Em CADA split de musculação, adicione 2-3 exercícios de cardio ({cardio_display}) ao FINAL da lista de exercícios.
- Estes exercícios finais devem ter muscle_group: "cardio" e seguir o formato padrão.
- Varie a intensidade do cardio entre os dias: em um split use cardio de alta intensidade (tiros, HIIT), em outro use cardio de baixa/moderada intensidade (zona 2, ritmo constante).
- Exemplo alta intensidade: {{"name": "{cardio_display} - Tiros (HIIT)", "sets": 1, "reps": "15-20min", "rest_seconds": 0, "muscle_group": "cardio", "tutorial": "Alterne 30seg de alta intensidade com 60seg de recuperação. Total 15-20 minutos."}}
- Exemplo baixa intensidade: {{"name": "{cardio_display} - Zona 2 (moderado)", "sets": 1, "reps": "20-30min", "rest_seconds": 0, "muscle_group": "cardio", "tutorial": "Mantenha ritmo constante e confortável, onde consiga conversar. Frequência cardíaca em zona 2."}}
- O ÚLTIMO dia da semana deve ser um dia de DESCANSO (split_label: "Descanso", com exercícios leves ou nenhum exercício).
"""
            elif cardio_mode == "hibrido_alternado":
                cardio_text = f"""
TREINO HÍBRIDO ALTERNADO (1 dia musculação, 1 dia cardio):
- Adicione um item extra no array "splits" com split_label: "Cardio", split_name: "{cardio_display}".
- O cardio DEVE usar o mesmo formato "exercises": name, sets (1), reps (duração), rest_seconds, muscle_group ("cardio"), tutorial.
- VARIE a intensidade do cardio entre os dias: alterne entre alta intensidade (tiros/HIIT) e baixa/moderada intensidade (zona 2).
- Inclua 4-5 exercícios por dia de cardio: aquecimento, blocos de intensidade variada, desaquecimento.
- Exemplo para dia de alta intensidade: {{"name": "{cardio_display} - Tiros", "sets": 1, "reps": "30seg sprint + 60seg descanso x8", "rest_seconds": 0, "muscle_group": "cardio", "tutorial": "Sprint máximo por 30 segundos, descanse 60 segundos caminhando. Repita 8 vezes."}}
- Exemplo para dia de zona 2: {{"name": "{cardio_display} - Zona 2", "sets": 1, "reps": "30-40min", "rest_seconds": 0, "muscle_group": "cardio", "tutorial": "Ritmo constante onde consegue manter uma conversa. Foco em resistência aeróbica."}}
- O ÚLTIMO dia da semana DEVE ser de DESCANSO absoluto ou ativo (caminhada leve de 20-30min). Use split_label: "Descanso".
- Padrão ideal: Seg musculação, Ter cardio alta intensidade, Qua musculação, Qui cardio zona 2, Sex musculação, Sáb cardio moderado, Dom descanso.
"""

        prompt = f"""Você é um personal trainer certificado. Gere APENAS os treinos BASE de cada divisão ({split_type}) em formato JSON compacto.

NÃO gere todos os dias do ciclo. Gere apenas 1 treino para cada letra da divisão + notas de progressão semanal.

PARÂMETROS:
- Objetivo: {gen_data.objective}
- Nível: {gen_data.level}
- Divisão: {split_type} ({len(split_labels)} treinos base)
- Ciclo: {cycle_weeks} semana(s), {days_per_week} dias/semana
{health_text}
DIVISÕES:
{split_text}
{cardio_text}

FORMATO JSON OBRIGATÓRIO:
{{
  "name": "Treino {split_type} - {gen_data.objective.capitalize()}",
  "description": "Descrição breve",
  "plan_duration": "ciclo",
  "split_type": "{split_type}",
  "cycle_weeks": {cycle_weeks},
  "training_days_per_week": {days_per_week},
  "splits": [
    {{
      "split_label": "A",
      "split_name": "Peito e Tríceps",
      "exercises": [
        {{
          "name": "Supino Reto",
          "sets": 4,
          "reps": 10,
          "weight": "adequado",
          "rest_seconds": 90,
          "muscle_group": "peito",
          "tutorial": "Instrução concisa de execução em 1-2 frases."
        }}
      ]
    }}
  ],
  "weekly_progression": [
    {{
      "week": 1,
      "focus": "Adaptação e técnica",
      "notes": "Carga moderada, foco na execução correta"
    }},
    {{
      "week": 2,
      "focus": "Aumento de volume",
      "notes": "Aumente 1-2 reps por exercício"
    }}
  ]
}}

REGRAS:
- Retorne APENAS JSON válido, sem markdown, sem texto extra.
- Gere 1 treino por letra ({', '.join(split_labels)}).
- Tutorial: máximo 2 frases curtas por exercício.
- {len(split_labels) * 5} a {len(split_labels) * 7} exercícios no total (4-6 por split para iniciante, 5-7 intermediário, 6-8 avançado).
- Gere {cycle_weeks} itens em weekly_progression.
- rest_seconds: 60s leves, 90s moderados, 120s compostos pesados."""

    else:
        # --- PERIOD-BASED GENERATION (existing flow) ---
        duration_instructions = {
            "dia": "Crie um treino para UM DIA ÚNICO. Liste os exercícios em um único bloco.",
            "semana": "Crie um treino para UMA SEMANA COMPLETA (segunda a sexta, 5 dias). Organize por dia da semana com exercícios diferentes para cada dia, alternando grupos musculares.",
            "mes": "Crie um plano de treino para UM MÊS (4 semanas). Organize em 4 semanas com progressão de carga/volume. Cada semana deve ter 5 dias de treino.",
            "ciclo": "Crie um ciclo de treino periodizado (8-12 semanas). Organize em fases: Adaptação (2 semanas), Hipertrofia (4 semanas), Força (3 semanas), Deload (1 semana). Cada fase com treinos específicos."
        }
        
        muscle_groups_text = ""
        if gen_data.muscle_groups and len(gen_data.muscle_groups) > 0:
            muscle_groups_text = f"\nGrupos musculares prioritários: {', '.join(gen_data.muscle_groups)}"
        
        prompt = f"""Você é um personal trainer certificado. Gere um plano de treino completo em formato JSON.

PARÂMETROS:
- Objetivo: {gen_data.objective}
- Nível: {gen_data.level}
- Duração: {gen_data.duration}{muscle_groups_text}
{health_text}
{duration_instructions.get(gen_data.duration, duration_instructions['dia'])}

Para CADA exercício, inclua um tutorial descritivo de execução em 1-2 frases curtas.

FORMATO JSON OBRIGATÓRIO:
{{
  "name": "Nome do plano de treino",
  "description": "Descrição breve do objetivo",
  "plan_duration": "{gen_data.duration}",
  "days": [
    {{
      "day_name": "dia1",
      "day_label": "Segunda - Peito e Tríceps",
      "exercises": [
        {{
          "name": "Supino Reto com Barra",
          "sets": 4,
          "reps": 10,
          "weight": "adequado ao nível",
          "rest_seconds": 90,
          "muscle_group": "peito",
          "tutorial": "Deite no banco, desça a barra ao peito controladamente e empurre para cima. Inspire ao descer, expire ao subir."
        }}
      ]
    }}
  ]
}}

{"Retorne apenas 1 dia no array 'days'." if gen_data.duration == "dia" else ""}
{"Retorne 5 dias (seg-sex) no array 'days'." if gen_data.duration == "semana" else ""}
{"Organize semanas: dias como 'sem1_dia1', 'sem1_dia2', etc." if gen_data.duration == "mes" else ""}
{"Organize por fases: dias como 'fase1_sem1_dia1', etc." if gen_data.duration == "ciclo" else ""}

IMPORTANTE: 
- Retorne APENAS JSON válido, sem markdown, sem texto extra.
- Tutorial: máximo 2 frases curtas por exercício.
- Adapte ao nível ({gen_data.level}).
- rest_seconds: 60s leves, 90s moderados, 120s compostos pesados."""

    # Helper to clean and parse JSON from AI response
    def _clean_and_parse_json(text: str) -> dict:
        """Robustly clean and parse JSON from AI response text"""
        import re
        cleaned = text.strip()
        # Remove markdown code blocks
        if cleaned.startswith("```"):
            cleaned = cleaned.split("\n", 1)[1] if "\n" in cleaned else cleaned[3:]
        if cleaned.endswith("```"):
            cleaned = cleaned[:-3].strip()
        if cleaned.startswith("json"):
            cleaned = cleaned[4:].strip()
        # Remove any trailing text after the last }
        last_brace = cleaned.rfind("}")
        if last_brace != -1 and last_brace < len(cleaned) - 1:
            cleaned = cleaned[:last_brace + 1]
        # Try to find JSON object if there's leading text
        first_brace = cleaned.find("{")
        if first_brace > 0:
            cleaned = cleaned[first_brace:]
        # Fix common JSON issues: trailing commas before } or ]
        cleaned = re.sub(r',\s*}', '}', cleaned)
        cleaned = re.sub(r',\s*]', ']', cleaned)
        return json.loads(cleaned)

    # Call LLM (Gemini via HTTP + TorGPT fallback)
    plan_data = None
    try:
        logging.info("Generating workout plan via LLM")
        response_text = await call_llm(
            prompt + "\n\nResponda APENAS com JSON puro, sem markdown, sem texto extra.",
            f"workout_{user.user_id}",
            "Você é um personal trainer profissional certificado. Sempre responda SOMENTE em JSON válido, sem nenhum texto adicional.",
            user_id=user.user_id
        )
        
        if not response_text or not response_text.strip():
            raise HTTPException(status_code=500, detail="Resposta vazia da IA. Tente novamente.")
        
        response_text = response_text.strip()
        if response_text.startswith("⚠"):
            raise HTTPException(status_code=500, detail=response_text)
        
        plan_data = _clean_and_parse_json(response_text)
        logging.info("Successfully parsed workout plan")
        
    except json.JSONDecodeError as e:
        logging.error(f"JSON parse error: {e}")
        logging.error(f"Response text (first 500 chars): {response_text[:500] if response_text else 'N/A'}")
        raise HTTPException(status_code=500, detail="Erro ao processar resposta da IA. Tente novamente.")
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Workout generation failed: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao gerar treino: {str(e)}. Tente novamente.")
    
    try:
        # Create the plan document
        plan_id = f"plan_{uuid.uuid4().hex[:12]}"
        
        # For tipo_treino: expand splits into days on the server side
        days = []
        weekly_progression = plan_data.get("weekly_progression", [])
        
        if gen_data.generation_mode == "tipo_treino" and plan_data.get("splits"):
            splits = plan_data["splits"]
            split_labels_ai = [s.get("split_label", f"S{i}") for i, s in enumerate(splits)]
            days_per_week = gen_data.training_days_per_week or 5
            cycle_weeks_count = gen_data.cycle_weeks or 4
            
            # Build rotation pattern using split labels from AI response
            # Filter out cardio and rest splits for the main rotation
            main_splits = [s for s in splits if s.get("split_label", "").lower() not in ("cardio", "descanso")]
            cardio_split = next((s for s in splits if s.get("split_label", "").lower() == "cardio"), None)
            rest_split = next((s for s in splits if s.get("split_label", "").lower() == "descanso"), None)
            
            muscle_day_counter = 0
            cardio_mode = gen_data.cardio_mode or "hibrido"
            
            for week in range(1, cycle_weeks_count + 1):
                week_progression = next((wp for wp in weekly_progression if wp.get("week") == week), None)
                progression_note = week_progression.get("notes", "") if week_progression else ""
                progression_focus = week_progression.get("focus", "") if week_progression else ""
                
                for day_in_week in range(1, days_per_week + 1):
                    is_rest_day = False
                    is_cardio_day = False
                    
                    if cardio_mode == "hibrido_alternado" and gen_data.include_cardio and cardio_split:
                        # Pattern: Muscle, Cardio, Muscle, Cardio, ..., Rest (last day)
                        if day_in_week == days_per_week:
                            is_rest_day = True
                        elif day_in_week % 2 == 0:
                            is_cardio_day = True
                        # Odd days (1, 3, 5, ...) are muscle days
                    elif cardio_mode == "hibrido" and gen_data.include_cardio:
                        # Hybrid mode: last day is rest, all others are muscle+cardio
                        if day_in_week == days_per_week:
                            is_rest_day = True
                    
                    if is_rest_day:
                        if rest_split:
                            current_split = rest_split
                        else:
                            current_split = {"exercises": [{"name": "Descanso ativo - Caminhada leve", "sets": 1, "reps": "20-30min", "rest_seconds": 0, "muscle_group": "descanso", "tutorial": "Caminhada leve para recuperação ativa. Mantenha ritmo tranquilo."}]}
                        label = "Descanso"
                        split_name = "Descanso / Recuperação"
                    elif is_cardio_day and cardio_split:
                        current_split = cardio_split
                        label = "Cardio"
                        split_name = cardio_split.get("split_name", "Cardio")
                    else:
                        split_idx = muscle_day_counter % len(main_splits) if main_splits else 0
                        current_split = main_splits[split_idx] if main_splits else {"exercises": []}
                        label = current_split.get("split_label", "?")
                        split_name = current_split.get("split_name", "")
                        muscle_day_counter += 1
                    
                    day_label = f"Semana {week} - Dia {day_in_week}: Treino {label} - {split_name}"
                    
                    days.append({
                        "day_name": f"sem{week}_dia{day_in_week}",
                        "day_label": day_label,
                        "split_label": label,
                        "week": week,
                        "exercises": current_split.get("exercises", []),
                        "progression_focus": progression_focus,
                        "progression_notes": progression_note,
                    })
        else:
            # Period-based: days come directly from AI response
            days = plan_data.get("days", [])
        
        # Flatten exercises for backward compatibility
        all_exercises = []
        for day in days:
            for ex in day.get("exercises", []):
                all_exercises.append(ex)
        
        # Determine plan_duration
        plan_duration = gen_data.duration
        if gen_data.generation_mode == "tipo_treino":
            plan_duration = "ciclo"
        
        plan_doc = {
            "plan_id": plan_id,
            "user_id": user.user_id,
            "name": plan_data.get("name", f"Treino {gen_data.objective} - {gen_data.level}"),
            "description": plan_data.get("description", ""),
            "exercises": all_exercises,
            "plan_duration": plan_duration,
            "generated_by_ai": True,
            "days": days,
            "weekly_progression": weekly_progression,
            "objective": gen_data.objective,
            "level": gen_data.level,
            "generation_mode": gen_data.generation_mode,
            "split_type": gen_data.split_type if gen_data.generation_mode == "tipo_treino" else None,
            "split_config": gen_data.split_config if gen_data.generation_mode == "tipo_treino" else None,
            "training_days_per_week": gen_data.training_days_per_week if gen_data.generation_mode == "tipo_treino" else None,
            "cycle_weeks": gen_data.cycle_weeks if gen_data.generation_mode == "tipo_treino" else None,
            "include_cardio": gen_data.include_cardio if gen_data.generation_mode == "tipo_treino" else False,
            "cardio_type": gen_data.cardio_type if gen_data.generation_mode == "tipo_treino" and gen_data.include_cardio else None,
            "cardio_mode": gen_data.cardio_mode if gen_data.generation_mode == "tipo_treino" and gen_data.include_cardio else None,
            "health_condition": gen_data.health_condition if gen_data.health_condition else None,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.workout_plans.insert_one(plan_doc)
        plan_doc.pop('_id', None)
        
        # Award XP for generating a plan
        xp_earned = 5
        new_xp = user.xp + xp_earned
        new_rank = calculate_rank(new_xp)
        await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
        
        plan_doc['created_at'] = datetime.fromisoformat(plan_doc['created_at'])
        
        return {
            "success": True,
            "plan": plan_doc,
            "xp_earned": xp_earned,
            "new_xp": new_xp,
            "new_rank": new_rank
        }
        
    except Exception as e:
        logging.error(f"Failed to save workout plan: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao salvar treino: {str(e)[:100]}")



# ========== IMPROVE WORKOUT (MELHORAR TREINO) ==========
@api_router.post("/workout-plans/{plan_id}/improve")
async def improve_workout_plan(request: Request, plan_id: str, session_token: Optional[str] = Cookie(None)):
    """AI analyzes a completed workout plan and generates an improved version"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Get the original plan
    plan = await db.workout_plans.find_one({"plan_id": plan_id, "user_id": user.user_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Plano não encontrado")
    
    # Get workout logs/sessions for this plan to understand user performance
    sessions = await db.workout_sessions.find(
        {"user_id": user.user_id, "plan_id": plan_id},
        {"_id": 0}
    ).to_list(100)
    
    daily_statuses = await db.daily_workout_status.find(
        {"user_id": user.user_id, "plan_id": plan_id},
        {"_id": 0}
    ).to_list(100)
    
    # Build performance summary
    completed_sessions = [s for s in sessions if s.get("status") == "completed"]
    total_sessions = len(sessions)
    
    performance_summary = f"Total de sessões: {total_sessions}, Completadas: {len(completed_sessions)}"
    if completed_sessions:
        avg_duration = sum(s.get("duration_minutes", 0) for s in completed_sessions) / len(completed_sessions)
        performance_summary += f", Duração média: {avg_duration:.0f} min"
    
    completed_days = len([d for d in daily_statuses if d.get("completed")])
    performance_summary += f", Dias concluídos: {completed_days}"
    
    # Build exercises summary from the plan
    exercises_summary = []
    splits_data = []
    if plan.get("generation_mode") == "tipo_treino" and plan.get("split_config"):
        for s in plan["split_config"]:
            split_exercises = []
            for day in plan.get("days", []):
                if day.get("split_label") == s.get("label"):
                    split_exercises = day.get("exercises", [])
                    break
            splits_data.append({
                "label": s.get("label"),
                "name": s.get("name"),
                "exercises": [{"name": ex.get("name"), "sets": ex.get("sets"), "reps": ex.get("reps"), "muscle_group": ex.get("muscle_group")} for ex in split_exercises[:8]]
            })
    else:
        for day in plan.get("days", [])[:5]:
            exercises_summary.append({
                "day": day.get("day_label", ""),
                "exercises": [{"name": ex.get("name"), "sets": ex.get("sets"), "reps": ex.get("reps")} for ex in day.get("exercises", [])[:8]]
            })

    plan_info = f"""
Plano: {plan.get('name', 'Treino')}
Objetivo: {plan.get('objective', 'hipertrofia')}
Nível: {plan.get('level', 'intermediario')}
Tipo: {plan.get('generation_mode', 'periodo')}
Divisão: {plan.get('split_type', 'N/A')}
Duração: {plan.get('plan_duration', 'ciclo')}
Semanas: {plan.get('cycle_weeks', 'N/A')}
Dias/semana: {plan.get('training_days_per_week', 'N/A')}
Performance: {performance_summary}
"""
    
    if splits_data:
        for sd in splits_data:
            plan_info += f"\nTreino {sd['label']} ({sd['name']}): {json.dumps([e['name'] for e in sd['exercises']], ensure_ascii=False)}"
    elif exercises_summary:
        for es in exercises_summary[:3]:
            plan_info += f"\n{es['day']}: {json.dumps([e['name'] for e in es['exercises']], ensure_ascii=False)}"
    
    is_split_mode = plan.get("generation_mode") == "tipo_treino"
    format_type = "splits" if is_split_mode else "days"
    format_instruction = "Gere no formato de splits (mesmo formato do plano atual)." if is_split_mode else "Gere no formato de days."
    
    if is_split_mode:
        item_example = '"split_label": "A", "split_name": "Nome",'
    else:
        item_example = '"day_name": "dia1", "day_label": "Nome do dia",'
    
    num_splits = len(splits_data) if splits_data else "a mesma quantidade de"
    cycle_wk_count = plan.get('cycle_weeks', 4)
    
    prompt = f"""Você é um personal trainer experiente. O aluno completou um ciclo de treino e precisa de EVOLUÇÃO.

PLANO ATUAL:
{plan_info}

Como um personal trainer faria, analise o treino atual e gere uma VERSÃO MELHORADA:
1. Substitua exercícios que podem ter estagnado por variações mais desafiadoras
2. Aumente volume ou intensidade gradualmente (mais séries, mais reps, ou mais carga)
3. Adicione exercícios complementares ou variações
4. Mantenha a estrutura geral (mesma divisão) mas evolua o conteúdo
5. Diversifique os estímulos musculares

{format_instruction}

FORMATO JSON OBRIGATÓRIO (use a chave "{format_type}"):
{{
  "name": "Nome do plano evoluído (v2, v3, etc)",
  "description": "Descrição das mudanças e evolução aplicada",
  "improvements_summary": "Resumo das melhorias em 2-3 frases",
  "{format_type}": [
    {{
      {item_example}
      "exercises": [
        {{
          "name": "Nome do exercício",
          "sets": 4,
          "reps": 12,
          "weight": "adequado",
          "rest_seconds": 90,
          "muscle_group": "grupo",
          "tutorial": "Instrução concisa em 1-2 frases."
        }}
      ]
    }}
  ],
  "weekly_progression": [
    {{"week": 1, "focus": "Foco", "notes": "Notas"}}
  ]
}}

REGRAS:
- Retorne APENAS JSON válido.
- Tutorial: máximo 2 frases por exercício.
- Mantenha {num_splits} splits/dias.
- Substitua pelo menos 30-40% dos exercícios por variações.
- rest_seconds: 60s leves, 90s moderados, 120s compostos.
- Gere {cycle_wk_count} itens em weekly_progression."""
    
    system_msg = "Você é um personal trainer certificado. Retorne apenas JSON válido sem markdown."
    
    try:
        response_text = None
        last_error = None
        
        for attempt in range(3):
            try:
                model_to_use = GEMINI_MODEL if attempt < 2 else GEMINI_FALLBACK_MODEL
                ai_response = gemini_client.models.generate_content(
                    model=model_to_use,
                    contents=prompt,
                    config=types.GenerateContentConfig(
                        system_instruction=system_msg,
                        temperature=0.7,
                        max_output_tokens=16384,
                    )
                )
                response_text = ai_response.text.strip()
                if response_text.startswith("```"):
                    response_text = response_text.split("\n", 1)[1] if "\n" in response_text else response_text[3:]
                    response_text = response_text.rsplit("```", 1)[0].strip()
                
                improved_data = json.loads(response_text)
                break
            except Exception as e:
                last_error = str(e)
                continue
        else:
            raise HTTPException(status_code=500, detail=f"Erro ao melhorar treino: {last_error}")
        
        # Create a new plan from the improved data
        new_plan_id = f"plan_{uuid.uuid4().hex[:12]}"
        
        # Build days from improved splits or use days directly
        new_days = []
        new_weekly_progression = improved_data.get("weekly_progression", [])
        
        if plan.get("generation_mode") == "tipo_treino" and improved_data.get("splits"):
            new_splits = improved_data["splits"]
            main_splits = [s for s in new_splits if s.get("split_label", "").lower() != "cardio"]
            cardio_split = next((s for s in new_splits if s.get("split_label", "").lower() == "cardio"), None)
            days_per_week = plan.get("training_days_per_week", 5)
            cycle_wks = plan.get("cycle_weeks", 4)
            day_counter = 0
            
            for week in range(1, cycle_wks + 1):
                wp = next((w for w in new_weekly_progression if w.get("week") == week), None)
                for day_in_week in range(1, days_per_week + 1):
                    split_idx = day_counter % len(main_splits)
                    current = main_splits[split_idx]
                    label = current.get("split_label", "?")
                    sname = current.get("split_name", "")
                    day_counter += 1
                    
                    new_days.append({
                        "day_name": f"sem{week}_dia{day_in_week}",
                        "day_label": f"Semana {week} - Dia {day_in_week}: Treino {label} - {sname}",
                        "split_label": label,
                        "week": week,
                        "exercises": current.get("exercises", []),
                        "progression_focus": wp.get("focus", "") if wp else "",
                        "progression_notes": wp.get("notes", "") if wp else "",
                    })
        else:
            new_days = improved_data.get("days", [])
        
        all_exercises = []
        for d in new_days:
            all_exercises.extend(d.get("exercises", []))
        
        new_plan_doc = {
            "plan_id": new_plan_id,
            "user_id": user.user_id,
            "name": improved_data.get("name", f"{plan.get('name', 'Treino')} - Evoluído"),
            "description": improved_data.get("description", ""),
            "exercises": all_exercises,
            "plan_duration": plan.get("plan_duration", "ciclo"),
            "generated_by_ai": True,
            "improved_from": plan_id,
            "improvements_summary": improved_data.get("improvements_summary", ""),
            "days": new_days,
            "weekly_progression": new_weekly_progression,
            "objective": plan.get("objective", "hipertrofia"),
            "level": plan.get("level", "intermediario"),
            "generation_mode": plan.get("generation_mode", "tipo_treino"),
            "split_type": plan.get("split_type"),
            "split_config": plan.get("split_config"),
            "training_days_per_week": plan.get("training_days_per_week"),
            "cycle_weeks": plan.get("cycle_weeks"),
            "include_cardio": plan.get("include_cardio", False),
            "cardio_type": plan.get("cardio_type"),
            "cardio_mode": plan.get("cardio_mode"),
            "health_condition": plan.get("health_condition"),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        new_plan_doc['created_at'] = datetime.fromisoformat(new_plan_doc['created_at'])
        await db.workout_plans.insert_one(new_plan_doc)
        
        xp_earned = 3
        new_xp = user.xp + xp_earned
        new_rank = calculate_rank(new_xp)
        await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
        
        return {
            "success": True,
            "plan": {**new_plan_doc, "_id": None},
            "improvements_summary": improved_data.get("improvements_summary", ""),
            "xp_earned": xp_earned,
            "new_xp": new_xp,
            "new_rank": new_rank
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Failed to improve workout plan: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao melhorar treino: {str(e)[:100]}")



# ========== WORKOUT SESSION ENDPOINTS ==========
@api_router.post("/workout-sessions/start")
async def start_workout_session(request: Request, session_token: Optional[str] = Cookie(None)):
    """Start an active workout session from a plan"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    body = await request.json()
    plan_id = body.get("plan_id")
    day_index = body.get("day_index", 0)
    rest_timer_seconds = body.get("rest_timer_seconds", 60)
    
    if not plan_id:
        raise HTTPException(status_code=400, detail="plan_id é obrigatório")
    
    # Check for existing active session
    active = await db.workout_sessions.find_one({
        "user_id": user.user_id, 
        "status": "active"
    }, {"_id": 0})
    if active:
        raise HTTPException(status_code=409, detail="Já existe uma sessão de treino ativa. Finalize ou abandone a sessão atual.")
    
    # Get the plan
    plan = await db.workout_plans.find_one({"plan_id": plan_id, "user_id": user.user_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Plano de treino não encontrado")
    
    # Get exercises for the session (either from specific day or all exercises)
    days = plan.get("days") or []
    if days and day_index < len(days):
        session_exercises = days[day_index].get("exercises", [])
        day_label = days[day_index].get("day_label", f"Dia {day_index + 1}")
    else:
        session_exercises = plan.get("exercises", [])
        day_label = plan.get("name", "Treino")
    
    # Prepare exercises with tracking fields
    exercises = []
    for ex in session_exercises:
        exercises.append({
            "name": ex.get("name", ""),
            "sets": ex.get("sets", 3),
            "reps": ex.get("reps", 12),
            "weight": ex.get("weight", ""),
            "rest_seconds": ex.get("rest_seconds", rest_timer_seconds),
            "muscle_group": ex.get("muscle_group", ""),
            "tutorial": ex.get("tutorial", ""),
            "video_url": ex.get("video_url", ""),
            "completed": False,
            "sets_completed": 0,
            "time_spent_seconds": 0
        })
    
    session_id = f"session_{uuid.uuid4().hex[:12]}"
    session_doc = {
        "session_id": session_id,
        "user_id": user.user_id,
        "plan_id": plan_id,
        "plan_name": f"{plan.get('name', 'Treino')} - {day_label}",
        "status": "active",
        "started_at": datetime.now(timezone.utc).isoformat(),
        "completed_at": None,
        "total_duration_seconds": 0,
        "exercises": exercises,
        "current_exercise_idx": 0,
        "rest_timer_seconds": rest_timer_seconds,
        "feedback": None,
        "day_index": day_index
    }
    
    await db.workout_sessions.insert_one(session_doc)
    session_doc.pop('_id', None)
    
    return session_doc


@api_router.get("/workout-sessions/active")
async def get_active_session(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get current active workout session"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    session = await db.workout_sessions.find_one({
        "user_id": user.user_id,
        "status": "active"
    }, {"_id": 0})
    
    if not session:
        return {"active": False, "session": None}
    
    return {"active": True, "session": session}


@api_router.patch("/workout-sessions/{session_id}/exercise/{exercise_idx}")
async def update_session_exercise(request: Request, session_id: str, exercise_idx: int, session_token: Optional[str] = Cookie(None)):
    """Update exercise progress in active session"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    body = await request.json()
    
    session = await db.workout_sessions.find_one({
        "session_id": session_id, 
        "user_id": user.user_id,
        "status": "active"
    }, {"_id": 0})
    
    if not session:
        raise HTTPException(status_code=404, detail="Sessão não encontrada ou já finalizada")
    
    exercises = session.get("exercises", [])
    if exercise_idx < 0 or exercise_idx >= len(exercises):
        raise HTTPException(status_code=400, detail="Índice de exercício inválido")
    
    # Update exercise fields
    if "completed" in body:
        exercises[exercise_idx]["completed"] = body["completed"]
    if "sets_completed" in body:
        exercises[exercise_idx]["sets_completed"] = body["sets_completed"]
    if "sets_data" in body:
        exercises[exercise_idx]["sets_data"] = body["sets_data"]
        # Auto-calculate sets_completed from sets_data
        completed_sets = sum(1 for s in body["sets_data"] if s.get("completed"))
        exercises[exercise_idx]["sets_completed"] = completed_sets
    if "time_spent_seconds" in body:
        exercises[exercise_idx]["time_spent_seconds"] = body["time_spent_seconds"]
    if "weight" in body:
        exercises[exercise_idx]["weight"] = body["weight"]
    
    # Handle per-set RPE updates from sets_data
    if "sets_data" in body:
        for sd in body["sets_data"]:
            if "rpe" in sd:
                # already saved via sets_data above
                pass
    
    update_fields = {"exercises": exercises}
    if "current_exercise_idx" in body:
        update_fields["current_exercise_idx"] = body["current_exercise_idx"]
    
    await db.workout_sessions.update_one(
        {"session_id": session_id},
        {"$set": update_fields}
    )
    
    session["exercises"] = exercises
    if "current_exercise_idx" in body:
        session["current_exercise_idx"] = body["current_exercise_idx"]
    
    return session


@api_router.post("/workout-sessions/{session_id}/complete")
async def complete_workout_session(request: Request, session_id: str, session_token: Optional[str] = Cookie(None)):
    """Complete a workout session with feedback"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    body = await request.json()
    
    session = await db.workout_sessions.find_one({
        "session_id": session_id,
        "user_id": user.user_id,
        "status": "active"
    }, {"_id": 0})
    
    if not session:
        raise HTTPException(status_code=404, detail="Sessão não encontrada ou já finalizada")
    
    completed_at = datetime.now(timezone.utc).isoformat()
    started_at = session.get("started_at", completed_at)
    
    # Calculate total duration
    start_time = datetime.fromisoformat(started_at.replace('Z', '+00:00'))
    end_time = datetime.fromisoformat(completed_at.replace('Z', '+00:00'))
    total_duration_seconds = int((end_time - start_time).total_seconds())
    
    # Count completed exercises
    exercises = session.get("exercises", [])
    completed_count = sum(1 for ex in exercises if ex.get("completed"))
    total_count = len(exercises)
    
    # Build feedback
    feedback = {
        "difficulty": body.get("difficulty", 3),  # 1-5
        "feeling": body.get("feeling", ""),  # ótimo, bom, regular, cansado, exausto
        "notes": body.get("notes", ""),
        "completed_exercises": completed_count,
        "total_exercises": total_count
    }
    
    # Calculate XP
    base_xp = 10
    exercise_bonus = completed_count * 2
    duration_bonus = (total_duration_seconds // 900) * 5  # +5 XP every 15 min
    xp_earned = base_xp + exercise_bonus + duration_bonus
    
    # Update session
    await db.workout_sessions.update_one(
        {"session_id": session_id},
        {"$set": {
            "status": "completed",
            "completed_at": completed_at,
            "total_duration_seconds": total_duration_seconds,
            "feedback": feedback
        }}
    )
    
    # Also log as a workout
    log_id = f"workout_{uuid.uuid4().hex[:12]}"
    duration_minutes = max(1, total_duration_seconds // 60)
    workout_doc = {
        "log_id": log_id,
        "user_id": user.user_id,
        "plan_id": session.get("plan_id"),
        "activity_type": "weightlifting",
        "name": session.get("plan_name", "Treino"),
        "duration_minutes": duration_minutes,
        "calories": int(duration_minutes * 6),
        "exercises_completed": [
            {**ex, "completed": ex.get("completed", False)} 
            for ex in exercises
        ],
        "notes": feedback.get("notes", ""),
        "xp_earned": xp_earned,
        "completed": True,
        "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "session_id": session_id,
        "created_at": completed_at
    }
    await db.workout_logs.insert_one(workout_doc)
    
    # Award XP
    new_xp = user.xp + xp_earned
    new_rank = calculate_rank(new_xp)
    await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
    
    return {
        "success": True,
        "session_id": session_id,
        "total_duration_seconds": total_duration_seconds,
        "total_duration_minutes": duration_minutes,
        "completed_exercises": completed_count,
        "total_exercises": total_count,
        "xp_earned": xp_earned,
        "new_xp": new_xp,
        "new_rank": new_rank,
        "feedback": feedback
    }


@api_router.get("/workouts/next-loads")
async def get_next_workout_loads(request: Request, plan_id: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    """Calculate suggested next weights for each exercise based on last session performance."""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Find the plan
    query = {"user_id": user.user_id}
    if plan_id:
        query["plan_id"] = plan_id
    plan = await db.workout_plans.find_one(query, {"_id": 0})
    if not plan:
        return {"suggestions": []}
    
    # Get all exercises from all days
    all_exercises = []
    days = plan.get("days", [])
    if days:
        for day in days:
            for ex in day.get("exercises", []):
                all_exercises.append({"name": ex.get("name", ""), "day_label": day.get("day_label", ""), "sets": ex.get("sets", 0), "reps": ex.get("reps", ""), "current_weight": ex.get("weight", "")})
    else:
        for ex in plan.get("exercises", []):
            all_exercises.append({"name": ex.get("name", ""), "day_label": "", "sets": ex.get("sets", 0), "reps": ex.get("reps", ""), "current_weight": ex.get("weight", "")})
    
    # Get the most recent workout log
    latest_log = await db.workout_logs.find_one(
        {"user_id": user.user_id, "completed": True},
        sort=[("created_at", -1)]
    )
    
    suggestions = []
    for ex in all_exercises:
        name = ex.get("name", "")
        if not name:
            continue
        
        suggested = {**ex, "next_weight": None, "reason": "", "progress_possible": True}
        
        if latest_log:
            # Find this exercise in the latest log
            for logged_ex in latest_log.get("exercises_completed", []):
                if logged_ex.get("name", "").strip().lower() == name.strip().lower():
                    sets_data = logged_ex.get("sets_data", [])
                    if sets_data:
                        all_completed = all(s.get("completed", False) for s in sets_data)
                        if all_completed and len(sets_data) >= int(ex.get("sets", 1)):
                            current_weight = logged_ex.get("weight", "")
                            if isinstance(current_weight, (int, float)):
                                increment = 2.5 if current_weight > 20 else 1.0
                                suggested["next_weight"] = round(current_weight + increment, 1)
                                suggested["reason"] = f"Aumento de {increment}kg (completou {len(sets_data)}/{ex.get('sets', 0)} séries)"
                            else:
                                # Try parsing from sets_data
                                weights_with_reps = [(s.get("weight", 0), s.get("reps", 0)) for s in sets_data if s.get("completed")]
                                if weights_with_reps:
                                    avg_weight = sum(w for w, r in weights_with_reps) / len(weights_with_reps)
                                    increment = 2.5 if avg_weight > 20 else 1.0
                                    suggested["next_weight"] = round(avg_weight + increment, 1)
                                    suggested["reason"] = f"Completou todas as séries. Sugerido +{increment}kg"
                        else:
                            suggested["reason"] = "Mantenha o peso atual - ainda não completou todas as séries"
                            suggested["progress_possible"] = False
                            suggested["next_weight"] = current_weight if isinstance(current_weight, (int, float)) else None
                    break
        
        suggestions.append(suggested)
    
    return {"suggestions": suggestions, "plan_id": plan.get("plan_id"), "plan_name": plan.get("name", "")}


@api_router.get("/workouts/exercise-history")
async def get_exercise_history(request: Request, exercise_name: str, session_token: Optional[str] = Cookie(None)):
    """Get the last 5 workout logs containing a specific exercise."""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if not exercise_name:
        return {"history": []}
    
    # Search in workout_logs
    logs = await db.workout_logs.find({
        "user_id": user.user_id,
        "completed": True,
        "exercises_completed": {"$elemMatch": {"name": {"$regex": exercise_name, "$options": "i"}}}
    }, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
    
    history = []
    for log in logs:
        for ex in log.get("exercises_completed", []):
            if exercise_name.lower() in ex.get("name", "").lower():
                history.append({
                    "date": log.get("date", ""),
                    "log_name": log.get("name", ""),
                    "exercise_name": ex.get("name", ""),
                    "sets_data": ex.get("sets_data", []),
                    "weight": ex.get("weight", ""),
                    "reps": ex.get("reps", ""),
                    "sets_completed": ex.get("sets_completed", 0)
                })
                break
    
    return {"history": history}


@api_router.post("/workouts/calculate-warmup")
async def calculate_warmup(request: Request, session_token: Optional[str] = Cookie(None)):
    """Calculate warmup sets based on working weight."""
    auth_header = request.headers.get("Authorization")
    await get_current_user(authorization=auth_header, session_token=session_token)
    
    body = await request.json()
    working_weight = body.get("weight", 0)
    if not working_weight or working_weight <= 0:
        return {"warmup_sets": []}
    
    warmup_percentages = [0.5, 0.6, 0.7, 0.8]
    warmup_reps = [8, 6, 4, 2]
    
    warmup_sets = []
    for i, (pct, reps) in enumerate(zip(warmup_percentages, warmup_reps)):
        w = round(working_weight * pct, 1)
        if w > 0:
            warmup_sets.append({
                "set_number": i + 1,
                "percentage": f"{int(pct * 100)}%",
                "weight": w,
                "reps": reps,
                "label": f"{int(pct * 100)}% x {reps}" if w < working_weight else "Trabalho"
            })
    
    warmup_sets.append({
        "set_number": len(warmup_sets) + 1,
        "percentage": "100%",
        "weight": working_weight,
        "reps": "Trabalho",
        "label": "Trabalho"
    })
    
    return {"warmup_sets": warmup_sets}


@api_router.post("/workout-sessions/{session_id}/abandon")
async def abandon_workout_session(request: Request, session_id: str, session_token: Optional[str] = Cookie(None)):
    """Abandon an active workout session"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.workout_sessions.update_one(
        {"session_id": session_id, "user_id": user.user_id, "status": "active"},
        {"$set": {
            "status": "abandoned",
            "completed_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Sessão não encontrada")
    
    return {"message": "Sessão abandonada"}


@api_router.get("/workout-sessions")
async def get_workout_sessions(request: Request, limit: int = 20, session_token: Optional[str] = Cookie(None)):
    """Get workout session history"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    sessions = await db.workout_sessions.find(
        {"user_id": user.user_id, "status": {"$in": ["completed", "abandoned"]}},
        {"_id": 0}
    ).sort("started_at", -1).to_list(limit)
    
    return sessions


# ========== NOTIFICATION ENDPOINTS ==========
@api_router.get("/notifications")
async def get_notifications(request: Request, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    notifications = await db.notifications.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    for notif in notifications:
        if isinstance(notif['created_at'], str):
            notif['created_at'] = datetime.fromisoformat(notif['created_at'])
    return notifications

@api_router.post("/notifications")
async def create_notification(request: Request, notif_data: NotificationCreate, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    notification_id = f"notif_{uuid.uuid4().hex[:12]}"
    notification_doc = {
        "notification_id": notification_id,
        "user_id": user.user_id,
        "title": notif_data.title,
        "message": notif_data.message,
        "type": notif_data.type,
        "category": notif_data.category,
        "scheduled_time": notif_data.scheduled_time,
        "repeat": notif_data.repeat,
        "repeat_days": notif_data.repeat_days,
        "enabled": True,
        "channels": notif_data.channels,
        "last_sent": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.notifications.insert_one(notification_doc)
    notification_doc.pop('_id', None)  # Remove MongoDB ObjectId
    notification_doc['created_at'] = datetime.fromisoformat(notification_doc['created_at'])
    return Notification(**notification_doc)

@api_router.patch("/notifications/{notification_id}")
async def update_notification(request: Request, notification_id: str, notif_data: NotificationCreate, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    update_data = {
        "title": notif_data.title,
        "message": notif_data.message,
        "type": notif_data.type,
        "category": notif_data.category,
        "scheduled_time": notif_data.scheduled_time,
        "repeat": notif_data.repeat,
        "repeat_days": notif_data.repeat_days,
        "channels": notif_data.channels
    }
    
    result = await db.notifications.update_one(
        {"notification_id": notification_id, "user_id": user.user_id},
        {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    return {"message": "Notification updated"}

@api_router.patch("/notifications/{notification_id}/toggle")
async def toggle_notification(request: Request, notification_id: str, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    notif = await db.notifications.find_one({"notification_id": notification_id, "user_id": user.user_id}, {"_id": 0})
    if not notif:
        raise HTTPException(status_code=404, detail="Notification not found")
    
    new_enabled = not notif['enabled']
    await db.notifications.update_one(
        {"notification_id": notification_id},
        {"$set": {"enabled": new_enabled}}
    )
    return {"message": "Notification toggled", "enabled": new_enabled}

@api_router.delete("/notifications/{notification_id}")
async def delete_notification(request: Request, notification_id: str, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.notifications.delete_one({"notification_id": notification_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    return {"message": "Notification deleted"}

@api_router.get("/notifications/pending")
async def get_pending_notifications(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get notifications that should be triggered now"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    current_time = datetime.now(timezone.utc).strftime("%H:%M")
    current_day = datetime.now(timezone.utc).strftime("%A").lower()
    
    # Find enabled notifications for current time
    notifications = await db.notifications.find({
        "user_id": user.user_id,
        "enabled": True,
        "scheduled_time": current_time
    }, {"_id": 0}).to_list(100)
    
    pending = []
    for notif in notifications:
        should_send = False
        if notif['repeat'] == "none":
            should_send = True
        elif notif['repeat'] == "daily":
            should_send = True
        elif notif['repeat'] == "weekly":
            if current_day in [d.lower() for d in notif.get('repeat_days', [])]:
                should_send = True
        elif notif['repeat'] == "custom":
            if current_day in [d.lower() for d in notif.get('repeat_days', [])]:
                should_send = True
        
        if should_send:
            pending.append(notif)
    
    return pending

@api_router.post("/notifications/{notification_id}/send")
async def mark_notification_sent(request: Request, notification_id: str, channel: str, session_token: Optional[str] = Cookie(None)):
    """Mark a notification as sent and log it"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Update last_sent timestamp
    await db.notifications.update_one(
        {"notification_id": notification_id, "user_id": user.user_id},
        {"$set": {"last_sent": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Log the notification send
    log_id = f"nlog_{uuid.uuid4().hex[:12]}"
    log_doc = {
        "log_id": log_id,
        "notification_id": notification_id,
        "user_id": user.user_id,
        "sent_at": datetime.now(timezone.utc).isoformat(),
        "channel": channel,
        "status": "sent"
    }
    await db.notification_logs.insert_one(log_doc)
    
    return {"message": "Notification marked as sent", "log_id": log_id}

# ========== NOTIFICATION TEMPLATES ==========
@api_router.get("/notification-templates")
async def get_notification_templates():
    """Get predefined notification templates"""
    templates = [
        {
            "id": "hydration",
            "title": "💧 Hora de Beber Água",
            "message": "Lembre-se de se manter hidratado! Beba um copo de água.",
            "category": "hydration",
            "suggested_times": ["08:00", "10:00", "12:00", "14:00", "16:00", "18:00", "20:00"],
            "repeat": "daily"
        },
        {
            "id": "workout",
            "title": "💪 Hora do Treino",
            "message": "Não esqueça do seu treino de hoje! Bora mover o corpo!",
            "category": "workout",
            "suggested_times": ["06:00", "07:00", "18:00", "19:00"],
            "repeat": "custom"
        },
        {
            "id": "morning_tasks",
            "title": "📋 Revisão Matinal",
            "message": "Bom dia! Hora de revisar suas tarefas do dia.",
            "category": "task",
            "suggested_times": ["07:00", "08:00"],
            "repeat": "daily"
        },
        {
            "id": "evening_review",
            "title": "🌙 Revisão Noturna",
            "message": "Como foi seu dia? Hora de revisar o progresso e planejar amanhã.",
            "category": "task",
            "suggested_times": ["21:00", "22:00"],
            "repeat": "daily"
        },
        {
            "id": "habit_check",
            "title": "✅ Verificar Hábitos",
            "message": "Já completou seus hábitos de hoje?",
            "category": "habit",
            "suggested_times": ["20:00"],
            "repeat": "daily"
        },
        {
            "id": "stretch",
            "title": "🧘 Hora de Alongar",
            "message": "Faça uma pausa e alongue-se por 5 minutos.",
            "category": "workout",
            "suggested_times": ["10:00", "15:00"],
            "repeat": "daily"
        },
        {
            "id": "posture",
            "title": "🪑 Verificar Postura",
            "message": "Corrija sua postura! Costas retas, ombros relaxados.",
            "category": "custom",
            "suggested_times": ["09:00", "11:00", "14:00", "16:00"],
            "repeat": "daily"
        }
    ]
    return templates

# ========== BODY MEASUREMENTS ENDPOINTS ==========
@api_router.get("/body-measurements")
async def get_body_measurements(request: Request, limit: int = 30, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    measurements = await db.body_measurements.find(
        {"user_id": user.user_id}, {"_id": 0}
    ).sort("date", -1).limit(limit).to_list(limit)
    
    return measurements

@api_router.get("/body-measurements/latest")
async def get_latest_measurement(request: Request, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    measurement = await db.body_measurements.find_one(
        {"user_id": user.user_id}, {"_id": 0}, sort=[("date", -1)]
    )
    
    return measurement

@api_router.post("/body-measurements")
async def create_body_measurement(request: Request, data: BodyMeasurementCreate, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    measurement_id = f"measure_{uuid.uuid4().hex[:12]}"
    
    # Calcular IMC se altura e peso foram fornecidos
    bmi = None
    if data.weight_kg and data.height_cm:
        height_m = data.height_cm / 100
        bmi = round(data.weight_kg / (height_m ** 2), 1)
    
    measurement_doc = {
        "measurement_id": measurement_id,
        "user_id": user.user_id,
        "date": data.date,
        "weight_kg": data.weight_kg,
        "body_fat_percentage": data.body_fat_percentage,
        "muscle_mass_kg": data.muscle_mass_kg,
        "bone_mass_kg": data.bone_mass_kg,
        "water_percentage": data.water_percentage,
        "visceral_fat": data.visceral_fat,
        "metabolic_age": data.metabolic_age,
        "bmr_kcal": data.bmr_kcal,
        "height_cm": data.height_cm,
        "neck_cm": data.neck_cm,
        "shoulders_cm": data.shoulders_cm,
        "chest_cm": data.chest_cm,
        "waist_cm": data.waist_cm,
        "abdomen_cm": data.abdomen_cm,
        "hips_cm": data.hips_cm,
        "left_arm_cm": data.left_arm_cm,
        "right_arm_cm": data.right_arm_cm,
        "left_forearm_cm": data.left_forearm_cm,
        "right_forearm_cm": data.right_forearm_cm,
        "left_thigh_cm": data.left_thigh_cm,
        "right_thigh_cm": data.right_thigh_cm,
        "left_calf_cm": data.left_calf_cm,
        "right_calf_cm": data.right_calf_cm,
        "bmi": bmi,
        "notes": data.notes,
        "source": data.source,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.body_measurements.insert_one(measurement_doc)
    measurement_doc.pop('_id', None)
    
    return measurement_doc

@api_router.delete("/body-measurements/{measurement_id}")
async def delete_body_measurement(request: Request, measurement_id: str, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.body_measurements.delete_one({"measurement_id": measurement_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Measurement not found")
    
    return {"message": "Measurement deleted"}

@api_router.get("/body-measurements/evolution")
async def get_body_evolution(request: Request, months: int = 6, session_token: Optional[str] = Cookie(None)):
    """Get body measurement evolution over time"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    start_date = (datetime.now() - timedelta(days=months * 30)).strftime("%Y-%m-%d")
    
    measurements = await db.body_measurements.find(
        {"user_id": user.user_id, "date": {"$gte": start_date}}, {"_id": 0}
    ).sort("date", 1).to_list(1000)
    
    # Calculate changes
    if len(measurements) >= 2:
        first = measurements[0]
        last = measurements[-1]
        
        changes = {}
        for field in ["weight_kg", "body_fat_percentage", "muscle_mass_kg", "waist_cm", "bmi"]:
            if first.get(field) and last.get(field):
                changes[field] = round(last[field] - first[field], 2)
    else:
        changes = {}
    
    return {
        "measurements": measurements,
        "changes": changes,
        "total_records": len(measurements)
    }

# ========== PDF ANALYSIS ENDPOINT ==========
@api_router.post("/body-measurements/analyze-pdf")
async def analyze_pdf_measurement(
    request: Request,
    file: UploadFile = File(...),
    session_token: Optional[str] = Cookie(None)
):
    """Analyze a PDF file containing body measurement data using AI"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files are accepted")
    
    # Read file content
    content = await file.read()
    file_base64 = base64.b64encode(content).decode('utf-8')
    
    # Use Gemini to analyze the PDF
    try:
        system_message = """Você é um especialista em análise de avaliações físicas e bioimpedância.
            Analise o documento e extraia TODOS os dados disponíveis.
            Responda APENAS em formato JSON válido com os campos encontrados.
            Use os seguintes nomes de campos (deixe null se não encontrado):
            - weight_kg, height_cm, body_fat_percentage, muscle_mass_kg
            - bone_mass_kg, water_percentage, visceral_fat, metabolic_age, bmr_kcal
            - neck_cm, shoulders_cm, chest_cm, waist_cm, abdomen_cm, hips_cm
            - left_arm_cm, right_arm_cm, left_forearm_cm, right_forearm_cm
            - left_thigh_cm, right_thigh_cm, left_calf_cm, right_calf_cm
            - date (formato YYYY-MM-DD), notes (observações relevantes)
            - recommendations (array de recomendações baseadas nos dados)"""
        
        # Upload file for Gemini using new SDK
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp_file:
            tmp_file.write(content)
            tmp_path = tmp_file.name
        
        uploaded_file = gemini_client.files.upload(file=tmp_path)
        
        response = gemini_client.models.generate_content(
            model=GEMINI_MODEL,
            contents=[
                types.Part.from_uri(file_uri=uploaded_file.uri, mime_type="application/pdf"),
                "Analise este documento de avaliação física/bioimpedância e extraia todos os dados em JSON:"
            ],
            config=types.GenerateContentConfig(
                system_instruction=system_message
            )
        )
        
        # Clean up temp file
        os.unlink(tmp_path)
        
        # Try to parse JSON from response
        try:
            # Remove markdown code blocks if present
            json_str = response.text.strip()
            if json_str.startswith("```json"):
                json_str = json_str[7:]
            if json_str.startswith("```"):
                json_str = json_str[3:]
            if json_str.endswith("```"):
                json_str = json_str[:-3]
            
            extracted_data = json.loads(json_str.strip())
        except json.JSONDecodeError:
            # If JSON parsing fails, return raw analysis
            extracted_data = {"raw_analysis": response.text, "parse_error": True}
        
        return {
            "success": True,
            "extracted_data": extracted_data,
            "filename": file.filename
        }
        
    except Exception as e:
        logging.error(f"PDF analysis failed: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to analyze PDF: {str(e)}")

# ========== AI RECOMMENDATIONS ==========
@api_router.get("/body-measurements/recommendations")
async def get_workout_recommendations(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get AI-powered workout and health recommendations based on body measurements"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Get latest measurements
    measurements = await db.body_measurements.find(
        {"user_id": user.user_id}, {"_id": 0}
    ).sort("date", -1).limit(5).to_list(5)
    
    # Get recent workouts
    workouts = await db.workout_logs.find(
        {"user_id": user.user_id}, {"_id": 0}
    ).sort("date", -1).limit(10).to_list(10)
    
    if not measurements:
        return {
            "recommendations": ["Registre suas medidas corporais para receber recomendações personalizadas."],
            "based_on": "no_data"
        }
    
    latest = measurements[0]
    
    prompt = f"""Com base nos seguintes dados corporais do usuário, forneça recomendações personalizadas de treino e saúde:

MEDIDAS ATUAIS:
- Peso: {latest.get('weight_kg', 'N/A')} kg
- Altura: {latest.get('height_cm', 'N/A')} cm
- IMC: {latest.get('bmi', 'N/A')}
- Gordura corporal: {latest.get('body_fat_percentage', 'N/A')}%
- Massa muscular: {latest.get('muscle_mass_kg', 'N/A')} kg
- Cintura: {latest.get('waist_cm', 'N/A')} cm
- Gordura visceral: {latest.get('visceral_fat', 'N/A')}

HISTÓRICO DE TREINOS (últimos 10):
{json.dumps([{"name": w.get("name"), "type": w.get("activity_type"), "duration": w.get("duration_minutes")} for w in workouts], indent=2)}

Forneça:
1. 3-5 recomendações específicas de treino
2. Dicas de nutrição
3. Áreas de foco prioritárias
4. Metas sugeridas para os próximos 30 dias

Responda em português de forma direta e motivadora."""

    try:
        response = await call_llm(
            prompt=prompt,
            session_id=f"recommendations_{user.user_id}",
            system_message="Você é um personal trainer e nutricionista experiente. Forneça recomendações práticas e motivadoras.",
            user_id=user.user_id
        )
        return {
            "recommendations": response,
            "based_on": latest,
            "workouts_analyzed": len(workouts)
        }
    except Exception as e:
        logging.error(f"Recommendations generation failed: {e}")
        return {
            "recommendations": "Não foi possível gerar recomendações no momento. Tente novamente mais tarde.",
            "error": str(e)
        }

# ========== MOTIVATIONAL QUOTES ==========
@api_router.get("/motivational-quote")
async def get_motivational_quote(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get a personalized motivational quote - one per day, resets at 5:00 AM"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Calculate the "motivational day" - resets at 5:00 AM
    now = datetime.now()
    if now.hour < 5:
        # Before 5 AM, still "yesterday's" motivational day
        motivational_date = (now - timedelta(days=1)).strftime("%Y-%m-%d")
    else:
        motivational_date = now.strftime("%Y-%m-%d")
    
    # Check if we already have a quote for this motivational day
    cached_quote = await db.daily_quotes.find_one({
        "user_id": user.user_id,
        "motivational_date": motivational_date
    }, {"_id": 0})
    
    if cached_quote:
        cached_text = cached_quote.get("quote", "")
        if cached_text.startswith("⚠"):
            await db.daily_quotes.delete_one({"user_id": user.user_id, "motivational_date": motivational_date})
        else:
            return {
                "quote": cached_text,
                "motivational_date": motivational_date,
                "cached": True,
                "context": cached_quote.get("context", {})
            }
    
    # Generate a new quote for today
    today = datetime.now().strftime("%Y-%m-%d")
    week_ago = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
    hour = datetime.now().hour
    
    workouts_this_week = await db.workout_logs.count_documents({
        "user_id": user.user_id,
        "date": {"$gte": week_ago},
        "completed": True
    })
    
    habits_today = await db.habit_logs.count_documents({
        "user_id": user.user_id,
        "date": today,
        "completed": True
    })
    
    # Determine time of day
    if hour < 12:
        time_of_day = "manhã"
    elif hour < 18:
        time_of_day = "tarde"
    else:
        time_of_day = "noite"
    
    prompt = f"""Gere UMA frase motivacional ÚNICA, CRIATIVA e IMPACTANTE.

CONTEXTO:
- Nome: {user.name}
- Hora do dia: {time_of_day}
- Treinos esta semana: {workouts_this_week}
- Atividades hoje: {habits_today}

ESTILOS POSSÍVEIS (escolha um aleatoriamente):
1. Frase filosófica profunda sobre disciplina e crescimento
2. Citação inspiradora no estilo de grandes líderes ou atletas
3. Metáfora poderosa sobre superação
4. Desafio direto e provocativo
5. Reflexão sobre mentalidade de guerreiro/campeão
6. Frase sobre consistência e processo
7. Motivação brutal e direta estilo militar
8. Insight sobre autoconhecimento e evolução
9. Comparação inspiradora com a natureza ou elementos
10. Frase sobre legado e propósito

REGRAS:
- Máximo 2 linhas
- Seja CRIATIVO e ORIGINAL - evite clichês
- Pode ou não mencionar o nome "{user.name}"
- Use 1-2 emojis impactantes (🔥💪⚡🎯🏆🦁⚔️🌟💎🚀)
- A frase deve causar IMPACTO e fazer a pessoa querer agir
- Varie entre tom filosófico, agressivo, reflexivo ou desafiador
- NÃO precisa falar de XP, patente ou progresso no app

Responda APENAS com a frase, sem explicações."""

    # Fallback quotes used when AI is unavailable
    fallback_quotes = [
        "🔥 A dor do treino é temporária. A dor do arrependimento é permanente.",
        "⚔️ Guerreiros não nascem. São forjados no fogo da disciplina diária.",
        "🦁 Seja a pessoa que você precisava quando era mais novo.",
        "💎 Diamantes são apenas pedras que não desistiram sob pressão.",
        "🎯 Enquanto outros dormem, você constrói seu império.",
        "⚡ Sua única competição é quem você era ontem.",
        "🏆 Champions são feitos quando ninguém está olhando.",
        "🚀 Conforto é a morte lenta dos seus sonhos. Acorde!",
        "💪 Seu corpo pode quase tudo. É sua mente que você precisa convencer.",
        "🌟 A excelência não é um ato, é um hábito. Que hábito você está construindo?"
    ]

    try:
        response = await call_llm(
            prompt=prompt,
            session_id=f"motivation_{user.user_id}_{motivational_date}",
            system_message="Você é um mestre motivacional que combina sabedoria filosófica, mentalidade de elite atlética e coaching de alta performance. Suas frases são impactantes, únicas e memoráveis.",
            user_id=user.user_id
        )
        
        quote_text = response.strip()
        
        # If LLM returned an error message, don't cache it — use fallback
        if quote_text.startswith("⚠"):
            await db.daily_quotes.delete_one({"user_id": user.user_id, "motivational_date": motivational_date})
            return {
                "quote": random.choice(fallback_quotes),
                "motivational_date": motivational_date,
                "fallback": True
            }
        
        # Cache the quote for this motivational day
        await db.daily_quotes.update_one(
            {"user_id": user.user_id, "motivational_date": motivational_date},
            {"$set": {
                "user_id": user.user_id,
                "motivational_date": motivational_date,
                "quote": quote_text,
                "context": {
                    "workouts_this_week": workouts_this_week,
                    "habits_today": habits_today,
                    "time_of_day": time_of_day
                },
                "created_at": datetime.now(timezone.utc).isoformat()
            }},
            upsert=True
        )
        
        return {
            "quote": quote_text,
            "motivational_date": motivational_date,
            "cached": False,
            "context": {
                "workouts_this_week": workouts_this_week,
                "habits_today": habits_today
            }
        }
        
    except Exception as e:
        logging.error(f"Quote generation failed: {e}")
        return {
            "quote": random.choice(fallback_quotes),
            "motivational_date": motivational_date,
            "fallback": True
        }

# ========== DAILY WORKOUT STATUS ==========
@api_router.get("/daily-workout-status/{plan_id}")
async def get_daily_workout_status(request: Request, plan_id: str, session_token: Optional[str] = Cookie(None)):
    """Get the exercise completion status for a plan on today's date"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    today = datetime.now().strftime("%Y-%m-%d")
    
    status = await db.daily_workout_status.find_one({
        "user_id": user.user_id,
        "plan_id": plan_id,
        "date": today
    }, {"_id": 0})
    
    return status or {"exercises_status": {}, "completed": False}

@api_router.post("/daily-workout-status/{plan_id}/toggle/{exercise_idx}")
async def toggle_daily_exercise(request: Request, plan_id: str, exercise_idx: int, session_token: Optional[str] = Cookie(None)):
    """Toggle an exercise completion status for today"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    today = datetime.now().strftime("%Y-%m-%d")
    
    # Get or create today's status
    status = await db.daily_workout_status.find_one({
        "user_id": user.user_id,
        "plan_id": plan_id,
        "date": today
    })
    
    if not status:
        status = {
            "status_id": f"dws_{uuid.uuid4().hex[:12]}",
            "user_id": user.user_id,
            "plan_id": plan_id,
            "date": today,
            "exercises_status": {},
            "completed": False,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.daily_workout_status.insert_one(status)
    
    # Toggle the exercise
    exercise_key = str(exercise_idx)
    current_status = status.get("exercises_status", {}).get(exercise_key, False)
    new_status = not current_status
    
    await db.daily_workout_status.update_one(
        {"user_id": user.user_id, "plan_id": plan_id, "date": today},
        {
            "$set": {
                f"exercises_status.{exercise_key}": new_status,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    # Get updated status
    updated = await db.daily_workout_status.find_one({
        "user_id": user.user_id,
        "plan_id": plan_id,
        "date": today
    }, {"_id": 0})
    
    return updated

@api_router.post("/daily-workout-status/{plan_id}/complete")
async def complete_daily_workout(request: Request, plan_id: str, data: dict, session_token: Optional[str] = Cookie(None)):
    """Complete a daily workout and log it"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    today = datetime.now().strftime("%Y-%m-%d")
    
    # Get the plan
    plan = await db.workout_plans.find_one({"plan_id": plan_id, "user_id": user.user_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    # Get today's status
    status = await db.daily_workout_status.find_one({
        "user_id": user.user_id,
        "plan_id": plan_id,
        "date": today
    }, {"_id": 0})
    
    exercises_status = status.get("exercises_status", {}) if status else {}
    
    # Build exercises_completed list
    exercises_completed = []
    for idx, ex in enumerate(plan.get("exercises", [])):
        exercises_completed.append({
            **ex,
            "completed": exercises_status.get(str(idx), False)
        })
    
    completed_count = sum(1 for ex in exercises_completed if ex.get("completed"))
    total_exercises = len(exercises_completed)
    
    # Calculate XP based on completion
    base_xp = 10
    completion_bonus = int((completed_count / total_exercises) * 30) if total_exercises > 0 else 0
    duration_bonus = (data.get("duration_minutes", 30) // 15) * 5
    total_xp = base_xp + completion_bonus + duration_bonus
    
    # Estimate calories if not provided (rough estimate based on duration and activity)
    calories = data.get("calories")
    if not calories:
        # Average 5-8 calories per minute for strength training
        calories = int(data.get("duration_minutes", 30) * 6)
    
    # Create workout log
    log_id = f"workout_{uuid.uuid4().hex[:12]}"
    workout_doc = {
        "log_id": log_id,
        "user_id": user.user_id,
        "plan_id": plan_id,
        "activity_type": "weightlifting",
        "name": plan.get("name", "Treino"),
        "duration_minutes": data.get("duration_minutes", 30),
        "calories": calories,
        "exercises_completed": exercises_completed,
        "notes": data.get("notes", ""),
        "xp_earned": total_xp,
        "completed": True,
        "date": today,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.workout_logs.insert_one(workout_doc)
    
    # Update user XP
    new_xp = user.xp + total_xp
    new_rank = calculate_rank(new_xp)
    await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
    
    # Mark daily status as completed
    await db.daily_workout_status.update_one(
        {"user_id": user.user_id, "plan_id": plan_id, "date": today},
        {"$set": {"completed": True, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    workout_doc.pop('_id', None)
    return {
        **workout_doc,
        "new_xp": new_xp,
        "new_rank": new_rank,
        "exercises_completed_count": completed_count,
        "total_exercises": total_exercises
    }

@api_router.post("/daily-workout-status/{plan_id}/reset")
async def reset_daily_workout(request: Request, plan_id: str, session_token: Optional[str] = Cookie(None)):
    """Reset today's workout status for a plan"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    today = datetime.now().strftime("%Y-%m-%d")
    
    await db.daily_workout_status.update_one(
        {"user_id": user.user_id, "plan_id": plan_id, "date": today},
        {
            "$set": {
                "exercises_status": {},
                "completed": False,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    return {"message": "Daily workout status reset", "date": today}

# ========== NUTRITION MODELS ==========
class Meal(BaseModel):
    model_config = ConfigDict(extra="ignore")
    meal_id: str
    user_id: str
    name: str
    meal_type: str  # breakfast, lunch, dinner, snack
    foods: List[Dict[str, Any]] = []  # [{name, calories, protein, carbs, fat, quantity, unit}]
    total_calories: int = 0
    total_protein: float = 0
    total_carbs: float = 0
    total_fat: float = 0
    date: str
    notes: Optional[str] = None
    created_at: datetime

class MealCreate(BaseModel):
    name: str
    meal_type: str
    foods: List[Dict[str, Any]] = []
    date: str
    notes: Optional[str] = None

class NutritionGoal(BaseModel):
    model_config = ConfigDict(extra="ignore")
    goal_id: str
    user_id: str
    daily_calories: int = 2000
    daily_protein: float = 150
    daily_carbs: float = 250
    daily_fat: float = 65
    water_goal_ml: int = 2000
    created_at: datetime
    updated_at: datetime

class NutritionGoalCreate(BaseModel):
    daily_calories: int = 2000
    daily_protein: float = 150
    daily_carbs: float = 250
    daily_fat: float = 65
    water_goal_ml: int = 2000

class WaterLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    log_id: str
    user_id: str
    amount_ml: int
    date: str
    created_at: datetime

class Diet(BaseModel):
    model_config = ConfigDict(extra="ignore")
    diet_id: str
    user_id: str
    name: str
    description: Optional[str] = None
    diet_type: str  # cutting, bulking, maintenance, keto, low_carb, etc.
    meals_plan: List[Dict[str, Any]] = []  # [{meal_type, suggested_foods, target_calories}]
    active: bool = True
    start_date: str
    end_date: Optional[str] = None
    created_at: datetime

class DietCreate(BaseModel):
    name: str
    description: Optional[str] = None
    diet_type: str
    meals_plan: List[Dict[str, Any]] = []
    start_date: str
    end_date: Optional[str] = None

class Recipe(BaseModel):
    model_config = ConfigDict(extra="ignore")
    recipe_id: str
    user_id: str
    name: str
    description: Optional[str] = None
    ingredients: List[Dict[str, Any]] = []  # [{name, quantity, unit}]
    instructions: List[str] = []
    prep_time_minutes: int = 0
    cook_time_minutes: int = 0
    servings: int = 1
    calories_per_serving: int = 0
    protein_per_serving: float = 0
    carbs_per_serving: float = 0
    fat_per_serving: float = 0
    tags: List[str] = []  # healthy, quick, high-protein, etc.
    ai_generated: bool = False
    created_at: datetime

# ========== STUDY MODELS ==========
class StudyArea(BaseModel):
    model_config = ConfigDict(extra="ignore")
    area_id: str
    user_id: str
    name: str  # Faculdade, Concursos, Trabalho, Outros
    description: Optional[str] = None
    color: str = "#007AFF"
    icon: str = "book"
    order: int = 0
    created_at: datetime

class StudyAreaCreate(BaseModel):
    name: str
    description: Optional[str] = None
    color: str = "#007AFF"
    icon: str = "book"

class StudyProgram(BaseModel):
    model_config = ConfigDict(extra="ignore")
    program_id: str
    user_id: str
    area_id: str
    name: str  # Ex: "Curso de Direito", "Concurso TRF5"
    description: Optional[str] = None
    color: str = "#007AFF"
    icon: str = "book"
    target_date: Optional[str] = None  # Meta date (exam date, graduation, etc.)
    status: str = "active"  # active, completed, paused
    total_questions: int = 0
    correct_questions: int = 0
    created_at: datetime

class StudyProgramCreate(BaseModel):
    area_id: str
    name: str
    description: Optional[str] = None
    color: str = "#007AFF"
    icon: str = "book"
    target_date: Optional[str] = None

class Notebook(BaseModel):
    model_config = ConfigDict(extra="ignore")
    notebook_id: str
    user_id: str
    area_id: str
    program_id: Optional[str] = None  # Optional link to a program
    name: str  # Matéria/Assunto
    description: Optional[str] = None
    color: str = "#007AFF"
    tags: List[str] = []
    total_study_time_minutes: int = 0
    total_questions: int = 0
    correct_questions: int = 0
    created_at: datetime

class NotebookCreate(BaseModel):
    area_id: str
    program_id: Optional[str] = None
    name: str
    description: Optional[str] = None
    color: str = "#007AFF"
    tags: List[str] = []

class QuestionLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    log_id: str
    user_id: str
    notebook_id: str
    program_id: Optional[str] = None
    total: int = 0
    correct: int = 0
    incorrect: int = 0
    source: str = "manual"  # manual, quiz, ai
    date: str
    created_at: datetime

class QuestionLogCreate(BaseModel):
    notebook_id: str
    total: int
    correct: int
    source: str = "manual"

class FocusSession(BaseModel):
    model_config = ConfigDict(extra="ignore")
    focus_id: str
    user_id: str
    notebook_id: Optional[str] = None
    focus_minutes: int = 25
    break_minutes: int = 5
    completed: bool = False
    date: str
    notes: Optional[str] = None
    xp_earned: int = 0
    created_at: datetime

class FocusSessionCreate(BaseModel):
    notebook_id: Optional[str] = None
    focus_minutes: int = 25
    break_minutes: int = 5
    notes: Optional[str] = None

class StudyNote(BaseModel):
    model_config = ConfigDict(extra="ignore")
    note_id: str
    user_id: str
    notebook_id: str
    title: str
    content: str
    tags: List[str] = []
    links: List[Dict[str, str]] = []  # [{title, url}]
    attachments: List[Dict[str, Any]] = []  # [{name, type, url/data}]
    created_at: datetime
    updated_at: datetime

class StudyNoteCreate(BaseModel):
    notebook_id: str
    title: str
    content: str
    tags: List[str] = []
    links: List[Dict[str, str]] = []

class StudyTask(BaseModel):
    model_config = ConfigDict(extra="ignore")
    task_id: str
    user_id: str
    notebook_id: Optional[str] = None
    title: str
    description: Optional[str] = None
    task_type: str  # reading, exercise, review, project, exam
    recurrence: str = "once"  # once, daily, weekly, monthly
    deadline: Optional[str] = None
    reminder: Optional[str] = None
    completed: bool = False
    completed_at: Optional[str] = None
    last_completed_date: Optional[str] = None  # For recurring tasks
    priority: str = "medium"
    estimated_minutes: int = 30
    actual_minutes: int = 0
    notes: Optional[str] = None
    xp_reward: int = 20
    created_at: datetime

class StudyTaskCreate(BaseModel):
    notebook_id: Optional[str] = None
    title: str
    description: Optional[str] = None
    task_type: str = "reading"
    recurrence: str = "once"  # once, daily, weekly, monthly
    deadline: Optional[str] = None
    reminder: Optional[str] = None
    priority: str = "medium"
    estimated_minutes: int = 30

class StudySession(BaseModel):
    model_config = ConfigDict(extra="ignore")
    session_id: str
    user_id: str
    notebook_id: str
    duration_minutes: int
    date: str
    notes: Optional[str] = None
    xp_earned: int = 0
    created_at: datetime

class StudySessionCreate(BaseModel):
    notebook_id: str
    duration_minutes: int
    date: str
    notes: Optional[str] = None

class StudySchedule(BaseModel):
    model_config = ConfigDict(extra="ignore")
    schedule_id: str
    user_id: str
    notebook_id: str
    day_of_week: str  # monday, tuesday, etc.
    start_time: str  # HH:MM
    end_time: str  # HH:MM
    repeat: bool = True
    created_at: datetime

class StudyScheduleCreate(BaseModel):
    notebook_id: str
    day_of_week: str
    start_time: str
    end_time: str
    repeat: bool = True

class Flashcard(BaseModel):
    model_config = ConfigDict(extra="ignore")
    flashcard_id: str
    user_id: str
    notebook_id: str
    deck_name: str
    front: str  # Question
    back: str  # Answer
    tags: List[str] = []
    # Spaced Repetition Fields
    ease_factor: float = 2.5
    interval_days: int = 1
    repetitions: int = 0
    next_review: str  # Date
    last_review: Optional[str] = None
    created_at: datetime

class FlashcardCreate(BaseModel):
    notebook_id: str
    deck_name: str
    front: str
    back: str
    tags: List[str] = []

class FlashcardReview(BaseModel):
    quality: int  # 0-5 (0=forgot, 5=perfect)

class Quiz(BaseModel):
    model_config = ConfigDict(extra="ignore")
    quiz_id: str
    user_id: str
    notebook_id: str
    title: str
    questions: List[Dict[str, Any]] = []  # [{question, options, correct_answer, explanation}]
    ai_generated: bool = False
    created_at: datetime

class QuizCreate(BaseModel):
    notebook_id: str
    title: str
    questions: List[Dict[str, Any]] = []

class QuizAttempt(BaseModel):
    model_config = ConfigDict(extra="ignore")
    attempt_id: str
    user_id: str
    quiz_id: str
    score: float
    answers: List[Dict[str, Any]] = []  # [{question_idx, selected_answer, correct}]
    completed_at: datetime

# ========== SIMULADO MODELS ==========

class SimuladoCreate(BaseModel):
    title: str
    description: Optional[str] = ""
    banca: Optional[str] = None  # CESPE, FCC, VUNESP, FGV, etc.
    disciplina: Optional[str] = None  # Direito, Português, Matemática, etc.
    concurso: Optional[str] = None  # TRF, TJ, Receita Federal, etc.
    question_type: str = "multipla_escolha"  # multipla_escolha, certo_errado, misto
    num_questions: int = 10
    difficulty: str = "medio"  # facil, medio, dificil, misto
    area_id: Optional[str] = None
    program_id: Optional[str] = None

class SimuladoSubmit(BaseModel):
    answers: List[Dict[str, Any]]  # [{question_idx: int, selected_answer: str}]
    time_spent_seconds: int = 0



class StudyStreak(BaseModel):
    model_config = ConfigDict(extra="ignore")
    streak_id: str
    user_id: str
    current_streak: int = 0
    best_streak: int = 0
    last_study_date: Optional[str] = None
    total_study_days: int = 0
    created_at: datetime

class StudyStats(BaseModel):
    model_config = ConfigDict(extra="ignore")
    stats_id: str
    user_id: str
    notebook_id: str
    total_time_minutes: int = 0
    sessions_count: int = 0
    flashcards_reviewed: int = 0
    quizzes_completed: int = 0
    average_quiz_score: float = 0
    tasks_completed: int = 0

# ========== NUTRITION ENDPOINTS ==========

@api_router.get("/nutrition/meals")
async def get_meals(request: Request, date: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    """Get meals for a specific date or all meals"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    query = {"user_id": user.user_id}
    if date:
        query["date"] = date
    
    meals = await db.meals.find(query, {"_id": 0}).to_list(1000)
    return meals

@api_router.post("/nutrition/meals")
async def create_meal(request: Request, meal_data: MealCreate, session_token: Optional[str] = Cookie(None)):
    """Create a new meal"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Calculate totals from foods
    total_calories = sum(f.get("calories", 0) * f.get("quantity", 1) for f in meal_data.foods)
    total_protein = sum(f.get("protein", 0) * f.get("quantity", 1) for f in meal_data.foods)
    total_carbs = sum(f.get("carbs", 0) * f.get("quantity", 1) for f in meal_data.foods)
    total_fat = sum(f.get("fat", 0) * f.get("quantity", 1) for f in meal_data.foods)
    
    meal_id = f"meal_{uuid.uuid4().hex[:12]}"
    meal_doc = {
        "meal_id": meal_id,
        "user_id": user.user_id,
        "name": meal_data.name,
        "meal_type": meal_data.meal_type,
        "foods": meal_data.foods,
        "total_calories": int(total_calories),
        "total_protein": round(total_protein, 1),
        "total_carbs": round(total_carbs, 1),
        "total_fat": round(total_fat, 1),
        "date": meal_data.date,
        "notes": meal_data.notes,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.meals.insert_one(meal_doc)
    meal_doc.pop('_id', None)
    return meal_doc

@api_router.delete("/nutrition/meals/{meal_id}")
async def delete_meal(request: Request, meal_id: str, session_token: Optional[str] = Cookie(None)):
    """Delete a meal"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.meals.delete_one({"meal_id": meal_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Meal not found")
    return {"message": "Meal deleted"}

@api_router.post("/nutrition/estimate-food")
async def estimate_food_nutrition(request: Request, session_token: Optional[str] = Cookie(None)):
    """Use AI to estimate nutritional values for a food item"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    body = await request.json()
    food_name = body.get("food_name", "").strip()
    quantity = body.get("quantity", "").strip()
    
    if not food_name:
        raise HTTPException(status_code=400, detail="Nome do alimento é obrigatório")
    if not quantity:
        raise HTTPException(status_code=400, detail="Quantidade/peso é obrigatório")
    
    prompt = f"""Analise o seguinte alimento e estime os valores nutricionais com precisão.

Alimento: {food_name}
Quantidade/Peso: {quantity}

Retorne APENAS um JSON válido (sem markdown, sem explicação) com esta estrutura exata:
{{
  "food_name": "nome do alimento formatado",
  "quantity": "{quantity}",
  "calories": número inteiro (kcal),
  "protein": número decimal (gramas),
  "carbs": número decimal (gramas),
  "fat": número decimal (gramas),
  "fiber": número decimal (gramas),
  "sodium": número decimal (mg),
  "sugar": número decimal (gramas)
}}

Use valores baseados em tabelas nutricionais brasileiras (TACO) quando possível.
Considere a quantidade informada para calcular os valores proporcionais.
Se for um prato composto (ex: "prato feito"), estime os ingredientes típicos.
Retorne SOMENTE o JSON, nada mais."""

    system_msg = "Você é um nutricionista especialista em tabelas nutricionais brasileiras. Retorne apenas JSON válido sem markdown."
    
    try:
        response = await call_llm(prompt, f"food_estimate_{user.user_id}_{uuid.uuid4().hex[:6]}", system_msg, user_id=user.user_id)
        
        # Parse JSON from response
        json_str = response.strip()
        if json_str.startswith("```"):
            json_str = json_str.split("\n", 1)[1] if "\n" in json_str else json_str[3:]
            json_str = json_str.rsplit("```", 1)[0]
        json_str = json_str.strip()
        
        nutrition_data = json.loads(json_str)
        
        return {
            "success": True,
            "food_name": nutrition_data.get("food_name", food_name),
            "quantity": nutrition_data.get("quantity", quantity),
            "calories": int(nutrition_data.get("calories", 0)),
            "protein": round(float(nutrition_data.get("protein", 0)), 1),
            "carbs": round(float(nutrition_data.get("carbs", 0)), 1),
            "fat": round(float(nutrition_data.get("fat", 0)), 1),
            "fiber": round(float(nutrition_data.get("fiber", 0)), 1),
            "sodium": round(float(nutrition_data.get("sodium", 0)), 1),
            "sugar": round(float(nutrition_data.get("sugar", 0)), 1)
        }
    except json.JSONDecodeError:
        return {
            "success": False,
            "error": "Não foi possível estimar os nutrientes. Tente novamente ou insira manualmente.",
            "food_name": food_name,
            "quantity": quantity,
            "calories": 0, "protein": 0, "carbs": 0, "fat": 0, "fiber": 0, "sodium": 0, "sugar": 0
        }
    except Exception as e:
        return {
            "success": False,
            "error": f"Erro ao estimar nutrientes: {str(e)}",
            "food_name": food_name,
            "quantity": quantity,
            "calories": 0, "protein": 0, "carbs": 0, "fat": 0, "fiber": 0, "sodium": 0, "sugar": 0
        }


@api_router.post("/nutrition/estimate-foods-batch")
async def estimate_foods_batch(request: Request, session_token: Optional[str] = Cookie(None)):
    """Use AI to estimate nutritional values for multiple food items at once"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    body = await request.json()
    foods = body.get("foods", [])
    
    if not foods or len(foods) == 0:
        raise HTTPException(status_code=400, detail="Lista de alimentos é obrigatória")
    
    if len(foods) > 20:
        raise HTTPException(status_code=400, detail="Máximo de 20 alimentos por vez")
    
    # Build the food list for the prompt
    food_list_text = ""
    for i, food in enumerate(foods):
        name = food.get("food_name", "").strip()
        qty = food.get("quantity", "").strip()
        if name and qty:
            food_list_text += f"{i+1}. {name} - {qty}\n"
    
    if not food_list_text:
        raise HTTPException(status_code=400, detail="Nenhum alimento válido na lista")
    
    prompt = f"""Analise os seguintes alimentos e estime os valores nutricionais de CADA UM com precisão.

ALIMENTOS:
{food_list_text}

Retorne APENAS um JSON válido (sem markdown, sem explicação) com esta estrutura exata:
{{
  "foods": [
    {{
      "index": 0,
      "food_name": "nome do alimento formatado",
      "quantity": "quantidade informada",
      "calories": número inteiro (kcal),
      "protein": número decimal (gramas),
      "carbs": número decimal (gramas),
      "fat": número decimal (gramas),
      "fiber": número decimal (gramas),
      "sodium": número decimal (mg),
      "sugar": número decimal (gramas)
    }}
  ]
}}

Use valores baseados em tabelas nutricionais brasileiras (TACO) quando possível.
Considere a quantidade informada para calcular os valores proporcionais.
Se for um prato composto (ex: "prato feito"), estime os ingredientes típicos.
Retorne UM item para CADA alimento listado, na mesma ordem.
Retorne SOMENTE o JSON, nada mais."""

    system_msg = "Você é um nutricionista especialista em tabelas nutricionais brasileiras. Retorne apenas JSON válido sem markdown."
    
    try:
        response = await call_llm(prompt, f"food_batch_{user.user_id}_{uuid.uuid4().hex[:6]}", system_msg, user_id=user.user_id)
        
        # Parse JSON from response
        json_str = response.strip()
        if json_str.startswith("```"):
            json_str = json_str.split("\n", 1)[1] if "\n" in json_str else json_str[3:]
            json_str = json_str.rsplit("```", 1)[0]
        json_str = json_str.strip()
        
        result = json.loads(json_str)
        estimated_foods = result.get("foods", [])
        
        # Normalize the response
        normalized = []
        for i, food in enumerate(foods):
            # Find matching estimation (by index or position)
            est = next((e for e in estimated_foods if e.get("index") == i), None)
            if not est and i < len(estimated_foods):
                est = estimated_foods[i]
            
            if est:
                normalized.append({
                    "index": i,
                    "success": True,
                    "food_name": est.get("food_name", food.get("food_name", "")),
                    "quantity": est.get("quantity", food.get("quantity", "")),
                    "calories": int(est.get("calories", 0)),
                    "protein": round(float(est.get("protein", 0)), 1),
                    "carbs": round(float(est.get("carbs", 0)), 1),
                    "fat": round(float(est.get("fat", 0)), 1),
                    "fiber": round(float(est.get("fiber", 0)), 1),
                    "sodium": round(float(est.get("sodium", 0)), 1),
                    "sugar": round(float(est.get("sugar", 0)), 1)
                })
            else:
                normalized.append({
                    "index": i,
                    "success": False,
                    "food_name": food.get("food_name", ""),
                    "quantity": food.get("quantity", ""),
                    "calories": 0, "protein": 0, "carbs": 0, "fat": 0, "fiber": 0, "sodium": 0, "sugar": 0
                })
        
        return {
            "success": True,
            "foods": normalized
        }
    except json.JSONDecodeError:
        return {
            "success": False,
            "error": "Não foi possível estimar os nutrientes. Tente novamente.",
            "foods": []
        }
    except Exception as e:
        return {
            "success": False,
            "error": f"Erro ao estimar nutrientes: {str(e)}",
            "foods": []
        }



@api_router.get("/nutrition/goals")
async def get_nutrition_goals(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get user's nutrition goals"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    goals = await db.nutrition_goals.find_one({"user_id": user.user_id}, {"_id": 0})
    if not goals:
        # Create default goals
        goal_id = f"ngoal_{uuid.uuid4().hex[:12]}"
        goals = {
            "goal_id": goal_id,
            "user_id": user.user_id,
            "daily_calories": 2000,
            "daily_protein": 150,
            "daily_carbs": 250,
            "daily_fat": 65,
            "water_goal_ml": 2000,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.nutrition_goals.insert_one(goals)
        goals.pop('_id', None)
    return goals

@api_router.put("/nutrition/goals")
async def update_nutrition_goals(request: Request, goal_data: NutritionGoalCreate, session_token: Optional[str] = Cookie(None)):
    """Update user's nutrition goals"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    existing = await db.nutrition_goals.find_one({"user_id": user.user_id}, {"_id": 0})
    
    if existing:
        await db.nutrition_goals.update_one(
            {"user_id": user.user_id},
            {"$set": {
                "daily_calories": goal_data.daily_calories,
                "daily_protein": goal_data.daily_protein,
                "daily_carbs": goal_data.daily_carbs,
                "daily_fat": goal_data.daily_fat,
                "water_goal_ml": goal_data.water_goal_ml,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    else:
        goal_id = f"ngoal_{uuid.uuid4().hex[:12]}"
        await db.nutrition_goals.insert_one({
            "goal_id": goal_id,
            "user_id": user.user_id,
            **goal_data.model_dump(),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        })
    
    updated = await db.nutrition_goals.find_one({"user_id": user.user_id}, {"_id": 0})
    return updated

@api_router.get("/nutrition/water")
async def get_water_logs(request: Request, date: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    """Get water logs for a date"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if not date:
        date = datetime.now().strftime("%Y-%m-%d")
    
    logs = await db.water_logs.find({"user_id": user.user_id, "date": date}, {"_id": 0}).to_list(100)
    total = sum(log.get("amount_ml", 0) for log in logs)
    return {"logs": logs, "total_ml": total, "date": date}

@api_router.post("/nutrition/water")
async def log_water(request: Request, amount_ml: int, date: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    """Log water intake"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if not date:
        date = datetime.now().strftime("%Y-%m-%d")
    
    log_id = f"water_{uuid.uuid4().hex[:12]}"
    log_doc = {
        "log_id": log_id,
        "user_id": user.user_id,
        "amount_ml": amount_ml,
        "date": date,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.water_logs.insert_one(log_doc)
    log_doc.pop('_id', None)
    return log_doc

@api_router.get("/nutrition/stats")
async def get_nutrition_stats(request: Request, date: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    """Get nutrition statistics for a date"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if not date:
        date = datetime.now().strftime("%Y-%m-%d")
    
    # Get meals for the date
    meals = await db.meals.find({"user_id": user.user_id, "date": date}, {"_id": 0}).to_list(100)
    
    # Calculate totals
    total_calories = sum(m.get("total_calories", 0) for m in meals)
    total_protein = sum(m.get("total_protein", 0) for m in meals)
    total_carbs = sum(m.get("total_carbs", 0) for m in meals)
    total_fat = sum(m.get("total_fat", 0) for m in meals)
    
    # Get water
    water_data = await db.water_logs.find({"user_id": user.user_id, "date": date}, {"_id": 0}).to_list(100)
    total_water = sum(w.get("amount_ml", 0) for w in water_data)
    
    # Get goals
    goals = await db.nutrition_goals.find_one({"user_id": user.user_id}, {"_id": 0})
    if not goals:
        goals = {"daily_calories": 2000, "daily_protein": 150, "daily_carbs": 250, "daily_fat": 65, "water_goal_ml": 2000}
    
    return {
        "date": date,
        "consumed": {
            "calories": total_calories,
            "protein": round(total_protein, 1),
            "carbs": round(total_carbs, 1),
            "fat": round(total_fat, 1),
            "water_ml": total_water
        },
        "goals": goals,
        "meals_count": len(meals),
        "remaining": {
            "calories": goals.get("daily_calories", 2000) - total_calories,
            "protein": round(goals.get("daily_protein", 150) - total_protein, 1),
            "carbs": round(goals.get("daily_carbs", 250) - total_carbs, 1),
            "fat": round(goals.get("daily_fat", 65) - total_fat, 1),
            "water_ml": goals.get("water_goal_ml", 2000) - total_water
        }
    }

@api_router.get("/nutrition/diets")
async def get_diets(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get user's diets"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    diets = await db.diets.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    return diets

@api_router.post("/nutrition/diets")
async def create_diet(request: Request, diet_data: DietCreate, session_token: Optional[str] = Cookie(None)):
    """Create a new diet plan"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    diet_id = f"diet_{uuid.uuid4().hex[:12]}"
    diet_doc = {
        "diet_id": diet_id,
        "user_id": user.user_id,
        "name": diet_data.name,
        "description": diet_data.description,
        "diet_type": diet_data.diet_type,
        "meals_plan": diet_data.meals_plan,
        "active": True,
        "start_date": diet_data.start_date,
        "end_date": diet_data.end_date,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.diets.insert_one(diet_doc)
    diet_doc.pop('_id', None)
    return diet_doc

@api_router.delete("/nutrition/diets/{diet_id}")
async def delete_diet(request: Request, diet_id: str, session_token: Optional[str] = Cookie(None)):
    """Delete a diet"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.diets.delete_one({"diet_id": diet_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Diet not found")
    return {"message": "Diet deleted"}

@api_router.get("/nutrition/recipes")
async def get_recipes(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get user's saved recipes"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    recipes = await db.recipes.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    return recipes

@api_router.post("/nutrition/recipes/suggest")
async def suggest_recipe(request: Request, preferences: dict, session_token: Optional[str] = Cookie(None)):
    """Get AI-suggested recipe based on preferences"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Get user's nutrition goals for context
    goals = await db.nutrition_goals.find_one({"user_id": user.user_id}, {"_id": 0})
    
    goal_info = ""
    if goals:
        goal_info = f"""
Metas nutricionais do usuário:
- Calorias diárias: {goals.get('daily_calories', 2000)} kcal
- Proteína: {goals.get('daily_protein', 150)}g
- Carboidratos: {goals.get('daily_carbs', 250)}g
- Gordura: {goals.get('daily_fat', 65)}g
"""
    
    diet_type = preferences.get("diet_type", "")
    meal_type = preferences.get("meal_type", "")
    ingredients = preferences.get("available_ingredients", [])
    restrictions = preferences.get("restrictions", [])
    cuisine = preferences.get("cuisine", "")
    max_time = preferences.get("max_prep_time_minutes", 60)
    
    prompt = f"""Sugira uma receita saudável com as seguintes preferências:
{goal_info}
- Tipo de refeição: {meal_type or 'qualquer'}
- Tipo de dieta: {diet_type or 'balanceada'}
- Ingredientes disponíveis: {', '.join(ingredients) if ingredients else 'qualquer'}
- Restrições alimentares: {', '.join(restrictions) if restrictions else 'nenhuma'}
- Culinária preferida: {cuisine or 'qualquer'}
- Tempo máximo de preparo: {max_time} minutos

Forneça a resposta em formato JSON com a seguinte estrutura:
{{
    "name": "Nome da Receita",
    "description": "Breve descrição",
    "ingredients": [{{"name": "ingrediente", "quantity": "quantidade", "unit": "unidade"}}],
    "instructions": ["Passo 1", "Passo 2"],
    "prep_time_minutes": 15,
    "cook_time_minutes": 30,
    "servings": 4,
    "calories_per_serving": 350,
    "protein_per_serving": 25,
    "carbs_per_serving": 40,
    "fat_per_serving": 12,
    "tags": ["saudável", "rápido"],
    "tips": "Dica extra"
}}"""
    
    if not gemini_client:
        raise HTTPException(status_code=503, detail="Serviço de IA não disponível")
    
    try:
        # Call Gemini directly with response_mime_type for guaranteed JSON output
        response = gemini_client.models.generate_content(
            model=GEMINI_MODEL,
            contents=prompt,
            config=types.GenerateContentConfig(
                system_instruction="Você é um nutricionista e chef experiente. Forneça receitas saudáveis e práticas. Sempre responda em JSON válido.",
                response_mime_type="application/json",
                response_schema={
                    "type": "OBJECT",
                    "required": ["name", "description", "ingredients", "instructions"],
                    "properties": {
                        "name": {"type": "STRING"},
                        "description": {"type": "STRING"},
                        "ingredients": {
                            "type": "ARRAY",
                            "items": {
                                "type": "OBJECT",
                                "properties": {
                                    "name": {"type": "STRING"},
                                    "quantity": {"type": "STRING"},
                                    "unit": {"type": "STRING"}
                                }
                            }
                        },
                        "instructions": {
                            "type": "ARRAY",
                            "items": {"type": "STRING"}
                        },
                        "prep_time_minutes": {"type": "INTEGER"},
                        "cook_time_minutes": {"type": "INTEGER"},
                        "servings": {"type": "INTEGER"},
                        "calories_per_serving": {"type": "INTEGER"},
                        "protein_per_serving": {"type": "INTEGER"},
                        "carbs_per_serving": {"type": "INTEGER"},
                        "fat_per_serving": {"type": "INTEGER"},
                        "tags": {
                            "type": "ARRAY",
                            "items": {"type": "STRING"}
                        },
                        "tips": {"type": "STRING"}
                    }
                }
            )
        )
        
        response_text = response.text.strip()
        
        # Parse JSON - should be valid since we used response_mime_type
        try:
            recipe_data = json.loads(response_text)
        except json.JSONDecodeError:
            # Fallback: clean markdown code blocks and retry
            cleaned = response_text
            if cleaned.startswith("```json"):
                cleaned = cleaned[7:]
            if cleaned.startswith("```"):
                cleaned = cleaned[3:]
            if cleaned.endswith("```"):
                cleaned = cleaned[:-3]
            cleaned = cleaned.strip()
            # Try fixing common JSON issues
            import re
            cleaned = re.sub(r',\s*}', '}', cleaned)  # Remove trailing commas before }
            cleaned = re.sub(r',\s*]', ']', cleaned)  # Remove trailing commas before ]
            recipe_data = json.loads(cleaned)
        
        # Save recipe
        recipe_id = f"recipe_{uuid.uuid4().hex[:12]}"
        recipe_doc = {
            "recipe_id": recipe_id,
            "user_id": user.user_id,
            "name": recipe_data.get("name", "Receita Sugerida"),
            "description": recipe_data.get("description", ""),
            "ingredients": recipe_data.get("ingredients", []),
            "instructions": recipe_data.get("instructions", []),
            "prep_time_minutes": recipe_data.get("prep_time_minutes", 0),
            "cook_time_minutes": recipe_data.get("cook_time_minutes", 0),
            "servings": recipe_data.get("servings", 1),
            "calories_per_serving": recipe_data.get("calories_per_serving", 0),
            "protein_per_serving": recipe_data.get("protein_per_serving", 0),
            "carbs_per_serving": recipe_data.get("carbs_per_serving", 0),
            "fat_per_serving": recipe_data.get("fat_per_serving", 0),
            "tags": recipe_data.get("tags", []),
            "ai_generated": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.recipes.insert_one(recipe_doc)
        recipe_doc.pop('_id', None)
        recipe_doc["tips"] = recipe_data.get("tips", "")
        return recipe_doc
    except json.JSONDecodeError as je:
        logging.error(f"Recipe JSON parse error: {je}")
        raise HTTPException(status_code=500, detail="Erro ao interpretar resposta da IA. Tente novamente.")
    except Exception as e:
        logging.error(f"Recipe suggestion failed: {e}")
        raise HTTPException(status_code=500, detail=f"Falha ao gerar receita: {str(e)}")

@api_router.get("/nutrition/recipes/{recipe_id}")
async def get_recipe_detail(request: Request, recipe_id: str, session_token: Optional[str] = Cookie(None)):
    """Get full recipe details"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    recipe = await db.recipes.find_one({"recipe_id": recipe_id, "user_id": user.user_id}, {"_id": 0})
    if not recipe:
        raise HTTPException(status_code=404, detail="Recipe not found")
    return recipe

@api_router.delete("/nutrition/recipes/{recipe_id}")
async def delete_recipe(request: Request, recipe_id: str, session_token: Optional[str] = Cookie(None)):
    """Delete a recipe"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.recipes.delete_one({"recipe_id": recipe_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Recipe not found")
    return {"message": "Recipe deleted"}

# ========== IMPORT MEAL PLAN ==========

@api_router.post("/nutrition/import-plan")
async def import_meal_plan(
    request: Request,
    file: UploadFile = File(...),
    session_token: Optional[str] = Cookie(None)
):
    """Import a meal plan from PDF or image file using AI extraction"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if not gemini_client:
        raise HTTPException(status_code=503, detail="Serviço de IA indisponível")
    
    allowed_types = ["application/pdf", "image/jpeg", "image/png", "image/webp", "image/heic"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Formato não suportado. Envie PDF, JPG, PNG ou WEBP.")
    
    import tempfile
    import os
    content = await file.read()
    suffix = ".pdf" if "pdf" in file.content_type else ".jpg" if "jpeg" in file.content_type else ".png" if "png" in file.content_type else ".webp"
    
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        
        uploaded_file = gemini_client.files.upload(file=tmp_path)
        
        mime = file.content_type
        prompt = """Analise este plano alimentar/dieta e extraia TODAS as refeições em formato JSON estruturado.

Para cada refeição, extraia:
- meal_type: "breakfast" (café da manhã), "lunch" (almoço), "dinner" (jantar), "snack" (lanche)
- name: nome da refeição
- time: horário sugerido (ex: "07:00")
- foods: lista de alimentos com quantidade
- calories: calorias estimadas (número)
- protein: proteína em gramas (número)
- carbs: carboidratos em gramas (número)
- fat: gordura em gramas (número)
- fiber: fibra em gramas (número, opcional)
- notes: observações adicionais

Também extraia informações gerais do plano:
- plan_name: nome do plano
- goal: objetivo (emagrecimento, hipertrofia, saúde, etc.)
- daily_calories: meta calórica diária total
- daily_protein: meta de proteína diária
- daily_carbs: meta de carboidratos diária
- daily_fat: meta de gordura diária
- restrictions: restrições alimentares mencionadas
- tips: dicas do nutricionista

Responda APENAS com JSON válido neste formato:
{
  "plan_name": "Nome do Plano",
  "goal": "objetivo",
  "daily_calories": 2000,
  "daily_protein": 150,
  "daily_carbs": 200,
  "daily_fat": 70,
  "restrictions": ["restrição 1"],
  "tips": ["dica 1", "dica 2"],
  "meals": [
    {
      "meal_type": "breakfast",
      "name": "Café da Manhã",
      "time": "07:00",
      "foods": [{"name": "Ovos mexidos", "quantity": "3 unidades", "calories": 210}],
      "calories": 350,
      "protein": 25,
      "carbs": 30,
      "fat": 15,
      "notes": ""
    }
  ]
}

IMPORTANTE: Retorne APENAS o JSON, sem markdown, sem ```json."""

        response = gemini_client.models.generate_content(
            model=GEMINI_MODEL,
            contents=[
                types.Part.from_uri(file_uri=uploaded_file.uri, mime_type=mime),
                prompt
            ],
            config=types.GenerateContentConfig(
                system_instruction="Você é um nutricionista especialista. Extraia com precisão todas as informações do plano alimentar."
            )
        )
        
        response_text = response.text.strip()
        if response_text.startswith("```"):
            response_text = response_text.split("\n", 1)[1] if "\n" in response_text else response_text[3:]
        if response_text.endswith("```"):
            response_text = response_text[:-3].strip()
        if response_text.startswith("json"):
            response_text = response_text[4:].strip()
        
        plan_data = json.loads(response_text)
        
        # Save the imported plan
        plan_id = f"mealplan_{uuid.uuid4().hex[:12]}"
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        
        plan_doc = {
            "plan_id": plan_id,
            "user_id": user.user_id,
            "name": plan_data.get("plan_name", "Plano Importado"),
            "goal": plan_data.get("goal", ""),
            "daily_calories": plan_data.get("daily_calories", 0),
            "daily_protein": plan_data.get("daily_protein", 0),
            "daily_carbs": plan_data.get("daily_carbs", 0),
            "daily_fat": plan_data.get("daily_fat", 0),
            "restrictions": plan_data.get("restrictions", []),
            "tips": plan_data.get("tips", []),
            "meals": plan_data.get("meals", []),
            "source": "imported",
            "source_filename": file.filename,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.meal_plans.insert_one(plan_doc)
        plan_doc.pop('_id', None)
        
        # Also create individual meal entries for today
        meals_created = 0
        for meal in plan_data.get("meals", []):
            meal_id = f"meal_{uuid.uuid4().hex[:12]}"
            meal_doc = {
                "meal_id": meal_id,
                "user_id": user.user_id,
                "date": today,
                "meal_type": meal.get("meal_type", "snack"),
                "name": meal.get("name", "Refeição importada"),
                "foods": meal.get("foods", []),
                "calories": meal.get("calories", 0),
                "protein": meal.get("protein", 0),
                "carbs": meal.get("carbs", 0),
                "fat": meal.get("fat", 0),
                "fiber": meal.get("fiber", 0),
                "notes": meal.get("notes", ""),
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.meals.insert_one(meal_doc)
            meals_created += 1
        
        # Update nutrition goals if plan has daily targets
        if plan_data.get("daily_calories"):
            await db.nutrition_goals.update_one(
                {"user_id": user.user_id},
                {"$set": {
                    "calories": plan_data.get("daily_calories", 2000),
                    "protein": plan_data.get("daily_protein", 150),
                    "carbs": plan_data.get("daily_carbs", 200),
                    "fat": plan_data.get("daily_fat", 70),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }},
                upsert=True
            )
        
        # Award XP
        xp_earned = 10
        new_xp = user.xp + xp_earned
        new_rank = calculate_rank(new_xp)
        await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
        
        os.unlink(tmp_path)
        
        return {
            "success": True,
            "plan": plan_doc,
            "meals_created": meals_created,
            "goals_updated": bool(plan_data.get("daily_calories")),
            "xp_earned": xp_earned
        }
        
    except json.JSONDecodeError as e:
        os.unlink(tmp_path)
        raise HTTPException(status_code=422, detail=f"Não foi possível extrair dados do arquivo. Tente com outro formato. Erro: {str(e)}")
    except Exception as e:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        logging.error(f"Import meal plan failed: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao importar plano: {str(e)}")

# ========== STUDY ENDPOINTS ==========

@api_router.get("/study/areas")
async def get_study_areas(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get all study areas"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    areas = await db.study_areas.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    
    # Create default areas if none exist
    if not areas:
        default_areas = [
            {"name": "Faculdade", "color": "#007AFF", "icon": "graduation-cap", "order": 0},
            {"name": "Concursos", "color": "#10B981", "icon": "file-text", "order": 1},
            {"name": "Trabalho", "color": "#F59E0B", "icon": "briefcase", "order": 2},
            {"name": "Outros", "color": "#8B5CF6", "icon": "folder", "order": 3}
        ]
        for i, area in enumerate(default_areas):
            area_id = f"area_{uuid.uuid4().hex[:12]}"
            area_doc = {
                "area_id": area_id,
                "user_id": user.user_id,
                **area,
                "description": None,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.study_areas.insert_one(area_doc)
            area_doc.pop('_id', None)
            areas.append(area_doc)
    
    return areas

@api_router.post("/study/areas")
async def create_study_area(request: Request, area_data: StudyAreaCreate, session_token: Optional[str] = Cookie(None)):
    """Create a new study area"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Get max order
    existing = await db.study_areas.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    max_order = max([a.get("order", 0) for a in existing], default=-1) + 1
    
    area_id = f"area_{uuid.uuid4().hex[:12]}"
    area_doc = {
        "area_id": area_id,
        "user_id": user.user_id,
        "name": area_data.name,
        "description": area_data.description,
        "color": area_data.color,
        "icon": area_data.icon,
        "order": max_order,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.study_areas.insert_one(area_doc)
    area_doc.pop('_id', None)
    return area_doc

@api_router.delete("/study/areas/{area_id}")
async def delete_study_area(request: Request, area_id: str, session_token: Optional[str] = Cookie(None)):
    """Delete a study area"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.study_areas.delete_one({"area_id": area_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Area not found")
    
    # Also delete related notebooks
    await db.notebooks.delete_many({"area_id": area_id, "user_id": user.user_id})
    # Also delete related programs
    await db.study_programs.delete_many({"area_id": area_id, "user_id": user.user_id})
    
    return {"message": "Area deleted"}

# ========== STUDY PROGRAMS ==========

@api_router.get("/study/programs")
async def get_study_programs(request: Request, area_id: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    """Get study programs, optionally filtered by area"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    query = {"user_id": user.user_id}
    if area_id:
        query["area_id"] = area_id
    
    programs = await db.study_programs.find(query, {"_id": 0}).to_list(100)
    
    # Enrich with notebook counts and question stats
    for prog in programs:
        nb_count = await db.notebooks.count_documents({"program_id": prog["program_id"], "user_id": user.user_id})
        prog["notebooks_count"] = nb_count
        # Get aggregated questions from notebooks in this program
        nbs = await db.notebooks.find({"program_id": prog["program_id"], "user_id": user.user_id}, {"_id": 0}).to_list(100)
        prog["total_questions"] = sum(n.get("total_questions", 0) for n in nbs)
        prog["correct_questions"] = sum(n.get("correct_questions", 0) for n in nbs)
        total_time = sum(n.get("total_study_time_minutes", 0) for n in nbs)
        prog["total_study_time_minutes"] = total_time
    
    return programs

@api_router.post("/study/programs")
async def create_study_program(request: Request, program_data: StudyProgramCreate, session_token: Optional[str] = Cookie(None)):
    """Create a new study program/course"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    program_id = f"prog_{uuid.uuid4().hex[:12]}"
    program_doc = {
        "program_id": program_id,
        "user_id": user.user_id,
        "area_id": program_data.area_id,
        "name": program_data.name,
        "description": program_data.description,
        "color": program_data.color,
        "icon": program_data.icon,
        "target_date": program_data.target_date,
        "status": "active",
        "total_questions": 0,
        "correct_questions": 0,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.study_programs.insert_one(program_doc)
    program_doc.pop('_id', None)
    return program_doc

@api_router.patch("/study/programs/{program_id}")
async def update_study_program(request: Request, program_id: str, data: dict, session_token: Optional[str] = Cookie(None)):
    """Update a study program"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    update_fields = {}
    for field in ["name", "description", "color", "icon", "target_date", "status"]:
        if field in data:
            update_fields[field] = data[field]
    
    if update_fields:
        await db.study_programs.update_one(
            {"program_id": program_id, "user_id": user.user_id},
            {"$set": update_fields}
        )
    
    updated = await db.study_programs.find_one({"program_id": program_id, "user_id": user.user_id}, {"_id": 0})
    if not updated:
        raise HTTPException(status_code=404, detail="Program not found")
    return updated

@api_router.delete("/study/programs/{program_id}")
async def delete_study_program(request: Request, program_id: str, session_token: Optional[str] = Cookie(None)):
    """Delete a study program and its notebooks"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.study_programs.delete_one({"program_id": program_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Program not found")
    
    # Delete related notebooks and their content
    notebooks = await db.notebooks.find({"program_id": program_id, "user_id": user.user_id}, {"_id": 0}).to_list(100)
    for nb in notebooks:
        nb_id = nb.get("notebook_id")
        await db.study_notes.delete_many({"notebook_id": nb_id, "user_id": user.user_id})
        await db.flashcards.delete_many({"notebook_id": nb_id, "user_id": user.user_id})
        await db.quizzes.delete_many({"notebook_id": nb_id, "user_id": user.user_id})
    await db.notebooks.delete_many({"program_id": program_id, "user_id": user.user_id})
    
    return {"message": "Program deleted"}

# ========== IMPORT EDITAL - AI STUDY PROGRAM GENERATOR ==========

@api_router.post("/study/programs/import-edital")
async def import_edital(
    request: Request,
    file: UploadFile = File(...),
    area_id: str = Form(...),
    target_date: Optional[str] = Form(None),
    hours_per_day: float = Form(4.0),
    days_per_week: int = Form(5),
    session_token: Optional[str] = Cookie(None)
):
    """Upload a PDF of an edital (public exam notice) and generate a complete study program with AI"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Apenas arquivos PDF são aceitos")
    
    if not gemini_client:
        raise HTTPException(status_code=500, detail="Serviço de IA indisponível")
    
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Arquivo muito grande. Limite de 20MB.")
    
    try:
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp_file:
            tmp_file.write(content)
            tmp_path = tmp_file.name
        
        uploaded_file = gemini_client.files.upload(file=tmp_path)
        
        system_msg = f"""Você é um especialista em concursos públicos brasileiros e planejamento de estudos.
Analise o edital do concurso contido neste PDF e extraia TODAS as informações relevantes para criar um programa de estudos completo.

ATENÇÃO ESPECIAL: Extraia o CONTEÚDO PROGRAMÁTICO COMPLETO de cada disciplina. A maioria dos editais traz uma seção de "Conteúdo Programático" ou "Programa" listando todos os assuntos cobrados em cada matéria. Extraia TODOS esses assuntos fielmente.

O aluno tem {hours_per_day} horas disponíveis por dia, {days_per_week} dias por semana para estudar.
{"A data da prova é: " + target_date + ". Considere o tempo disponível até a prova para o cronograma." if target_date else "Não há data definida para a prova."}

Responda APENAS com JSON válido no formato abaixo. NÃO inclua texto antes ou depois do JSON.

{{
  "concurso": {{
    "nome": "Nome completo do concurso",
    "orgao": "Órgão/instituição",
    "banca": "Banca organizadora",
    "cargo": "Cargo(s) principal(is)",
    "vagas": "Número de vagas se informado",
    "remuneracao": "Remuneração se informada",
    "escolaridade": "Nível de escolaridade exigido",
    "data_prova": "Data da prova se informada no edital"
  }},
  "disciplinas": [
    {{
      "nome": "Nome da Disciplina/Matéria",
      "peso": 3,
      "num_questoes": 10,
      "topicos": ["Tópico resumido 1", "Tópico resumido 2"],
      "conteudo_programatico": [
        {{
          "assunto": "Nome do assunto/tópico principal",
          "subtopicos": ["Subtópico 1", "Subtópico 2", "Subtópico 3"]
        }}
      ],
      "dificuldade": "alta",
      "dicas_estudo": "Dica específica para esta matéria",
      "recursos_recomendados": "Livros, materiais recomendados"
    }}
  ],
  "cronograma_semanal": [
    {{
      "dia": "Segunda",
      "blocos": [
        {{
          "disciplina": "Nome da Disciplina",
          "duracao_minutos": 120,
          "tipo_estudo": "Teoria + Questões",
          "prioridade": "alta",
          "assuntos_foco": ["Assunto principal a estudar neste bloco"]
        }}
      ]
    }}
  ],
  "estrategia": {{
    "resumo": "Resumo da estratégia de estudo recomendada",
    "fase_1": "Descrição da primeira fase de estudo (base teórica)",
    "fase_2": "Descrição da segunda fase (aprofundamento + questões)",
    "fase_3": "Descrição da terceira fase (revisão + simulados)",
    "dicas_gerais": ["Dica 1", "Dica 2", "Dica 3"],
    "materias_prioritarias": ["Matéria com maior peso/importância"],
    "horas_semanais_total": {hours_per_day * days_per_week}
  }}
}}

REGRAS IMPORTANTES:
- Extraia TODAS as disciplinas/matérias mencionadas no edital
- O campo "peso" deve refletir a importância relativa (1-5, sendo 5 o mais importante) baseado no número de questões e peso na nota
- "num_questoes" é o número de questões da disciplina conforme o edital
- O cronograma deve distribuir as matérias pela semana priorizando as de maior peso
- A soma dos minutos por dia deve respeitar o limite de {int(hours_per_day * 60)} minutos
- Inclua APENAS {days_per_week} dias no cronograma (ex: Segunda a Sexta se 5 dias)
- Alterne matérias pesadas com leves no mesmo dia
- Reserve tempo para revisão e questões
- CONTEÚDO PROGRAMÁTICO: Este é o campo MAIS IMPORTANTE. Extraia FIELMENTE todos os assuntos do conteúdo programático de cada disciplina conforme listado no edital. Cada assunto principal deve virar um item em "conteudo_programatico" com seus respectivos subtópicos. Se o edital não tiver conteúdo programático detalhado, deixe o array vazio [].
- "topicos" é um RESUMO dos principais tópicos (máximo 10 itens simples). "conteudo_programatico" é a LISTA COMPLETA E DETALHADA dos assuntos cobrados.
- "assuntos_foco" nos blocos do cronograma deve indicar quais assuntos específicos do conteúdo programático o aluno deve focar naquele bloco de estudo.
- Tudo em português brasileiro
- "dificuldade" deve ser: "baixa", "media" ou "alta"
- Os dias devem ser: "Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"
"""

        response = gemini_client.models.generate_content(
            model=GEMINI_MODEL,
            contents=[
                types.Part.from_uri(file_uri=uploaded_file.uri, mime_type="application/pdf"),
                "Analise este edital de concurso e gere um programa de estudos completo em JSON conforme as instruções:"
            ],
            config=types.GenerateContentConfig(
                system_instruction=system_msg
            )
        )
        
        os.unlink(tmp_path)
        
        json_str = response.text.strip()
        if json_str.startswith("```json"):
            json_str = json_str[7:]
        if json_str.startswith("```"):
            json_str = json_str[3:]
        if json_str.endswith("```"):
            json_str = json_str[:-3]
        
        parsed = json.loads(json_str.strip())
        
        concurso_info = parsed.get("concurso", {})
        disciplinas = parsed.get("disciplinas", [])
        cronograma = parsed.get("cronograma_semanal", [])
        estrategia = parsed.get("estrategia", {})
        
        if not disciplinas:
            raise HTTPException(status_code=400, detail="Não foi possível extrair disciplinas do edital. Verifique se o PDF contém o conteúdo programático.")
        
        # Create the study program
        program_id = f"prog_{uuid.uuid4().hex[:12]}"
        program_name = concurso_info.get("nome", "Programa do Concurso")
        if concurso_info.get("cargo"):
            program_name = f"{concurso_info['nome']} - {concurso_info['cargo']}"
        
        program_doc = {
            "program_id": program_id,
            "user_id": user.user_id,
            "area_id": area_id,
            "name": program_name[:100],
            "description": f"Programa gerado a partir do edital. Banca: {concurso_info.get('banca', 'N/A')} | Órgão: {concurso_info.get('orgao', 'N/A')}",
            "color": "#8B5CF6",
            "icon": "file-text",
            "target_date": target_date or concurso_info.get("data_prova"),
            "status": "active",
            "total_questions": 0,
            "correct_questions": 0,
            "source_type": "edital_import",
            "edital_data": {
                "concurso": concurso_info,
                "estrategia": estrategia,
                "total_disciplinas": len(disciplinas),
                "hours_per_day": hours_per_day,
                "days_per_week": days_per_week,
                "pdf_filename": file.filename
            },
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.study_programs.insert_one(program_doc)
        program_doc.pop('_id', None)
        
        # Create notebooks for each discipline
        created_notebooks = []
        color_palette = ["#007AFF", "#34C759", "#FF9500", "#FF3B30", "#AF52DE", "#5AC8FA", "#FF2D55", "#FFCC00", "#30D158", "#64D2FF", "#BF5AF2", "#FF6482"]
        
        for i, disc in enumerate(disciplinas):
            nb_id = f"nb_{uuid.uuid4().hex[:12]}"
            nb_color = color_palette[i % len(color_palette)]
            
            nb_doc = {
                "notebook_id": nb_id,
                "user_id": user.user_id,
                "area_id": area_id,
                "program_id": program_id,
                "name": disc.get("nome", f"Disciplina {i+1}"),
                "description": disc.get("dicas_estudo", ""),
                "color": nb_color,
                "tags": disc.get("topicos", [])[:10],
                "total_study_time_minutes": 0,
                "total_questions": 0,
                "correct_questions": 0,
                "weight": disc.get("peso", 1),
                "num_questoes_edital": disc.get("num_questoes", 0),
                "dificuldade": disc.get("dificuldade", "media"),
                "topicos": disc.get("topicos", []),
                "conteudo_programatico": disc.get("conteudo_programatico", []),
                "recursos_recomendados": disc.get("recursos_recomendados", ""),
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.notebooks.insert_one(nb_doc)
            nb_doc.pop('_id', None)
            created_notebooks.append(nb_doc)
        
        # Create study schedule entries from cronograma
        created_schedules = []
        day_map = {
            "Segunda": "monday", "Terça": "tuesday", "Quarta": "wednesday",
            "Quinta": "thursday", "Sexta": "friday", "Sábado": "saturday", "Domingo": "sunday"
        }
        
        for day_entry in cronograma:
            dia = day_entry.get("dia", "")
            day_of_week = day_map.get(dia, dia.lower())
            blocos = day_entry.get("blocos", [])
            
            current_hour = 8
            current_min = 0
            
            for bloco in blocos:
                disc_name = bloco.get("disciplina", "")
                duracao = bloco.get("duracao_minutos", 60)
                
                matching_nb = None
                for nb in created_notebooks:
                    if nb["name"].lower() == disc_name.lower():
                        matching_nb = nb
                        break
                if not matching_nb:
                    for nb in created_notebooks:
                        if disc_name.lower() in nb["name"].lower() or nb["name"].lower() in disc_name.lower():
                            matching_nb = nb
                            break
                
                if matching_nb:
                    start_time = f"{current_hour:02d}:{current_min:02d}"
                    end_min_total = current_min + duracao
                    end_hour = current_hour + end_min_total // 60
                    end_min = end_min_total % 60
                    end_time = f"{end_hour:02d}:{end_min:02d}"
                    
                    sched_id = f"sched_{uuid.uuid4().hex[:12]}"
                    sched_doc = {
                        "schedule_id": sched_id,
                        "user_id": user.user_id,
                        "notebook_id": matching_nb["notebook_id"],
                        "program_id": program_id,
                        "day_of_week": day_of_week,
                        "start_time": start_time,
                        "end_time": end_time,
                        "repeat": True,
                        "tipo_estudo": bloco.get("tipo_estudo", "Teoria + Questões"),
                        "prioridade": bloco.get("prioridade", "media"),
                        "assuntos_foco": bloco.get("assuntos_foco", []),
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db.study_schedules.insert_one(sched_doc)
                    sched_doc.pop('_id', None)
                    created_schedules.append(sched_doc)
                    
                    current_hour = end_hour
                    current_min = end_min
        
        # Award XP
        xp_earned = 25
        await award_xp(user.user_id, xp_earned)
        
        return {
            "success": True,
            "program": program_doc,
            "concurso": concurso_info,
            "disciplinas": created_notebooks,
            "cronograma": cronograma,
            "estrategia": estrategia,
            "schedules_created": len(created_schedules),
            "xp_earned": xp_earned,
            "message": f"Programa de estudos criado com {len(disciplinas)} disciplinas e cronograma semanal de {len(created_schedules)} blocos!"
        }
        
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar resposta da IA. Tente novamente. Detalhe: {str(e)}")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao analisar edital: {str(e)}")


@api_router.get("/study/programs/{program_id}/cronograma")
async def get_program_cronograma(request: Request, program_id: str, session_token: Optional[str] = Cookie(None)):
    """Get the study schedule/cronograma for a specific program"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    program = await db.study_programs.find_one({"program_id": program_id, "user_id": user.user_id}, {"_id": 0})
    if not program:
        raise HTTPException(status_code=404, detail="Programa não encontrado")
    
    schedules = await db.study_schedules.find(
        {"program_id": program_id, "user_id": user.user_id}, {"_id": 0}
    ).to_list(200)
    
    notebooks = await db.notebooks.find(
        {"program_id": program_id, "user_id": user.user_id}, {"_id": 0}
    ).to_list(100)
    
    nb_lookup = {nb["notebook_id"]: nb for nb in notebooks}
    
    for sched in schedules:
        nb = nb_lookup.get(sched.get("notebook_id"))
        if nb:
            sched["disciplina_nome"] = nb["name"]
            sched["disciplina_color"] = nb.get("color", "#007AFF")
            sched["weight"] = nb.get("weight", 1)
            # Include assuntos_foco if available
            if not sched.get("assuntos_foco"):
                sched["assuntos_foco"] = []
    
    days_order = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
    day_labels = {"monday": "Segunda", "tuesday": "Terça", "wednesday": "Quarta", "thursday": "Quinta", "friday": "Sexta", "saturday": "Sábado", "sunday": "Domingo"}
    
    cronograma_by_day = []
    for day in days_order:
        day_schedules = [s for s in schedules if s.get("day_of_week") == day]
        if day_schedules:
            day_schedules.sort(key=lambda x: x.get("start_time", "00:00"))
            total_minutes = 0
            for s in day_schedules:
                try:
                    start_parts = s.get("start_time", "00:00").split(":")
                    end_parts = s.get("end_time", "00:00").split(":")
                    start_min = int(start_parts[0]) * 60 + int(start_parts[1])
                    end_min = int(end_parts[0]) * 60 + int(end_parts[1])
                    total_minutes += (end_min - start_min)
                except (ValueError, IndexError):
                    pass
            cronograma_by_day.append({
                "day": day,
                "day_label": day_labels.get(day, day),
                "blocos": day_schedules,
                "total_minutes": total_minutes
            })
    
    weight_summary = []
    total_weight = sum(nb.get("weight", 1) for nb in notebooks)
    for nb in sorted(notebooks, key=lambda x: x.get("weight", 1), reverse=True):
        w = nb.get("weight", 1)
        weight_summary.append({
            "disciplina": nb["name"],
            "peso": w,
            "percentual": round((w / total_weight * 100) if total_weight > 0 else 0, 1),
            "num_questoes": nb.get("num_questoes_edital", 0),
            "dificuldade": nb.get("dificuldade", "media"),
            "color": nb.get("color", "#007AFF"),
            "topicos": nb.get("topicos", []),
            "conteudo_programatico": nb.get("conteudo_programatico", [])
        })
    
    return {
        "program": program,
        "cronograma": cronograma_by_day,
        "disciplinas": weight_summary,
        "notebooks": notebooks,
        "estrategia": program.get("edital_data", {}).get("estrategia", {}),
        "total_schedules": len(schedules)
    }


@api_router.get("/study/programs/{program_id}/edital-verticalizado")
async def get_edital_verticalizado(request: Request, program_id: str, session_token: Optional[str] = Cookie(None)):
    """Generate a verticalized edital view - organized list of all disciplines and their detailed content/topics"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    program = await db.study_programs.find_one({"program_id": program_id, "user_id": user.user_id}, {"_id": 0})
    if not program:
        raise HTTPException(status_code=404, detail="Programa não encontrado")
    
    notebooks = await db.notebooks.find(
        {"program_id": program_id, "user_id": user.user_id}, {"_id": 0}
    ).to_list(100)
    
    if not notebooks:
        raise HTTPException(status_code=404, detail="Nenhuma disciplina encontrada para este programa")
    
    concurso_info = program.get("edital_data", {}).get("concurso", {})
    cargo_info = program.get("edital_data", {}).get("cargo_selecionado", {})
    
    # Build verticalized view
    disciplinas_verticalizadas = []
    total_assuntos = 0
    
    # Check if all num_questoes_edital are the same (AI likely used total instead of per-discipline)
    all_questoes = [nb.get("num_questoes_edital", 0) for nb in notebooks]
    all_same_questoes = len(set(all_questoes)) == 1 and all_questoes[0] > 0 and len(all_questoes) > 1
    
    for nb in sorted(notebooks, key=lambda x: x.get("weight", 1), reverse=True):
        conteudo = nb.get("conteudo_programatico", [])
        topicos = nb.get("topicos", [])
        
        # Count total assuntos
        assuntos_count = len(conteudo)
        subtopicos_count = sum(len(item.get("subtopicos", [])) for item in conteudo)
        total_assuntos += assuntos_count + subtopicos_count
        
        disc_entry = {
            "notebook_id": nb.get("notebook_id"),
            "nome": nb.get("name", ""),
            "peso": nb.get("weight", 1),
            "num_questoes": 0 if all_same_questoes else nb.get("num_questoes_edital", 0),
            "dificuldade": nb.get("dificuldade", "media"),
            "grupo": nb.get("grupo", ""),
            "color": nb.get("color", "#007AFF"),
            "topicos": topicos,
            "conteudo_programatico": conteudo,
            "total_assuntos": assuntos_count,
            "total_subtopicos": subtopicos_count,
            "study_hours": round(nb.get("total_study_time_minutes", 0) / 60, 1),
            "total_questions_answered": nb.get("total_questions", 0),
            "accuracy": round((nb.get("correct_questions", 0) / nb.get("total_questions", 1) * 100) if nb.get("total_questions", 0) > 0 else 0, 1)
        }
        disciplinas_verticalizadas.append(disc_entry)
    
    return {
        "success": True,
        "program_id": program_id,
        "program_name": program.get("name", ""),
        "concurso": concurso_info,
        "cargo": cargo_info.get("nome", concurso_info.get("cargo", "")),
        "banca": concurso_info.get("banca", ""),
        "orgao": concurso_info.get("orgao", ""),
        "target_date": program.get("target_date"),
        "total_disciplinas": len(disciplinas_verticalizadas),
        "total_assuntos": total_assuntos,
        "disciplinas": disciplinas_verticalizadas
    }



@api_router.post("/study/programs/{program_id}/update-disciplinas")
async def update_program_disciplinas(request: Request, program_id: str, data: dict, session_token: Optional[str] = Cookie(None)):
    """Batch update disciplines (weight, difficulty, user_difficulty) and optionally regenerate schedule"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    program = await db.study_programs.find_one({"program_id": program_id, "user_id": user.user_id}, {"_id": 0})
    if not program:
        raise HTTPException(status_code=404, detail="Programa não encontrado")
    
    # Update program name if provided
    if data.get("program_name"):
        await db.study_programs.update_one(
            {"program_id": program_id, "user_id": user.user_id},
            {"$set": {"name": data["program_name"]}}
        )
    
    # Update each discipline
    disciplinas = data.get("disciplinas", [])
    updated_count = 0
    for disc in disciplinas:
        nb_id = disc.get("notebook_id")
        if not nb_id:
            continue
        update_fields = {}
        for field in ["weight", "dificuldade", "user_difficulty", "name", "topicos"]:
            if field in disc:
                update_fields[field] = disc[field]
        if update_fields:
            await db.notebooks.update_one(
                {"notebook_id": nb_id, "user_id": user.user_id},
                {"$set": update_fields}
            )
            updated_count += 1
    
    # Regenerate schedule if requested
    if data.get("regenerate_schedule"):
        # Delete old schedules for this program
        await db.study_schedules.delete_many({"program_id": program_id, "user_id": user.user_id})
        
        hours_per_day = data.get("hours_per_day", program.get("edital_data", {}).get("hours_per_day", 4))
        days_per_week = data.get("days_per_week", program.get("edital_data", {}).get("days_per_week", 5))
        
        # Get updated notebooks
        notebooks = await db.notebooks.find(
            {"program_id": program_id, "user_id": user.user_id}, {"_id": 0}
        ).to_list(100)
        
        if notebooks:
            # Calculate time allocation based on weights and user_difficulty
            total_weight = 0
            for nb in notebooks:
                w = nb.get("weight", 1)
                ud = nb.get("user_difficulty", "media")
                multiplier = 1.3 if ud == "alta" else (1.0 if ud == "media" else 0.8)
                nb["effective_weight"] = w * multiplier
                total_weight += nb["effective_weight"]
            
            minutes_per_day = int(hours_per_day * 60)
            day_names = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"][:days_per_week]
            day_labels_pt = {"monday": "Segunda", "tuesday": "Terça", "wednesday": "Quarta", "thursday": "Quinta", "friday": "Sexta", "saturday": "Sábado", "sunday": "Domingo"}
            
            # Sort by effective weight descending
            sorted_nbs = sorted(notebooks, key=lambda x: x.get("effective_weight", 1), reverse=True)
            
            # Distribute subjects across days
            created_schedules = 0
            block_count = 0
            for day_idx, day in enumerate(day_names):
                current_hour = 8
                current_min = 0
                remaining_minutes = minutes_per_day
                
                # Assign 2-3 subjects per day, rotating through
                subjects_today = []
                for i, nb in enumerate(sorted_nbs):
                    if i % days_per_week == day_idx or (i + 1) % days_per_week == day_idx:
                        subjects_today.append(nb)
                
                if not subjects_today:
                    subjects_today = [sorted_nbs[day_idx % len(sorted_nbs)]]
                
                # Allocate time proportionally
                today_total_weight = sum(s.get("effective_weight", 1) for s in subjects_today)
                for nb in subjects_today:
                    if remaining_minutes <= 0:
                        break
                    proportion = nb.get("effective_weight", 1) / today_total_weight if today_total_weight > 0 else 1
                    duracao = max(30, min(int(minutes_per_day * proportion), remaining_minutes))
                    duracao = (duracao // 15) * 15  # Round to 15-min blocks
                    if duracao <= 0:
                        continue
                    
                    start_time = f"{current_hour:02d}:{current_min:02d}"
                    end_min_total = current_min + duracao
                    end_hour = current_hour + end_min_total // 60
                    end_min = end_min_total % 60
                    end_time = f"{end_hour:02d}:{end_min:02d}"
                    
                    ud = nb.get("user_difficulty", "media")
                    # Cycle through study types: Teoria, Questões, Revisão
                    study_types_cycle = ["📖 Teoria", "📝 Questões", "🔄 Revisão"]
                    tipo = study_types_cycle[block_count % 3]
                    if ud == "alta":
                        # For difficult subjects, more questions and review
                        study_types_hard = ["📖 Teoria", "📝 Questões", "📝 Questões", "🔄 Revisão"]
                        tipo = study_types_hard[block_count % 4]
                    
                    sched_id = f"sched_{uuid.uuid4().hex[:12]}"
                    sched_doc = {
                        "schedule_id": sched_id,
                        "user_id": user.user_id,
                        "notebook_id": nb["notebook_id"],
                        "program_id": program_id,
                        "day_of_week": day,
                        "start_time": start_time,
                        "end_time": end_time,
                        "repeat": True,
                        "tipo_estudo": tipo,
                        "prioridade": "alta" if ud == "alta" else ("media" if ud == "media" else "baixa"),
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db.study_schedules.insert_one(sched_doc)
                    created_schedules += 1
                    block_count += 1
                    
                    current_hour = end_hour
                    current_min = end_min
                    remaining_minutes -= duracao
        
        return {"success": True, "updated": updated_count, "schedules_regenerated": created_schedules, "message": f"{updated_count} disciplinas atualizadas e cronograma regenerado com {created_schedules} blocos!"}
    
    return {"success": True, "updated": updated_count, "message": f"{updated_count} disciplinas atualizadas!"}


@api_router.get("/study/programs/{program_id}/study-indicators")
async def get_program_study_indicators(request: Request, program_id: str, session_token: Optional[str] = Cookie(None)):
    """Get study progress indicators per discipline for a program"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    notebooks = await db.notebooks.find(
        {"program_id": program_id, "user_id": user.user_id}, {"_id": 0}
    ).to_list(100)
    
    indicators = []
    for nb in notebooks:
        nb_id = nb["notebook_id"]
        
        # Get question logs for this notebook
        q_logs = await db.question_logs.find(
            {"notebook_id": nb_id, "user_id": user.user_id}, {"_id": 0}
        ).to_list(1000)
        
        total_q = sum(q.get("total", 0) for q in q_logs)
        correct_q = sum(q.get("correct", 0) for q in q_logs)
        accuracy = round((correct_q / total_q * 100) if total_q > 0 else 0, 1)
        
        # Get focus sessions for this notebook
        sessions = await db.study_sessions.find(
            {"notebook_id": nb_id, "user_id": user.user_id}, {"_id": 0}
        ).to_list(1000)
        total_study_minutes = sum(s.get("duration_minutes", 0) for s in sessions)
        
        # Get flashcard count and due count
        flashcards = await db.flashcards.find(
            {"notebook_id": nb_id, "user_id": user.user_id}, {"_id": 0}
        ).to_list(1000)
        today_str = datetime.now().strftime("%Y-%m-%d")
        due_cards = sum(1 for f in flashcards if f.get("next_review", "") <= today_str)
        
        # Get notes count
        notes_count = await db.study_notes.count_documents({"notebook_id": nb_id, "user_id": user.user_id})
        
        # Calculate progress estimate
        edital_questions = nb.get("num_questoes_edital", 0)
        target_questions = max(edital_questions * 10, 100)  # Target: 10x the edital questions
        q_progress = min(100, round((total_q / target_questions * 100) if target_questions > 0 else 0))
        
        indicators.append({
            "notebook_id": nb_id,
            "name": nb["name"],
            "color": nb.get("color", "#007AFF"),
            "weight": nb.get("weight", 1),
            "dificuldade": nb.get("dificuldade", "media"),
            "user_difficulty": nb.get("user_difficulty", "media"),
            "num_questoes_edital": edital_questions,
            "total_questions_answered": total_q,
            "correct_questions": correct_q,
            "accuracy": accuracy,
            "total_study_minutes": total_study_minutes + nb.get("total_study_time_minutes", 0),
            "study_hours": round((total_study_minutes + nb.get("total_study_time_minutes", 0)) / 60, 1),
            "flashcards_total": len(flashcards),
            "flashcards_due": due_cards,
            "notes_count": notes_count,
            "sessions_count": len(sessions),
            "question_progress": q_progress,
            "topicos": nb.get("topicos", []),
        })
    
    # Sort by weight descending
    indicators.sort(key=lambda x: x.get("weight", 1), reverse=True)
    
    return {"program_id": program_id, "indicators": indicators}


# ========== QUESTION TRACKING ==========

@api_router.post("/study/questions/log")
async def log_questions(request: Request, data: QuestionLogCreate, session_token: Optional[str] = Cookie(None)):
    """Log questions answered for a notebook/subject"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    today = datetime.now().strftime("%Y-%m-%d")
    incorrect = data.total - data.correct
    
    # Get notebook to find program_id
    notebook = await db.notebooks.find_one({"notebook_id": data.notebook_id, "user_id": user.user_id}, {"_id": 0})
    program_id = notebook.get("program_id") if notebook else None
    
    log_id = f"qlog_{uuid.uuid4().hex[:12]}"
    log_doc = {
        "log_id": log_id,
        "user_id": user.user_id,
        "notebook_id": data.notebook_id,
        "program_id": program_id,
        "total": data.total,
        "correct": data.correct,
        "incorrect": incorrect,
        "source": data.source,
        "date": today,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.question_logs.insert_one(log_doc)
    
    # Update notebook question counters
    await db.notebooks.update_one(
        {"notebook_id": data.notebook_id, "user_id": user.user_id},
        {"$inc": {"total_questions": data.total, "correct_questions": data.correct}}
    )
    
    # Award XP: 2 XP per correct answer
    xp_earned = data.correct
    new_xp = user.xp + xp_earned
    new_rank = calculate_rank(new_xp)
    await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
    
    # Update study streak
    await update_study_streak(user.user_id)
    
    log_doc.pop('_id', None)
    log_doc["xp_earned"] = xp_earned
    log_doc["new_xp"] = new_xp
    return log_doc

@api_router.get("/study/questions/stats")
async def get_question_stats(request: Request, notebook_id: Optional[str] = None, program_id: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    """Get question statistics"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    query = {"user_id": user.user_id}
    if notebook_id:
        query["notebook_id"] = notebook_id
    if program_id:
        query["program_id"] = program_id
    
    logs = await db.question_logs.find(query, {"_id": 0}).to_list(5000)
    
    total = sum(log.get("total", 0) for log in logs)
    correct = sum(log.get("correct", 0) for log in logs)
    incorrect = sum(log.get("incorrect", 0) for log in logs)
    
    # Stats by date (last 30 days)
    daily_stats = {}
    for log in logs:
        date = log.get("date", "")
        if date not in daily_stats:
            daily_stats[date] = {"total": 0, "correct": 0, "incorrect": 0}
        daily_stats[date]["total"] += log.get("total", 0)
        daily_stats[date]["correct"] += log.get("correct", 0)
        daily_stats[date]["incorrect"] += log.get("incorrect", 0)
    
    # Stats by notebook
    by_notebook = {}
    for log in logs:
        nb_id = log.get("notebook_id", "")
        if nb_id not in by_notebook:
            by_notebook[nb_id] = {"total": 0, "correct": 0, "incorrect": 0}
        by_notebook[nb_id]["total"] += log.get("total", 0)
        by_notebook[nb_id]["correct"] += log.get("correct", 0)
        by_notebook[nb_id]["incorrect"] += log.get("incorrect", 0)
    
    return {
        "total_questions": total,
        "correct": correct,
        "incorrect": incorrect,
        "accuracy": round((correct / total * 100), 1) if total > 0 else 0,
        "daily_stats": daily_stats,
        "by_notebook": by_notebook
    }

# ========== FOCUS/POMODORO ==========

@api_router.post("/study/focus/complete")
async def complete_focus_session(request: Request, data: FocusSessionCreate, session_token: Optional[str] = Cookie(None)):
    """Complete a focus/pomodoro session"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    today = datetime.now().strftime("%Y-%m-%d")
    
    # XP: 10 XP per 25 min completed
    xp_earned = max(3, (data.focus_minutes // 25) * 5)
    
    focus_id = f"focus_{uuid.uuid4().hex[:12]}"
    focus_doc = {
        "focus_id": focus_id,
        "user_id": user.user_id,
        "notebook_id": data.notebook_id,
        "focus_minutes": data.focus_minutes,
        "break_minutes": data.break_minutes,
        "completed": True,
        "date": today,
        "notes": data.notes,
        "xp_earned": xp_earned,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.focus_sessions.insert_one(focus_doc)
    
    # Update notebook study time if linked
    if data.notebook_id:
        await db.notebooks.update_one(
            {"notebook_id": data.notebook_id, "user_id": user.user_id},
            {"$inc": {"total_study_time_minutes": data.focus_minutes}}
        )
    
    # Award XP
    new_xp = user.xp + xp_earned
    new_rank = calculate_rank(new_xp)
    await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
    
    # Update study streak
    await update_study_streak(user.user_id)
    
    focus_doc.pop('_id', None)
    focus_doc["new_xp"] = new_xp
    focus_doc["new_rank"] = new_rank
    return focus_doc

@api_router.get("/study/focus/stats")
async def get_focus_stats(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get focus/pomodoro statistics"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    today = datetime.now().strftime("%Y-%m-%d")
    
    # Today's sessions
    today_sessions = await db.focus_sessions.find({"user_id": user.user_id, "date": today}, {"_id": 0}).to_list(100)
    
    # Last 7 days
    seven_days_ago = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
    week_sessions = await db.focus_sessions.find({
        "user_id": user.user_id,
        "date": {"$gte": seven_days_ago}
    }, {"_id": 0}).to_list(500)
    
    # All time
    all_sessions = await db.focus_sessions.find({"user_id": user.user_id}, {"_id": 0}).to_list(5000)
    
    return {
        "today": {
            "sessions": len(today_sessions),
            "total_minutes": sum(s.get("focus_minutes", 0) for s in today_sessions),
            "xp_earned": sum(s.get("xp_earned", 0) for s in today_sessions)
        },
        "week": {
            "sessions": len(week_sessions),
            "total_minutes": sum(s.get("focus_minutes", 0) for s in week_sessions),
            "daily_minutes": {s.get("date"): 0 for s in week_sessions}  # Will be enriched below
        },
        "all_time": {
            "sessions": len(all_sessions),
            "total_minutes": sum(s.get("focus_minutes", 0) for s in all_sessions),
            "total_hours": round(sum(s.get("focus_minutes", 0) for s in all_sessions) / 60, 1)
        }
    }

# ========== AI STUDY ASSISTANT ==========

@api_router.post("/study/ai-chat")
async def study_ai_chat(request: Request, data: dict, session_token: Optional[str] = Cookie(None)):
    """AI study assistant - contextual help for studying"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    message = data.get("message", "")
    notebook_id = data.get("notebook_id")
    context_type = data.get("context_type", "general")  # general, explain, quiz_help, summarize, motivate
    
    # Build context from notebook if provided
    context = ""
    if notebook_id:
        notebook = await db.notebooks.find_one({"notebook_id": notebook_id, "user_id": user.user_id}, {"_id": 0})
        if notebook:
            context += f"\nMatéria: {notebook.get('name', '')}"
            notes = await db.study_notes.find({"notebook_id": notebook_id, "user_id": user.user_id}, {"_id": 0}).to_list(10)
            if notes:
                context += "\n\nNotas do aluno:\n"
                for note in notes[:5]:
                    context += f"- {note.get('title', '')}: {note.get('content', '')[:300]}\n"
    
    # Get study stats for motivation context
    streak = await db.study_streaks.find_one({"user_id": user.user_id}, {"_id": 0})
    streak_info = f"Streak atual: {streak.get('current_streak', 0)} dias" if streak else "Sem streak"
    
    system_messages = {
        "general": "Você é um tutor de estudos inteligente e paciente. Ajude o aluno a entender conceitos, organize seus estudos e dê dicas práticas. Seja conciso e use linguagem clara em português.",
        "explain": "Você é um professor especialista. Explique conceitos de forma clara, use exemplos práticos e analogias simples. Responda em português.",
        "quiz_help": "Você é um especialista em preparação para provas. Ajude a resolver questões, explique a lógica por trás das respostas e dê dicas para questões similares. Responda em português.",
        "summarize": "Você é um especialista em resumos e técnicas de estudo. Crie resumos objetivos e organizados. Use bullet points e destaque conceitos-chave. Responda em português.",
        "motivate": f"Você é um coach motivacional de estudos. O aluno tem {streak_info}. Motive-o a continuar estudando, dê dicas de produtividade e foco. Seja energético e positivo. Responda em português."
    }
    
    system_msg = system_messages.get(context_type, system_messages["general"])
    
    full_prompt = f"{context}\n\nPergunta do aluno: {message}" if context else message
    
    try:
        response = await call_llm(
            full_prompt,
            session_id=f"study_{user.user_id}",
            system_message=system_msg,
            user_id=user.user_id
        )
        return {"response": response, "context_type": context_type}
    except Exception as e:
        logging.error(f"Study AI chat failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/study/notebooks")
async def get_notebooks(request: Request, area_id: Optional[str] = None, program_id: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    """Get notebooks, optionally filtered by area or program"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    query = {"user_id": user.user_id}
    if area_id:
        query["area_id"] = area_id
    if program_id:
        query["program_id"] = program_id
    
    notebooks = await db.notebooks.find(query, {"_id": 0}).to_list(1000)
    return notebooks

@api_router.post("/study/notebooks")
async def create_notebook(request: Request, notebook_data: NotebookCreate, session_token: Optional[str] = Cookie(None)):
    """Create a new notebook/subject"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    notebook_id = f"notebook_{uuid.uuid4().hex[:12]}"
    notebook_doc = {
        "notebook_id": notebook_id,
        "user_id": user.user_id,
        "area_id": notebook_data.area_id,
        "program_id": notebook_data.program_id,
        "name": notebook_data.name,
        "description": notebook_data.description,
        "color": notebook_data.color,
        "tags": notebook_data.tags,
        "total_study_time_minutes": 0,
        "total_questions": 0,
        "correct_questions": 0,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.notebooks.insert_one(notebook_doc)
    notebook_doc.pop('_id', None)
    return notebook_doc

@api_router.patch("/study/notebooks/{notebook_id}")
async def update_notebook(request: Request, notebook_id: str, data: dict, session_token: Optional[str] = Cookie(None)):
    """Update a notebook"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    update_fields = {}
    for field in ["name", "description", "color", "tags", "area_id", "program_id",
                   "weight", "dificuldade", "user_difficulty", "topicos", "conteudo_programatico", "recursos_recomendados", "num_questoes_edital"]:
        if field in data:
            update_fields[field] = data[field]
    
    if update_fields:
        await db.notebooks.update_one(
            {"notebook_id": notebook_id, "user_id": user.user_id},
            {"$set": update_fields}
        )
    
    updated = await db.notebooks.find_one({"notebook_id": notebook_id, "user_id": user.user_id}, {"_id": 0})
    return updated

@api_router.delete("/study/notebooks/{notebook_id}")
async def delete_notebook(request: Request, notebook_id: str, session_token: Optional[str] = Cookie(None)):
    """Delete a notebook"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.notebooks.delete_one({"notebook_id": notebook_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Notebook not found")
    
    # Delete related notes, flashcards, etc.
    await db.study_notes.delete_many({"notebook_id": notebook_id, "user_id": user.user_id})
    await db.flashcards.delete_many({"notebook_id": notebook_id, "user_id": user.user_id})
    await db.quizzes.delete_many({"notebook_id": notebook_id, "user_id": user.user_id})
    
    return {"message": "Notebook deleted"}

@api_router.get("/study/notes")
async def get_notes(request: Request, notebook_id: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    """Get study notes"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    query = {"user_id": user.user_id}
    if notebook_id:
        query["notebook_id"] = notebook_id
    
    notes = await db.study_notes.find(query, {"_id": 0}).to_list(1000)
    return notes

@api_router.post("/study/notes")
async def create_note(request: Request, note_data: StudyNoteCreate, session_token: Optional[str] = Cookie(None)):
    """Create a new study note"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    note_id = f"note_{uuid.uuid4().hex[:12]}"
    note_doc = {
        "note_id": note_id,
        "user_id": user.user_id,
        "notebook_id": note_data.notebook_id,
        "title": note_data.title,
        "content": note_data.content,
        "tags": note_data.tags,
        "links": note_data.links,
        "attachments": [],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.study_notes.insert_one(note_doc)
    note_doc.pop('_id', None)
    return note_doc

@api_router.patch("/study/notes/{note_id}")
async def update_note(request: Request, note_id: str, data: dict, session_token: Optional[str] = Cookie(None)):
    """Update a note"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    update_fields = {"updated_at": datetime.now(timezone.utc).isoformat()}
    for field in ["title", "content", "tags", "links"]:
        if field in data:
            update_fields[field] = data[field]
    
    await db.study_notes.update_one(
        {"note_id": note_id, "user_id": user.user_id},
        {"$set": update_fields}
    )
    
    updated = await db.study_notes.find_one({"note_id": note_id, "user_id": user.user_id}, {"_id": 0})
    return updated

@api_router.delete("/study/notes/{note_id}")
async def delete_note(request: Request, note_id: str, session_token: Optional[str] = Cookie(None)):
    """Delete a note"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.study_notes.delete_one({"note_id": note_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Note not found")
    return {"message": "Note deleted"}

@api_router.post("/study/notes/{note_id}/upload")
async def upload_attachment(
    request: Request,
    note_id: str,
    file: UploadFile = File(...),
    session_token: Optional[str] = Cookie(None)
):
    """Upload an attachment to a note"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Check note exists
    note = await db.study_notes.find_one({"note_id": note_id, "user_id": user.user_id}, {"_id": 0})
    if not note:
        raise HTTPException(status_code=404, detail="Note not found")
    
    # Allowed types
    allowed_types = ["application/pdf", "image/jpeg", "image/png", "image/gif", "image/webp",
                    "application/vnd.ms-powerpoint", "application/vnd.openxmlformats-officedocument.presentationml.presentation"]
    
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="File type not allowed. Use PDF, images, or presentations.")
    
    # Read and encode
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:  # 10MB limit
        raise HTTPException(status_code=400, detail="File size must be less than 10MB")
    
    file_base64 = base64.b64encode(content).decode('utf-8')
    
    attachment = {
        "attachment_id": f"att_{uuid.uuid4().hex[:12]}",
        "name": file.filename,
        "type": file.content_type,
        "data": f"data:{file.content_type};base64,{file_base64}",
        "size": len(content),
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.study_notes.update_one(
        {"note_id": note_id, "user_id": user.user_id},
        {"$push": {"attachments": attachment}}
    )
    
    return {"message": "Attachment uploaded", "attachment": {**attachment, "data": "[base64 data]"}}

@api_router.get("/study/tasks")
async def get_study_tasks(request: Request, notebook_id: Optional[str] = None, completed: Optional[bool] = None, session_token: Optional[str] = Cookie(None)):
    """Get study tasks"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    query = {"user_id": user.user_id}
    if notebook_id:
        query["notebook_id"] = notebook_id
    if completed is not None:
        query["completed"] = completed
    
    tasks = await db.study_tasks.find(query, {"_id": 0}).to_list(1000)
    
    # For recurring tasks, check if completed today
    today = datetime.now().strftime("%Y-%m-%d")
    for task in tasks:
        recurrence = task.get("recurrence", "once")
        if recurrence != "once":
            last_completed = task.get("last_completed_date")
            if last_completed == today:
                task["completed_today"] = True
            else:
                task["completed_today"] = False
        else:
            task["completed_today"] = task.get("completed", False)
    
    return tasks

@api_router.post("/study/tasks")
async def create_study_task(request: Request, task_data: StudyTaskCreate, session_token: Optional[str] = Cookie(None)):
    """Create a study task"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    xp_reward = 5 if task_data.priority == "low" else 10 if task_data.priority == "medium" else 15
    
    task_id = f"stask_{uuid.uuid4().hex[:12]}"
    task_doc = {
        "task_id": task_id,
        "user_id": user.user_id,
        "notebook_id": task_data.notebook_id,
        "title": task_data.title,
        "description": task_data.description,
        "task_type": task_data.task_type,
        "recurrence": task_data.recurrence,
        "deadline": task_data.deadline,
        "reminder": task_data.reminder,
        "completed": False,
        "completed_at": None,
        "last_completed_date": None,
        "priority": task_data.priority,
        "estimated_minutes": task_data.estimated_minutes,
        "actual_minutes": 0,
        "notes": None,
        "xp_reward": xp_reward,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.study_tasks.insert_one(task_doc)
    task_doc.pop('_id', None)
    return task_doc

@api_router.patch("/study/tasks/{task_id}")
async def update_study_task(request: Request, task_id: str, data: dict, session_token: Optional[str] = Cookie(None)):
    """Update a study task"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    task = await db.study_tasks.find_one({"task_id": task_id, "user_id": user.user_id}, {"_id": 0})
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    update_fields = {}
    for field in ["title", "description", "deadline", "reminder", "priority", "estimated_minutes", "actual_minutes", "notes", "recurrence"]:
        if field in data:
            update_fields[field] = data[field]
    
    # Handle completion toggle
    if "completed" in data:
        new_completed = data["completed"]
        was_completed = task.get("completed", False)
        recurrence = task.get("recurrence", "once")
        today = datetime.now().strftime("%Y-%m-%d")
        
        if new_completed and not was_completed:
            # Completing task - award XP
            xp = task.get("xp_reward", 20)
            new_xp = user.xp + xp
            new_rank = calculate_rank(new_xp)
            await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
            
            # Update study streak
            await update_study_streak(user.user_id)
            
            if recurrence == "once":
                # One-time task - mark as completed permanently
                update_fields["completed"] = True
                update_fields["completed_at"] = datetime.now(timezone.utc).isoformat()
            else:
                # Recurring task - mark last completed date but keep available
                update_fields["last_completed_date"] = today
                update_fields["completed_at"] = datetime.now(timezone.utc).isoformat()
                # For recurring tasks, completed stays False so it shows up again
                update_fields["completed"] = False
            
        elif not new_completed and was_completed:
            # Uncompleting task - deduct XP (only for non-recurring)
            if recurrence == "once":
                update_fields["completed"] = False
                update_fields["completed_at"] = None
                
                xp = task.get("xp_reward", 20)
                new_xp = max(0, user.xp - xp)
                new_rank = calculate_rank(new_xp)
                await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
    
    if update_fields:
        await db.study_tasks.update_one(
            {"task_id": task_id, "user_id": user.user_id},
            {"$set": update_fields}
        )
    
    updated = await db.study_tasks.find_one({"task_id": task_id, "user_id": user.user_id}, {"_id": 0})
    return updated

@api_router.delete("/study/tasks/{task_id}")
async def delete_study_task(request: Request, task_id: str, session_token: Optional[str] = Cookie(None)):
    """Delete a study task"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.study_tasks.delete_one({"task_id": task_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Task not found")
    return {"message": "Task deleted"}

@api_router.get("/study/sessions")
async def get_study_sessions(request: Request, notebook_id: Optional[str] = None, date: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    """Get study sessions"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    query = {"user_id": user.user_id}
    if notebook_id:
        query["notebook_id"] = notebook_id
    if date:
        query["date"] = date
    
    sessions = await db.study_sessions.find(query, {"_id": 0}).to_list(1000)
    return sessions

@api_router.post("/study/sessions")
async def create_study_session(request: Request, session_data: StudySessionCreate, session_token: Optional[str] = Cookie(None)):
    """Log a study session"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Calculate XP based on duration
    xp_earned = (session_data.duration_minutes // 15) * 10  # 10 XP per 15 minutes
    
    session_id = f"ssession_{uuid.uuid4().hex[:12]}"
    session_doc = {
        "session_id": session_id,
        "user_id": user.user_id,
        "notebook_id": session_data.notebook_id,
        "duration_minutes": session_data.duration_minutes,
        "date": session_data.date,
        "notes": session_data.notes,
        "xp_earned": xp_earned,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.study_sessions.insert_one(session_doc)
    
    # Update notebook study time
    await db.notebooks.update_one(
        {"notebook_id": session_data.notebook_id, "user_id": user.user_id},
        {"$inc": {"total_study_time_minutes": session_data.duration_minutes}}
    )
    
    # Award XP
    new_xp = user.xp + xp_earned
    new_rank = calculate_rank(new_xp)
    await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
    
    # Update study streak
    await update_study_streak(user.user_id)
    
    session_doc.pop('_id', None)
    session_doc["new_xp"] = new_xp
    session_doc["new_rank"] = new_rank
    return session_doc

async def update_study_streak(user_id: str):
    """Update user's study streak"""
    today = datetime.now().strftime("%Y-%m-%d")
    yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    
    streak = await db.study_streaks.find_one({"user_id": user_id}, {"_id": 0})
    
    if not streak:
        streak_id = f"streak_{uuid.uuid4().hex[:12]}"
        streak = {
            "streak_id": streak_id,
            "user_id": user_id,
            "current_streak": 1,
            "best_streak": 1,
            "last_study_date": today,
            "total_study_days": 1,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.study_streaks.insert_one(streak)
        return
    
    last_date = streak.get("last_study_date")
    
    if last_date == today:
        return  # Already studied today
    
    if last_date == yesterday:
        # Continue streak
        new_streak = streak.get("current_streak", 0) + 1
        best_streak = max(streak.get("best_streak", 0), new_streak)
        await db.study_streaks.update_one(
            {"user_id": user_id},
            {"$set": {
                "current_streak": new_streak,
                "best_streak": best_streak,
                "last_study_date": today,
                "total_study_days": streak.get("total_study_days", 0) + 1
            }}
        )
    else:
        # Streak broken, start new
        await db.study_streaks.update_one(
            {"user_id": user_id},
            {"$set": {
                "current_streak": 1,
                "last_study_date": today,
                "total_study_days": streak.get("total_study_days", 0) + 1
            }}
        )

@api_router.get("/study/streak")
async def get_study_streak(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get user's study streak"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    streak = await db.study_streaks.find_one({"user_id": user.user_id}, {"_id": 0})
    if not streak:
        return {"current_streak": 0, "best_streak": 0, "total_study_days": 0}
    return streak

@api_router.get("/study/schedule")
async def get_study_schedule(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get study schedule"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    schedules = await db.study_schedules.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    return schedules

@api_router.post("/study/schedule")
async def create_study_schedule(request: Request, schedule_data: StudyScheduleCreate, session_token: Optional[str] = Cookie(None)):
    """Create a study schedule entry"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    schedule_id = f"sched_{uuid.uuid4().hex[:12]}"
    schedule_doc = {
        "schedule_id": schedule_id,
        "user_id": user.user_id,
        "notebook_id": schedule_data.notebook_id,
        "day_of_week": schedule_data.day_of_week,
        "start_time": schedule_data.start_time,
        "end_time": schedule_data.end_time,
        "repeat": schedule_data.repeat,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.study_schedules.insert_one(schedule_doc)
    schedule_doc.pop('_id', None)
    return schedule_doc

@api_router.delete("/study/schedule/{schedule_id}")
async def delete_study_schedule(request: Request, schedule_id: str, session_token: Optional[str] = Cookie(None)):
    """Delete a schedule entry"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.study_schedules.delete_one({"schedule_id": schedule_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Schedule not found")
    return {"message": "Schedule deleted"}

@api_router.get("/study/flashcards")
async def get_flashcards(request: Request, notebook_id: Optional[str] = None, deck_name: Optional[str] = None, due_only: bool = False, session_token: Optional[str] = Cookie(None)):
    """Get flashcards"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    query = {"user_id": user.user_id}
    if notebook_id:
        query["notebook_id"] = notebook_id
    if deck_name:
        query["deck_name"] = deck_name
    if due_only:
        today = datetime.now().strftime("%Y-%m-%d")
        query["next_review"] = {"$lte": today}
    
    flashcards = await db.flashcards.find(query, {"_id": 0}).to_list(1000)
    return flashcards

@api_router.post("/study/flashcards")
async def create_flashcard(request: Request, card_data: FlashcardCreate, session_token: Optional[str] = Cookie(None)):
    """Create a flashcard"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    today = datetime.now().strftime("%Y-%m-%d")
    
    flashcard_id = f"flash_{uuid.uuid4().hex[:12]}"
    card_doc = {
        "flashcard_id": flashcard_id,
        "user_id": user.user_id,
        "notebook_id": card_data.notebook_id,
        "deck_name": card_data.deck_name,
        "front": card_data.front,
        "back": card_data.back,
        "tags": card_data.tags,
        "ease_factor": 2.5,
        "interval_days": 1,
        "repetitions": 0,
        "next_review": today,
        "last_review": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.flashcards.insert_one(card_doc)
    card_doc.pop('_id', None)
    return card_doc

@api_router.post("/study/flashcards/{flashcard_id}/review")
async def review_flashcard(request: Request, flashcard_id: str, review: FlashcardReview, session_token: Optional[str] = Cookie(None)):
    """Review a flashcard using SM-2 algorithm for spaced repetition"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    card = await db.flashcards.find_one({"flashcard_id": flashcard_id, "user_id": user.user_id}, {"_id": 0})
    if not card:
        raise HTTPException(status_code=404, detail="Flashcard not found")
    
    quality = review.quality
    today = datetime.now().strftime("%Y-%m-%d")
    
    # SM-2 Algorithm
    ease_factor = card.get("ease_factor", 2.5)
    interval = card.get("interval_days", 1)
    repetitions = card.get("repetitions", 0)
    
    if quality < 3:
        # Failed - reset
        repetitions = 0
        interval = 1
    else:
        if repetitions == 0:
            interval = 1
        elif repetitions == 1:
            interval = 6
        else:
            interval = int(interval * ease_factor)
        
        repetitions += 1
    
    # Update ease factor
    ease_factor = ease_factor + (0.1 - (5 - quality) * (0.08 + (5 - quality) * 0.02))
    ease_factor = max(1.3, ease_factor)  # Minimum 1.3
    
    next_review = (datetime.now() + timedelta(days=interval)).strftime("%Y-%m-%d")
    
    await db.flashcards.update_one(
        {"flashcard_id": flashcard_id, "user_id": user.user_id},
        {"$set": {
            "ease_factor": ease_factor,
            "interval_days": interval,
            "repetitions": repetitions,
            "next_review": next_review,
            "last_review": today
        }}
    )
    
    # Award XP for reviewing
    xp_earned = 5 if quality >= 3 else 2
    new_xp = user.xp + xp_earned
    new_rank = calculate_rank(new_xp)
    await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
    
    return {
        "message": "Card reviewed",
        "next_review": next_review,
        "interval_days": interval,
        "ease_factor": ease_factor,
        "xp_earned": xp_earned,
        "new_xp": new_xp
    }

@api_router.delete("/study/flashcards/{flashcard_id}")
async def delete_flashcard(request: Request, flashcard_id: str, session_token: Optional[str] = Cookie(None)):
    """Delete a flashcard"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.flashcards.delete_one({"flashcard_id": flashcard_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Flashcard not found")
    return {"message": "Flashcard deleted"}

@api_router.post("/study/flashcards/generate")
async def generate_flashcards(request: Request, data: dict, session_token: Optional[str] = Cookie(None)):
    """Generate flashcards from a note using AI"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    note_id = data.get("note_id")
    count = data.get("count", 5)
    
    note = await db.study_notes.find_one({"note_id": note_id, "user_id": user.user_id}, {"_id": 0})
    if not note:
        raise HTTPException(status_code=404, detail="Note not found")
    
    prompt = f"""Com base no seguinte conteúdo de estudo, gere {count} flashcards para memorização.
    
Conteúdo:
{note.get('title', '')}
{note.get('content', '')}

Gere flashcards no seguinte formato JSON:
[
    {{"front": "Pergunta ou conceito", "back": "Resposta ou explicação"}},
    ...
]

Regras:
- Foque nos conceitos mais importantes
- Perguntas devem ser claras e objetivas
- Respostas devem ser concisas mas completas
- Use a técnica de perguntas ativas (não apenas definições)
"""
    
    try:
        response = await call_llm(
            prompt,
            session_id=user.user_id,
            system_message="Você é um especialista em técnicas de memorização e aprendizado. Crie flashcards efetivos para estudo.",
            user_id=user.user_id
        )
        
        import re
        json_match = re.search(r'\[[\s\S]*\]', response)
        if json_match:
            cards_data = json.loads(json_match.group())
            
            created_cards = []
            today = datetime.now().strftime("%Y-%m-%d")
            
            for card in cards_data:
                flashcard_id = f"flash_{uuid.uuid4().hex[:12]}"
                card_doc = {
                    "flashcard_id": flashcard_id,
                    "user_id": user.user_id,
                    "notebook_id": note.get("notebook_id"),
                    "deck_name": note.get("title", "AI Generated"),
                    "front": card.get("front", ""),
                    "back": card.get("back", ""),
                    "tags": ["ai-generated"],
                    "ease_factor": 2.5,
                    "interval_days": 1,
                    "repetitions": 0,
                    "next_review": today,
                    "last_review": None,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.flashcards.insert_one(card_doc)
                card_doc.pop('_id', None)
                created_cards.append(card_doc)
            
            return {"message": f"{len(created_cards)} flashcards created", "flashcards": created_cards}
        else:
            return {"message": response}
    except Exception as e:
        logging.error(f"Flashcard generation failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/study/quizzes")
async def get_quizzes(request: Request, notebook_id: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    """Get quizzes"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    query = {"user_id": user.user_id}
    if notebook_id:
        query["notebook_id"] = notebook_id
    
    quizzes = await db.quizzes.find(query, {"_id": 0}).to_list(100)
    return quizzes

@api_router.post("/study/quizzes")
async def create_quiz(request: Request, quiz_data: QuizCreate, session_token: Optional[str] = Cookie(None)):
    """Create a quiz"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    quiz_id = f"quiz_{uuid.uuid4().hex[:12]}"
    quiz_doc = {
        "quiz_id": quiz_id,
        "user_id": user.user_id,
        "notebook_id": quiz_data.notebook_id,
        "title": quiz_data.title,
        "questions": quiz_data.questions,
        "ai_generated": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.quizzes.insert_one(quiz_doc)
    quiz_doc.pop('_id', None)
    return quiz_doc

@api_router.post("/study/quizzes/generate")
async def generate_quiz(request: Request, data: dict, session_token: Optional[str] = Cookie(None)):
    """Generate a quiz from notes using AI"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    notebook_id = data.get("notebook_id")
    question_count = data.get("count", 5)
    
    # Get notes from notebook
    notes = await db.study_notes.find({"notebook_id": notebook_id, "user_id": user.user_id}, {"_id": 0}).to_list(50)
    
    if not notes:
        raise HTTPException(status_code=404, detail="No notes found in this notebook")
    
    content = "\n\n".join([f"## {n.get('title', '')}\n{n.get('content', '')}" for n in notes])
    
    prompt = f"""Com base no seguinte conteúdo, gere {question_count} perguntas de múltipla escolha para um quiz.

Conteúdo:
{content[:5000]}

Gere no seguinte formato JSON:
[
    {{
        "question": "Pergunta aqui",
        "options": ["A) opção 1", "B) opção 2", "C) opção 3", "D) opção 4"],
        "correct_answer": "A",
        "explanation": "Explicação da resposta correta"
    }},
    ...
]

Regras:
- Perguntas devem testar compreensão, não apenas memorização
- Todas as opções devem ser plausíveis
- Explicações devem ser educativas
"""
    
    try:
        response = await call_llm(
            prompt,
            session_id=user.user_id,
            system_message="Você é um professor experiente. Crie questões que avaliem compreensão profunda do conteúdo.",
            user_id=user.user_id
        )
        
        import re
        json_match = re.search(r'\[[\s\S]*\]', response)
        if json_match:
            questions = json.loads(json_match.group())
            
            # Get notebook info
            notebook = await db.notebooks.find_one({"notebook_id": notebook_id}, {"_id": 0})
            
            quiz_id = f"quiz_{uuid.uuid4().hex[:12]}"
            quiz_doc = {
                "quiz_id": quiz_id,
                "user_id": user.user_id,
                "notebook_id": notebook_id,
                "title": f"Quiz: {notebook.get('name', 'Matéria')}",
                "questions": questions,
                "ai_generated": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.quizzes.insert_one(quiz_doc)
            quiz_doc.pop('_id', None)
            return quiz_doc
        else:
            return {"message": response}
    except Exception as e:
        logging.error(f"Quiz generation failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/study/quizzes/{quiz_id}/attempt")
async def submit_quiz_attempt(request: Request, quiz_id: str, data: dict, session_token: Optional[str] = Cookie(None)):
    """Submit a quiz attempt"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    quiz = await db.quizzes.find_one({"quiz_id": quiz_id, "user_id": user.user_id}, {"_id": 0})
    if not quiz:
        raise HTTPException(status_code=404, detail="Quiz not found")
    
    answers = data.get("answers", [])  # [{question_idx, selected_answer}]
    
    # Calculate score
    correct = 0
    results = []
    for ans in answers:
        q_idx = ans.get("question_idx", 0)
        selected = ans.get("selected_answer", "")
        
        if q_idx < len(quiz.get("questions", [])):
            question = quiz["questions"][q_idx]
            is_correct = selected.upper() == question.get("correct_answer", "").upper()
            if is_correct:
                correct += 1
            results.append({
                "question_idx": q_idx,
                "selected_answer": selected,
                "correct": is_correct,
                "correct_answer": question.get("correct_answer"),
                "explanation": question.get("explanation")
            })
    
    total = len(quiz.get("questions", []))
    score = (correct / total * 100) if total > 0 else 0
    
    # Save attempt
    attempt_id = f"attempt_{uuid.uuid4().hex[:12]}"
    attempt_doc = {
        "attempt_id": attempt_id,
        "user_id": user.user_id,
        "quiz_id": quiz_id,
        "score": score,
        "answers": results,
        "completed_at": datetime.now(timezone.utc).isoformat()
    }
    await db.quiz_attempts.insert_one(attempt_doc)
    
    # Award XP based on score
    xp_earned = int(score / 10) * 3  # Up to 30 XP for perfect score
    new_xp = user.xp + xp_earned
    new_rank = calculate_rank(new_xp)
    await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
    
    # Update study streak
    await update_study_streak(user.user_id)
    
    attempt_doc.pop('_id', None)
    attempt_doc["correct_count"] = correct
    attempt_doc["total_questions"] = total
    attempt_doc["xp_earned"] = xp_earned
    attempt_doc["new_xp"] = new_xp
    return attempt_doc

@api_router.delete("/study/quizzes/{quiz_id}")
async def delete_quiz(request: Request, quiz_id: str, session_token: Optional[str] = Cookie(None)):
    """Delete a quiz"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.quizzes.delete_one({"quiz_id": quiz_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Quiz not found")
    return {"message": "Quiz deleted"}

@api_router.get("/study/stats")
async def get_study_stats(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get comprehensive study statistics"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Get notebooks with study time
    notebooks = await db.notebooks.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    
    # Get sessions for last 30 days
    thirty_days_ago = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    sessions = await db.study_sessions.find({
        "user_id": user.user_id,
        "date": {"$gte": thirty_days_ago}
    }, {"_id": 0}).to_list(1000)
    
    # Get streak
    streak = await db.study_streaks.find_one({"user_id": user.user_id}, {"_id": 0})
    
    # Get flashcard stats
    flashcards = await db.flashcards.find({"user_id": user.user_id}, {"_id": 0}).to_list(1000)
    due_flashcards = [f for f in flashcards if f.get("next_review", "") <= datetime.now().strftime("%Y-%m-%d")]
    
    # Get quiz attempts
    attempts = await db.quiz_attempts.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    avg_score = sum(a.get("score", 0) for a in attempts) / len(attempts) if attempts else 0
    
    # Get completed tasks
    tasks = await db.study_tasks.find({"user_id": user.user_id, "completed": True}, {"_id": 0}).to_list(1000)
    
    # Calculate time by notebook
    time_by_notebook = {}
    for nb in notebooks:
        time_by_notebook[nb.get("name", "Unknown")] = nb.get("total_study_time_minutes", 0)
    
    # Daily study time for last 7 days
    daily_time = {}
    for s in sessions:
        date = s.get("date", "")
        daily_time[date] = daily_time.get(date, 0) + s.get("duration_minutes", 0)
    
    total_time = sum(nb.get("total_study_time_minutes", 0) for nb in notebooks)
    
    return {
        "total_study_time_minutes": total_time,
        "total_study_time_hours": round(total_time / 60, 1),
        "notebooks_count": len(notebooks),
        "time_by_notebook": time_by_notebook,
        "daily_time_last_7_days": daily_time,
        "streak": {
            "current": streak.get("current_streak", 0) if streak else 0,
            "best": streak.get("best_streak", 0) if streak else 0,
            "total_days": streak.get("total_study_days", 0) if streak else 0
        },
        "flashcards": {
            "total": len(flashcards),
            "due_today": len(due_flashcards),
            "mastered": len([f for f in flashcards if f.get("interval_days", 0) > 21])
        },
        "quizzes": {
            "total_attempts": len(attempts),
            "average_score": round(avg_score, 1)
        },
        "tasks_completed": len(tasks)
    }

@api_router.post("/study/ai-suggestions")
async def get_ai_study_suggestions(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get AI suggestions for study improvement based on spaced repetition"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Gather user's study data
    stats = await get_study_stats(request, session_token)
    
    # Get due flashcards
    today = datetime.now().strftime("%Y-%m-%d")
    due_flashcards = await db.flashcards.find({
        "user_id": user.user_id,
        "next_review": {"$lte": today}
    }, {"_id": 0}).to_list(100)
    
    # Get pending tasks
    pending_tasks = await db.study_tasks.find({
        "user_id": user.user_id,
        "completed": False
    }, {"_id": 0}).to_list(50)
    
    # Get notebooks needing attention (least studied)
    notebooks = await db.notebooks.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    notebooks_sorted = sorted(notebooks, key=lambda x: x.get("total_study_time_minutes", 0))
    
    prompt = f"""Com base nos dados de estudo do usuário, forneça sugestões personalizadas:

ESTATÍSTICAS:
- Tempo total de estudo: {stats.get('total_study_time_hours', 0)} horas
- Streak atual: {stats.get('streak', {}).get('current', 0)} dias
- Flashcards para revisar hoje: {len(due_flashcards)}
- Tarefas pendentes: {len(pending_tasks)}
- Média nos quizzes: {stats.get('quizzes', {}).get('average_score', 0)}%

MATÉRIAS MENOS ESTUDADAS:
{', '.join([f"{nb.get('name')} ({nb.get('total_study_time_minutes', 0)} min)" for nb in notebooks_sorted[:3]])}

TAREFAS PENDENTES COM DEADLINE PRÓXIMO:
{[f"{t.get('title')} - {t.get('deadline', 'Sem prazo')}" for t in pending_tasks[:5]]}

Forneça:
1. Uma análise breve do progresso
2. 3-5 sugestões práticas para melhorar
3. Quais matérias precisam de mais atenção
4. Dicas de repetição espaçada baseadas nos dados
5. Motivação personalizada

Responda de forma concisa e motivadora em português."""

    try:
        response = await call_llm(
            prompt,
            session_id=user.user_id,
            system_message="Você é um coach de estudos especializado em técnicas de aprendizado eficiente e repetição espaçada. Seja motivador e prático.",
            user_id=user.user_id
        )
        
        return {
            "suggestions": response,
            "due_flashcards_count": len(due_flashcards),
            "pending_tasks_count": len(pending_tasks),
            "least_studied_notebooks": [nb.get("name") for nb in notebooks_sorted[:3]]
        }
    except Exception as e:
        logging.error(f"AI suggestions failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))



# ========== PDF CONTENT ANALYSIS ENDPOINT ==========

@api_router.post("/study/content/analyze-pdf")
async def analyze_content_pdf(
    request: Request,
    file: UploadFile = File(...),
    notebook_id: Optional[str] = Form(None),
    generate_notes: bool = Form(True),
    generate_flashcards: bool = Form(True),
    generate_quiz: bool = Form(True),
    num_flashcards: int = Form(10),
    num_quiz_questions: int = Form(5),
    session_token: Optional[str] = Cookie(None)
):
    """Analyze a PDF of study content and generate review notes, flashcards, and quizzes"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Apenas arquivos PDF são aceitos")
    
    if not gemini_client:
        raise HTTPException(status_code=500, detail="Serviço de IA indisponível")
    
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Arquivo muito grande. Limite de 20MB.")
    
    try:
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp_file:
            tmp_file.write(content)
            tmp_path = tmp_file.name
        
        uploaded_file = gemini_client.files.upload(file=tmp_path)
        
        # Build generation instructions
        gen_parts = []
        if generate_notes:
            gen_parts.append("""
"review_notes": {
  "title": "Título da revisão",
  "summary": "Resumo completo e detalhado do conteúdo (mínimo 500 palavras), cobrindo TODOS os tópicos principais",
  "key_topics": ["Tópico 1", "Tópico 2", ...],
  "important_points": ["Ponto importante 1", "Ponto importante 2", ...],
  "study_tips": ["Dica de estudo 1", "Dica 2", ...]
}""")
        
        if generate_flashcards:
            gen_parts.append(f"""
"flashcards": [
  {{"front": "Pergunta/conceito", "back": "Resposta/definição", "deck_name": "Tema"}},
  ... (gere exatamente {num_flashcards} flashcards cobrindo os conceitos mais importantes)
]""")
        
        if generate_quiz:
            gen_parts.append(f"""
"quiz": {{
  "title": "Quiz sobre o conteúdo",
  "questions": [
    {{
      "question_text": "Pergunta sobre o conteúdo",
      "options": ["A) opção", "B) opção", "C) opção", "D) opção", "E) opção"],
      "correct_answer": "A",
      "explanation": "Explicação da resposta"
    }},
    ... (gere exatamente {num_quiz_questions} questões de múltipla escolha)
  ]
}}""")
        
        system_msg = f"""Você é um especialista em educação e criação de materiais de estudo.
Analise o conteúdo do documento PDF e gere materiais de estudo de alta qualidade.

Identifique os temas principais, conceitos-chave, definições importantes e relações entre os tópicos.

Responda APENAS com JSON válido contendo:
{{
  {",".join(gen_parts)}
}}

REGRAS:
- A revisão deve ser completa e educativa
- Flashcards devem cobrir conceitos-chave e definições
- Quiz deve ter questões que testem compreensão real do conteúdo
- Tudo em português
- JSON deve ser válido e bem formatado"""

        response = gemini_client.models.generate_content(
            model=GEMINI_MODEL,
            contents=[
                types.Part.from_uri(file_uri=uploaded_file.uri, mime_type="application/pdf"),
                "Analise este documento de estudo e gere materiais de revisão completos em JSON:"
            ],
            config=types.GenerateContentConfig(
                system_instruction=system_msg
            )
        )
        
        os.unlink(tmp_path)
        
        json_str = response.text.strip()
        if json_str.startswith("```json"):
            json_str = json_str[7:]
        if json_str.startswith("```"):
            json_str = json_str[3:]
        if json_str.endswith("```"):
            json_str = json_str[:-3]
        
        parsed = json.loads(json_str.strip())
        
        results = {"filename": file.filename, "notebook_id": notebook_id}
        xp_earned = 0
        
        # Save review notes
        if generate_notes and "review_notes" in parsed:
            review = parsed["review_notes"]
            note_id = f"note_{uuid.uuid4().hex[:12]}"
            topics = review.get("key_topics", [])
            important = review.get("important_points", [])
            tips = review.get("study_tips", [])
            
            content_text = f"# {review.get('title', 'Revisão')}\n\n"
            content_text += f"{review.get('summary', '')}\n\n"
            if important:
                content_text += "## Pontos Importantes\n" + "\n".join(f"• {p}" for p in important) + "\n\n"
            if tips:
                content_text += "## Dicas de Estudo\n" + "\n".join(f"• {t}" for t in tips) + "\n"
            
            note_doc = {
                "note_id": note_id,
                "user_id": user.user_id,
                "notebook_id": notebook_id,
                "title": f"📝 Revisão: {review.get('title', file.filename)}",
                "content": content_text,
                "tags": topics[:10],
                "links": [],
                "ai_generated": True,
                "source_pdf": file.filename,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.study_notes.insert_one(note_doc)
            note_doc.pop("_id", None)
            results["note"] = note_doc
            xp_earned += 5
        
        # Save flashcards
        if generate_flashcards and "flashcards" in parsed:
            saved_flashcards = []
            for fc in parsed["flashcards"]:
                fc_id = f"fc_{uuid.uuid4().hex[:12]}"
                fc_doc = {
                    "flashcard_id": fc_id,
                    "user_id": user.user_id,
                    "notebook_id": notebook_id,
                    "front": fc.get("front", ""),
                    "back": fc.get("back", ""),
                    "deck_name": fc.get("deck_name", "PDF Import"),
                    "ease_factor": 2.5,
                    "interval_days": 0,
                    "repetitions": 0,
                    "next_review": datetime.now().strftime("%Y-%m-%d"),
                    "ai_generated": True,
                    "source_pdf": file.filename,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.flashcards.insert_one(fc_doc)
                fc_doc.pop("_id", None)
                saved_flashcards.append(fc_doc)
            results["flashcards"] = saved_flashcards
            results["flashcards_count"] = len(saved_flashcards)
            xp_earned += len(saved_flashcards)
        
        # Save quiz
        if generate_quiz and "quiz" in parsed:
            quiz_data = parsed["quiz"]
            quiz_id = f"quiz_{uuid.uuid4().hex[:12]}"
            quiz_doc = {
                "quiz_id": quiz_id,
                "user_id": user.user_id,
                "notebook_id": notebook_id,
                "title": quiz_data.get("title", f"Quiz: {file.filename}"),
                "questions": quiz_data.get("questions", []),
                "ai_generated": True,
                "source_pdf": file.filename,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.quizzes.insert_one(quiz_doc)
            quiz_doc.pop("_id", None)
            results["quiz"] = quiz_doc
            xp_earned += 5
        
        # Award XP
        if xp_earned > 0:
            await award_xp(user.user_id, xp_earned)
        results["xp_earned"] = xp_earned
        
        generated_items = []
        if "note" in results: generated_items.append("revisão")
        if "flashcards" in results: generated_items.append(f"{results['flashcards_count']} flashcards")
        if "quiz" in results: generated_items.append("quiz")
        
        results["message"] = f"Conteúdo analisado! Gerado: {', '.join(generated_items)}. +{xp_earned} XP"
        results["success"] = True
        
        return results
        
    except json.JSONDecodeError as e:
        logging.error(f"Failed to parse content analysis JSON: {e}")
        raise HTTPException(status_code=500, detail="Erro ao processar o conteúdo do PDF. Tente novamente.")
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Content PDF analysis failed: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao analisar PDF: {str(e)}")


# ========== SIMULADOS ENDPOINTS ==========

@api_router.get("/study/simulados")
async def get_simulados(request: Request, area_id: Optional[str] = None, program_id: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    """Get all simulados for the user"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    query = {"user_id": user.user_id}
    if area_id:
        query["area_id"] = area_id
    if program_id:
        query["program_id"] = program_id
    
    simulados = await db.simulados.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)
    
    # Enrich with attempt data
    for sim in simulados:
        attempts = await db.simulado_attempts.find(
            {"simulado_id": sim["simulado_id"], "user_id": user.user_id}, {"_id": 0}
        ).sort("completed_at", -1).to_list(50)
        sim["attempts_count"] = len(attempts)
        sim["best_score"] = max((a.get("score", 0) for a in attempts), default=0)
        sim["last_attempt"] = attempts[0] if attempts else None
        # Remove questions from list view to save bandwidth
        sim["questions_count"] = len(sim.get("questions", []))
        sim.pop("questions", None)
    
    return simulados


@api_router.get("/study/simulados/stats")
async def get_simulado_stats(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get simulado statistics"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    attempts = await db.simulado_attempts.find(
        {"user_id": user.user_id}, {"_id": 0}
    ).to_list(5000)
    
    simulados = await db.simulados.find(
        {"user_id": user.user_id}, {"_id": 0}
    ).to_list(200)
    
    if not attempts:
        return {
            "total_simulados": len(simulados),
            "total_attempts": 0,
            "average_score": 0,
            "best_score": 0,
            "total_questions_answered": 0,
            "total_correct": 0,
            "accuracy_rate": 0,
            "total_time_minutes": 0,
            "by_banca": {},
            "by_disciplina": {},
            "by_concurso": {},
            "recent_attempts": []
        }
    
    total_correct = sum(a.get("correct_count", 0) for a in attempts)
    total_questions = sum(a.get("total_questions", 0) for a in attempts)
    total_time = sum(a.get("time_spent_seconds", 0) for a in attempts)
    scores = [a.get("score", 0) for a in attempts]
    
    # Stats by banca
    by_banca = {}
    by_disciplina = {}
    by_concurso = {}
    
    for a in attempts:
        banca = a.get("banca", "Outros")
        disciplina = a.get("disciplina", "Outros")
        concurso = a.get("concurso", "Outros")
        
        if banca:
            if banca not in by_banca:
                by_banca[banca] = {"attempts": 0, "total_score": 0, "total_correct": 0, "total_questions": 0}
            by_banca[banca]["attempts"] += 1
            by_banca[banca]["total_score"] += a.get("score", 0)
            by_banca[banca]["total_correct"] += a.get("correct_count", 0)
            by_banca[banca]["total_questions"] += a.get("total_questions", 0)
        
        if disciplina:
            if disciplina not in by_disciplina:
                by_disciplina[disciplina] = {"attempts": 0, "total_score": 0, "total_correct": 0, "total_questions": 0}
            by_disciplina[disciplina]["attempts"] += 1
            by_disciplina[disciplina]["total_score"] += a.get("score", 0)
            by_disciplina[disciplina]["total_correct"] += a.get("correct_count", 0)
            by_disciplina[disciplina]["total_questions"] += a.get("total_questions", 0)
        
        if concurso:
            if concurso not in by_concurso:
                by_concurso[concurso] = {"attempts": 0, "total_score": 0, "total_correct": 0, "total_questions": 0}
            by_concurso[concurso]["attempts"] += 1
            by_concurso[concurso]["total_score"] += a.get("score", 0)
            by_concurso[concurso]["total_correct"] += a.get("correct_count", 0)
            by_concurso[concurso]["total_questions"] += a.get("total_questions", 0)
    
    # Calculate averages
    for group in [by_banca, by_disciplina, by_concurso]:
        for key in group:
            group[key]["avg_score"] = round(group[key]["total_score"] / group[key]["attempts"], 1) if group[key]["attempts"] > 0 else 0
            group[key]["accuracy"] = round(group[key]["total_correct"] / group[key]["total_questions"] * 100, 1) if group[key]["total_questions"] > 0 else 0
    
    # Recent attempts (last 10)
    recent = sorted(attempts, key=lambda x: x.get("completed_at", ""), reverse=True)[:10]
    
    return {
        "total_simulados": len(simulados),
        "total_attempts": len(attempts),
        "average_score": round(sum(scores) / len(scores), 1) if scores else 0,
        "best_score": round(max(scores), 1) if scores else 0,
        "total_questions_answered": total_questions,
        "total_correct": total_correct,
        "accuracy_rate": round(total_correct / total_questions * 100, 1) if total_questions > 0 else 0,
        "total_time_minutes": round(total_time / 60, 1),
        "by_banca": by_banca,
        "by_disciplina": by_disciplina,
        "by_concurso": by_concurso,
        "recent_attempts": recent
    }


@api_router.get("/study/simulados/{simulado_id}")
async def get_simulado(request: Request, simulado_id: str, session_token: Optional[str] = Cookie(None)):
    """Get a single simulado with all questions"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    simulado = await db.simulados.find_one(
        {"simulado_id": simulado_id, "user_id": user.user_id}, {"_id": 0}
    )
    if not simulado:
        raise HTTPException(status_code=404, detail="Simulado not found")
    
    # Get attempts
    attempts = await db.simulado_attempts.find(
        {"simulado_id": simulado_id, "user_id": user.user_id}, {"_id": 0}
    ).sort("completed_at", -1).to_list(50)
    simulado["attempts"] = attempts
    
    return simulado


@api_router.post("/study/simulados/import-pdf")
async def import_simulado_pdf(
    request: Request,
    file: UploadFile = File(...),
    title: str = Form("Simulado Importado"),
    banca: Optional[str] = Form(None),
    disciplina: Optional[str] = Form(None),
    concurso: Optional[str] = Form(None),
    question_type: str = Form("multipla_escolha"),
    area_id: Optional[str] = Form(None),
    program_id: Optional[str] = Form(None),
    session_token: Optional[str] = Cookie(None)
):
    """Import a PDF with questions/answer sheet and create a simulado"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Apenas arquivos PDF são aceitos")
    
    content = await file.read()
    
    if len(content) > 20 * 1024 * 1024:  # 20MB limit
        raise HTTPException(status_code=400, detail="Arquivo muito grande. Limite de 20MB.")
    
    try:
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp_file:
            tmp_file.write(content)
            tmp_path = tmp_file.name
        
        uploaded_file = gemini_client.files.upload(file=tmp_path)
        
        type_instruction = ""
        if question_type == "multipla_escolha":
            type_instruction = """Cada questão DEVE ter exatamente 5 alternativas (A, B, C, D, E).
O campo "correct_answer" deve ser a LETRA da alternativa correta (ex: "A", "B", "C", "D" ou "E")."""
        elif question_type == "certo_errado":
            type_instruction = """Cada questão é do tipo CERTO ou ERRADO.
O campo "options" deve ser ["Certo", "Errado"].
O campo "correct_answer" deve ser "Certo" ou "Errado"."""
        else:
            type_instruction = """As questões podem ser de múltipla escolha (5 alternativas A-E) ou certo/errado.
Para múltipla escolha: options com 5 alternativas, correct_answer = letra (A-E).
Para certo/errado: options = ["Certo", "Errado"], correct_answer = "Certo" ou "Errado".
Adicione o campo "type": "multipla_escolha" ou "certo_errado" em cada questão."""
        
        system_msg = f"""Você é um especialista em extrair questões de provas e concursos de documentos PDF.
Sua tarefa é analisar o documento e extrair TODAS as questões encontradas, incluindo TEXTOS BASE / TEXTOS DE APOIO.
{type_instruction}

REGRAS:
- Extraia TODAS as questões do documento, sem pular nenhuma
- Se houver gabarito no documento, use-o para determinar a resposta correta
- Se não houver gabarito, analise e determine a resposta correta
- Mantenha a numeração original das questões
- Preserve o texto completo de cada questão e alternativas
- Se possível, identifique a disciplina/matéria de cada questão
- Adicione uma breve explicação para cada resposta correta

TEXTOS BASE / TEXTOS DE APOIO (MUITO IMPORTANTE):
- Muitas questões de provas (especialmente de Língua Portuguesa, interpretação de texto, legislação, etc.) possuem um TEXTO BASE (texto de apoio, trecho, fragmento, excerto, poema, etc.) que precede as questões
- O texto base é um trecho ou passagem que o candidato precisa ler para responder as questões relacionadas
- Exemplos de cabeçalhos de texto base: "Texto para as questões X a Y", "Leia o texto a seguir", "Com base no texto abaixo", "Texto I", "Texto II", etc.
- EXTRAIA INTEGRALMENTE o texto base associado a cada questão no campo "texto_base"
- Se várias questões se referem ao MESMO texto base, REPITA o texto base completo em CADA uma dessas questões
- Se a questão NÃO possui texto base, deixe o campo "texto_base" como null ou string vazia ""
- O campo "texto_base" deve conter APENAS o texto de apoio, NÃO o enunciado da questão em si
- Preserve a formatação original do texto base (parágrafos, versos de poemas, citações, etc.)

Responda APENAS com um JSON válido no formato:
{{
  "questions": [
    {{
      "question_number": 1,
      "texto_base": "Texto de apoio completo que precede a questão, se houver. Null se não houver.",
      "question_text": "Texto completo do enunciado da questão (sem o texto base)",
      "options": ["A) texto", "B) texto", "C) texto", "D) texto", "E) texto"],
      "correct_answer": "A",
      "explanation": "Breve explicação",
      "disciplina": "Matéria identificada",
      "type": "multipla_escolha"
    }}
  ],
  "metadata": {{
    "total_questions": 10,
    "banca_detected": "Nome da banca se identificada",
    "concurso_detected": "Nome do concurso se identificado",
    "year_detected": "Ano da prova se identificado"
  }}
}}"""

        response = gemini_client.models.generate_content(
            model=GEMINI_MODEL,
            contents=[
                types.Part.from_uri(file_uri=uploaded_file.uri, mime_type="application/pdf"),
                "Extraia todas as questões deste documento de prova/simulado e retorne em JSON:"
            ],
            config=types.GenerateContentConfig(
                system_instruction=system_msg
            )
        )
        
        # Clean up temp file
        os.unlink(tmp_path)
        
        # Parse JSON response
        json_str = response.text.strip()
        if json_str.startswith("```json"):
            json_str = json_str[7:]
        if json_str.startswith("```"):
            json_str = json_str[3:]
        if json_str.endswith("```"):
            json_str = json_str[:-3]
        
        parsed = json.loads(json_str.strip())
        questions = parsed.get("questions", [])
        metadata = parsed.get("metadata", {})
        
        if not questions:
            raise HTTPException(status_code=400, detail="Não foi possível extrair questões do PDF. Verifique se o documento contém questões válidas.")
        
        # Use detected metadata if user didn't provide
        final_banca = banca or metadata.get("banca_detected")
        final_concurso = concurso or metadata.get("concurso_detected")
        
        simulado_id = f"sim_{uuid.uuid4().hex[:12]}"
        simulado_doc = {
            "simulado_id": simulado_id,
            "user_id": user.user_id,
            "title": title,
            "description": f"Importado de: {file.filename}",
            "source_type": "pdf_import",
            "banca": final_banca,
            "disciplina": disciplina,
            "concurso": final_concurso,
            "question_type": question_type,
            "difficulty": "misto",
            "questions": questions,
            "questions_count": len(questions),
            "area_id": area_id,
            "program_id": program_id,
            "pdf_filename": file.filename,
            "year_detected": metadata.get("year_detected"),
            "status": "ready",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.simulados.insert_one(simulado_doc)
        simulado_doc.pop("_id", None)
        
        return {
            "success": True,
            "simulado": simulado_doc,
            "message": f"Simulado criado com {len(questions)} questões extraídas do PDF!"
        }
        
    except json.JSONDecodeError as e:
        logging.error(f"Failed to parse PDF questions JSON: {e}")
        raise HTTPException(status_code=500, detail="Erro ao interpretar as questões do PDF. Tente novamente.")
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"PDF simulado import failed: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao processar PDF: {str(e)}")


@api_router.post("/study/simulados/generate")
async def generate_simulado(request: Request, data: SimuladoCreate, session_token: Optional[str] = Cookie(None)):
    """Generate a simulado with AI based on banca/disciplina/concurso"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if not gemini_client:
        raise HTTPException(status_code=500, detail="Serviço de IA indisponível")
    
    num_q = data.num_questions
    
    type_instruction = ""
    if data.question_type == "multipla_escolha":
        type_instruction = """Todas as questões devem ser de MÚLTIPLA ESCOLHA com exatamente 5 alternativas (A, B, C, D, E).
O campo "correct_answer" deve ser a LETRA da alternativa correta (ex: "A", "B", "C", "D" ou "E").
O campo "type" deve ser "multipla_escolha"."""
    elif data.question_type == "certo_errado":
        type_instruction = """Todas as questões devem ser do tipo CERTO ou ERRADO (estilo CESPE/CEBRASPE).
O campo "options" deve ser ["Certo", "Errado"].
O campo "correct_answer" deve ser "Certo" ou "Errado".
O campo "type" deve ser "certo_errado"."""
    else:
        type_instruction = """Misture questões de múltipla escolha (5 alternativas A-E) e certo/errado.
Para múltipla escolha: options com 5 alternativas, correct_answer = letra (A-E), type = "multipla_escolha".
Para certo/errado: options = ["Certo", "Errado"], correct_answer = "Certo" ou "Errado", type = "certo_errado"."""
    
    difficulty_instruction = ""
    if data.difficulty == "facil":
        difficulty_instruction = "Nível FÁCIL: questões básicas e conceituais."
    elif data.difficulty == "medio":
        difficulty_instruction = "Nível MÉDIO: questões intermediárias que exigem compreensão aprofundada."
    elif data.difficulty == "dificil":
        difficulty_instruction = "Nível DIFÍCIL: questões complexas, com pegadinhas e que exigem raciocínio avançado."
    else:
        difficulty_instruction = "Misture questões de diferentes níveis de dificuldade (fácil, médio e difícil)."
    
    banca_info = f"Banca: {data.banca}. Siga o ESTILO e formato típico desta banca." if data.banca else "Sem banca específica."
    disciplina_info = f"Disciplina: {data.disciplina}." if data.disciplina else ""
    concurso_info = f"Concurso: {data.concurso}." if data.concurso else ""
    
    system_msg = f"""Você é um especialista em elaboração de questões para concursos públicos brasileiros.
Gere questões ORIGINAIS, realistas e de alta qualidade, no estilo de provas reais.

CONTEXTO:
{banca_info}
{disciplina_info}
{concurso_info}
{difficulty_instruction}

{type_instruction}

REGRAS:
- Gere exatamente {num_q} questões
- As questões devem ser originais mas no estilo de questões reais de concursos
- Cada questão deve ter enunciado claro e completo
- As alternativas devem ser plausíveis (não deve ser óbvio qual é a correta)
- A explicação deve ser detalhada e educativa
- Identifique a subdisciplina/tópico de cada questão
- Questões devem cobrir diferentes tópicos dentro da disciplina
- Para questões de interpretação de texto, inclua um "texto_base" (trecho, fragmento, artigo de lei, etc.) que o candidato deve ler para responder
- Pelo menos 20-30% das questões devem ter texto_base quando a disciplina envolver interpretação, legislação ou jurisprudência
- Se a questão não precisar de texto base, use null no campo texto_base

Responda APENAS com JSON válido no formato:
{{
  "questions": [
    {{
      "question_number": 1,
      "texto_base": "Texto de apoio/trecho para leitura, se aplicável. Null se não houver.",
      "question_text": "Texto completo da questão",
      "options": ["A) texto", "B) texto", "C) texto", "D) texto", "E) texto"],
      "correct_answer": "A",
      "explanation": "Explicação detalhada da resposta correta",
      "disciplina": "{data.disciplina or 'Geral'}",
      "subdisciplina": "Tópico específico",
      "difficulty": "medio",
      "type": "multipla_escolha"
    }}
  ]
}}"""

    prompt = f"Gere {num_q} questões de simulado para concurso público com as seguintes especificações:\n"
    if data.banca:
        prompt += f"- Banca: {data.banca}\n"
    if data.disciplina:
        prompt += f"- Disciplina: {data.disciplina}\n"
    if data.concurso:
        prompt += f"- Concurso: {data.concurso}\n"
    prompt += f"- Tipo: {data.question_type}\n- Dificuldade: {data.difficulty}\n"
    prompt += "\nRetorne APENAS o JSON com as questões."
    
    try:
        response = gemini_client.models.generate_content(
            model=GEMINI_MODEL,
            contents=prompt,
            config=types.GenerateContentConfig(
                system_instruction=system_msg
            )
        )
        
        json_str = response.text.strip()
        if json_str.startswith("```json"):
            json_str = json_str[7:]
        if json_str.startswith("```"):
            json_str = json_str[3:]
        if json_str.endswith("```"):
            json_str = json_str[:-3]
        
        parsed = json.loads(json_str.strip())
        questions = parsed.get("questions", [])
        
        if not questions:
            raise HTTPException(status_code=500, detail="A IA não conseguiu gerar as questões. Tente novamente.")
        
        simulado_id = f"sim_{uuid.uuid4().hex[:12]}"
        simulado_doc = {
            "simulado_id": simulado_id,
            "user_id": user.user_id,
            "title": data.title,
            "description": data.description or f"Simulado gerado por IA - {data.banca or ''} {data.disciplina or ''} {data.concurso or ''}".strip(),
            "source_type": "ai_generated",
            "banca": data.banca,
            "disciplina": data.disciplina,
            "concurso": data.concurso,
            "question_type": data.question_type,
            "difficulty": data.difficulty,
            "questions": questions,
            "questions_count": len(questions),
            "area_id": data.area_id,
            "program_id": data.program_id,
            "status": "ready",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.simulados.insert_one(simulado_doc)
        simulado_doc.pop("_id", None)
        
        # Award XP for creating simulado
        xp_earned = 5
        await award_xp(user.user_id, xp_earned)
        
        return {
            "success": True,
            "simulado": simulado_doc,
            "message": f"Simulado gerado com {len(questions)} questões! +{xp_earned} XP",
            "xp_earned": xp_earned
        }
        
    except json.JSONDecodeError as e:
        logging.error(f"Failed to parse generated simulado JSON: {e}")
        raise HTTPException(status_code=500, detail="Erro ao gerar simulado. Tente novamente.")
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Simulado generation failed: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao gerar simulado: {str(e)}")


@api_router.post("/study/simulados/{simulado_id}/submit")
async def submit_simulado(request: Request, simulado_id: str, submission: SimuladoSubmit, session_token: Optional[str] = Cookie(None)):
    """Submit answers for a simulado and get correction"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    simulado = await db.simulados.find_one(
        {"simulado_id": simulado_id, "user_id": user.user_id}, {"_id": 0}
    )
    if not simulado:
        raise HTTPException(status_code=404, detail="Simulado not found")
    
    questions = simulado.get("questions", [])
    answers = submission.answers
    
    # Correct answers
    results = []
    correct_count = 0
    by_disciplina = {}
    
    for ans in answers:
        q_idx = ans.get("question_idx", 0)
        selected = ans.get("selected_answer", "")
        
        if q_idx < 0 or q_idx >= len(questions):
            continue
        
        question = questions[q_idx]
        correct_answer = question.get("correct_answer", "")
        is_correct = selected.strip().upper() == correct_answer.strip().upper()
        
        if is_correct:
            correct_count += 1
        
        disc = question.get("disciplina", "Geral")
        if disc not in by_disciplina:
            by_disciplina[disc] = {"total": 0, "correct": 0}
        by_disciplina[disc]["total"] += 1
        if is_correct:
            by_disciplina[disc]["correct"] += 1
        
        results.append({
            "question_idx": q_idx,
            "question_number": question.get("question_number", q_idx + 1),
            "selected_answer": selected,
            "correct_answer": correct_answer,
            "is_correct": is_correct,
            "explanation": question.get("explanation", ""),
            "disciplina": disc
        })
    
    total_answered = len(results)
    total_questions = len(questions)
    score = round((correct_count / total_answered * 100), 1) if total_answered > 0 else 0
    
    # Calculate by_disciplina percentages
    for disc in by_disciplina:
        t = by_disciplina[disc]["total"]
        c = by_disciplina[disc]["correct"]
        by_disciplina[disc]["accuracy"] = round(c / t * 100, 1) if t > 0 else 0
    
    attempt_id = f"sattempt_{uuid.uuid4().hex[:12]}"
    attempt_doc = {
        "attempt_id": attempt_id,
        "simulado_id": simulado_id,
        "user_id": user.user_id,
        "title": simulado.get("title", ""),
        "banca": simulado.get("banca"),
        "disciplina": simulado.get("disciplina"),
        "concurso": simulado.get("concurso"),
        "answers": results,
        "score": score,
        "correct_count": correct_count,
        "total_questions": total_questions,
        "total_answered": total_answered,
        "unanswered": total_questions - total_answered,
        "time_spent_seconds": submission.time_spent_seconds,
        "by_disciplina": by_disciplina,
        "completed_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.simulado_attempts.insert_one(attempt_doc)
    attempt_doc.pop("_id", None)
    
    # Award XP (2 XP per correct answer)
    xp_earned = correct_count * 2
    new_xp, new_rank = await award_xp(user.user_id, xp_earned)
    attempt_doc["xp_earned"] = xp_earned
    attempt_doc["new_xp"] = new_xp
    
    # Update study streak
    await update_study_streak(user.user_id)
    
    # Log questions for stats
    await db.question_logs.insert_one({
        "log_id": f"qlog_{uuid.uuid4().hex[:12]}",
        "user_id": user.user_id,
        "notebook_id": None,
        "simulado_id": simulado_id,
        "total": total_answered,
        "correct": correct_count,
        "source": "simulado",
        "banca": simulado.get("banca"),
        "disciplina": simulado.get("disciplina"),
        "concurso": simulado.get("concurso"),
        "date": datetime.now().strftime("%Y-%m-%d"),
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return attempt_doc


@api_router.get("/study/simulados/{simulado_id}/results")
async def get_simulado_results(request: Request, simulado_id: str, session_token: Optional[str] = Cookie(None)):
    """Get all attempt results for a simulado"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    attempts = await db.simulado_attempts.find(
        {"simulado_id": simulado_id, "user_id": user.user_id}, {"_id": 0}
    ).sort("completed_at", -1).to_list(50)
    
    return attempts


@api_router.delete("/study/simulados/{simulado_id}")
async def delete_simulado(request: Request, simulado_id: str, session_token: Optional[str] = Cookie(None)):
    """Delete a simulado and its attempts"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.simulados.delete_one({"simulado_id": simulado_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Simulado not found")
    
    # Delete all attempts
    await db.simulado_attempts.delete_many({"simulado_id": simulado_id, "user_id": user.user_id})
    
    return {"message": "Simulado e tentativas excluídos com sucesso"}


# ========== ANALYZE EDITAL (MULTI-CARGO) ==========

@api_router.post("/study/programs/analyze-edital")
async def analyze_edital_cargos(
    request: Request,
    file: UploadFile = File(...),
    session_token: Optional[str] = Cookie(None)
):
    """Analyze an edital PDF and return available cargos/positions before generating the program"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Apenas arquivos PDF são aceitos")
    
    if not gemini_client:
        raise HTTPException(status_code=500, detail="Serviço de IA indisponível")
    
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Arquivo muito grande. Limite de 20MB.")
    
    try:
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp_file:
            tmp_file.write(content)
            tmp_path = tmp_file.name
        
        uploaded_file = gemini_client.files.upload(file=tmp_path)
        
        system_msg = """Você é um especialista em concursos públicos brasileiros.
Analise o edital do concurso contido neste PDF e identifique:
1. Se há MÚLTIPLOS CARGOS/POSIÇÕES disponíveis
2. Informações gerais do concurso
3. Para cada cargo: as disciplinas, pesos, número de questões E o conteúdo programático completo

ATENÇÃO: Extraia FIELMENTE o CONTEÚDO PROGRAMÁTICO de cada disciplina. A maioria dos editais traz uma seção listando todos os assuntos cobrados. Extraia TODOS esses assuntos.

Responda APENAS com JSON válido no formato abaixo. NÃO inclua texto antes ou depois do JSON.

{
  "concurso": {
    "nome": "Nome completo do concurso",
    "orgao": "Órgão/instituição",
    "banca": "Banca organizadora"
  },
  "multiple_cargos": true,
  "cargos": [
    {
      "nome": "Nome do Cargo",
      "vagas": "Número de vagas",
      "remuneracao": "Remuneração",
      "escolaridade": "Nível exigido",
      "disciplinas": [
        {
          "nome": "Nome da Disciplina",
          "peso": 3,
          "num_questoes": 10,
          "grupo": "Conhecimentos Gerais",
          "topicos": ["Tópico resumido 1", "Tópico resumido 2"],
          "conteudo_programatico": [
            {
              "assunto": "Nome do assunto principal",
              "subtopicos": ["Subtópico 1", "Subtópico 2"]
            }
          ]
        }
      ]
    }
  ]
}

REGRAS:
- Se o edital tem APENAS UM CARGO, defina "multiple_cargos" como false e coloque apenas 1 cargo no array
- Se há múltiplos cargos com disciplinas DIFERENTES, defina "multiple_cargos" como true
- O "peso" deve refletir exatamente o peso descrito no edital (se o edital diz peso 2, coloque 2)
- Se o edital define pesos por grupo (ex: Conhecimentos Gerais peso 1, Conhecimentos Específicos peso 2), aplique o peso do grupo a cada disciplina daquele grupo
- Extraia TODOS os cargos disponíveis
- CONTEÚDO PROGRAMÁTICO: Extraia FIELMENTE todos os assuntos listados no conteúdo programático de cada disciplina. Cada assunto principal deve virar um item em "conteudo_programatico" com subtópicos quando houver. Se o edital não detalhar o conteúdo programático, deixe o array vazio.
- "topicos" é um resumo (máximo 10 itens). "conteudo_programatico" é a lista COMPLETA e DETALHADA.
- Tudo em português brasileiro
"""

        response = gemini_client.models.generate_content(
            model=GEMINI_MODEL,
            contents=[
                types.Part.from_uri(file_uri=uploaded_file.uri, mime_type="application/pdf"),
                "Analise este edital de concurso e identifique os cargos disponíveis e suas disciplinas:"
            ],
            config=types.GenerateContentConfig(
                system_instruction=system_msg
            )
        )
        
        os.unlink(tmp_path)
        
        json_str = response.text.strip()
        if json_str.startswith("```json"):
            json_str = json_str[7:]
        if json_str.startswith("```"):
            json_str = json_str[3:]
        if json_str.endswith("```"):
            json_str = json_str[:-3]
        
        parsed = json.loads(json_str.strip())
        
        # Store the analysis temporarily for the user
        analysis_id = f"edital_analysis_{uuid.uuid4().hex[:12]}"
        
        # Upload file again for later use
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp_file2:
            tmp_file2.write(content)
            tmp_path2 = tmp_file2.name
        
        # Store analysis with PDF content for later
        analysis_doc = {
            "analysis_id": analysis_id,
            "user_id": user.user_id,
            "concurso": parsed.get("concurso", {}),
            "multiple_cargos": parsed.get("multiple_cargos", False),
            "cargos": parsed.get("cargos", []),
            "pdf_filename": file.filename,
            "pdf_content_b64": base64.b64encode(content).decode('utf-8'),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "expires_at": (datetime.now(timezone.utc) + timedelta(hours=1)).isoformat()
        }
        await db.edital_analyses.insert_one(analysis_doc)
        
        os.unlink(tmp_path2)
        
        return {
            "success": True,
            "analysis_id": analysis_id,
            "concurso": parsed.get("concurso", {}),
            "multiple_cargos": parsed.get("multiple_cargos", False),
            "cargos": parsed.get("cargos", []),
            "message": f"Edital analisado! {'Encontrados ' + str(len(parsed.get('cargos', []))) + ' cargos.' if parsed.get('multiple_cargos') else 'Cargo único identificado.'}"
        }
        
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="Erro ao processar resposta da IA. Tente novamente.")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao analisar edital: {str(e)}")


@api_router.post("/study/programs/import-edital-with-cargo")
async def import_edital_with_cargo(
    request: Request,
    data: dict,
    session_token: Optional[str] = Cookie(None)
):
    """Create study program from a previously analyzed edital, selecting a specific cargo"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    analysis_id = data.get("analysis_id")
    cargo_index = data.get("cargo_index", 0)
    area_id = data.get("area_id")
    target_date = data.get("target_date")
    hours_per_day = data.get("hours_per_day", 4.0)
    days_per_week = data.get("days_per_week", 5)
    
    if not analysis_id or not area_id:
        raise HTTPException(status_code=400, detail="analysis_id e area_id são obrigatórios")
    
    # Get stored analysis
    analysis = await db.edital_analyses.find_one({"analysis_id": analysis_id, "user_id": user.user_id}, {"_id": 0})
    if not analysis:
        raise HTTPException(status_code=404, detail="Análise não encontrada. Faça upload do edital novamente.")
    
    cargos = analysis.get("cargos", [])
    if not cargos or cargo_index >= len(cargos):
        raise HTTPException(status_code=400, detail="Cargo inválido")
    
    selected_cargo = cargos[cargo_index]
    concurso_info = analysis.get("concurso", {})
    concurso_info["cargo"] = selected_cargo.get("nome", "")
    concurso_info["vagas"] = selected_cargo.get("vagas", "")
    concurso_info["remuneracao"] = selected_cargo.get("remuneracao", "")
    concurso_info["escolaridade"] = selected_cargo.get("escolaridade", "")
    
    disciplinas = selected_cargo.get("disciplinas", [])
    if not disciplinas:
        raise HTTPException(status_code=400, detail="Nenhuma disciplina encontrada para este cargo")
    
    # Now use AI to generate cronograma based on the selected cargo's disciplines
    if not gemini_client:
        raise HTTPException(status_code=500, detail="Serviço de IA indisponível")
    
    try:
        disc_list = json.dumps(disciplinas, ensure_ascii=False)
        
        system_msg = f"""Você é um MESTRE em planejamento de estudos para concursos públicos brasileiros, com experiência em coaching de aprovados.
Crie um cronograma semanal IMPECÁVEL e OTIMIZADO, considerando neurociência da aprendizagem.

Disciplinas e seus pesos: {disc_list}

Configuração do aluno:
- {hours_per_day} horas por dia, {days_per_week} dias por semana
{"- Data da prova: " + target_date if target_date else "- Sem data definida."}

Responda APENAS com JSON:
{{
  "cronograma_semanal": [
    {{
      "dia": "Segunda",
      "frase_motivacional": "Frase motivacional curta para o dia",
      "blocos": [
        {{
          "disciplina": "Nome exato da disciplina",
          "duracao_minutos": 120,
          "tipo_estudo": "Teoria + Questões",
          "prioridade": "alta"
        }}
      ]
    }}
  ],
  "materias_por_dia_sugerido": 3,
  "inclui_redacao": true,
  "ciclo_revisao": "A cada 3 dias, reserve 30-45min para revisão das matérias estudadas nos dias anteriores",
  "estrategia": {{
    "resumo": "Resumo da estratégia geral",
    "fase_1": "Fase 1: Base teórica (primeiros 30% do tempo)",
    "fase_2": "Fase 2: Aprofundamento + questões (40% do tempo)",
    "fase_3": "Fase 3: Revisão intensiva + simulados (30% final)",
    "dicas_gerais": ["Dica 1", "Dica 2", "Dica 3"],
    "materias_prioritarias": ["Matéria de maior peso"],
    "plano_revisao": "Explicação do ciclo de revisão espaçada"
  }}
}}

REGRAS CRÍTICAS:
1. PESOS: Matérias com maior peso devem ter MAIS tempo (proporcional ao peso). Ex: peso 3 = ~3x mais tempo que peso 1
2. REDAÇÃO: Se o concurso exige redação (comum em concursos de nível superior), inclua 1-2 blocos semanais de "Redação" com tipo_estudo "Prática de Redação"
3. REVISÃO: Inclua blocos de "Revisão Geral" a cada 2-3 dias (30-45min) para fixação por repetição espaçada
4. ALTERNÂNCIA: Nunca coloque duas matérias pesadas/densas em sequência - intercale com matérias mais leves
5. MATÉRIAS POR DIA: Sugira 2-4 matérias por dia (ideal 3), nunca mais que 4 para manter foco
6. BLOCOS: Cada bloco entre 45-120min. Intervalos de 10-15min entre blocos (implícito)
7. LIMITE: Máximo {int(hours_per_day * 60)} minutos por dia de estudo efetivo
8. DIAS: Usar apenas {days_per_week} dias. Dias: Segunda, Terça, Quarta, Quinta, Sexta, Sábado, Domingo
9. MOTIVAÇÃO: Cada dia deve ter uma frase motivacional ÚNICA e IMPACTANTE (curta, 1 linha)
10. DESCANSO: Se {days_per_week} < 7, os dias livres são para descanso/lazer
11. CONSISTÊNCIA: O nome da disciplina nos blocos deve ser EXATAMENTE igual ao nome na lista de disciplinas"""
        
        response = gemini_client.models.generate_content(
            model=GEMINI_MODEL,
            contents=f"Gere cronograma para estas disciplinas de concurso: {disc_list}",
            config=types.GenerateContentConfig(
                system_instruction=system_msg
            )
        )
        
        json_str = response.text.strip()
        if json_str.startswith("```json"):
            json_str = json_str[7:]
        if json_str.startswith("```"):
            json_str = json_str[3:]
        if json_str.endswith("```"):
            json_str = json_str[:-3]
        
        parsed = json.loads(json_str.strip())
        cronograma = parsed.get("cronograma_semanal", [])
        estrategia = parsed.get("estrategia", {})
        
        # Create program
        program_id = f"prog_{uuid.uuid4().hex[:12]}"
        program_name = f"{concurso_info.get('nome', 'Concurso')} - {selected_cargo.get('nome', 'Cargo')}"
        
        program_doc = {
            "program_id": program_id,
            "user_id": user.user_id,
            "area_id": area_id,
            "name": program_name[:100],
            "description": f"Banca: {concurso_info.get('banca', 'N/A')} | Órgão: {concurso_info.get('orgao', 'N/A')} | Cargo: {selected_cargo.get('nome', '')}",
            "color": "#8B5CF6",
            "icon": "file-text",
            "target_date": target_date or concurso_info.get("data_prova"),
            "status": "active",
            "total_questions": 0,
            "correct_questions": 0,
            "source_type": "edital_import",
            "edital_data": {
                "concurso": concurso_info,
                "cargo_selecionado": selected_cargo,
                "estrategia": estrategia,
                "total_disciplinas": len(disciplinas),
                "hours_per_day": hours_per_day,
                "days_per_week": days_per_week,
                "pdf_filename": analysis.get("pdf_filename", ""),
                "analysis_id": analysis_id
            },
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.study_programs.insert_one(program_doc)
        program_doc.pop('_id', None)
        
        # Create notebooks
        created_notebooks = []
        color_palette = ["#007AFF", "#34C759", "#FF9500", "#FF3B30", "#AF52DE", "#5AC8FA", "#FF2D55", "#FFCC00", "#30D158", "#64D2FF", "#BF5AF2", "#FF6482"]
        
        for i, disc in enumerate(disciplinas):
            nb_id = f"nb_{uuid.uuid4().hex[:12]}"
            nb_doc = {
                "notebook_id": nb_id,
                "user_id": user.user_id,
                "area_id": area_id,
                "program_id": program_id,
                "name": disc.get("nome", f"Disciplina {i+1}"),
                "description": disc.get("grupo", ""),
                "color": color_palette[i % len(color_palette)],
                "tags": disc.get("topicos", [])[:10],
                "total_study_time_minutes": 0,
                "total_questions": 0,
                "correct_questions": 0,
                "weight": disc.get("peso", 1),
                "num_questoes_edital": disc.get("num_questoes", 0),
                "dificuldade": "media",
                "grupo": disc.get("grupo", ""),
                "topicos": disc.get("topicos", []),
                "conteudo_programatico": disc.get("conteudo_programatico", []),
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.notebooks.insert_one(nb_doc)
            nb_doc.pop('_id', None)
            created_notebooks.append(nb_doc)
        
        # Create schedules
        created_schedules = []
        day_map = {
            "Segunda": "monday", "Terça": "tuesday", "Quarta": "wednesday",
            "Quinta": "thursday", "Sexta": "friday", "Sábado": "saturday", "Domingo": "sunday"
        }
        
        for day_entry in cronograma:
            dia = day_entry.get("dia", "")
            day_of_week = day_map.get(dia, dia.lower())
            blocos = day_entry.get("blocos", [])
            current_hour = 8
            current_min = 0
            
            for bloco in blocos:
                disc_name = bloco.get("disciplina", "")
                duracao = bloco.get("duracao_minutos", 60)
                
                matching_nb = None
                for nb in created_notebooks:
                    if nb["name"].lower() == disc_name.lower():
                        matching_nb = nb
                        break
                if not matching_nb:
                    for nb in created_notebooks:
                        if disc_name.lower() in nb["name"].lower() or nb["name"].lower() in disc_name.lower():
                            matching_nb = nb
                            break
                
                if matching_nb:
                    start_time = f"{current_hour:02d}:{current_min:02d}"
                    end_min_total = current_min + duracao
                    end_hour = current_hour + end_min_total // 60
                    end_min = end_min_total % 60
                    end_time = f"{end_hour:02d}:{end_min:02d}"
                    
                    sched_id = f"sched_{uuid.uuid4().hex[:12]}"
                    sched_doc = {
                        "schedule_id": sched_id,
                        "user_id": user.user_id,
                        "notebook_id": matching_nb["notebook_id"],
                        "program_id": program_id,
                        "day_of_week": day_of_week,
                        "start_time": start_time,
                        "end_time": end_time,
                        "repeat": True,
                        "tipo_estudo": bloco.get("tipo_estudo", "Teoria + Questões"),
                        "prioridade": bloco.get("prioridade", "media"),
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db.study_schedules.insert_one(sched_doc)
                    sched_doc.pop('_id', None)
                    created_schedules.append(sched_doc)
                    
                    current_hour = end_hour
                    current_min = end_min
        
        # Award XP
        xp_earned = 25
        await award_xp(user.user_id, xp_earned)
        
        # Clean up analysis
        await db.edital_analyses.delete_one({"analysis_id": analysis_id})
        
        return {
            "success": True,
            "program": program_doc,
            "concurso": concurso_info,
            "cargo": selected_cargo,
            "disciplinas": created_notebooks,
            "cronograma": cronograma,
            "estrategia": estrategia,
            "schedules_created": len(created_schedules),
            "xp_earned": xp_earned,
            "message": f"Programa criado para cargo '{selected_cargo.get('nome', '')}' com {len(disciplinas)} disciplinas!"
        }
        
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="Erro ao processar resposta da IA. Tente novamente.")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao criar programa: {str(e)}")


# ========== STUDY AI CHAT WITH FILE UPLOAD ==========

@api_router.post("/study/ai-chat-with-file")
async def study_ai_chat_with_file(
    request: Request,
    file: UploadFile = File(...),
    message: str = Form(""),
    context_type: str = Form("summarize"),
    notebook_id: Optional[str] = Form(None),
    session_token: Optional[str] = Cookie(None)
):
    """AI study chat with file upload (PDF or image) for summarization"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if not gemini_client:
        raise HTTPException(status_code=500, detail="Serviço de IA indisponível")
    
    allowed_types = ["application/pdf", "image/jpeg", "image/png", "image/gif", "image/webp"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Tipo de arquivo não suportado. Use PDF ou imagens (JPG, PNG, GIF, WebP).")
    
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Arquivo muito grande. Limite de 20MB.")
    
    try:
        import tempfile
        suffix = '.pdf' if file.content_type == 'application/pdf' else '.jpg'
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp_file:
            tmp_file.write(content)
            tmp_path = tmp_file.name
        
        uploaded_gemini_file = gemini_client.files.upload(file=tmp_path)
        
        context_prompts = {
            "summarize": "Faça um resumo completo e organizado do conteúdo deste documento. Use tópicos, subtópicos e destaque os pontos-chave. Responda em português.",
            "explain": "Explique o conteúdo deste documento de forma didática e detalhada. Use exemplos quando possível. Responda em português.",
            "quiz_help": "A partir do conteúdo deste documento, crie 5 questões de estudo com respostas. Responda em português.",
            "general": "Analise o conteúdo deste documento e responda à pergunta do aluno. Responda em português.",
            "mindmap": "Analise o conteúdo e crie uma estrutura de mapa mental. Responda em português."
        }
        
        system_msg = context_prompts.get(context_type, context_prompts["general"])
        user_prompt = message if message else "Analise este documento e faça um resumo detalhado."
        
        response = gemini_client.models.generate_content(
            model=GEMINI_MODEL,
            contents=[
                types.Part.from_uri(file_uri=uploaded_gemini_file.uri, mime_type=file.content_type),
                user_prompt
            ],
            config=types.GenerateContentConfig(
                system_instruction=system_msg
            )
        )
        
        os.unlink(tmp_path)
        
        return {
            "response": response.text,
            "context_type": context_type,
            "filename": file.filename,
            "file_type": file.content_type
        }
        
    except Exception as e:
        logging.error(f"Study AI chat with file failed: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao processar arquivo: {str(e)}")


# ========== MIND MAP GENERATION ==========

@api_router.post("/study/mindmap/generate")
async def generate_mindmap(
    request: Request,
    file: Optional[UploadFile] = File(None),
    text: Optional[str] = Form(None),
    topic: Optional[str] = Form(None),
    notebook_id: Optional[str] = Form(None),
    session_token: Optional[str] = Cookie(None)
):
    """Generate a mind map structure from PDF, image, text, or topic"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if not gemini_client:
        raise HTTPException(status_code=500, detail="Serviço de IA indisponível")
    
    system_msg = """Você é um especialista em criar mapas mentais estruturados para estudo.
Analise o conteúdo fornecido e crie um mapa mental hierárquico completo.

Responda APENAS com JSON válido neste formato:
{
  "title": "Título central do mapa mental",
  "nodes": [
    {
      "id": "1",
      "label": "Tópico Principal 1",
      "color": "#007AFF",
      "children": [
        {
          "id": "1.1",
          "label": "Subtópico 1.1",
          "children": [
            {
              "id": "1.1.1",
              "label": "Detalhe 1.1.1",
              "children": []
            }
          ]
        }
      ]
    }
  ]
}

REGRAS:
- Crie pelo menos 4-6 nós principais
- Cada nó pode ter 2-4 filhos
- Máximo 3 níveis de profundidade
- Use cores variadas (#007AFF, #34C759, #FF9500, #FF3B30, #AF52DE, #5AC8FA, #FF2D55)
- Textos curtos e objetivos
- Organize logicamente por temas/categorias
- Tudo em português"""
    
    try:
        contents = []
        
        if file:
            file_content = await file.read()
            if len(file_content) > 20 * 1024 * 1024:
                raise HTTPException(status_code=400, detail="Arquivo muito grande. Limite de 20MB.")
            
            import tempfile
            suffix = '.pdf' if file.content_type == 'application/pdf' else '.jpg'
            with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp_file:
                tmp_file.write(file_content)
                tmp_path = tmp_file.name
            
            uploaded_gemini_file = gemini_client.files.upload(file=tmp_path)
            contents.append(types.Part.from_uri(file_uri=uploaded_gemini_file.uri, mime_type=file.content_type))
            contents.append("Crie um mapa mental completo a partir do conteúdo deste documento:")
            os.unlink(tmp_path)
        elif text:
            contents.append(f"Crie um mapa mental completo sobre o seguinte conteúdo:\n\n{text}")
        elif topic:
            contents.append(f"Crie um mapa mental completo sobre o tema: {topic}")
        elif notebook_id:
            # Get notes from notebook
            notes = await db.study_notes.find(
                {"notebook_id": notebook_id, "user_id": user.user_id}, {"_id": 0}
            ).to_list(20)
            if notes:
                notes_text = "\n\n".join([f"## {n.get('title', '')}\n{n.get('content', '')}" for n in notes])
                contents.append(f"Crie um mapa mental completo baseado nestas notas de estudo:\n\n{notes_text}")
            else:
                nb = await db.notebooks.find_one({"notebook_id": notebook_id, "user_id": user.user_id}, {"_id": 0})
                topic_name = nb.get("name", "Matéria") if nb else "Matéria"
                contents.append(f"Crie um mapa mental completo sobre: {topic_name}")
        else:
            raise HTTPException(status_code=400, detail="Forneça um arquivo, texto, tópico ou notebook_id")
        
        response = gemini_client.models.generate_content(
            model=GEMINI_MODEL,
            contents=contents,
            config=types.GenerateContentConfig(
                system_instruction=system_msg
            )
        )
        
        json_str = response.text.strip()
        if json_str.startswith("```json"):
            json_str = json_str[7:]
        if json_str.startswith("```"):
            json_str = json_str[3:]
        if json_str.endswith("```"):
            json_str = json_str[:-3]
        
        mindmap_data = json.loads(json_str.strip())
        
        # Save mind map
        mindmap_id = f"mm_{uuid.uuid4().hex[:12]}"
        mindmap_doc = {
            "mindmap_id": mindmap_id,
            "user_id": user.user_id,
            "notebook_id": notebook_id,
            "title": mindmap_data.get("title", "Mapa Mental"),
            "data": mindmap_data,
            "source": "file" if file else ("text" if text else ("topic" if topic else "notebook")),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.mindmaps.insert_one(mindmap_doc)
        mindmap_doc.pop('_id', None)
        
        # Award XP
        await award_xp(user.user_id, 10)
        
        return {
            "success": True,
            "mindmap_id": mindmap_id,
            "mindmap": mindmap_data,
            "xp_earned": 10,
            "message": "Mapa mental gerado com sucesso!"
        }
        
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="Erro ao processar mapa mental. Tente novamente.")
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Mind map generation failed: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao gerar mapa mental: {str(e)}")


@api_router.get("/study/mindmaps")
async def get_mindmaps(request: Request, notebook_id: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    """Get user's mind maps"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    query = {"user_id": user.user_id}
    if notebook_id:
        query["notebook_id"] = notebook_id
    
    mindmaps = await db.mindmaps.find(query, {"_id": 0}).to_list(50)
    return mindmaps


@api_router.delete("/study/mindmaps/{mindmap_id}")
async def delete_mindmap(request: Request, mindmap_id: str, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.mindmaps.delete_one({"mindmap_id": mindmap_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Mapa mental não encontrado")
    return {"message": "Mapa mental excluído"}


# ========== PROGRESS HISTORY / COMPARATOR ==========

@api_router.get("/study/programs/{program_id}/progress-history")
async def get_progress_history(
    request: Request,
    program_id: str,
    days: int = 30,
    session_token: Optional[str] = Cookie(None)
):
    """Get progress history for all disciplines in a program, for comparison over time"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    notebooks = await db.notebooks.find(
        {"program_id": program_id, "user_id": user.user_id}, {"_id": 0}
    ).to_list(100)
    
    if not notebooks:
        return {"program_id": program_id, "history": [], "notebooks": []}
    
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    
    history_data = {}
    
    for nb in notebooks:
        nb_id = nb["notebook_id"]
        nb_name = nb["name"]
        nb_color = nb.get("color", "#007AFF")
        
        # Get question logs
        q_logs = await db.question_logs.find({
            "notebook_id": nb_id,
            "user_id": user.user_id,
            "created_at": {"$gte": cutoff}
        }, {"_id": 0}).to_list(1000)
        
        # Get focus sessions
        sessions = await db.study_sessions.find({
            "notebook_id": nb_id,
            "user_id": user.user_id,
            "created_at": {"$gte": cutoff}
        }, {"_id": 0}).to_list(1000)
        
        # Aggregate by date
        for log in q_logs:
            date = log.get("date", log.get("created_at", "")[:10])
            if date not in history_data:
                history_data[date] = {}
            if nb_name not in history_data[date]:
                history_data[date][nb_name] = {"questions": 0, "correct": 0, "minutes": 0, "color": nb_color}
            history_data[date][nb_name]["questions"] += log.get("total", 0)
            history_data[date][nb_name]["correct"] += log.get("correct", 0)
        
        for sess in sessions:
            date = sess.get("date", sess.get("created_at", "")[:10])
            if date not in history_data:
                history_data[date] = {}
            if nb_name not in history_data[date]:
                history_data[date][nb_name] = {"questions": 0, "correct": 0, "minutes": 0, "color": nb_color}
            history_data[date][nb_name]["minutes"] += sess.get("duration_minutes", 0)
    
    # Format for chart
    sorted_dates = sorted(history_data.keys())
    chart_data = []
    cumulative = {}
    
    for date in sorted_dates:
        entry = {"date": date}
        for nb in notebooks:
            nb_name = nb["name"]
            day_data = history_data[date].get(nb_name, {"questions": 0, "correct": 0, "minutes": 0})
            
            if nb_name not in cumulative:
                cumulative[nb_name] = {"questions": 0, "correct": 0, "minutes": 0}
            cumulative[nb_name]["questions"] += day_data["questions"]
            cumulative[nb_name]["correct"] += day_data["correct"]
            cumulative[nb_name]["minutes"] += day_data["minutes"]
            
            entry[f"{nb_name}_questoes"] = cumulative[nb_name]["questions"]
            entry[f"{nb_name}_acerto"] = round((cumulative[nb_name]["correct"] / cumulative[nb_name]["questions"] * 100) if cumulative[nb_name]["questions"] > 0 else 0, 1)
            entry[f"{nb_name}_horas"] = round(cumulative[nb_name]["minutes"] / 60, 1)
        chart_data.append(entry)
    
    nb_info = [{"name": nb["name"], "color": nb.get("color", "#007AFF"), "weight": nb.get("weight", 1)} for nb in notebooks]
    
    return {
        "program_id": program_id,
        "history": chart_data,
        "notebooks": nb_info,
        "days": days
    }


# ========== SCHEDULE-BASED NOTIFICATIONS ==========

@api_router.post("/study/programs/{program_id}/create-reminders")
async def create_schedule_reminders(
    request: Request,
    program_id: str,
    data: dict,
    session_token: Optional[str] = Cookie(None)
):
    """Create notifications/reminders from schedule blocks"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    program = await db.study_programs.find_one({"program_id": program_id, "user_id": user.user_id}, {"_id": 0})
    if not program:
        raise HTTPException(status_code=404, detail="Programa não encontrado")
    
    schedules = await db.study_schedules.find(
        {"program_id": program_id, "user_id": user.user_id}, {"_id": 0}
    ).to_list(200)
    
    notebooks = await db.notebooks.find(
        {"program_id": program_id, "user_id": user.user_id}, {"_id": 0}
    ).to_list(100)
    nb_lookup = {nb["notebook_id"]: nb for nb in notebooks}
    
    reminder_minutes_before = data.get("minutes_before", 5)
    include_end_reminder = data.get("include_end_reminder", False)
    
    day_map_reverse = {
        "monday": "Seg", "tuesday": "Ter", "wednesday": "Qua",
        "thursday": "Qui", "friday": "Sex", "saturday": "Sáb", "sunday": "Dom"
    }
    
    created = 0
    for sched in schedules:
        nb = nb_lookup.get(sched.get("notebook_id"))
        disc_name = nb["name"] if nb else "Matéria"
        day_label = day_map_reverse.get(sched.get("day_of_week", ""), "")
        
        # Calculate reminder time (X minutes before start)
        start_time = sched.get("start_time", "08:00")
        try:
            parts = start_time.split(":")
            total_mins = int(parts[0]) * 60 + int(parts[1]) - reminder_minutes_before
            if total_mins < 0:
                total_mins = 0
            reminder_time = f"{total_mins // 60:02d}:{total_mins % 60:02d}"
        except (ValueError, IndexError):
            reminder_time = start_time
        
        # Map day_of_week to repeat_days
        day_of_week = sched.get("day_of_week", "")
        
        notification_id = f"notif_{uuid.uuid4().hex[:12]}"
        notif_doc = {
            "notification_id": notification_id,
            "user_id": user.user_id,
            "title": f"Hora de estudar: {disc_name}",
            "message": f"{day_label} {start_time} - {sched.get('end_time', '')} | {sched.get('tipo_estudo', 'Estudo')}",
            "type": "reminder",
            "category": "study",
            "scheduled_time": reminder_time,
            "repeat": "weekly",
            "repeat_days": [day_of_week],
            "enabled": True,
            "channels": ["in_app", "browser"],
            "last_sent": None,
            "program_id": program_id,
            "schedule_id": sched.get("schedule_id"),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.notifications.insert_one(notif_doc)
        created += 1
    
    return {
        "success": True,
        "created": created,
        "message": f"{created} lembretes criados para o cronograma de estudos!"
    }


# ========== FIX: PENDING NOTIFICATIONS WITH TIMEZONE ==========

@api_router.get("/notifications/check")
async def check_notifications(request: Request, timezone_offset: int = 0, session_token: Optional[str] = Cookie(None)):
    """Check for pending notifications considering user timezone"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Use timezone offset from client to calculate local time
    utc_now = datetime.now(timezone.utc)
    user_local = utc_now - timedelta(minutes=timezone_offset)
    current_time = user_local.strftime("%H:%M")
    current_day = user_local.strftime("%A").lower()
    
    # Find enabled notifications matching current time (with 2 minute window)
    try:
        h, m = map(int, current_time.split(":"))
        time_start_mins = h * 60 + m - 1
        time_end_mins = h * 60 + m + 1
        
        times_to_check = []
        for t_mins in range(max(0, time_start_mins), min(1440, time_end_mins + 1)):
            times_to_check.append(f"{t_mins // 60:02d}:{t_mins % 60:02d}")
    except (ValueError, IndexError):
        times_to_check = [current_time]
    
    notifications = await db.notifications.find({
        "user_id": user.user_id,
        "enabled": True,
        "scheduled_time": {"$in": times_to_check}
    }, {"_id": 0}).to_list(100)
    
    pending = []
    for notif in notifications:
        should_send = False
        if notif['repeat'] == "none":
            if not notif.get('last_sent'):
                should_send = True
        elif notif['repeat'] == "daily":
            # Check if not already sent today
            last_sent = notif.get('last_sent')
            if not last_sent or last_sent[:10] != user_local.strftime("%Y-%m-%d"):
                should_send = True
        elif notif['repeat'] in ["weekly", "custom"]:
            if current_day in [d.lower() for d in notif.get('repeat_days', [])]:
                last_sent = notif.get('last_sent')
                if not last_sent or last_sent[:10] != user_local.strftime("%Y-%m-%d"):
                    should_send = True
        
        if should_send:
            pending.append(notif)
            # Mark as sent
            await db.notifications.update_one(
                {"notification_id": notif["notification_id"]},
                {"$set": {"last_sent": datetime.now(timezone.utc).isoformat()}}
            )
    
    return pending


# ========== REDAÇÃO (ESSAY) SECTION ==========

@api_router.post("/study/redacao/correct")
async def correct_essay(
    request: Request,
    file: UploadFile = File(...),
    instructions: str = Form(""),
    session_token: Optional[str] = Cookie(None)
):
    """AI correction of an essay from PDF, image, txt or docx"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if not gemini_client:
        raise HTTPException(status_code=500, detail="Serviço de IA indisponível")
    
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Arquivo muito grande. Limite de 20MB.")
    
    try:
        import tempfile
        ext = os.path.splitext(file.filename)[1].lower()
        
        text_content = None
        if ext in ['.txt']:
            text_content = content.decode('utf-8', errors='replace')
        
        system_msg = """Você é um professor especialista em redação para concursos públicos brasileiros (CESPE, FGV, FCC, VUNESP).
Analise a redação enviada e faça uma correção detalhada seguindo os critérios de avaliação de concursos:

Responda APENAS com JSON válido:
{
  "nota_geral": 8.5,
  "nota_maxima": 10,
  "competencias": [
    {
      "nome": "Domínio da Norma Culta",
      "nota": 8,
      "nota_maxima": 10,
      "comentario": "Boa gramática no geral, mas..."
    },
    {
      "nome": "Compreensão do Tema",
      "nota": 9,
      "nota_maxima": 10,
      "comentario": "Demonstrou bom entendimento..."
    },
    {
      "nome": "Argumentação",
      "nota": 7,
      "nota_maxima": 10,
      "comentario": "Argumentos válidos porém..."
    },
    {
      "nome": "Coesão e Coerência",
      "nota": 8,
      "nota_maxima": 10,
      "comentario": "Boa conexão entre parágrafos..."
    },
    {
      "nome": "Proposta de Intervenção",
      "nota": 9,
      "nota_maxima": 10,
      "comentario": "Proposta viável e detalhada..."
    }
  ],
  "pontos_fortes": ["Ponto forte 1", "Ponto forte 2"],
  "pontos_melhorar": ["Ponto a melhorar 1", "Ponto a melhorar 2"],
  "erros_gramaticais": [
    {"trecho": "texto original", "correcao": "texto corrigido", "explicacao": "Regra gramatical"}
  ],
  "dicas_estrategicas": ["Dica 1", "Dica 2", "Dica 3"],
  "texto_reescrito_sugestao": "Versão melhorada do primeiro parágrafo...",
  "nivel": "Bom"
}"""

        contents = []
        if text_content:
            contents.append(f"Corrija esta redação:\n\n{text_content}\n\n{instructions}")
        else:
            suffix = ext if ext in ['.pdf'] else '.jpg'
            with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                tmp.write(content)
                tmp_path = tmp.name
            mime = file.content_type or ('application/pdf' if ext == '.pdf' else 'image/jpeg')
            uploaded = gemini_client.files.upload(file=tmp_path)
            contents.append(types.Part.from_uri(file_uri=uploaded.uri, mime_type=mime))
            contents.append(f"Corrija esta redação detalhadamente. {instructions}")
            os.unlink(tmp_path)

        response = gemini_client.models.generate_content(
            model=GEMINI_MODEL,
            contents=contents,
            config=types.GenerateContentConfig(system_instruction=system_msg)
        )

        json_str = response.text.strip()
        if json_str.startswith("```json"): json_str = json_str[7:]
        if json_str.startswith("```"): json_str = json_str[3:]
        if json_str.endswith("```"): json_str = json_str[:-3]
        correction = json.loads(json_str.strip())

        # Save correction
        correction_id = f"red_{uuid.uuid4().hex[:12]}"
        correction_doc = {
            "correction_id": correction_id,
            "user_id": user.user_id,
            "filename": file.filename,
            "correction": correction,
            "instructions": instructions,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.redacoes.insert_one(correction_doc)
        await award_xp(user.user_id, 15)

        return {"success": True, "correction_id": correction_id, "correction": correction, "xp_earned": 15}

    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="Erro ao processar correção. Tente novamente.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro: {str(e)}")


@api_router.get("/study/redacao/history")
async def get_essay_history(request: Request, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    redacoes = await db.redacoes.find({"user_id": user.user_id}, {"_id": 0}).sort("created_at", -1).to_list(50)
    return redacoes


@api_router.post("/study/redacao/random-theme")
async def random_essay_theme(request: Request, data: dict = {}, session_token: Optional[str] = Cookie(None)):
    """Generate a random essay theme likely to appear in exams"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)

    concurso_type = data.get("concurso_type", "geral")

    prompt = f"""Sorteie UM tema de redação que tenha alta probabilidade de cair em provas de concurso público ({concurso_type}).
Responda APENAS com JSON:
{{
  "tema": "O tema completo da redação",
  "tipo_texto": "Dissertativo-Argumentativo",
  "banca_relacionada": "CESPE/CEBRASPE",
  "contexto": "Breve contexto sobre o tema e por que é relevante para concursos",
  "textos_motivadores": ["Texto motivador 1 (trecho real ou adaptado)", "Texto motivador 2"],
  "dicas": ["Dica para abordar este tema 1", "Dica 2"],
  "temas_relacionados": ["Tema relacionado 1", "Tema 2"],
  "nivel_dificuldade": "Médio"
}}

Use temas ATUAIS e RELEVANTES de {datetime.now().year}. Varie entre:
- Saúde pública, Educação, Tecnologia, Meio ambiente, Segurança, Direitos humanos, Economia, Cidadania digital"""

    try:
        response = await call_llm(prompt, f"essay_theme_{user.user_id}", "Você é especialista em redação para concursos públicos brasileiros.", user_id=user.user_id)
        json_str = response.strip()
        if json_str.startswith("```json"): json_str = json_str[7:]
        if json_str.startswith("```"): json_str = json_str[3:]
        if json_str.endswith("```"): json_str = json_str[:-3]
        theme = json.loads(json_str.strip())
        return {"success": True, "theme": theme}
    except:
        return {"success": True, "theme": {
            "tema": "O papel da tecnologia na promoção da inclusão social no Brasil",
            "tipo_texto": "Dissertativo-Argumentativo",
            "banca_relacionada": "Diversas",
            "contexto": "A transformação digital e seus impactos na sociedade brasileira",
            "textos_motivadores": ["A inclusão digital é fundamental para o exercício pleno da cidadania no século XXI."],
            "dicas": ["Aborde os desafios de acesso à tecnologia em regiões remotas", "Mencione políticas públicas existentes"],
            "temas_relacionados": ["Exclusão digital", "Direito à informação"],
            "nivel_dificuldade": "Médio"
        }}


# ========== GENERAL INTEGRATED CHAT ==========

@api_router.post("/chat/general")
async def general_integrated_chat(request: Request, data: dict, session_token: Optional[str] = Cookie(None)):
    """General chat that integrates with all app features - can create recipes, workouts, cronograms, etc."""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)

    content = data.get("content", "").strip()
    if not content:
        raise HTTPException(status_code=400, detail="Mensagem vazia")

    # Save user message
    msg_id = f"msg_{uuid.uuid4().hex[:12]}"
    user_msg = {
        "message_id": msg_id,
        "user_id": user.user_id,
        "role": "user",
        "content": content,
        "chat_type": "general",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.chat_messages.insert_one(user_msg.copy())

    content_lower = content.lower()

    # Detect intent with expanded keywords
    recipe_kw = ["receita", "recipe", "cozinhar", "preparar comida", "fazer um prato", "sugerir comida", "o que comer", "cardápio", "refeição"]
    workout_kw = ["treino", "exercício", "exercicio", "ficha de treino", "workout", "malhar", "academia", "musculação", "ficha"]
    study_kw = ["cronograma de estudo", "plano de estudo", "matéria de concurso"]
    
    # Financial transaction keywords (register money in/out)
    finance_income_kw = ['ganhei', 'recebi', 'entrou', 'salário', 'salario', 'renda', 'recebimento', 
                         'depósito', 'deposito', 'pix recebido', 'crédito', 'credito', 'freelance', 
                         'bônus', 'bonus', 'comissão', 'comissao', 'vendi', 'entrada de']
    finance_expense_kw = ['gastei', 'paguei', 'comprei', 'compra de', 'boleto', 'parcela', 'débito', 
                          'debito', 'saída', 'saida', 'pix enviado', 'transferi', 'conta de luz',
                          'conta de água', 'aluguel de', 'supermercado', 'mercado', 'restaurante']
    finance_report_kw = ['relatório financeiro', 'relatorio financeiro', 'resumo financeiro', 'como estão minhas finanças',
                         'como está meu saldo', 'quanto gastei', 'quanto ganhei', 'balanço financeiro',
                         'extrato', 'situação financeira', 'previsão financeira', 'projeção financeira']
    
    # Task/Goal keywords
    task_kw = ['tarefa', 'task', 'adicionar tarefa', 'criar tarefa', 'nova tarefa', 'to-do', 'todo',
               'lembrete', 'reminder', 'preciso fazer', 'tenho que fazer', 'não esquecer']
    goal_kw = ['criar meta', 'definir meta', 'quero alcançar', 'nova meta',
               'minha meta é', 'minha meta e', 'definir objetivo', 'novo objetivo', 'criar objetivo']
    
    import re
    amount_pattern = r'(?:R\$\s*)?(\d+(?:[.,]\d{1,2})?)'
    amount_matches = re.findall(amount_pattern, content)
    has_amount = len(amount_matches) > 0

    intent = "general"
    if any(k in content_lower for k in recipe_kw): intent = "recipe"
    elif any(k in content_lower for k in workout_kw): intent = "workout"
    elif has_amount and any(k in content_lower for k in finance_income_kw) and any(k in content_lower for k in finance_expense_kw): intent = "finance_mixed"
    elif has_amount and any(k in content_lower for k in finance_income_kw): intent = "finance_income"
    elif has_amount and any(k in content_lower for k in finance_expense_kw): intent = "finance_expense"
    elif any(k in content_lower for k in finance_report_kw): intent = "finance_report"
    elif any(k in content_lower for k in task_kw): intent = "task"
    elif any(k in content_lower for k in goal_kw): intent = "goal"
    elif any(k in content_lower for k in study_kw): intent = "study"
    elif any(k in content_lower for k in ['gasto', 'despesa', 'receita financeira', 'saldo', 'orçamento', 'economizar', 'investir']): intent = "finance_general"

    saved_item = None
    ai_response_text = ""

    try:
        if intent == "recipe":
            prompt = f"""O usuário pediu: "{content}"
Gere uma receita completa. Responda em texto normal formatado com markdown.
Inclua: nome, tempo de preparo, ingredientes, modo de preparo, informações nutricionais.
Ao final, inclua um bloco JSON separado com:
```json
{{"name": "Nome da Receita", "description": "Descrição curta", "prep_time_minutes": 15, "cook_time_minutes": 30, "servings": 4, "calories_per_serving": 350, "protein_per_serving": 25, "carbs_per_serving": 30, "fat_per_serving": 12, "ingredients": ["ingrediente 1", "ingrediente 2"], "instructions": ["Passo 1", "Passo 2"], "meal_type": "lunch", "diet_type": "balanced"}}
```"""
            response = await call_llm(prompt, f"general_chat_{user.user_id}", "Você é um chef e nutricionista. Gere receitas detalhadas e saudáveis.", user_id=user.user_id)
            ai_response_text = response

            # Try to extract and save recipe
            try:
                if "```json" in response:
                    json_part = response.split("```json")[1].split("```")[0].strip()
                    recipe_data = json.loads(json_part)
                    recipe_id = f"recipe_{uuid.uuid4().hex[:12]}"
                    recipe_doc = {
                        "recipe_id": recipe_id,
                        "user_id": user.user_id,
                        **recipe_data,
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db.recipes.insert_one(recipe_doc)
                    saved_item = {"type": "recipe", "id": recipe_id, "name": recipe_data.get("name", "")}
            except: pass

        elif intent == "workout":
            prompt = f"""O usuário pediu: "{content}"
Gere um plano de treino completo. Responda em texto normal formatado com markdown.
Ao final, inclua um bloco JSON:
```json
{{"name": "Nome do Treino", "description": "Descrição", "exercises": [{{"name": "Exercício", "sets": 3, "reps": "12", "weight": "", "notes": "Observações"}}]}}
```"""
            response = await call_llm(prompt, f"general_chat_{user.user_id}", "Você é um personal trainer especialista.", user_id=user.user_id)
            ai_response_text = response

            try:
                if "```json" in response:
                    json_part = response.split("```json")[1].split("```")[0].strip()
                    plan_data = json.loads(json_part)
                    plan_id = f"plan_{uuid.uuid4().hex[:12]}"
                    plan_doc = {
                        "plan_id": plan_id,
                        "user_id": user.user_id,
                        **plan_data,
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db.workout_plans.insert_one(plan_doc)
                    saved_item = {"type": "workout", "id": plan_id, "name": plan_data.get("name", "")}
            except: pass

        # FINANCE: Register Income
        elif intent == "finance_income":
            prompt = f'''Extraia TODAS as receitas/entradas financeiras desta mensagem.
Mensagem: "{content}"
Responda APENAS com JSON array:
[{{"amount": 5000.0, "category": "salário", "description": "salário mensal"}}]
Categorias: salário, freelance, investimentos, vendas, reembolso, outros
Extraia o valor numérico exato. SOMENTE o JSON array.'''
            response = await call_llm(prompt, f"general_fin_{user.user_id}", user_id=user.user_id)
            try:
                clean = response.strip()
                if "```" in clean: clean = clean.split("```")[1].replace("json", "").strip()
                start_i = clean.find('['); end_i = clean.rfind(']') + 1
                items = json.loads(clean[start_i:end_i]) if start_i != -1 and end_i > start_i else []
                if not isinstance(items, list): items = [items]
                
                registered = []
                for td in items:
                    amt = float(td.get("amount", 0))
                    if amt > 0:
                        tid = f"trans_{uuid.uuid4().hex[:12]}"
                        tdoc = {"transaction_id": tid, "user_id": user.user_id, "type": "income", "amount": amt,
                                "category": td.get("category", "outros"), "description": td.get("description", ""),
                                "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"), "created_at": datetime.now(timezone.utc).isoformat()}
                        await db.transactions.insert_one(tdoc)
                        registered.append({"type": "income", "amount": amt, "category": td.get("category", "outros"), "description": td.get("description", "")})
                
                if registered:
                    saved_item = {"type": "transactions", "items": registered}
                    total = sum(r["amount"] for r in registered)
                    ai_response_text = f"✅ **{len(registered)} receita(s) registrada(s)!**\n\n"
                    for r in registered:
                        ai_response_text += f"💰 +R$ {r['amount']:.2f} | {r['category']} | {r['description']}\n"
                    ai_response_text += f"\n**Total registrado: R$ {total:.2f}**\n\n💡 As transações já estão disponíveis na aba Finanças!"
                else:
                    ai_response_text = "Não consegui identificar o valor. Pode repetir? Ex: 'Recebi 3000 de salário'"
            except Exception as e:
                ai_response_text = f"Não consegui processar. Tente: 'Recebi 3000 de salário'. Erro: {str(e)}"

        # FINANCE: Register Expense
        elif intent == "finance_expense":
            prompt = f'''Extraia TODAS as despesas/gastos desta mensagem.
Mensagem: "{content}"
Responda APENAS com JSON array:
[{{"amount": 150.0, "category": "alimentação", "description": "supermercado"}}]
Categorias: alimentação, transporte, moradia, saúde, educação, lazer, outros
Extraia o valor numérico exato. SOMENTE o JSON array.'''
            response = await call_llm(prompt, f"general_fin_{user.user_id}", user_id=user.user_id)
            try:
                clean = response.strip()
                if "```" in clean: clean = clean.split("```")[1].replace("json", "").strip()
                start_i = clean.find('['); end_i = clean.rfind(']') + 1
                items = json.loads(clean[start_i:end_i]) if start_i != -1 and end_i > start_i else []
                if not isinstance(items, list): items = [items]
                
                registered = []
                for td in items:
                    amt = float(td.get("amount", 0))
                    if amt > 0:
                        tid = f"trans_{uuid.uuid4().hex[:12]}"
                        tdoc = {"transaction_id": tid, "user_id": user.user_id, "type": "expense", "amount": amt,
                                "category": td.get("category", "outros"), "description": td.get("description", ""),
                                "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"), "created_at": datetime.now(timezone.utc).isoformat()}
                        await db.transactions.insert_one(tdoc)
                        registered.append({"type": "expense", "amount": amt, "category": td.get("category", "outros"), "description": td.get("description", "")})
                
                if registered:
                    saved_item = {"type": "transactions", "items": registered}
                    total = sum(r["amount"] for r in registered)
                    ai_response_text = f"✅ **{len(registered)} despesa(s) registrada(s)!**\n\n"
                    for r in registered:
                        ai_response_text += f"🔴 -R$ {r['amount']:.2f} | {r['category']} | {r['description']}\n"
                    ai_response_text += f"\n**Total registrado: R$ {total:.2f}**\n\n💡 As transações já estão disponíveis na aba Finanças!"
                else:
                    ai_response_text = "Não consegui identificar o valor. Pode repetir? Ex: 'Gastei 50 no mercado'"
            except Exception as e:
                ai_response_text = f"Não consegui processar. Tente: 'Gastei 150 no supermercado'. Erro: {str(e)}"

        # FINANCE: Mixed transactions
        elif intent == "finance_mixed":
            prompt = f'''Extraia TODAS as transações financeiras desta mensagem (receitas E despesas).
Mensagem: "{content}"
Responda APENAS com JSON array:
[{{"type": "income", "amount": 5000.0, "category": "salário", "description": "salário"}}, {{"type": "expense", "amount": 50.0, "category": "alimentação", "description": "mercado"}}]
Categorias receita: salário, freelance, investimentos, vendas, reembolso, outros
Categorias despesa: alimentação, transporte, moradia, saúde, educação, lazer, outros
SOMENTE o JSON array.'''
            response = await call_llm(prompt, f"general_fin_{user.user_id}", user_id=user.user_id)
            try:
                clean = response.strip()
                if "```" in clean: clean = clean.split("```")[1].replace("json", "").strip()
                start_i = clean.find('['); end_i = clean.rfind(']') + 1
                items = json.loads(clean[start_i:end_i]) if start_i != -1 and end_i > start_i else []
                if not isinstance(items, list): items = [items]
                
                registered = []
                for td in items:
                    amt = float(td.get("amount", 0))
                    t_type = td.get("type", "expense")
                    if t_type not in ["income", "expense"]: t_type = "expense"
                    if amt > 0:
                        tid = f"trans_{uuid.uuid4().hex[:12]}"
                        tdoc = {"transaction_id": tid, "user_id": user.user_id, "type": t_type, "amount": amt,
                                "category": td.get("category", "outros"), "description": td.get("description", ""),
                                "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"), "created_at": datetime.now(timezone.utc).isoformat()}
                        await db.transactions.insert_one(tdoc)
                        registered.append({"type": t_type, "amount": amt, "category": td.get("category", "outros"), "description": td.get("description", "")})
                
                if registered:
                    saved_item = {"type": "transactions", "items": registered}
                    total_in = sum(r["amount"] for r in registered if r["type"] == "income")
                    total_out = sum(r["amount"] for r in registered if r["type"] == "expense")
                    ai_response_text = f"✅ **{len(registered)} transação(ões) registrada(s)!**\n\n"
                    for r in registered:
                        emoji = "💰" if r["type"] == "income" else "🔴"
                        sinal = "+" if r["type"] == "income" else "-"
                        ai_response_text += f"{emoji} {sinal}R$ {r['amount']:.2f} | {r['category']} | {r['description']}\n"
                    if total_in > 0: ai_response_text += f"\n💚 Total receitas: R$ {total_in:.2f}"
                    if total_out > 0: ai_response_text += f"\n🔴 Total despesas: R$ {total_out:.2f}"
                    ai_response_text += "\n\n💡 As transações já estão disponíveis na aba Finanças!"
                else:
                    ai_response_text = "Não consegui identificar os valores. Pode detalhar melhor?"
            except:
                ai_response_text = "Não consegui processar as transações mistas. Tente uma de cada vez."

        # FINANCE: Report
        elif intent == "finance_report":
            current_month = datetime.now(timezone.utc).strftime("%Y-%m")
            transactions_list = await db.transactions.find(
                {"user_id": user.user_id, "date": {"$regex": f"^{current_month}"}}, {"_id": 0}
            ).to_list(500)
            total_income = sum(t['amount'] for t in transactions_list if t['type'] == 'income')
            total_expense = sum(t['amount'] for t in transactions_list if t['type'] == 'expense')
            bal = total_income - total_expense
            
            expense_by_cat = {}
            for t in transactions_list:
                if t['type'] == 'expense':
                    expense_by_cat[t['category']] = expense_by_cat.get(t['category'], 0) + t['amount']
            
            context_str = f"""Dados financeiros do mês ({current_month}):
- Receitas: R$ {total_income:.2f}
- Despesas: R$ {total_expense:.2f}
- Saldo: R$ {bal:.2f}
- Transações: {len(transactions_list)}
- Despesas por categoria: {json.dumps(expense_by_cat, ensure_ascii=False)}"""
            
            prompt = f"""{context_str}

O usuário pediu: "{content}"

Gere um relatório/análise financeira detalhado com base nos dados acima. Inclua insights, dicas e sugestões. Use markdown para formatar. Seja objetivo e prático."""
            
            response = await call_llm(prompt, f"general_fin_{user.user_id}", "Você é um consultor financeiro pessoal especialista.", user_id=user.user_id)
            ai_response_text = response

        # TASK: Create task
        elif intent == "task":
            prompt = f'''O usuário quer criar uma tarefa/lembrete. Extraia as informações:
Mensagem: "{content}"
Responda SOMENTE com JSON:
{{"title": "Título da tarefa", "description": "Descrição detalhada", "due_date": "YYYY-MM-DD ou null", "priority": "high/medium/low", "category": "pessoal/trabalho/estudos/saúde/outros"}}'''
            response = await call_llm(prompt, f"general_task_{user.user_id}", user_id=user.user_id)
            try:
                clean = response.strip()
                if "```" in clean: clean = clean.split("```")[1].replace("json", "").strip()
                start_i = clean.find('{'); end_i = clean.rfind('}') + 1
                task_data = json.loads(clean[start_i:end_i])
                
                task_id = f"task_{uuid.uuid4().hex[:12]}"
                task_doc = {
                    "task_id": task_id,
                    "user_id": user.user_id,
                    "title": task_data.get("title", content[:50]),
                    "description": task_data.get("description", ""),
                    "due_date": task_data.get("due_date"),
                    "priority": task_data.get("priority", "medium"),
                    "category": task_data.get("category", "pessoal"),
                    "completed": False,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.tasks.insert_one(task_doc)
                saved_item = {"type": "task", "id": task_id, "title": task_data.get("title", "")}
                
                ai_response_text = f"""✅ **Tarefa criada com sucesso!**

📋 **{task_data.get('title', '')}**
{f"📝 {task_data.get('description', '')}" if task_data.get('description') else ""}
{f"📅 Prazo: {task_data.get('due_date')}" if task_data.get('due_date') else ""}
🔴 Prioridade: {task_data.get('priority', 'medium')}

💡 A tarefa está disponível na aba Tarefas!"""
            except:
                ai_response_text = "Não consegui criar a tarefa. Tente: 'Criar tarefa: Estudar direito constitucional até sexta'"

        # GOAL: Create goal
        elif intent == "goal":
            prompt = f'''O usuário quer definir uma meta/objetivo. Extraia as informações:
Mensagem: "{content}"
Responda SOMENTE com JSON:
{{"title": "Título da meta", "description": "Descrição", "target_date": "YYYY-MM-DD ou null", "category": "financeiro/saúde/estudos/carreira/pessoal/outros", "target_value": null, "current_value": 0}}'''
            response = await call_llm(prompt, f"general_goal_{user.user_id}", user_id=user.user_id)
            try:
                clean = response.strip()
                if "```" in clean: clean = clean.split("```")[1].replace("json", "").strip()
                start_i = clean.find('{'); end_i = clean.rfind('}') + 1
                goal_data = json.loads(clean[start_i:end_i])
                
                goal_id = f"goal_{uuid.uuid4().hex[:12]}"
                goal_doc = {
                    "goal_id": goal_id,
                    "user_id": user.user_id,
                    "title": goal_data.get("title", content[:50]),
                    "description": goal_data.get("description", ""),
                    "target_date": goal_data.get("target_date"),
                    "category": goal_data.get("category", "pessoal"),
                    "target_value": goal_data.get("target_value"),
                    "current_value": goal_data.get("current_value", 0),
                    "completed": False,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.goals.insert_one(goal_doc)
                saved_item = {"type": "goal", "id": goal_id, "title": goal_data.get("title", "")}
                
                ai_response_text = f"""✅ **Meta criada com sucesso!**

🎯 **{goal_data.get('title', '')}**
{f"📝 {goal_data.get('description', '')}" if goal_data.get('description') else ""}
{f"📅 Prazo: {goal_data.get('target_date')}" if goal_data.get('target_date') else ""}
📂 Categoria: {goal_data.get('category', 'pessoal')}

💡 A meta está disponível na aba Metas!"""
            except:
                ai_response_text = "Não consegui criar a meta. Tente: 'Minha meta é economizar 5000 até dezembro'"

        else:
            # General / finance_general / study - enriched with DEEP FULL app context
            current_month = datetime.now(timezone.utc).strftime("%Y-%m")
            today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
            now_utc = datetime.now(timezone.utc)
            week_ago_str = (now_utc - timedelta(days=7)).strftime("%Y-%m-%d")
            last_month = (now_utc.replace(day=1) - timedelta(days=1)).strftime("%Y-%m")
            
            # === USER PROFILE CONTEXT ===
            user_doc = await db.users.find_one({"user_id": user.user_id}, {"_id": 0, "name": 1, "email": 1, "birth_date": 1, "bio": 1, "xp": 1, "rank": 1, "health_condition": 1, "created_at": 1})
            user_name = user_doc.get("name", "Usuário") if user_doc else "Usuário"
            user_rank = user_doc.get("rank", "Recruta") if user_doc else "Recruta"
            user_xp = user_doc.get("xp", 0) if user_doc else 0
            user_bio = user_doc.get("bio", "") if user_doc else ""
            user_birth = user_doc.get("birth_date", "") if user_doc else ""
            health_cond = user_doc.get("health_condition", "") if user_doc else ""
            user_age = ""
            if user_birth:
                try:
                    bd = datetime.fromisoformat(user_birth)
                    user_age = str(now_utc.year - bd.year - ((now_utc.month, now_utc.day) < (bd.month, bd.day)))
                except: pass
            
            # === FINANCE CONTEXT (current + last month for trends) ===
            fin_trans = await db.transactions.find({"user_id": user.user_id, "date": {"$regex": f"^{current_month}"}}, {"_id": 0}).to_list(100)
            total_in = sum(t['amount'] for t in fin_trans if t['type'] == 'income')
            total_out = sum(t['amount'] for t in fin_trans if t['type'] == 'expense')
            expense_cats = {}
            for t in fin_trans:
                if t['type'] == 'expense':
                    expense_cats[t['category']] = expense_cats.get(t['category'], 0) + t['amount']
            top_expenses = sorted(expense_cats.items(), key=lambda x: x[1], reverse=True)[:5]
            
            # Last month comparison
            last_month_trans = await db.transactions.find({"user_id": user.user_id, "date": {"$regex": f"^{last_month}"}}, {"_id": 0}).to_list(100)
            lm_in = sum(t['amount'] for t in last_month_trans if t['type'] == 'income')
            lm_out = sum(t['amount'] for t in last_month_trans if t['type'] == 'expense')
            
            budgets = await db.budgets.find({"user_id": user.user_id, "month": current_month}, {"_id": 0}).to_list(20)
            budget_alerts = []
            budget_status = []
            for b in budgets:
                spent = expense_cats.get(b.get("category", ""), 0)
                if b.get("amount", 0) > 0:
                    pct = int(spent / b["amount"] * 100)
                    budget_status.append(f"{b['category']}: R$ {spent:.0f}/{b['amount']:.0f} ({pct}%)")
                    if pct >= 80:
                        budget_alerts.append(f"{b['category']}: {pct}% usado (R$ {spent:.0f}/{b['amount']:.0f})")
            
            # Credit cards
            credit_cards = await db.credit_cards.find({"user_id": user.user_id}, {"_id": 0, "name": 1, "limit": 1, "current_balance": 1}).to_list(10)
            cc_info = []
            for cc in credit_cards:
                cc_info.append(f"{cc.get('name','Cartão')}: R$ {cc.get('current_balance',0):.0f}/{cc.get('limit',0):.0f}")
            
            # Projections for next month
            next_month_dt = (now_utc.replace(day=28) + timedelta(days=4)).replace(day=1)
            next_month_str = next_month_dt.strftime("%Y-%m")
            projections = await db.projections.find({"user_id": user.user_id, "month": next_month_str}, {"_id": 0, "description": 1, "amount": 1, "type": 1}).to_list(20)
            proj_total = sum(p.get("amount", 0) for p in projections)
            
            # === GOALS CONTEXT ===
            goals = await db.goals.find({"user_id": user.user_id, "completed": False}, {"_id": 0, "title": 1, "category": 1, "target_date": 1, "target_value": 1, "current_value": 1}).to_list(10)
            completed_goals = await db.goals.count_documents({"user_id": user.user_id, "completed": True})
            goals_text = []
            for g in goals:
                prog = ""
                if g.get("target_value") and g.get("current_value") is not None:
                    pct = int((g["current_value"] / g["target_value"]) * 100) if g["target_value"] > 0 else 0
                    prog = f" ({pct}% concluído)"
                deadline = f" - prazo: {g['target_date']}" if g.get("target_date") else ""
                goals_text.append(f"• {g['title']} [{g.get('category','geral')}]{prog}{deadline}")
            
            # === WORKOUT CONTEXT (deep) ===
            recent_workouts = await db.workout_logs.find(
                {"user_id": user.user_id}, {"_id": 0, "date": 1, "duration_minutes": 1, "calories": 1, "exercises": 1}
            ).sort("created_at", -1).to_list(15)
            last_workout_date = recent_workouts[0].get("date", "nunca") if recent_workouts else "nunca"
            workouts_this_week = len([w for w in recent_workouts if w.get("date", "") >= week_ago_str])
            workouts_this_month = len([w for w in recent_workouts if w.get("date", "").startswith(current_month)])
            total_workout_min = sum(w.get("duration_minutes", 0) for w in recent_workouts[:7])
            total_workout_cals = sum(w.get("calories", 0) for w in recent_workouts[:7])
            
            workout_plans = await db.workout_plans.find({"user_id": user.user_id}, {"_id": 0, "name": 1, "objective": 1, "level": 1}).to_list(5)
            
            # Recent sessions with feedback
            recent_sessions = await db.workout_sessions.find(
                {"user_id": user.user_id, "status": "completed"}, {"_id": 0, "difficulty": 1, "feeling": 1, "notes": 1, "completed_at": 1}
            ).sort("completed_at", -1).to_list(5)
            avg_difficulty = sum(s.get("difficulty", 3) for s in recent_sessions) / len(recent_sessions) if recent_sessions else 0
            feelings = [s.get("feeling", "") for s in recent_sessions if s.get("feeling")]
            
            # === NUTRITION CONTEXT (deep) ===
            today_meals = await db.meals.find({"user_id": user.user_id, "date": today_str}, {"_id": 0}).to_list(20)
            today_cals = sum(m.get("calories", 0) for m in today_meals)
            today_protein = sum(m.get("protein", 0) for m in today_meals)
            today_carbs = sum(m.get("carbs", 0) for m in today_meals)
            today_fat = sum(m.get("fat", 0) for m in today_meals)
            nutrition_goals = await db.nutrition_goals.find_one({"user_id": user.user_id}, {"_id": 0})
            cal_goal = nutrition_goals.get("calories", 2000) if nutrition_goals else 2000
            protein_goal = nutrition_goals.get("protein", 150) if nutrition_goals else 150
            water_today = await db.water_logs.find_one({"user_id": user.user_id, "date": today_str}, {"_id": 0})
            water_ml = water_today.get("total_ml", 0) if water_today else 0
            water_goal = nutrition_goals.get("water_ml", 2500) if nutrition_goals else 2500
            
            # Week average nutrition
            week_meals = await db.meals.find({"user_id": user.user_id, "date": {"$gte": week_ago_str}}, {"_id": 0, "calories": 1, "date": 1}).to_list(100)
            meals_by_day = {}
            for m in week_meals:
                d = m.get("date", "")
                meals_by_day[d] = meals_by_day.get(d, 0) + m.get("calories", 0)
            avg_cals_week = int(sum(meals_by_day.values()) / max(len(meals_by_day), 1)) if meals_by_day else 0
            
            # Active diets
            active_diets = await db.diets.find({"user_id": user.user_id, "active": True}, {"_id": 0, "name": 1, "type": 1}).to_list(5)
            
            # === STUDY CONTEXT (deep) ===
            study_streak_doc = await db.study_streaks.find_one({"user_id": user.user_id}, {"_id": 0})
            study_streak = study_streak_doc.get("current_streak", 0) if study_streak_doc else 0
            longest_study_streak = study_streak_doc.get("longest_streak", 0) if study_streak_doc else 0
            focus_today = await db.focus_sessions.find({"user_id": user.user_id, "date": today_str}, {"_id": 0}).to_list(20)
            focus_min_today = sum(f.get("duration_minutes", 0) for f in focus_today)
            
            # Study programs
            study_programs = await db.study_programs.find({"user_id": user.user_id, "status": "active"}, {"_id": 0, "name": 1, "source_type": 1, "target_date": 1}).to_list(5)
            
            # Flashcards due for review
            flashcard_decks = await db.flashcard_decks.find({"user_id": user.user_id}, {"_id": 0, "name": 1, "cards": 1}).to_list(10)
            total_cards = 0
            cards_due = 0
            for deck in flashcard_decks:
                for card in deck.get("cards", []):
                    total_cards += 1
                    nr = card.get("next_review", "")
                    if nr and nr <= today_str:
                        cards_due += 1
            
            # Study notebooks with recent activity
            notebooks = await db.notebooks.find({"user_id": user.user_id}, {"_id": 0, "name": 1, "study_time_minutes": 1, "total_questions": 1, "correct_questions": 1}).to_list(20)
            top_subjects = sorted(notebooks, key=lambda n: n.get("study_time_minutes", 0), reverse=True)[:5]
            
            # Question stats
            total_q = sum(n.get("total_questions", 0) for n in notebooks)
            correct_q = sum(n.get("correct_questions", 0) for n in notebooks)
            accuracy = int(correct_q / total_q * 100) if total_q > 0 else 0
            
            # === HABITS & TASKS CONTEXT (deep) ===
            habits = await db.habits.find({"user_id": user.user_id}, {"_id": 0, "name": 1, "frequency": 1}).to_list(20)
            habits_completed = await db.habits.count_documents({"user_id": user.user_id, f"completions.{today_str}": True})
            tasks_total = await db.tasks.count_documents({"user_id": user.user_id})
            tasks_completed = await db.task_instances.count_documents({"user_id": user.user_id, "date": today_str, "completed": True})
            
            # Pending tasks
            pending_tasks = await db.tasks.find({"user_id": user.user_id}, {"_id": 0, "title": 1, "priority": 1}).to_list(10)
            high_priority_tasks = [t for t in pending_tasks if t.get("priority") == "high"]
            
            # Global streaks
            streaks_doc = await db.global_streaks.find_one({"user_id": user.user_id}, {"_id": 0})
            global_streak = streaks_doc.get("current_streak", 0) if streaks_doc else 0
            
            # === BUILD COMPREHENSIVE CONTEXT ===
            context_parts = []
            
            # User profile
            profile_line = f"👤 PERFIL: {user_name}"
            if user_age: profile_line += f", {user_age} anos"
            profile_line += f" | Rank: {user_rank} ({user_xp} XP) | Streak global: {global_streak} dias"
            if user_bio: profile_line += f"\n   Bio: {user_bio}"
            if health_cond: profile_line += f"\n   Condição de saúde: {health_cond}"
            context_parts.append(profile_line)
            
            # Finance
            fin_line = f"💰 FINANÇAS ({current_month}): Receitas R$ {total_in:.2f} | Despesas R$ {total_out:.2f} | Saldo R$ {total_in - total_out:.2f}"
            if lm_out > 0:
                change_pct = int(((total_out - lm_out) / lm_out) * 100) if lm_out > 0 else 0
                trend = "↑" if change_pct > 0 else "↓" if change_pct < 0 else "="
                fin_line += f"\n   Vs mês anterior: {trend} {abs(change_pct)}% nas despesas (era R$ {lm_out:.0f})"
            context_parts.append(fin_line)
            if top_expenses:
                context_parts.append(f"   Top gastos: {', '.join(f'{c}: R$ {v:.0f}' for c,v in top_expenses)}")
            if budget_status:
                context_parts.append(f"   Orçamentos: {' | '.join(budget_status)}")
            if budget_alerts:
                context_parts.append(f"   ⚠️ Alertas: {'; '.join(budget_alerts)}")
            if cc_info:
                context_parts.append(f"   💳 Cartões: {' | '.join(cc_info)}")
            if projections:
                context_parts.append(f"   📊 Projeções próximo mês: R$ {proj_total:.0f} em {len(projections)} itens")
            
            # Workouts
            wk_line = f"🏋️ TREINOS: Último treino: {last_workout_date} | Esta semana: {workouts_this_week} treinos ({total_workout_min}min, ~{total_workout_cals} kcal)"
            context_parts.append(wk_line)
            if workout_plans:
                plan_names = ', '.join(p.get('name','')[:25] + " (" + p.get('objective','') + ")" for p in workout_plans)
                context_parts.append(f"   Fichas: {plan_names}")
            if recent_sessions:
                context_parts.append(f"   Dificuldade média: {avg_difficulty:.1f}/5 | Sentimentos recentes: {', '.join(feelings[:3])}")
            
            # Nutrition
            nut_line = f"🍽️ NUTRIÇÃO HOJE: {today_cals}/{cal_goal} cal | {today_protein}g/{protein_goal}g prot | {today_carbs}g carbs | {today_fat}g gordura | {len(today_meals)} refeições"
            context_parts.append(nut_line)
            context_parts.append(f"   💧 Água: {water_ml}/{water_goal}ml | Média semanal: {avg_cals_week} cal/dia")
            if active_diets:
                context_parts.append(f"   Dietas ativas: {', '.join(d.get('name','') for d in active_diets)}")
            
            # Studies
            st_line = f"📚 ESTUDOS: Streak {study_streak} dias (recorde: {longest_study_streak}) | Foco hoje: {focus_min_today}min"
            context_parts.append(st_line)
            if study_programs:
                progs = ', '.join(f"{p.get('name','')[:30]}" + (f" (prova: {p['target_date']})" if p.get('target_date') else "") for p in study_programs)
                context_parts.append(f"   Programas ativos: {progs}")
            if total_q > 0:
                context_parts.append(f"   Questões: {total_q} total, {accuracy}% acerto")
            if cards_due > 0:
                context_parts.append(f"   ⚠️ {cards_due} flashcards pendentes para revisão hoje!")
            if top_subjects:
                subj_names = ', '.join(n.get('name','')[:20] + " (" + str(n.get('study_time_minutes',0)) + "min)" for n in top_subjects[:3])
                context_parts.append(f"   Top matérias: {subj_names}")
            
            # Tasks & Habits
            context_parts.append(f"✅ TAREFAS: {tasks_completed}/{tasks_total} completadas hoje | Hábitos: {habits_completed}/{len(habits)} hoje")
            if high_priority_tasks:
                context_parts.append(f"   🔴 Tarefas urgentes: {', '.join(t.get('title','')[:25] for t in high_priority_tasks[:3])}")
            
            # Goals
            if goals:
                context_parts.append(f"🎯 METAS ATIVAS ({len(goals)}, {completed_goals} concluídas):")
                for gt in goals_text[:5]:
                    context_parts.append(f"   {gt}")
            
            full_context = "\n".join(context_parts)
            
            # Build INTELLIGENT proactive suggestions based on comprehensive data
            proactive_hints = []
            
            # Workout proactivity
            if last_workout_date != "nunca":
                try:
                    days_since = (now_utc - datetime.fromisoformat(last_workout_date.replace("Z", "+00:00") if "T" in last_workout_date else last_workout_date + "T00:00:00+00:00")).days
                    if days_since >= 3:
                        proactive_hints.append(f"O usuário não treina há {days_since} dias - sugira gentilmente retomar")
                    if days_since == 0 and avg_difficulty >= 4:
                        proactive_hints.append("Treinou hoje com dificuldade alta - sugira descanso ou alongamento")
                except: pass
            elif workouts_this_month == 0:
                proactive_hints.append("Nenhum treino este mês - motive a começar uma rotina de exercícios")
            
            # Nutrition proactivity
            if today_cals > 0 and today_cals < cal_goal * 0.3 and now_utc.hour >= 14:
                proactive_hints.append(f"Já passa das 14h e consumiu apenas {today_cals} de {cal_goal} calorias - pergunte se está se alimentando bem")
            if today_protein > 0 and today_protein < protein_goal * 0.3 and now_utc.hour >= 16:
                proactive_hints.append(f"Proteína baixa hoje ({today_protein}g de {protein_goal}g) - sugira alimentos ricos em proteína")
            if water_ml < water_goal * 0.4 and now_utc.hour >= 12:
                proactive_hints.append(f"Apenas {water_ml}ml de {water_goal}ml de água - lembre de se hidratar")
            
            # Finance proactivity
            if budget_alerts:
                proactive_hints.append("Alguns orçamentos estão próximos do limite - alerte se relevante")
            if total_out > lm_out * 1.2 and lm_out > 0:
                proactive_hints.append(f"Gastos {int(((total_out-lm_out)/lm_out)*100)}% acima do mês anterior - sugira revisão de gastos")
            
            # Study proactivity
            if cards_due > 0:
                proactive_hints.append(f"{cards_due} flashcards pendentes para revisão - motivar a estudar")
            if study_streak > 0 and focus_min_today == 0 and now_utc.hour >= 18:
                proactive_hints.append(f"Streak de {study_streak} dias em risco! Não estudou hoje ainda")
            if study_programs:
                for sp in study_programs:
                    if sp.get("target_date"):
                        try:
                            td = datetime.fromisoformat(sp["target_date"])
                            days_left = (td - now_utc).days
                            if 0 < days_left <= 30:
                                proactive_hints.append(f"Prova de '{sp['name'][:30]}' em {days_left} dias - motivar intensificação dos estudos")
                        except: pass
            
            # Goal proactivity
            for g in goals:
                if g.get("target_date"):
                    try:
                        td = datetime.fromisoformat(g["target_date"])
                        days_left = (td - now_utc).days
                        if 0 < days_left <= 14:
                            proactive_hints.append(f"Meta '{g['title'][:25]}' vence em {days_left} dias")
                    except: pass
            
            # Tasks proactivity
            if high_priority_tasks and tasks_completed == 0 and now_utc.hour >= 15:
                proactive_hints.append(f"{len(high_priority_tasks)} tarefas urgentes pendentes e nenhuma concluída hoje")
            
            proactive_text = ""
            if proactive_hints:
                proactive_text = "\n\nSugestões proativas (mencione NATURALMENTE se relevante à conversa, não liste tudo de uma vez):\n- " + "\n- ".join(proactive_hints[:6])
            
            system = f"""Você é o SIRIUS 🐺, assistente pessoal integrado com visão TOTAL da vida do usuário. Você é como um mentor estratégico que conhece finanças, treinos, nutrição, estudos, metas e rotina do usuário.

PERSONALIDADE: Direto, inteligente, motivador mas realista. Como um lobo líder da matilha — leal, estratégico e focado em resultados. Use o nome do usuário quando relevante.

CAPACIDADES:
🏋️ TREINOS: Criar fichas, sugerir exercícios, analisar desempenho, ajustar planos
🍽️ ALIMENTAÇÃO: Receitas, dietas, análise nutricional, planos alimentares personalizados
📚 ESTUDOS: Cronogramas, revisão espaçada, dicas para concursos, análise de desempenho
💰 FINANÇAS: Registrar transações, analisar gastos, projeções, dicas de economia
✅ TAREFAS: Criar e organizar tarefas, sugerir prioridades
🎯 METAS: Definir objetivos, acompanhar progresso, sugerir ações

CONTEXTO COMPLETO DO USUÁRIO:
{full_context}
{proactive_text}

INSTRUÇÕES CRÍTICAS:
- SEMPRE use o contexto real do usuário para personalizar respostas
- Faça CONEXÕES INTELIGENTES entre módulos (ex: "Você treinou pesado hoje, que tal aumentar a proteína no jantar?")
- Se o usuário perguntar algo genérico, dê insights baseados nos dados dele
- Seja proativo: se notar padrões ou alertas, mencione naturalmente (mas não despeje tudo de uma vez)
- Para registrar transação: 'Gastei 50 no mercado' ou 'Recebi 3000 de salário'
- Para criar tarefa: 'Criar tarefa: ...'
- Para criar meta: 'Minha meta é ...'
- Responda em português, de forma objetiva, personalizada e motivadora
- Se perceber que o usuário está negligenciando alguma área, sugira de forma empática"""
            response = await call_llm(content, f"general_chat_{user.user_id}", system, user_id=user.user_id)
            ai_response_text = response

    except Exception:
        ai_response_text = "Desculpe, ocorreu um erro ao processar sua mensagem. Tente novamente."

    # Save AI response
    ai_msg_id = f"msg_{uuid.uuid4().hex[:12]}"
    ai_msg = {
        "message_id": ai_msg_id,
        "user_id": user.user_id,
        "role": "assistant",
        "content": ai_response_text,
        "chat_type": "general",
        "intent": intent,
        "saved_item": saved_item,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.chat_messages.insert_one(ai_msg.copy())

    return {
        "user_message": {k: v for k, v in user_msg.items() if k != '_id'},
        "ai_message": {k: v for k, v in ai_msg.items() if k != '_id'},
        "intent": intent,
        "saved_item": saved_item
    }


@api_router.get("/chat/general/messages")
async def get_general_messages(request: Request, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    messages = await db.chat_messages.find(
        {"user_id": user.user_id, "chat_type": "general"}, {"_id": 0}
    ).sort("created_at", 1).to_list(200)
    return messages


# ========== MONTHLY BILLS (CONTAS DO MÊS) ==========

@api_router.get("/finance/monthly-bills")
async def get_monthly_bills(request: Request, month: Optional[str] = None, session_token: Optional[str] = Cookie(None)):
    """Get monthly bills - auto-import from projections when current month"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)

    if not month:
        month = datetime.now(timezone.utc).strftime("%Y-%m")

    # Check if bills already exist for this month
    existing_bills = await db.monthly_bills.find(
        {"user_id": user.user_id, "month": month}, {"_id": 0}
    ).to_list(500)

    # Auto-import from projections if no bills exist
    if not existing_bills:
        projections = await db.projections.find(
            {"user_id": user.user_id, "month": month}, {"_id": 0}
        ).to_list(500)

        for proj in projections:
            bill_id = f"bill_{uuid.uuid4().hex[:12]}"
            bill_doc = {
                "bill_id": bill_id,
                "user_id": user.user_id,
                "month": month,
                "description": proj.get("description", ""),
                "amount": proj.get("amount", 0),
                "category": proj.get("category", "outros"),
                "paid": False,
                "paid_date": None,
                "source": "projection",
                "projection_id": proj.get("projection_id"),
                "card_id": proj.get("card_id"),
                "installment_info": f"{proj.get('installment_number', '')}/{proj.get('total_installments', '')}" if proj.get('installment_number') else None,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.monthly_bills.insert_one(bill_doc)
            existing_bills.append(bill_doc)

    # Calculate totals
    total = sum(b["amount"] for b in existing_bills)
    total_paid = sum(b["amount"] for b in existing_bills if b.get("paid"))
    total_pending = total - total_paid

    return {
        "month": month,
        "bills": existing_bills,
        "total": total,
        "total_paid": total_paid,
        "total_pending": total_pending,
        "count": len(existing_bills),
        "paid_count": sum(1 for b in existing_bills if b.get("paid"))
    }


@api_router.post("/finance/monthly-bills")
async def create_monthly_bill(request: Request, data: dict, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)

    bill_id = f"bill_{uuid.uuid4().hex[:12]}"
    bill_doc = {
        "bill_id": bill_id,
        "user_id": user.user_id,
        "month": data.get("month", datetime.now(timezone.utc).strftime("%Y-%m")),
        "description": data.get("description", ""),
        "amount": data.get("amount", 0),
        "category": data.get("category", "outros"),
        "paid": False,
        "paid_date": None,
        "source": "manual",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.monthly_bills.insert_one(bill_doc)
    bill_doc.pop("_id", None)
    return bill_doc


@api_router.patch("/finance/monthly-bills/{bill_id}/toggle")
async def toggle_bill_paid(request: Request, bill_id: str, session_token: Optional[str] = Cookie(None)):
    """Toggle bill paid status - affects balance"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)

    bill = await db.monthly_bills.find_one({"bill_id": bill_id, "user_id": user.user_id})
    if not bill:
        raise HTTPException(status_code=404, detail="Conta não encontrada")

    new_paid = not bill.get("paid", False)
    paid_date = datetime.now(timezone.utc).isoformat() if new_paid else None

    await db.monthly_bills.update_one(
        {"bill_id": bill_id},
        {"$set": {"paid": new_paid, "paid_date": paid_date}}
    )

    # If marking as paid, create an expense transaction
    if new_paid:
        tx_id = f"tx_{uuid.uuid4().hex[:12]}"
        tx_doc = {
            "transaction_id": tx_id,
            "user_id": user.user_id,
            "type": "expense",
            "amount": bill["amount"],
            "category": bill.get("category", "outros"),
            "description": f"[Conta Paga] {bill.get('description', '')}",
            "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "bill_id": bill_id,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.transactions.insert_one(tx_doc)
    else:
        # If unmarking, remove the auto-created transaction
        await db.transactions.delete_many({"bill_id": bill_id, "user_id": user.user_id})

    return {"bill_id": bill_id, "paid": new_paid, "message": "Conta marcada como paga!" if new_paid else "Pagamento desmarcado."}


@api_router.delete("/finance/monthly-bills/{bill_id}")
async def delete_monthly_bill(request: Request, bill_id: str, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    result = await db.monthly_bills.delete_one({"bill_id": bill_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Conta não encontrada")
    return {"message": "Conta removida"}


# ========== WORKOUT IMPORT FROM FILE ==========

@api_router.post("/workouts/import-plan")
async def import_workout_plan(
    request: Request,
    file: UploadFile = File(...),
    session_token: Optional[str] = Cookie(None)
):
    """Import a workout plan from PDF or image using AI extraction"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)

    if not gemini_client:
        raise HTTPException(status_code=500, detail="Serviço de IA indisponível")

    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Arquivo muito grande. Limite de 20MB.")

    try:
        import tempfile
        ext = os.path.splitext(file.filename)[1].lower()
        suffix = ext if ext in ['.pdf'] else '.jpg'

        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        mime = file.content_type or ('application/pdf' if ext == '.pdf' else 'image/jpeg')
        uploaded = gemini_client.files.upload(file=tmp_path)
        os.unlink(tmp_path)

        system_msg = """Analise esta ficha de treino e extraia TODOS os exercícios.
Responda APENAS com JSON:
{
  "name": "Nome do Treino (ex: Treino A - Peito/Tríceps)",
  "description": "Descrição breve",
  "exercises": [
    {"name": "Nome do exercício", "sets": 3, "reps": "12", "weight": "10kg", "notes": "Observações"}
  ]
}
REGRAS: Extraia TODOS os exercícios fielmente. Se não conseguir ler algo, indique com [ilegível]."""

        response = gemini_client.models.generate_content(
            model=GEMINI_MODEL,
            contents=[
                types.Part.from_uri(file_uri=uploaded.uri, mime_type=mime),
                "Extraia a ficha de treino deste documento:"
            ],
            config=types.GenerateContentConfig(system_instruction=system_msg)
        )

        json_str = response.text.strip()
        if json_str.startswith("```json"): json_str = json_str[7:]
        if json_str.startswith("```"): json_str = json_str[3:]
        if json_str.endswith("```"): json_str = json_str[:-3]
        plan_data = json.loads(json_str.strip())

        plan_id = f"plan_{uuid.uuid4().hex[:12]}"
        plan_doc = {
            "plan_id": plan_id,
            "user_id": user.user_id,
            "name": plan_data.get("name", f"Treino importado - {file.filename}"),
            "description": plan_data.get("description", "Importado via arquivo"),
            "exercises": plan_data.get("exercises", []),
            "source": "file_import",
            "source_filename": file.filename,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.workout_plans.insert_one(plan_doc)
        plan_doc.pop("_id", None)

        await award_xp(user.user_id, 10)

        return {
            "success": True,
            "plan": plan_doc,
            "message": f"Treino importado com {len(plan_data.get('exercises', []))} exercícios!",
            "xp_earned": 10
        }

    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="Erro ao processar ficha de treino. Tente novamente.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro: {str(e)}")


# ========== SAVED WORKOUT INSIGHTS ==========

@api_router.post("/workout-suggestions/save")
async def save_workout_suggestion(request: Request, data: dict, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)

    insight_id = f"wi_{uuid.uuid4().hex[:12]}"
    doc = {
        "insight_id": insight_id,
        "user_id": user.user_id,
        "title": data.get("title", "Sugestão de Treino"),
        "content": data.get("content", ""),
        "based_on": data.get("based_on", {}),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.workout_insights.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.get("/workout-suggestions/saved")
async def get_saved_insights(request: Request, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    insights = await db.workout_insights.find({"user_id": user.user_id}, {"_id": 0}).sort("created_at", -1).to_list(50)
    return insights


@api_router.delete("/workout-suggestions/saved/{insight_id}")
async def delete_saved_insight(request: Request, insight_id: str, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    result = await db.workout_insights.delete_one({"insight_id": insight_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Insight não encontrado")
    return {"message": "Insight removido"}


# ========== AI MEAL PLAN GENERATION ==========
@api_router.post("/nutrition/meal-plan/generate")
async def generate_meal_plan(request: Request, session_token: Optional[str] = Cookie(None)):
    """Generate a personalized meal plan with AI"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if not gemini_client:
        raise HTTPException(status_code=503, detail="Serviço de IA indisponível")
    
    body = await request.json()
    objective = body.get("objective", "saude")  # saude, emagrecimento, hipertrofia, definicao
    restrictions = body.get("restrictions", [])  # vegetariano, vegano, sem_gluten, sem_lactose, low_carb
    meals_per_day = body.get("meals_per_day", 5)
    duration = body.get("duration", "dia")  # dia, semana
    calories_target = body.get("calories_target", 0)  # 0 = auto
    
    restrictions_text = ""
    if restrictions:
        restrictions_text = f"\nRestrições alimentares: {', '.join(restrictions)}"
    
    calories_text = ""
    if calories_target > 0:
        calories_text = f"\nMeta calórica: {calories_target} kcal/dia"
    
    duration_instruction = "Crie um plano para UM DIA." if duration == "dia" else "Crie um plano para UMA SEMANA (segunda a domingo, 7 dias)."
    
    prompt = f"""Você é um nutricionista certificado. Gere um plano alimentar completo em JSON.

PARÂMETROS:
- Objetivo: {objective}
- Refeições por dia: {meals_per_day}{restrictions_text}{calories_text}
- {duration_instruction}

FORMATO JSON OBRIGATÓRIO:
{{
  "name": "Plano Alimentar - {objective}",
  "description": "Descrição breve",
  "calories_total": 2000,
  "macros": {{"protein_g": 150, "carbs_g": 200, "fat_g": 70, "fiber_g": 30}},
  "days": [
    {{
      "day_name": "dia1",
      "day_label": "Segunda-feira",
      "calories": 2000,
      "meals": [
        {{
          "meal_type": "café_da_manhã",
          "time": "07:00",
          "name": "Omelete de claras com aveia",
          "foods": [
            {{"name": "Clara de ovo", "quantity": "4 unidades", "calories": 68, "protein": 14, "carbs": 0, "fat": 0}},
            {{"name": "Aveia", "quantity": "40g", "calories": 140, "protein": 5, "carbs": 24, "fat": 3}}
          ],
          "total_calories": 208,
          "preparation": "Bata as claras, adicione sal e temperos. Cozinhe em frigideira antiaderente. Sirva com aveia cozida em água."
        }}
      ]
    }}
  ],
  "shopping_list": [
    {{"name": "Clara de ovo", "quantity": "20 unidades", "category": "proteínas"}},
    {{"name": "Aveia", "quantity": "200g", "category": "cereais"}}
  ],
  "tips": ["Beba no mínimo 2L de água por dia", "Evite comer 2h antes de dormir"]
}}

IMPORTANTE:
- Retorne APENAS o JSON, sem markdown, sem ```json
- Inclua lista de compras (shopping_list) completa
- Inclua dicas (tips) personalizadas ao objetivo
- Macros devem ser realistas e adaptados ao objetivo
- Cada refeição deve ter instrução de preparo
- {"Retorne apenas 1 dia" if duration == "dia" else "Retorne 7 dias"} no array days"""

    try:
        response = gemini_client.models.generate_content(
            model=GEMINI_MODEL,
            contents=prompt,
            config=types.GenerateContentConfig(
                system_instruction="Você é um nutricionista profissional. Sempre responda em JSON válido."
            )
        )
        
        response_text = response.text.strip()
        if response_text.startswith("```"):
            response_text = response_text.split("\n", 1)[1] if "\n" in response_text else response_text[3:]
        if response_text.endswith("```"):
            response_text = response_text[:-3].strip()
        if response_text.startswith("json"):
            response_text = response_text[4:].strip()
            
        plan_data = json.loads(response_text)
        
        plan_id = f"mealplan_{uuid.uuid4().hex[:12]}"
        plan_doc = {
            "plan_id": plan_id,
            "user_id": user.user_id,
            "type": "meal_plan",
            **plan_data,
            "objective": objective,
            "restrictions": restrictions,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.meal_plans.insert_one(plan_doc)
        plan_doc.pop('_id', None)
        
        xp_earned = 5
        new_xp = user.xp + xp_earned
        new_rank = calculate_rank(new_xp)
        await db.users.update_one({"user_id": user.user_id}, {"$set": {"xp": new_xp, "rank": new_rank}})
        
        return {"success": True, "plan": plan_doc, "xp_earned": xp_earned}
        
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="Erro ao processar resposta da IA. Tente novamente.")
    except Exception as e:
        logging.error(f"Meal plan generation failed: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao gerar plano alimentar: {str(e)[:100]}")


@api_router.get("/nutrition/meal-plans")
async def get_meal_plans(request: Request, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    plans = await db.meal_plans.find({"user_id": user.user_id}, {"_id": 0}).sort("created_at", -1).to_list(20)
    return plans


@api_router.delete("/nutrition/meal-plans/{plan_id}")
async def delete_meal_plan(request: Request, plan_id: str, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    result = await db.meal_plans.delete_one({"plan_id": plan_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Plano não encontrado")
    return {"success": True}


# ========== HEALTH CALCULATOR ==========
@api_router.post("/health/calculate")
async def calculate_health_metrics(request: Request, session_token: Optional[str] = Cookie(None)):
    """Calculate BMI, BMR, TDEE, and recommended macros"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    body = await request.json()
    weight = float(body.get("weight", 70))
    height = float(body.get("height", 170))  # cm
    age = int(body.get("age", 25))
    gender = body.get("gender", "male")  # male, female
    activity_level = body.get("activity_level", "moderate")  # sedentary, light, moderate, active, very_active
    objective = body.get("objective", "maintain")  # lose, maintain, gain
    
    # BMI
    height_m = height / 100
    bmi = round(weight / (height_m ** 2), 1)
    
    if bmi < 18.5:
        bmi_class = "Abaixo do peso"
    elif bmi < 25:
        bmi_class = "Peso normal"
    elif bmi < 30:
        bmi_class = "Sobrepeso"
    elif bmi < 35:
        bmi_class = "Obesidade grau I"
    elif bmi < 40:
        bmi_class = "Obesidade grau II"
    else:
        bmi_class = "Obesidade grau III"
    
    # BMR (Mifflin-St Jeor)
    if gender == "male":
        bmr = round(10 * weight + 6.25 * height - 5 * age + 5)
    else:
        bmr = round(10 * weight + 6.25 * height - 5 * age - 161)
    
    # TDEE
    activity_multipliers = {
        "sedentary": 1.2,
        "light": 1.375,
        "moderate": 1.55,
        "active": 1.725,
        "very_active": 1.9
    }
    tdee = round(bmr * activity_multipliers.get(activity_level, 1.55))
    
    # Calorie target based on objective
    if objective == "lose":
        calories_target = tdee - 500
    elif objective == "gain":
        calories_target = tdee + 300
    else:
        calories_target = tdee
    
    # Macros
    if objective == "gain":
        protein_g = round(weight * 2.0)
        fat_g = round(weight * 1.0)
        carbs_g = round((calories_target - (protein_g * 4 + fat_g * 9)) / 4)
    elif objective == "lose":
        protein_g = round(weight * 2.2)
        fat_g = round(weight * 0.8)
        carbs_g = round((calories_target - (protein_g * 4 + fat_g * 9)) / 4)
    else:
        protein_g = round(weight * 1.6)
        fat_g = round(weight * 1.0)
        carbs_g = round((calories_target - (protein_g * 4 + fat_g * 9)) / 4)
    
    # Ideal weight range (BMI 18.5-24.9)
    ideal_weight_min = round(18.5 * (height_m ** 2), 1)
    ideal_weight_max = round(24.9 * (height_m ** 2), 1)
    
    # Water intake recommendation
    water_liters = round(weight * 0.035, 1)
    
    result = {
        "bmi": bmi,
        "bmi_class": bmi_class,
        "bmr": bmr,
        "tdee": tdee,
        "calories_target": calories_target,
        "macros": {
            "protein_g": max(protein_g, 0),
            "carbs_g": max(carbs_g, 0),
            "fat_g": max(fat_g, 0),
            "protein_pct": round(protein_g * 4 / max(calories_target, 1) * 100),
            "carbs_pct": round(max(carbs_g, 0) * 4 / max(calories_target, 1) * 100),
            "fat_pct": round(fat_g * 9 / max(calories_target, 1) * 100)
        },
        "ideal_weight": {"min": ideal_weight_min, "max": ideal_weight_max},
        "water_liters": water_liters,
        "objective": objective,
        "input": {"weight": weight, "height": height, "age": age, "gender": gender, "activity_level": activity_level}
    }
    
    return result


# ========== DASHBOARD WEEKLY SUMMARY ==========
@api_router.get("/dashboard/weekly-summary")
async def get_weekly_summary(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get a comprehensive weekly summary across all modules"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    today = datetime.now(timezone.utc)
    week_start = (today - timedelta(days=today.weekday())).strftime("%Y-%m-%d")
    week_end = today.strftime("%Y-%m-%d")
    
    # Finance summary
    transactions = await db.transactions.find(
        {"user_id": user.user_id, "date": {"$gte": week_start, "$lte": week_end}},
        {"_id": 0}
    ).to_list(500)
    
    total_income = sum(t['amount'] for t in transactions if t.get('type') == 'income')
    total_expense = sum(t['amount'] for t in transactions if t.get('type') == 'expense')
    
    # Workouts summary
    workouts = await db.workout_logs.find(
        {"user_id": user.user_id, "date": {"$gte": week_start, "$lte": week_end}},
        {"_id": 0}
    ).to_list(100)
    
    total_workout_minutes = sum(w.get('duration_minutes', 0) for w in workouts)
    total_workout_calories = sum(w.get('calories', 0) for w in workouts)
    
    # Sessions
    sessions = await db.workout_sessions.find(
        {"user_id": user.user_id, "status": "completed", "started_at": {"$gte": week_start}},
        {"_id": 0}
    ).to_list(100)
    
    avg_difficulty = 0
    if sessions:
        difficulties = [s.get('feedback', {}).get('difficulty', 0) for s in sessions if s.get('feedback')]
        avg_difficulty = round(sum(difficulties) / max(len(difficulties), 1), 1)
    
    # Habits summary
    habits = await db.habits.find({"user_id": user.user_id}, {"_id": 0}).to_list(50)
    habits_completed_this_week = 0
    habits_total_possible = 0
    for h in habits:
        completions = h.get('completions', [])
        for c in completions:
            if c >= week_start and c <= week_end:
                habits_completed_this_week += 1
        habits_total_possible += 7
    
    habits_pct = round(habits_completed_this_week / max(habits_total_possible, 1) * 100)
    
    # Tasks summary
    tasks = await db.tasks.find({"user_id": user.user_id}, {"_id": 0}).to_list(200)
    tasks_completed = sum(1 for t in tasks if t.get('completed'))
    tasks_total = len(tasks)
    
    # Study summary
    study_sessions = await db.study_sessions.find(
        {"user_id": user.user_id, "date": {"$gte": week_start}},
        {"_id": 0}
    ).to_list(100)
    total_study_minutes = sum(s.get('duration_minutes', 0) for s in study_sessions)
    
    return {
        "period": {"start": week_start, "end": week_end},
        "finance": {
            "income": total_income,
            "expense": total_expense,
            "balance": total_income - total_expense,
            "transactions_count": len(transactions)
        },
        "workouts": {
            "count": len(workouts),
            "total_minutes": total_workout_minutes,
            "total_calories": total_workout_calories,
            "sessions_completed": len(sessions),
            "avg_difficulty": avg_difficulty
        },
        "habits": {
            "completed": habits_completed_this_week,
            "total_possible": habits_total_possible,
            "completion_pct": habits_pct,
            "active_habits": len(habits)
        },
        "tasks": {
            "completed": tasks_completed,
            "total": tasks_total,
            "completion_pct": round(tasks_completed / max(tasks_total, 1) * 100)
        },
        "study": {
            "sessions": len(study_sessions),
            "total_minutes": total_study_minutes
        },
        "xp": user.xp,
        "rank": user.rank
    }



# ========== UNIFIED STREAKS ==========
@api_router.get("/streaks/global")
async def get_global_streaks(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get unified streaks across all modules"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    today = datetime.now(timezone.utc)
    today_str = today.strftime("%Y-%m-%d")
    
    # Check last 120 days for activity
    active_days = set()
    module_days = {"tasks": set(), "habits": set(), "study": set(), "workouts": set(), "nutrition": set()}
    
    lookback_start = (today - timedelta(days=120)).strftime("%Y-%m-%d")
    
    # Tasks
    task_instances = await db.task_instances.find(
        {"user_id": user.user_id, "completed": True, "date": {"$gte": lookback_start}},
        {"_id": 0, "date": 1}
    ).to_list(5000)
    for t in task_instances:
        active_days.add(t["date"])
        module_days["tasks"].add(t["date"])
    
    # Habits
    habits = await db.habits.find({"user_id": user.user_id}, {"_id": 0, "completions": 1}).to_list(500)
    for h in habits:
        for c in h.get("completions", []):
            if c >= lookback_start:
                active_days.add(c)
                module_days["habits"].add(c)
    
    # Study
    study_sessions = await db.study_sessions.find(
        {"user_id": user.user_id, "date": {"$gte": lookback_start}},
        {"_id": 0, "date": 1}
    ).to_list(5000)
    for s in study_sessions:
        active_days.add(s["date"])
        module_days["study"].add(s["date"])
    
    # Workouts
    workout_logs = await db.workout_logs.find(
        {"user_id": user.user_id, "completed": True, "date": {"$gte": lookback_start}},
        {"_id": 0, "date": 1}
    ).to_list(1000)
    for w in workout_logs:
        active_days.add(w["date"])
        module_days["workouts"].add(w["date"])
    
    # Nutrition
    meals = await db.meals.find(
        {"user_id": user.user_id, "date": {"$gte": lookback_start}},
        {"_id": 0, "date": 1}
    ).to_list(5000)
    for m in meals:
        active_days.add(m["date"])
        module_days["nutrition"].add(m["date"])
    
    # Calculate current streak
    current_streak = 0
    check_date = today
    while True:
        ds = check_date.strftime("%Y-%m-%d")
        if ds in active_days:
            current_streak += 1
            check_date -= timedelta(days=1)
        else:
            # Allow today to be "not yet active" and still count yesterday's streak
            if ds == today_str and current_streak == 0:
                check_date -= timedelta(days=1)
                continue
            break
    
    # Calculate longest streak
    sorted_days = sorted(active_days)
    longest_streak = 0
    temp_streak = 0
    prev_date = None
    for ds in sorted_days:
        d = datetime.strptime(ds, "%Y-%m-%d")
        if prev_date and (d - prev_date).days == 1:
            temp_streak += 1
        else:
            temp_streak = 1
        longest_streak = max(longest_streak, temp_streak)
        prev_date = d
    
    # Calculate combo bonus (how many modules active today)
    today_modules = []
    for mod, days in module_days.items():
        if today_str in days:
            today_modules.append(mod)
    
    combo_count = len(today_modules)
    combo_bonus_xp = 0
    if combo_count >= 3:
        combo_bonus_xp = combo_count * 5
    elif combo_count >= 2:
        combo_bonus_xp = combo_count * 3
    
    # Last 7 days heatmap
    heatmap = []
    for i in range(6, -1, -1):
        d = (today - timedelta(days=i)).strftime("%Y-%m-%d")
        day_modules = []
        for mod, days in module_days.items():
            if d in days:
                day_modules.append(mod)
        heatmap.append({
            "date": d,
            "active": d in active_days,
            "modules": day_modules,
            "count": len(day_modules)
        })
    
    # Module-specific streaks
    module_streaks = {}
    for mod, days in module_days.items():
        mod_streak = 0
        cd = today
        while True:
            ds = cd.strftime("%Y-%m-%d")
            if ds in days:
                mod_streak += 1
                cd -= timedelta(days=1)
            else:
                if ds == today_str and mod_streak == 0:
                    cd -= timedelta(days=1)
                    continue
                break
        module_streaks[mod] = mod_streak
    
    return {
        "current_streak": current_streak,
        "longest_streak": longest_streak,
        "total_active_days": len(active_days),
        "today_modules": today_modules,
        "combo_count": combo_count,
        "combo_bonus_xp": combo_bonus_xp,
        "heatmap": heatmap,
        "module_streaks": module_streaks
    }


# ========== DAILY SUMMARY (AI) ==========
@api_router.get("/dashboard/daily-summary")
async def get_daily_summary(request: Request, session_token: Optional[str] = Cookie(None)):
    """Generate AI-powered daily briefing"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # Check cache first (valid for 4 hours)
    cached = await db.daily_summaries.find_one(
        {"user_id": user.user_id, "date": today_str},
        {"_id": 0}
    )
    if cached and cached.get("summary"):
        created = cached.get("created_at", "")
        if created:
            try:
                cache_time = datetime.fromisoformat(created.replace("Z", "+00:00"))
                if (datetime.now(timezone.utc) - cache_time).total_seconds() < 14400:
                    return cached
            except Exception:
                pass
    
    # Gather data for the summary
    tasks = await db.tasks.find({"user_id": user.user_id, "is_template": True}, {"_id": 0}).to_list(200)
    task_instances = await db.task_instances.find(
        {"user_id": user.user_id, "date": today_str}, {"_id": 0}
    ).to_list(500)
    completed_task_ids = set(i["task_id"] for i in task_instances if i.get("completed"))
    
    pending_tasks = [t for t in tasks if t["task_id"] not in completed_task_ids and t.get("recurrence") in ("daily", "once")]
    done_tasks = [t for t in tasks if t["task_id"] in completed_task_ids]
    
    habits = await db.habits.find({"user_id": user.user_id}, {"_id": 0}).to_list(50)
    habits_done = [h for h in habits if today_str in h.get("completions", [])]
    habits_pending = [h for h in habits if today_str not in h.get("completions", [])]
    
    study_sessions = await db.study_sessions.find(
        {"user_id": user.user_id, "date": today_str}, {"_id": 0}
    ).to_list(100)
    study_minutes = sum(s.get("duration_minutes", 0) for s in study_sessions)
    
    meals = await db.meals.find(
        {"user_id": user.user_id, "date": today_str}, {"_id": 0}
    ).to_list(20)
    total_calories = sum(m.get("total_calories", 0) for m in meals)
    
    workout_logs = await db.workout_logs.find(
        {"user_id": user.user_id, "date": today_str}, {"_id": 0}
    ).to_list(10)
    
    # Build context for AI
    context = f"""Dados do dia ({today_str}) do usuário {user.name}:
- Rank: {user.rank} | XP: {user.xp}
- Tarefas pendentes: {len(pending_tasks)} ({', '.join(t['title'] for t in pending_tasks[:5])})
- Tarefas concluídas: {len(done_tasks)}
- Hábitos pendentes: {len(habits_pending)} ({', '.join(h['name'] for h in habits_pending[:5])})
- Hábitos concluídos: {len(habits_done)}
- Estudo: {study_minutes} minutos em {len(study_sessions)} sessões
- Refeições: {len(meals)} registradas, {total_calories:.0f} kcal total
- Treinos: {len(workout_logs)} realizados"""
    
    prompt = f"""{context}

Gere um resumo diário motivacional e prático em PORTUGUÊS para este usuário.
Inclua:
1. Saudação personalizada usando o nome e rank
2. Resumo do progresso até agora no dia
3. O que ainda falta fazer (tarefas e hábitos pendentes)
4. Dica motivacional breve no estilo militar/disciplina
5. Uma sugestão de ação prioritária

Responda em formato JSON:
{{"greeting": "...", "progress_summary": "...", "pending_items": ["item1", "item2", ...], "motivation": "...", "priority_action": "...", "score": 0-100}}

O score deve refletir o progresso do dia (0 = nada feito, 100 = tudo feito)."""

    summary_data = None
    
    try:
        resp_text = await call_llm(
            prompt + "\n\nResponda APENAS com JSON puro, sem markdown, sem texto extra.",
            f"daily_summary_{user.user_id}",
            "Você é um assistente motivacional que gera resumos diários em JSON.",
            user_id=user.user_id
        )
        if resp_text and not resp_text.startswith("⚠"):
            first_brace = resp_text.find("{")
            if first_brace >= 0:
                resp_text = resp_text[first_brace:]
                last_brace = resp_text.rfind("}")
                if last_brace >= 0:
                    resp_text = resp_text[:last_brace + 1]
                summary_data = json.loads(resp_text)
    except Exception as e:
        logging.error(f"Daily summary AI error: {e}")
    
    # Fallback if AI fails
    if not summary_data:
        total_items = len(pending_tasks) + len(habits_pending) + len(done_tasks) + len(habits_done)
        done_items = len(done_tasks) + len(habits_done)
        score = round((done_items / max(total_items, 1)) * 100)
        
        summary_data = {
            "greeting": f"Bom dia, {user.name}! Seu rank atual é {user.rank}.",
            "progress_summary": f"Você já concluiu {len(done_tasks)} tarefas e {len(habits_done)} hábitos hoje. {'Estudou ' + str(study_minutes) + ' min. ' if study_minutes else ''}{'Treinou! ' if workout_logs else ''}{'Registrou ' + str(len(meals)) + ' refeições.' if meals else ''}",
            "pending_items": [t["title"] for t in pending_tasks[:5]] + [h["name"] for h in habits_pending[:5]],
            "motivation": "Disciplina é o que te move quando a motivação falta. Continue!",
            "priority_action": pending_tasks[0]["title"] if pending_tasks else (habits_pending[0]["name"] if habits_pending else "Dia limpo! Descanse ou avance no extra."),
            "score": score
        }
    
    result = {
        "user_id": user.user_id,
        "date": today_str,
        "summary": summary_data,
        "raw_data": {
            "tasks_pending": len(pending_tasks),
            "tasks_done": len(done_tasks),
            "habits_pending": len(habits_pending),
            "habits_done": len(habits_done),
            "study_minutes": study_minutes,
            "meals_count": len(meals),
            "calories": round(total_calories),
            "workouts_count": len(workout_logs)
        },
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Cache the summary
    await db.daily_summaries.update_one(
        {"user_id": user.user_id, "date": today_str},
        {"$set": result},
        upsert=True
    )
    
    return result



# ========== GLOBAL SEARCH ==========
@api_router.get("/search/global")
async def global_search(request: Request, q: str = "", session_token: Optional[str] = Cookie(None)):
    """Search across all user data"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if not q or len(q) < 2:
        return {"results": []}
    
    query_lower = q.lower()
    results = []
    
    # Search transactions
    transactions = await db.transactions.find(
        {"user_id": user.user_id, "$or": [
            {"description": {"$regex": q, "$options": "i"}},
            {"category": {"$regex": q, "$options": "i"}}
        ]},
        {"_id": 0}
    ).to_list(5)
    for t in transactions:
        results.append({
            "type": "transaction",
            "icon": "💰",
            "title": f"{t.get('description', '')} - R$ {t.get('amount', 0):.2f}",
            "subtitle": f"{t.get('category', '')} · {t.get('date', '')}",
            "link": "/finance"
        })
    
    # Search tasks
    tasks = await db.tasks.find(
        {"user_id": user.user_id, "title": {"$regex": q, "$options": "i"}},
        {"_id": 0}
    ).to_list(5)
    for t in tasks:
        results.append({
            "type": "task",
            "icon": "✅",
            "title": t.get('title', ''),
            "subtitle": f"{'Concluída' if t.get('completed') else 'Pendente'}",
            "link": "/tasks"
        })
    
    # Search habits
    habits = await db.habits.find(
        {"user_id": user.user_id, "name": {"$regex": q, "$options": "i"}},
        {"_id": 0}
    ).to_list(5)
    for h in habits:
        results.append({
            "type": "habit",
            "icon": "🔄",
            "title": h.get('name', ''),
            "subtitle": f"Streak: {h.get('streak', 0)} dias",
            "link": "/habits"
        })
    
    # Search workout plans
    plans = await db.workout_plans.find(
        {"user_id": user.user_id, "$or": [
            {"name": {"$regex": q, "$options": "i"}},
            {"description": {"$regex": q, "$options": "i"}}
        ]},
        {"_id": 0}
    ).to_list(5)
    for p in plans:
        results.append({
            "type": "workout_plan",
            "icon": "🏋️",
            "title": p.get('name', ''),
            "subtitle": f"{len(p.get('exercises', []))} exercícios",
            "link": "/workouts"
        })
    
    # Search notes
    notes = await db.study_notes.find(
        {"user_id": user.user_id, "$or": [
            {"title": {"$regex": q, "$options": "i"}},
            {"content": {"$regex": q, "$options": "i"}}
        ]},
        {"_id": 0}
    ).to_list(5)
    for n in notes:
        results.append({
            "type": "note",
            "icon": "📝",
            "title": n.get('title', ''),
            "subtitle": "Nota de estudo",
            "link": "/studies"
        })
    
    # Search goals
    goals = await db.goals.find(
        {"user_id": user.user_id, "title": {"$regex": q, "$options": "i"}},
        {"_id": 0}
    ).to_list(5)
    for g in goals:
        results.append({
            "type": "goal",
            "icon": "🎯",
            "title": g.get('title', ''),
            "subtitle": f"Progresso: {g.get('progress', 0)}%",
            "link": "/goals"
        })
    
    return {"results": results[:20]}


# ========== SMART REMINDERS ==========
@api_router.get("/reminders/smart")
async def get_smart_reminders(request: Request, session_token: Optional[str] = Cookie(None)):
    """Generate smart reminders based on user patterns"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    reminders = []
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    today_dt = datetime.now(timezone.utc)
    
    # Check workout frequency
    recent_workouts = await db.workout_logs.find(
        {"user_id": user.user_id, "date": {"$gte": (today_dt - timedelta(days=7)).strftime("%Y-%m-%d")}},
        {"_id": 0}
    ).to_list(50)
    
    if len(recent_workouts) == 0:
        reminders.append({
            "type": "workout",
            "icon": "🏋️",
            "message": "Você não treinou nos últimos 7 dias! Que tal retomar hoje?",
            "priority": "high",
            "action_link": "/workouts"
        })
    elif len(recent_workouts) < 3:
        reminders.append({
            "type": "workout",
            "icon": "💪",
            "message": f"Apenas {len(recent_workouts)} treino(s) esta semana. Tente manter ao menos 3x/semana!",
            "priority": "medium",
            "action_link": "/workouts"
        })
    
    # Check budget alerts
    month = today[:7]
    budgets = await db.budgets.find({"user_id": user.user_id, "month": month}, {"_id": 0}).to_list(20)
    transactions = await db.transactions.find(
        {"user_id": user.user_id, "date": {"$regex": f"^{month}"}, "type": "expense"},
        {"_id": 0}
    ).to_list(500)
    
    expense_by_cat = {}
    for t in transactions:
        cat = t.get('category', 'outros')
        expense_by_cat[cat] = expense_by_cat.get(cat, 0) + t.get('amount', 0)
    
    for b in budgets:
        cat = b.get('category', '')
        limit_val = b.get('limit', 0)
        spent = expense_by_cat.get(cat, 0)
        if limit_val > 0 and spent > 0:
            pct = spent / limit_val * 100
            if pct >= 90:
                reminders.append({
                    "type": "finance",
                    "icon": "⚠️",
                    "message": f"Orçamento de {cat}: {pct:.0f}% usado (R$ {spent:.2f} de R$ {limit_val:.2f})",
                    "priority": "high",
                    "action_link": "/finance"
                })
            elif pct >= 70:
                reminders.append({
                    "type": "finance",
                    "icon": "📊",
                    "message": f"Orçamento de {cat}: {pct:.0f}% usado. Fique atento!",
                    "priority": "medium",
                    "action_link": "/finance"
                })
    
    # Check habit streaks at risk
    habits = await db.habits.find({"user_id": user.user_id}, {"_id": 0}).to_list(50)
    for h in habits:
        completions = h.get('completions', [])
        if h.get('streak', 0) >= 3 and today not in completions:
            yesterday = (today_dt - timedelta(days=1)).strftime("%Y-%m-%d")
            if yesterday in completions:
                reminders.append({
                    "type": "habit",
                    "icon": "🔥",
                    "message": f"'{h.get('name', '')}': streak de {h.get('streak', 0)} dias em risco! Complete hoje.",
                    "priority": "high",
                    "action_link": "/habits"
                })
    
    # Check pending tasks
    pending_tasks = await db.tasks.find(
        {"user_id": user.user_id, "completed": False},
        {"_id": 0}
    ).to_list(100)
    
    overdue = [t for t in pending_tasks if t.get('due_date') and t['due_date'] < today]
    if overdue:
        reminders.append({
            "type": "task",
            "icon": "📋",
            "message": f"Você tem {len(overdue)} tarefa(s) atrasada(s)!",
            "priority": "high",
            "action_link": "/tasks"
        })
    
    # Sort by priority
    priority_order = {"high": 0, "medium": 1, "low": 2}
    reminders.sort(key=lambda r: priority_order.get(r.get('priority', 'low'), 2))
    
    return {"reminders": reminders}


# ========== SHOPPING LIST FROM RECIPES ==========
@api_router.post("/nutrition/shopping-list/generate")
async def generate_shopping_list(request: Request, session_token: Optional[str] = Cookie(None)):
    """Generate shopping list from meal plan or recipes"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    body = await request.json()
    plan_id = body.get("plan_id")
    recipe_ids = body.get("recipe_ids", [])
    
    items = {}
    
    if plan_id:
        plan = await db.meal_plans.find_one({"plan_id": plan_id, "user_id": user.user_id}, {"_id": 0})
        if plan and plan.get("shopping_list"):
            for item in plan["shopping_list"]:
                name = item.get("name", "").lower()
                if name in items:
                    items[name]["quantity"] += f" + {item.get('quantity', '')}"
                else:
                    items[name] = {
                        "name": item.get("name", ""),
                        "quantity": item.get("quantity", ""),
                        "category": item.get("category", "outros"),
                        "checked": False
                    }
    
    if recipe_ids:
        for rid in recipe_ids:
            recipe = await db.nutrition_recipes.find_one({"recipe_id": rid, "user_id": user.user_id}, {"_id": 0})
            if recipe:
                for ing in recipe.get("ingredients", []):
                    name = ing.lower() if isinstance(ing, str) else ing.get("name", "").lower()
                    if name not in items:
                        items[name] = {
                            "name": ing if isinstance(ing, str) else ing.get("name", ""),
                            "quantity": "" if isinstance(ing, str) else ing.get("quantity", ""),
                            "category": "ingredientes",
                            "checked": False
                        }
    
    shopping_list = list(items.values())
    
    list_id = f"shoplist_{uuid.uuid4().hex[:12]}"
    list_doc = {
        "list_id": list_id,
        "user_id": user.user_id,
        "items": shopping_list,
        "plan_id": plan_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.shopping_lists.insert_one(list_doc)
    list_doc.pop('_id', None)
    
    return {"success": True, "shopping_list": list_doc}


@api_router.get("/nutrition/shopping-lists")
async def get_shopping_lists(request: Request, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    lists = await db.shopping_lists.find({"user_id": user.user_id}, {"_id": 0}).sort("created_at", -1).to_list(10)
    return lists


@api_router.patch("/nutrition/shopping-lists/{list_id}/toggle/{item_idx}")
async def toggle_shopping_item(request: Request, list_id: str, item_idx: int, session_token: Optional[str] = Cookie(None)):
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    doc = await db.shopping_lists.find_one({"list_id": list_id, "user_id": user.user_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Lista não encontrada")
    
    items = doc.get("items", [])
    if 0 <= item_idx < len(items):
        items[item_idx]["checked"] = not items[item_idx].get("checked", False)
        await db.shopping_lists.update_one({"list_id": list_id}, {"$set": {"items": items}})
    
    return {"success": True, "items": items}


# ========== UNIFIED CALENDAR ==========
@api_router.get("/calendar/events")
async def get_calendar_events(request: Request, start: str = None, end: str = None, session_token: Optional[str] = Cookie(None)):
    """Get all events across modules for calendar view"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    if not start:
        today = datetime.now(timezone.utc)
        start = (today.replace(day=1)).strftime("%Y-%m-%d")
    if not end:
        today = datetime.now(timezone.utc)
        next_month = today.replace(day=28) + timedelta(days=4)
        end = next_month.replace(day=1).strftime("%Y-%m-%d")
    
    events = []
    
    # Tasks
    task_templates = await db.tasks.find({"user_id": user.user_id, "is_template": True}, {"_id": 0}).to_list(500)
    task_instances = await db.task_instances.find({
        "user_id": user.user_id,
        "date": {"$gte": start, "$lte": end}
    }, {"_id": 0}).to_list(5000)
    
    instance_map = {}
    for inst in task_instances:
        key = f"{inst['task_id']}_{inst['date']}"
        instance_map[key] = inst
    
    for t in task_templates:
        rec = t.get("recurrence", "once")
        if rec == "daily":
            d = datetime.strptime(start, "%Y-%m-%d")
            end_d = datetime.strptime(end, "%Y-%m-%d")
            while d <= end_d:
                ds = d.strftime("%Y-%m-%d")
                inst = instance_map.get(f"{t['task_id']}_{ds}")
                events.append({
                    "id": f"task_{t['task_id']}_{ds}",
                    "title": t["title"],
                    "date": ds,
                    "type": "task",
                    "color": "#007AFF",
                    "completed": inst.get("completed", False) if inst else False,
                    "priority": t.get("priority", "medium"),
                    "ref_id": t["task_id"]
                })
                d += timedelta(days=1)
        elif rec == "once":
            created = t.get("created_at", "")[:10]
            if start <= created <= end:
                inst = instance_map.get(f"{t['task_id']}_{created}")
                events.append({
                    "id": f"task_{t['task_id']}",
                    "title": t["title"],
                    "date": created,
                    "type": "task",
                    "color": "#007AFF",
                    "completed": inst.get("completed", False) if inst else False,
                    "priority": t.get("priority", "medium"),
                    "ref_id": t["task_id"]
                })
    
    # Habits
    habits = await db.habits.find({"user_id": user.user_id}, {"_id": 0}).to_list(500)
    for h in habits:
        for comp_date in h.get("completions", []):
            if start <= comp_date <= end:
                events.append({
                    "id": f"habit_{h['habit_id']}_{comp_date}",
                    "title": h["name"],
                    "date": comp_date,
                    "type": "habit",
                    "color": "#39FF14",
                    "completed": True,
                    "ref_id": h["habit_id"]
                })
    
    # Study sessions
    sessions = await db.study_sessions.find({
        "user_id": user.user_id,
        "date": {"$gte": start, "$lte": end}
    }, {"_id": 0}).to_list(5000)
    for s in sessions:
        events.append({
            "id": f"study_{s.get('session_id', '')}",
            "title": f"Estudo: {s.get('notebook_name', 'Sessão')}",
            "date": s["date"],
            "type": "study",
            "color": "#A855F7",
            "completed": True,
            "duration_minutes": s.get("duration_minutes", 0),
            "ref_id": s.get("session_id", "")
        })
    
    # Workouts
    workouts = await db.workout_logs.find({
        "user_id": user.user_id,
        "date": {"$gte": start, "$lte": end}
    }, {"_id": 0}).to_list(1000)
    for w in workouts:
        events.append({
            "id": f"workout_{w.get('log_id', '')}",
            "title": f"Treino: {w.get('plan_name', 'Sessão')}",
            "date": w["date"],
            "type": "workout",
            "color": "#EF4444",
            "completed": w.get("completed", False),
            "duration_minutes": w.get("duration_minutes", 0),
            "ref_id": w.get("log_id", "")
        })
    
    # Meals
    meals = await db.meals.find({
        "user_id": user.user_id,
        "date": {"$gte": start, "$lte": end}
    }, {"_id": 0}).to_list(5000)
    for m in meals:
        events.append({
            "id": f"meal_{m.get('meal_id', '')}",
            "title": f"{m.get('meal_type', 'Refeição').capitalize()}",
            "date": m["date"],
            "type": "meal",
            "color": "#22C55E",
            "completed": True,
            "calories": m.get("total_calories", 0),
            "ref_id": m.get("meal_id", "")
        })
    
    return {"events": events, "start": start, "end": end}



# ========== CROSS-MODULE SUGGESTIONS ==========
@api_router.get("/suggestions/cross-module")
async def get_cross_module_suggestions(request: Request, session_token: Optional[str] = Cookie(None)):
    """Get smart suggestions that connect different modules"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    suggestions = []
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    

# ===== EXPORT ENDPOINTS =====
from io import BytesIO
from fastapi.responses import StreamingResponse

@api_router.get("/export/finance/{format}")
async def export_finance(request: Request, format: str, session_token: Optional[str] = Cookie(None)):
    """Export financial data as PDF or Excel"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    transactions = await db.transactions.find({"user_id": user.user_id}, {"_id": 0}).sort("date", -1).to_list(5000)
    
    if format == "excel":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transações"
        
        # Header styling
        header_fill = PatternFill(start_color="007AFF", end_color="007AFF", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        
        headers = ["Data", "Descrição", "Categoria", "Tipo", "Valor (R$)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        
        total_income = 0
        total_expenses = 0
        for row, t in enumerate(transactions, 2):
            ws.cell(row=row, column=1, value=t.get("date", "")).border = thin_border
            ws.cell(row=row, column=2, value=t.get("description", "")).border = thin_border
            ws.cell(row=row, column=3, value=t.get("category", "")).border = thin_border
            tipo = "Receita" if t.get("type") == "income" else "Despesa"
            ws.cell(row=row, column=4, value=tipo).border = thin_border
            amount = t.get("amount", 0)
            ws.cell(row=row, column=5, value=amount).border = thin_border
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            if t.get("type") == "income":
                total_income += amount
            else:
                total_expenses += amount
        
        # Summary row
        summary_row = len(transactions) + 3
        ws.cell(row=summary_row, column=3, value="TOTAL RECEITAS:").font = Font(bold=True)
        ws.cell(row=summary_row, column=5, value=total_income).font = Font(bold=True, color="00AA00")
        ws.cell(row=summary_row, column=5).number_format = '#,##0.00'
        ws.cell(row=summary_row + 1, column=3, value="TOTAL DESPESAS:").font = Font(bold=True)
        ws.cell(row=summary_row + 1, column=5, value=total_expenses).font = Font(bold=True, color="FF0000")
        ws.cell(row=summary_row + 1, column=5).number_format = '#,##0.00'
        ws.cell(row=summary_row + 2, column=3, value="SALDO:").font = Font(bold=True)
        ws.cell(row=summary_row + 2, column=5, value=total_income - total_expenses).font = Font(bold=True)
        ws.cell(row=summary_row + 2, column=5).number_format = '#,##0.00'
        
        # Auto-fit columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
        
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=financas_sirius.xlsx"}
        )
    
    elif format == "pdf":
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        
        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm)
        styles = getSampleStyleSheet()
        elements = []
        
        title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=18, textColor=colors.HexColor('#007AFF'))
        elements.append(Paragraph("Relatório Financeiro - Sirius", title_style))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"Usuário: {user.name}", styles['Normal']))
        elements.append(Paragraph(f"Data: {datetime.now().strftime('%d/%m/%Y')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Table
        data = [["Data", "Descrição", "Categoria", "Tipo", "Valor"]]
        total_income = 0
        total_expenses = 0
        for t in transactions[:200]:
            tipo = "Receita" if t.get("type") == "income" else "Despesa"
            amount = t.get("amount", 0)
            data.append([
                t.get("date", ""),
                t.get("description", "")[:30],
                t.get("category", ""),
                tipo,
                f"R$ {amount:.2f}"
            ])
            if t.get("type") == "income":
                total_income += amount
            else:
                total_expenses += amount
        
        data.append(["", "", "", "RECEITAS:", f"R$ {total_income:.2f}"])
        data.append(["", "", "", "DESPESAS:", f"R$ {total_expenses:.2f}"])
        data.append(["", "", "", "SALDO:", f"R$ {total_income - total_expenses:.2f}"])
        
        table = Table(data, colWidths=[2.5*cm, 5*cm, 3*cm, 2.5*cm, 3*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007AFF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -4), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -4), [colors.white, colors.HexColor('#F5F5F5')]),
            ('FONTNAME', (3, -3), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(table)
        doc.build(elements)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=financas_sirius.pdf"}
        )
    
    raise HTTPException(status_code=400, detail="Formato deve ser 'excel' ou 'pdf'")


@api_router.get("/export/study/{format}")
async def export_study(request: Request, format: str, session_token: Optional[str] = Cookie(None)):
    """Export study data as PDF or Excel"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    notebooks = await db.notebooks.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    sessions = await db.study_sessions.find({"user_id": user.user_id}, {"_id": 0}).sort("date", -1).to_list(5000)
    flashcards = await db.flashcards.find({"user_id": user.user_id}, {"_id": 0}).to_list(5000)
    
    if format == "excel":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        
        # Sheet 1: Study Sessions
        ws1 = wb.active
        ws1.title = "Sessões de Estudo"
        header_fill = PatternFill(start_color="A855F7", end_color="A855F7", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        
        headers = ["Data", "Matéria", "Tipo", "Duração (min)", "Pomodoros"]
        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        for row, s in enumerate(sessions, 2):
            ws1.cell(row=row, column=1, value=s.get("date", "")).border = thin_border
            ws1.cell(row=row, column=2, value=s.get("notebook_name", "")).border = thin_border
            ws1.cell(row=row, column=3, value=s.get("session_type", "study")).border = thin_border
            ws1.cell(row=row, column=4, value=s.get("duration_minutes", 0)).border = thin_border
            ws1.cell(row=row, column=5, value=s.get("pomodoros", 0)).border = thin_border
        
        total_row = len(sessions) + 3
        ws1.cell(row=total_row, column=3, value="TOTAL:").font = Font(bold=True)
        ws1.cell(row=total_row, column=4, value=sum(s.get("duration_minutes", 0) for s in sessions)).font = Font(bold=True)
        
        # Sheet 2: Notebooks overview
        ws2 = wb.create_sheet("Matérias")
        headers2 = ["Matéria", "Área", "Flashcards", "Sessões"]
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        for row, nb in enumerate(notebooks, 2):
            nb_id = nb.get("notebook_id", "")
            nb_flashcards = len([f for f in flashcards if f.get("notebook_id") == nb_id])
            nb_sessions = len([s for s in sessions if s.get("notebook_id") == nb_id])
            ws2.cell(row=row, column=1, value=nb.get("name", "")).border = thin_border
            ws2.cell(row=row, column=2, value=nb.get("area_name", "")).border = thin_border
            ws2.cell(row=row, column=3, value=nb_flashcards).border = thin_border
            ws2.cell(row=row, column=4, value=nb_sessions).border = thin_border
        
        for ws in [ws1, ws2]:
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
        
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=estudos_sirius.xlsx"}
        )
    
    elif format == "pdf":
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        
        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm)
        styles = getSampleStyleSheet()
        elements = []
        
        title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=18, textColor=colors.HexColor('#A855F7'))
        elements.append(Paragraph("Relatório de Estudos - Sirius", title_style))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"Usuário: {user.name}", styles['Normal']))
        elements.append(Paragraph(f"Data: {datetime.now().strftime('%d/%m/%Y')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Notebooks summary
        elements.append(Paragraph("Resumo por Matéria", styles['Heading2']))
        nb_data = [["Matéria", "Área", "Flashcards", "Sessões"]]
        for nb in notebooks:
            nb_id = nb.get("notebook_id", "")
            nb_data.append([
                nb.get("name", "")[:25],
                nb.get("area_name", "")[:20],
                str(len([f for f in flashcards if f.get("notebook_id") == nb_id])),
                str(len([s for s in sessions if s.get("notebook_id") == nb_id]))
            ])
        
        table = Table(nb_data, colWidths=[5*cm, 4*cm, 3*cm, 3*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#A855F7')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))
        
        # Summary stats
        total_minutes = sum(s.get("duration_minutes", 0) for s in sessions)
        total_hours = total_minutes / 60
        elements.append(Paragraph(f"Total de horas estudadas: {total_hours:.1f}h ({total_minutes} min)", styles['Normal']))
        elements.append(Paragraph(f"Total de sessões: {len(sessions)}", styles['Normal']))
        elements.append(Paragraph(f"Total de flashcards: {len(flashcards)}", styles['Normal']))
        
        doc.build(elements)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=estudos_sirius.pdf"}
        )
    
    raise HTTPException(status_code=400, detail="Formato deve ser 'excel' ou 'pdf'")


@api_router.get("/export/nutrition/{format}")
async def export_nutrition(request: Request, format: str, session_token: Optional[str] = Cookie(None)):
    """Export nutrition data as PDF or Excel"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    meals = await db.meals.find({"user_id": user.user_id}, {"_id": 0}).sort("date", -1).to_list(5000)
    
    if format == "excel":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Refeições"
        header_fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        
        headers = ["Data", "Refeição", "Calorias", "Proteína(g)", "Carbos(g)", "Gordura(g)"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        for row, m in enumerate(meals, 2):
            ws.cell(row=row, column=1, value=m.get("date", "")).border = thin_border
            ws.cell(row=row, column=2, value=m.get("meal_type", "")).border = thin_border
            ws.cell(row=row, column=3, value=round(m.get("total_calories", 0), 1)).border = thin_border
            ws.cell(row=row, column=4, value=round(m.get("total_protein", 0), 1)).border = thin_border
            ws.cell(row=row, column=5, value=round(m.get("total_carbs", 0), 1)).border = thin_border
            ws.cell(row=row, column=6, value=round(m.get("total_fat", 0), 1)).border = thin_border
        
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 25)
        
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=nutricao_sirius.xlsx"}
        )
    
    elif format == "pdf":
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        
        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm)
        styles = getSampleStyleSheet()
        elements = []
        
        title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=18, textColor=colors.HexColor('#22C55E'))
        elements.append(Paragraph("Relatório Nutricional - Sirius", title_style))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"Usuário: {user.name}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        data = [["Data", "Refeição", "Calorias", "Proteína", "Carbos", "Gordura"]]
        for m in meals[:200]:
            data.append([
                m.get("date", ""),
                m.get("meal_type", ""),
                f"{m.get('total_calories', 0):.0f}",
                f"{m.get('total_protein', 0):.1f}g",
                f"{m.get('total_carbs', 0):.1f}g",
                f"{m.get('total_fat', 0):.1f}g",
            ])
        
        table = Table(data, colWidths=[2.5*cm, 3*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#22C55E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ]))
        elements.append(table)
        doc.build(elements)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=nutricao_sirius.pdf"}
        )
    
    raise HTTPException(status_code=400, detail="Formato deve ser 'excel' ou 'pdf'")


    # Check if user just completed a workout -> suggest meal
    recent_sessions = await db.workout_sessions.find(
        {"user_id": user.user_id, "status": "completed"},
        {"_id": 0}
    ).sort("completed_at", -1).to_list(1)
    
    if recent_sessions:
        last_session = recent_sessions[0]
        completed_at = last_session.get("completed_at", "")
        if completed_at and completed_at[:10] == today:
            suggestions.append({
                "type": "post_workout",
                "icon": "🍗",
                "title": "Refeição pós-treino",
                "message": "Você treinou hoje! Consuma proteínas e carboidratos nas próximas 2 horas para melhor recuperação.",
                "action": "Ir para Nutrição",
                "action_link": "/nutrition"
            })
    
    # Check if studying too long -> suggest break
    study_today = await db.study_sessions.find(
        {"user_id": user.user_id, "date": today},
        {"_id": 0}
    ).to_list(50)
    total_study_min = sum(s.get('duration_minutes', 0) for s in study_today)
    if total_study_min > 120:
        suggestions.append({
            "type": "study_break",
            "icon": "🧘",
            "title": "Hora de uma pausa",
            "message": f"Você já estudou {total_study_min} minutos hoje. Uma caminhada de 15 min melhora a concentração!",
            "action": "Ver treinos rápidos",
            "action_link": "/workouts"
        })
    
    # Check spending pattern -> suggest budget
    month = today[:7]
    month_expenses = await db.transactions.find(
        {"user_id": user.user_id, "date": {"$regex": f"^{month}"}, "type": "expense"},
        {"_id": 0}
    ).to_list(500)
    
    total_month_expense = sum(t.get('amount', 0) for t in month_expenses)
    day_of_month = int(today[8:10])
    if day_of_month > 0:
        daily_avg = total_month_expense / day_of_month
        projected_month = daily_avg * 30
        if projected_month > total_month_expense * 1.3 and day_of_month < 20:
            suggestions.append({
                "type": "finance_alert",
                "icon": "📊",
                "title": "Projeção de gastos",
                "message": f"Seus gastos projetados para o mês: R$ {projected_month:.2f}. Revise seus orçamentos!",
                "action": "Ver finanças",
                "action_link": "/finance"
            })
    
    return {"suggestions": suggestions}


# ========== TELEGRAM BOT INTEGRATION ==========
# import httpx
# import secrets
# from telegram import Bot, Update

# TELEGRAM_BOT_TOKEN = os.environ.get('TELEGRAM_BOT_TOKEN', '')
telegram_bot = None
TELEGRAM_BOT_TOKEN = ''
# if TELEGRAM_BOT_TOKEN:
#     try:
#         telegram_bot = Bot(token=TELEGRAM_BOT_TOKEN)
# except Exception as e:
#     logging.error(f"Failed to initialize Telegram bot: {e}")

@api_router.post("/telegram/setup-webhook")
async def setup_telegram_webhook(request: Request, session_token: Optional[str] = Cookie(None)):
    """Setup Telegram webhook - accepts backend_url in body or detects from request"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    if not telegram_bot:
        raise HTTPException(status_code=503, detail="Bot do Telegram não configurado")
    
    # Try to get backend URL from request body first
    backend_url = ''
    try:
        body = await request.json()
        backend_url = body.get("backend_url", "").rstrip("/")
    except:
        pass
    
    # Fallback: use BACKEND_PUBLIC_URL env var
    if not backend_url:
        backend_url = os.environ.get('BACKEND_PUBLIC_URL', '')
    
    # Fallback: try to detect from referer/origin (works when frontend proxies to backend)
    if not backend_url:
        origin = request.headers.get("origin", "")
        if not origin:
            referer = request.headers.get("referer", "")
            if referer:
                from urllib.parse import urlparse
                parsed = urlparse(referer)
                origin = f"{parsed.scheme}://{parsed.netloc}"
        backend_url = origin
    
    if not backend_url:
        raise HTTPException(status_code=400, detail="Não foi possível detectar a URL do backend. Configure BACKEND_PUBLIC_URL no servidor.")
    
    webhook_url = f"{backend_url}/api/telegram/webhook/{TELEGRAM_BOT_TOKEN}"
    
    try:
        async with httpx.AsyncClient() as hclient:
            resp = await hclient.post(
                f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/setWebhook",
                json={"url": webhook_url, "allowed_updates": ["message"]}
            )
            result = resp.json()
            if result.get("ok"):
                return {"success": True, "webhook_url": webhook_url}
            else:
                raise HTTPException(status_code=500, detail=f"Falha ao configurar webhook: {result.get('description')}")
    except httpx.HTTPError as e:
        raise HTTPException(status_code=500, detail=f"Erro de conexão: {str(e)}")

@api_router.post("/telegram/link")
async def link_telegram(request: Request, session_token: Optional[str] = Cookie(None)):
    """Generate a link code for the user to send to the Telegram bot"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    # Check if already linked
    existing = await db.telegram_links.find_one({"user_id": user.user_id, "status": "active"}, {"_id": 0})
    if existing:
        return {
            "already_linked": True,
            "chat_id": existing.get("chat_id"),
            "linked_at": existing.get("linked_at")
        }
    
    # Generate a 6-char code
    code = secrets.token_hex(3).upper()
    
    await db.telegram_link_codes.update_one(
        {"user_id": user.user_id},
        {"$set": {
            "user_id": user.user_id,
            "code": code,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=10)).isoformat()
        }},
        upsert=True
    )
    
    bot_info = None
    if telegram_bot:
        try:
            async with httpx.AsyncClient() as hclient:
                resp = await hclient.get(f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getMe")
                bot_data = resp.json()
                if bot_data.get("ok"):
                    bot_info = bot_data["result"]
        except:
            pass
    
    bot_username = bot_info.get("username", "") if bot_info else ""
    
    return {
        "code": code,
        "expires_in_minutes": 10,
        "bot_username": bot_username,
        "bot_link": f"https://t.me/{bot_username}" if bot_username else "",
        "instructions": f"Envie /vincular {code} para o bot @{bot_username} no Telegram"
    }

@api_router.post("/telegram/unlink")
async def unlink_telegram(request: Request, session_token: Optional[str] = Cookie(None)):
    """Unlink Telegram account"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    result = await db.telegram_links.update_one(
        {"user_id": user.user_id, "status": "active"},
        {"$set": {"status": "inactive", "unlinked_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"success": True, "was_linked": result.modified_count > 0}

@api_router.get("/telegram/status")
async def get_telegram_status(request: Request, session_token: Optional[str] = Cookie(None)):
    """Check Telegram link status"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    link = await db.telegram_links.find_one({"user_id": user.user_id, "status": "active"}, {"_id": 0})
    
    bot_username = ""
    if telegram_bot:
        try:
            async with httpx.AsyncClient() as hclient:
                resp = await hclient.get(f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getMe")
                bot_data = resp.json()
                if bot_data.get("ok"):
                    bot_username = bot_data["result"].get("username", "")
        except:
            pass
    
    return {
        "linked": link is not None,
        "chat_id": link.get("chat_id") if link else None,
        "linked_at": link.get("linked_at") if link else None,
        "telegram_name": link.get("telegram_name") if link else None,
        "bot_username": bot_username,
        "bot_configured": bool(TELEGRAM_BOT_TOKEN)
    }

async def handle_telegram_message(chat_id: int, text: str, telegram_name: str = ""):
    """Process incoming Telegram messages"""
    import random
    
    text = text.strip()
    
    # /start command
    if text.startswith("/start"):
        welcome = (
            "🌟 *Bem-vindo ao Sirius Bot!*\n\n"
            "Eu sou seu assistente pessoal do Sirius. Aqui você pode:\n\n"
            "📝 *Registrar transações:*\n"
            "• `Gastei 50 no mercado`\n"
            "• `Recebi 3000 de salário`\n\n"
            "📊 *Consultar dados:*\n"
            "• /resumo - Resumo do dia\n"
            "• /saldo - Saldo atual\n"
            "• /metas - Progresso das metas\n\n"
            "🔗 *Vincular conta:*\n"
            "• /vincular CODIGO - Vincule com sua conta Sirius\n\n"
            "❓ /ajuda - Ver todos os comandos"
        )
        await send_telegram_message(chat_id, welcome, parse_mode="Markdown")
        return
    
    # /ajuda command
    if text.startswith("/ajuda") or text.startswith("/help"):
        help_text = (
            "📋 *Comandos disponíveis:*\n\n"
            "🔗 `/vincular CODIGO` - Vincular conta Sirius\n"
            "💰 `/saldo` - Ver saldo atual\n"
            "📊 `/resumo` - Resumo financeiro do dia\n"
            "🎯 `/metas` - Progresso das metas\n"
            "📅 `/mes` - Resumo do mês\n"
            "💪 `/frase` - Frase motivacional\n"
            "❓ `/ajuda` - Esta mensagem\n\n"
            "💡 *Dica:* Envie mensagens naturais como:\n"
            "• `Gastei 150 no supermercado`\n"
            "• `Recebi 500 de freelance`\n"
            "• `Paguei 200 de luz e 100 de água`"
        )
        await send_telegram_message(chat_id, help_text, parse_mode="Markdown")
        return
    
    # /vincular command
    if text.startswith("/vincular"):
        parts = text.split()
        if len(parts) < 2:
            await send_telegram_message(chat_id, "⚠️ Use: /vincular CODIGO\n\nGere o código no app Sirius em Configurações > Telegram.")
            return
        
        code = parts[1].upper()
        link_code = await db.telegram_link_codes.find_one({"code": code}, {"_id": 0})
        
        if not link_code:
            await send_telegram_message(chat_id, "❌ Código inválido ou expirado. Gere um novo código no app.")
            return
        
        # Check expiration
        expires_at = datetime.fromisoformat(link_code["expires_at"].replace("Z", "+00:00"))
        if datetime.now(timezone.utc) > expires_at:
            await send_telegram_message(chat_id, "⏰ Código expirado. Gere um novo código no app.")
            await db.telegram_link_codes.delete_one({"code": code})
            return
        
        user_id = link_code["user_id"]
        
        # Deactivate any previous links
        await db.telegram_links.update_many(
            {"user_id": user_id, "status": "active"},
            {"$set": {"status": "inactive"}}
        )
        await db.telegram_links.update_many(
            {"chat_id": chat_id, "status": "active"},
            {"$set": {"status": "inactive"}}
        )
        
        # Create new link
        await db.telegram_links.insert_one({
            "user_id": user_id,
            "chat_id": chat_id,
            "telegram_name": telegram_name,
            "status": "active",
            "linked_at": datetime.now(timezone.utc).isoformat()
        })
        
        # Delete used code
        await db.telegram_link_codes.delete_one({"code": code})
        
        # Get user info
        user_doc = await db.users.find_one({"user_id": user_id}, {"_id": 0})
        name = user_doc.get("name", "Usuário") if user_doc else "Usuário"
        
        await send_telegram_message(
            chat_id,
            f"✅ *Conta vinculada com sucesso!*\n\nOlá, {name}! 🎉\nAgora você pode registrar transações e consultar dados diretamente aqui no Telegram."
            , parse_mode="Markdown"
        )
        return
    
    # For all other commands/messages, check if linked
    link = await db.telegram_links.find_one({"chat_id": chat_id, "status": "active"}, {"_id": 0})
    if not link:
        await send_telegram_message(
            chat_id,
            "🔒 Você precisa vincular sua conta primeiro!\n\n"
            "1️⃣ Abra o app Sirius\n"
            "2️⃣ Vá em Configurações > Telegram\n"
            "3️⃣ Clique em 'Vincular'\n"
            "4️⃣ Envie aqui: /vincular CODIGO"
        )
        return
    
    user_id = link["user_id"]
    user_doc = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not user_doc:
        await send_telegram_message(chat_id, "❌ Erro: usuário não encontrado.")
        return
    
    # /saldo command
    if text.startswith("/saldo"):
        today = datetime.now().strftime("%Y-%m-%d")
        month_start = datetime.now().strftime("%Y-%m-01")
        
        all_txns = await db.transactions.find({"user_id": user_id}).to_list(10000)
        total_income = sum(t.get("amount", 0) for t in all_txns if t.get("type") == "income")
        total_expense = sum(t.get("amount", 0) for t in all_txns if t.get("type") == "expense")
        balance = total_income - total_expense
        
        month_txns = [t for t in all_txns if t.get("date", "") >= month_start]
        month_income = sum(t.get("amount", 0) for t in month_txns if t.get("type") == "income")
        month_expense = sum(t.get("amount", 0) for t in month_txns if t.get("type") == "expense")
        
        emoji = "📈" if balance >= 0 else "📉"
        msg = (
            f"{emoji} *Seu Saldo*\n\n"
            f"💰 Saldo total: R$ {balance:,.2f}\n\n"
            f"📅 *Este mês:*\n"
            f"   ↗️ Receitas: R$ {month_income:,.2f}\n"
            f"   ↘️ Despesas: R$ {month_expense:,.2f}\n"
            f"   📊 Balanço: R$ {month_income - month_expense:,.2f}"
        )
        await send_telegram_message(chat_id, msg, parse_mode="Markdown")
        return
    
    # /resumo command
    if text.startswith("/resumo"):
        today = datetime.now().strftime("%Y-%m-%d")
        today_txns = await db.transactions.find({"user_id": user_id, "date": today}).to_list(100)
        
        if not today_txns:
            await send_telegram_message(chat_id, "📋 Nenhuma transação registrada hoje.\n\nEnvie algo como `Gastei 50 no almoço` para começar!")
            return
        
        incomes = [t for t in today_txns if t.get("type") == "income"]
        expenses = [t for t in today_txns if t.get("type") == "expense"]
        
        msg = f"📊 *Resumo de Hoje ({today})*\n\n"
        
        if incomes:
            total_in = sum(t.get("amount", 0) for t in incomes)
            msg += f"↗️ *Receitas:* R$ {total_in:,.2f}\n"
            for t in incomes:
                msg += f"   • {t.get('description', 'Sem descrição')} - R$ {t.get('amount', 0):,.2f}\n"
            msg += "\n"
        
        if expenses:
            total_exp = sum(t.get("amount", 0) for t in expenses)
            msg += f"↘️ *Despesas:* R$ {total_exp:,.2f}\n"
            for t in expenses:
                msg += f"   • {t.get('description', 'Sem descrição')} - R$ {t.get('amount', 0):,.2f}\n"
        
        await send_telegram_message(chat_id, msg, parse_mode="Markdown")
        return
    
    # /metas command
    if text.startswith("/metas"):
        goals = await db.goals.find({"user_id": user_id}).to_list(20)
        if not goals:
            await send_telegram_message(chat_id, "🎯 Nenhuma meta cadastrada.\nCrie metas no app Sirius!")
            return
        
        msg = "🎯 *Suas Metas*\n\n"
        for g in goals:
            progress = g.get("progress", 0)
            bar_filled = int(progress / 10)
            bar_empty = 10 - bar_filled
            bar = "█" * bar_filled + "░" * bar_empty
            msg += f"• {g.get('title', 'Meta')}\n  [{bar}] {progress}%\n\n"
        
        await send_telegram_message(chat_id, msg, parse_mode="Markdown")
        return
    
    # /mes command
    if text.startswith("/mes"):
        month_start = datetime.now().strftime("%Y-%m-01")
        month_txns = await db.transactions.find({
            "user_id": user_id,
            "date": {"$gte": month_start}
        }).to_list(1000)
        
        income = sum(t.get("amount", 0) for t in month_txns if t.get("type") == "income")
        expense = sum(t.get("amount", 0) for t in month_txns if t.get("type") == "expense")
        
        # Group expenses by category
        cat_totals = {}
        for t in month_txns:
            if t.get("type") == "expense":
                cat = t.get("category", "outros")
                cat_totals[cat] = cat_totals.get(cat, 0) + t.get("amount", 0)
        
        sorted_cats = sorted(cat_totals.items(), key=lambda x: x[1], reverse=True)
        
        msg = "📅 *Resumo do Mês*\n\n"
        msg += f"↗️ Receitas: R$ {income:,.2f}\n"
        msg += f"↘️ Despesas: R$ {expense:,.2f}\n"
        msg += f"📊 Balanço: R$ {income - expense:,.2f}\n\n"
        
        if sorted_cats:
            msg += "*Top despesas por categoria:*\n"
            for cat, total in sorted_cats[:5]:
                pct = (total / expense * 100) if expense > 0 else 0
                msg += f"   • {cat.title()}: R$ {total:,.2f} ({pct:.0f}%)\n"
        
        await send_telegram_message(chat_id, msg, parse_mode="Markdown")
        return
    
    # /frase command
    if text.startswith("/frase"):
        fallback_quotes = [
            "🔥 A dor do treino é temporária. A dor do arrependimento é permanente.",
            "⚔️ Guerreiros não nascem. São forjados no fogo da disciplina diária.",
            "🦁 Seja a pessoa que você precisava quando era mais novo.",
            "💎 Diamantes são apenas pedras que não desistiram sob pressão.",
            "🎯 Enquanto outros dormem, você constrói seu império.",
            "⚡ Sua única competição é quem você era ontem.",
            "🏆 Champions são feitos quando ninguém está olhando.",
            "🚀 Conforto é a morte lenta dos seus sonhos. Acorde!",
            "💪 Seu corpo pode quase tudo. É sua mente que você precisa convencer.",
            "🌟 A excelência não é um ato, é um hábito."
        ]
        try:
            quote_resp = await call_llm(
                "Gere UMA frase motivacional curta, impactante e única. Use 1-2 emojis. Máximo 2 linhas. Responda APENAS com a frase.",
                f"tg_motivation_{user_id}",
                "Você é um mestre motivacional."
            )
            await send_telegram_message(chat_id, quote_resp.strip())
        except:
            await send_telegram_message(chat_id, random.choice(fallback_quotes))
        return
    
    # Natural language transaction registration
    # Check if it looks like a transaction
    prompt = f"""Analise esta mensagem e determine se é um registro de transação financeira.
Mensagem: "{text}"

Se for uma ou mais transações, responda APENAS com um JSON array:
[{{"type": "income" ou "expense", "amount": valor_numerico, "description": "descrição curta", "category": "categoria"}}]

Categorias válidas: alimentação, transporte, moradia, saúde, educação, lazer, vestuário, investimentos, salário, freelance, outros

Se NÃO for uma transação, responda EXATAMENTE: NOT_TRANSACTION

Responda APENAS com o JSON array ou NOT_TRANSACTION, sem explicações."""

    try:
        ai_response = await call_llm(prompt, f"tg_parse_{user_id}", "Você é um parser de transações financeiras. Extraia dados com precisão.")
        ai_response = ai_response.strip()
        
        if "NOT_TRANSACTION" in ai_response:
            # General chat response
            chat_prompt = f"O usuário disse: '{text}'. Responda de forma breve e útil como assistente financeiro do Sirius. Máximo 3 linhas."
            chat_resp = await call_llm(chat_prompt, f"tg_chat_{user_id}", "Você é o assistente do Sirius, focado em finanças, produtividade e saúde.")
            await send_telegram_message(chat_id, chat_resp.strip())
            return
        
        # Parse JSON
        clean = ai_response.replace("```json", "").replace("```", "").strip()
        transactions = json.loads(clean)
        if not isinstance(transactions, list):
            transactions = [transactions]
        
        registered = []
        today = datetime.now().strftime("%Y-%m-%d")
        
        for txn in transactions:
            txn_id = f"txn_{uuid.uuid4().hex[:12]}"
            txn_doc = {
                "transaction_id": txn_id,
                "user_id": user_id,
                "type": txn.get("type", "expense"),
                "amount": float(txn.get("amount", 0)),
                "description": txn.get("description", "Sem descrição"),
                "category": txn.get("category", "outros"),
                "date": today,
                "source": "telegram",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.transactions.insert_one(txn_doc)
            registered.append(txn_doc)
        
        if len(registered) == 1:
            t = registered[0]
            emoji = "↗️" if t["type"] == "income" else "↘️"
            msg = f"✅ Registrado!\n\n{emoji} {t['description']}\n💰 R$ {t['amount']:,.2f}\n📂 {t['category'].title()}"
        else:
            total = sum(t["amount"] for t in registered)
            msg = f"✅ {len(registered)} transações registradas!\n\n"
            for i, t in enumerate(registered, 1):
                emoji = "↗️" if t["type"] == "income" else "↘️"
                msg += f"{i}. {emoji} {t['description']} - R$ {t['amount']:,.2f}\n"
            msg += f"\n💰 Total: R$ {total:,.2f}"
        
        await send_telegram_message(chat_id, msg)
    
    except json.JSONDecodeError:
        await send_telegram_message(chat_id, "🤖 Não consegui entender. Tente algo como:\n• `Gastei 50 no mercado`\n• `Recebi 3000 de salário`\n• /ajuda")
    except Exception as e:
        logging.error(f"Telegram message handling error: {e}")
        await send_telegram_message(chat_id, "⚠️ Ocorreu um erro. Tente novamente em instantes.")

async def send_telegram_message(chat_id: int, text: str, parse_mode: str = None):
    """Send a message via Telegram Bot API"""
    try:
        payload = {"chat_id": chat_id, "text": text}
        if parse_mode:
            payload["parse_mode"] = parse_mode
        
        async with httpx.AsyncClient() as hclient:
            resp = await hclient.post(
                f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage",
                json=payload,
                timeout=10
            )
            result = resp.json()
            if not result.get("ok"):
                # Retry without parse_mode if Markdown failed
                if parse_mode:
                    payload.pop("parse_mode")
                    await hclient.post(
                        f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage",
                        json=payload,
                        timeout=10
                    )
            return result
    except Exception as e:
        logging.error(f"Failed to send Telegram message: {e}")

async def send_telegram_daily_summary(user_id: str, chat_id: int):
    """Send daily summary to a linked Telegram user"""
    today = datetime.now().strftime("%Y-%m-%d")
    txns = await db.transactions.find({"user_id": user_id, "date": today}).to_list(100)
    
    if not txns:
        return
    
    incomes = sum(t.get("amount", 0) for t in txns if t.get("type") == "income")
    expenses = sum(t.get("amount", 0) for t in txns if t.get("type") == "expense")
    
    msg = f"🌙 *Resumo do dia ({today})*\n\n"
    msg += f"↗️ Receitas: R$ {incomes:,.2f}\n"
    msg += f"↘️ Despesas: R$ {expenses:,.2f}\n"
    msg += f"📊 Balanço do dia: R$ {incomes - expenses:,.2f}\n\n"
    msg += f"📝 {len(txns)} transações registradas"
    
    await send_telegram_message(chat_id, msg, parse_mode="Markdown")

# Telegram webhook endpoint (no auth required - Telegram calls this)
@app.post("/api/telegram/webhook/{token}")
async def telegram_webhook(token: str, request: Request):
    """Receive updates from Telegram"""
    if token != TELEGRAM_BOT_TOKEN:
        raise HTTPException(status_code=403, detail="Invalid token")
    
    try:
        update_data = await request.json()
        message = update_data.get("message", {})
        
        if not message:
            return {"ok": True}
        
        chat_id = message.get("chat", {}).get("id")
        text = message.get("text", "")
        first_name = message.get("from", {}).get("first_name", "")
        username = message.get("from", {}).get("username", "")
        telegram_name = first_name or username
        
        if chat_id and text:
            # Handle in background to respond quickly
            import asyncio
            asyncio.create_task(handle_telegram_message(chat_id, text, telegram_name))
        
        return {"ok": True}
    except Exception as e:
        logging.error(f"Telegram webhook error: {e}")
        return {"ok": True}  # Always return 200 to Telegram

@api_router.post("/telegram/send-daily-summaries")
async def trigger_daily_summaries(request: Request, session_token: Optional[str] = Cookie(None)):
    """Trigger daily summaries to all linked Telegram users"""
    auth_header = request.headers.get("Authorization")
    user = await get_current_user(authorization=auth_header, session_token=session_token)
    
    links = await db.telegram_links.find({"status": "active"}).to_list(1000)
    sent = 0
    for link in links:
        try:
            await send_telegram_daily_summary(link["user_id"], link["chat_id"])
            sent += 1
        except Exception as e:
            logging.error(f"Failed to send summary to {link.get('chat_id')}: {e}")
    
    return {"sent": sent, "total": len(links)}


# Include router AFTER all endpoints are defined
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_setup():
    """Auto-setup Telegram webhook on startup"""
    if TELEGRAM_BOT_TOKEN:
        try:
            # Priority: BACKEND_PUBLIC_URL > construct from REACT_APP_BACKEND_URL
            # On Railway, frontend and backend are separate services
            backend_url = os.environ.get('BACKEND_PUBLIC_URL', '')
            
            if not backend_url:
                # Try to detect from CORS_ORIGINS + common patterns
                cors_origins = os.environ.get('CORS_ORIGINS', '')
                for origin in cors_origins.split(','):
                    origin = origin.strip()
                    if origin and 'localhost' not in origin and '127.0.0.1' not in origin:
                        backend_url = origin
                        break
            
            if backend_url:
                webhook_url = f"{backend_url}/api/telegram/webhook/{TELEGRAM_BOT_TOKEN}"
                async with httpx.AsyncClient() as hclient:
                    resp = await hclient.post(
                        f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/setWebhook",
                        json={"url": webhook_url, "allowed_updates": ["message"]},
                        timeout=10
                    )
                    result = resp.json()
                    if result.get("ok"):
                        logging.info(f"Telegram webhook set to: {webhook_url}")
                    else:
                        logging.warning(f"Telegram webhook setup failed: {result}")
            else:
                logging.warning("No public URL found for Telegram webhook. Set BACKEND_PUBLIC_URL env var.")
        except Exception as e:
            logging.warning(f"Telegram webhook auto-setup failed: {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
